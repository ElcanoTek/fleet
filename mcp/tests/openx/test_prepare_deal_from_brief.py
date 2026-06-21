"""Tests for the high-level OpenX prepare/create/execute workflow."""

import json
import os
import sys
from pathlib import Path
from typing import Any

import httpx
import pytest
import respx
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from openx_mcp import (
    ox_create_prepared_deal,
    ox_execute_deal_from_prompt_inputs,
    ox_prepare_deal_from_brief,
    ox_prepare_deal_from_prompt_inputs,
    ox_validate_audience_geo_compatibility,
)

from .conftest import OPENX_GRAPHQL_ENDPOINT
from .fixtures import CREATE_DEAL_SUCCESS_RESPONSE, INTROSPECT_GEOGRAPHIC_ITEM_RESPONSE


class TestPrepareDealFromBrief:
    @pytest.mark.asyncio
    async def test_prepare_resolves_business_inputs(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            query = payload.get("query", "")
            variables = payload.get("variables", {})
            if "optionsByPath" not in query:
                raise AssertionError(f"Unexpected query during prepare: {query}")

            path = variables["path"]
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                assert variables["filter"] == {"demand_partner_id": "537073256"}
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "bidswitch-seats",
                                    "name": "Bidswitch Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393", "394"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            if path == "deal.package.targeting.technographic.language":
                return httpx.Response(
                    200,
                    json={"data": {"optionsByPath": [{"id": "es", "name": "Spanish", "path": path, "extra": {}}]}},
                )
            if path == "deal.package.targeting.domain.categories_iab_v2":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "18",
                                    "name": "Certified Pre-Owned Cars",
                                    "path": path,
                                    "extra": {"parent": "1"},
                                },
                                {"id": "32", "name": "Auto Parts", "path": path, "extra": {"parent": "1"}},
                                {"id": "30", "name": "Auto Buying and Selling", "path": path, "extra": {"parent": "1"}},
                                {"id": "25", "name": "Car Culture", "path": path, "extra": {"parent": "1"}},
                                {"id": "9", "name": "Coupe", "path": path, "extra": {"parent": "2"}},
                            ]
                        }
                    },
                )
            if path == "deal.package.targeting.audience.openaudience_custom":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "openaudience-123e4567-e89b-12d3-a456-426614174000",
                                    "name": "Cars & Auto_Chrysler Enthusiasts",
                                    "path": path,
                                    "extra": {},
                                }
                            ]
                        }
                    },
                )

            raise AssertionError(f"Unexpected options path: {path}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Elcano_OpenX_TRADR_TorqueDrive_Display_ELC07225_A0",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            end_date="2036-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            fee={"partner_name_or_id": "Elcano", "revenue_method": "PoM", "gross_share_percent": 40},
            domains_allowlist=["example.com", "EXAMPLE.com", "bad domain"],
            targeting={
                "channel": "DISPLAY",
                "device_types": ["DESKTOP", "MOBILE", "TABLET"],
                "geo": {"includes": {"country": "US"}},
                "language": "Spanish",
                "iab_categories": [
                    "Certified Pre-Owned Cars",
                    "Auto Parts",
                    "Auto Buying and Selling",
                    "Car Culture",
                    "Coupe",
                ],
                "audience": ["Cars & Auto_Chrysler Enthusiasts"],
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["blocking_issues"] == []
        assert result["resolved_entities"]["demand_partner"]["resolved_id"] == "537073256"
        assert result["resolved_entities"]["buyer_ids"] == ["393"]
        assert result["resolved_entities"]["fee_partner_id"] == "560610563"
        assert len(result["invalid_domains"]["allowlist"]) == 1

        create_args = result["create_args_preview"]
        assert create_args["deal_participants"] == [{"demand_partner": "537073256", "buyer_ids": ["393"]}]
        assert create_args["third_party_fees_config"]["gross_share"] == "0.4"
        assert create_args["third_party_fees_config"]["partner_id"] == "560610563"
        assert create_args["url_targeting"]["allowlist"] == ["example.com"]
        assert create_args["targeting"]["technographic"]["language"] == {"op": "INTERSECTS", "val": "es"}
        assert create_args["targeting"]["audience"]["openaudience_custom"] == {
            "op": "INTERSECTS",
            "val": "openaudience-123e4567-e89b-12d3-a456-426614174000",
        }
        assert create_args["targeting"]["domain"]["categories_iab_v2"] == {
            "op": "INTERSECTS",
            "val": "18,32,30,25,9",
        }
        assert result["resolved_entities"]["audience"] == {
            "op": "INTERSECTS",
            "val": "openaudience-123e4567-e89b-12d3-a456-426614174000",
        }

    @pytest.mark.asyncio
    async def test_prepare_sets_main_buyer_for_ttd_single_buyer(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            variables = payload.get("variables", {})
            path = variables["path"]
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "537073292", "name": "The Trade Desk - RTB", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                assert variables["filter"] == {"demand_partner_id": "537073292"}
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "ttd-seats",
                                    "name": "TTD Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["109"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            raise AssertionError(f"Unexpected options path: {path}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Reklaim_OpenX_TTD_Test_Display",
            currency="USD",
            deal_price=0.01,
            start_date="2026-05-04T00:00:00Z",
            package_name="TTD_Display_Targeting",
            demand_partner="The Trade Desk - RTB",
            buyer_ids=["109"],
            targeting={"channel": "DISPLAY", "geo": {"includes": {"country": "CA"}}},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["deal_participants"] == [
            {"demand_partner": "537073292", "buyer_ids": ["109"], "main_buyer_id": "109"}
        ]

    @pytest.mark.asyncio
    async def test_prepare_returns_structured_blockers(self, mock_openx_graphql: respx.MockRouter):
        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(
            return_value=httpx.Response(200, json={"data": {"optionsByPath": []}})
        )

        result = await ox_prepare_deal_from_brief(
            name="Blocked_Deal",
            currency="USD",
            deal_price=1.0,
            start_date="2026-03-30T00:00:00Z",
            package_name="Blocked_Package",
            demand_partner="Unknown Partner",
            targeting={"language": "Spanish"},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is False
        assert any(blocker["code"] == "demand_partner_unresolved" for blocker in result["blockers"])
        assert any(blocker["code"] == "unresolved_language" for blocker in result["blockers"])

    @pytest.mark.asyncio
    async def test_prepare_ignores_non_unique_targeting_override(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            raise AssertionError(f"Unexpected options path: {path}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Prepared_Unique_Targeting_Deal",
            currency="USD",
            deal_price=1.0,
            start_date="2026-03-30T00:00:00Z",
            package_name="Prepared_Unique_Targeting_Package",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": ["US"], "inter_dimension_operator": "OR"},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["targeting"].get("inter_dimension_operator") is None

    @pytest.mark.asyncio
    async def test_prepare_from_prompt_inputs_reads_attached_xlsx(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.append(["Sites"])
        worksheet.append(["https://example.com/path"])
        worksheet.append(["www.duplicate.com"])
        worksheet.append(["duplicate.com"])
        worksheet.append(["bad domain"])
        file_path = tmp_path / "automotivelist.xlsx"
        workbook.save(file_path)

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            query = payload.get("query", "")
            variables = payload.get("variables", {})
            if "optionsByPath" not in query:
                raise AssertionError(f"Unexpected query: {query}")
            path = variables["path"]
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                assert variables["filter"] == {"demand_partner_id": "537073256"}
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "bidswitch-seats",
                                    "name": "Bidswitch Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_prompt_inputs(
            name="Prompt_Input_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="537073256",
            buyer_ids=["393"],
            fee={"partner_name_or_id": "560610563", "revenue_method": "PoM", "gross_share_percent": 40},
            targeting={"channel": "DISPLAY", "device_types": ["DESKTOP"], "geo": {"includes": {"country": "US"}}},
            domain_file_path=str(file_path),
            domain_match_operator="blocklist",
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["extracted_domain_count"] == 2
        assert result["domain_source"]["sheet_name"] == "Sheet1"
        assert result["domain_source"]["column_name"] == "Sites"
        assert len(result["invalid_domain_rows"]) == 1
        assert result["create_args_preview"]["url_targeting"]["blocklist"] == ["duplicate.com", "example.com"]
        assert "allowlist" not in result["create_args_preview"]["url_targeting"]

    @pytest.mark.asyncio
    async def test_prepare_from_prompt_inputs_reads_attached_csv(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        file_path = tmp_path / "automotivelist.csv"
        file_path.write_text("domain\nexample.com\nwww.duplicate.com\nduplicate.com\nbad domain\n", encoding="utf-8")

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_prompt_inputs(
            name="CSV_Input_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="537073256",
            buyer_ids=["393"],
            targeting={"channel": "DISPLAY", "device_types": ["DESKTOP"], "geo": {"includes": {"country": "US"}}},
            domain_file_path=str(file_path),
            domain_column="domain",
            domain_match_operator="blocklist",
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["extracted_domain_count"] == 2
        assert result["domain_source"]["sheet_name"] is None
        assert result["domain_source"]["column_name"] == "domain"
        assert result["create_args_preview"]["url_targeting"]["blocklist"] == ["duplicate.com", "example.com"]
        assert "allowlist" not in result["create_args_preview"]["url_targeting"]

    @pytest.mark.asyncio
    async def test_prepare_from_prompt_inputs_allowlist_routes_to_whitelist(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        """An explicit ``domain_match_operator='allowlist'`` routes the attached file to
        ``url_targeting.allowlist`` (which the create layer maps to ``type='whitelist'`` on the
        wire). Allowlists are a first-class brief type — no warning, no privileged default."""

        file_path = tmp_path / "allowlist.csv"
        file_path.write_text("domain\nexample.com\nduplicate.com\n", encoding="utf-8")

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_prompt_inputs(
            name="Allowlist_Opt_In_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Allowlist_Opt_In_Package",
            demand_partner="537073256",
            buyer_ids=["393"],
            targeting={"channel": "DISPLAY", "device_types": ["DESKTOP"], "geo": {"includes": {"country": "US"}}},
            domain_file_path=str(file_path),
            domain_column="domain",
            domain_match_operator="allowlist",
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["url_targeting"]["allowlist"] == ["duplicate.com", "example.com"]
        assert "blocklist" not in result["create_args_preview"]["url_targeting"]
        # Allowlist is a first-class option; no "double-check" warning should be emitted.
        assert not any("ALLOWLIST" in warning for warning in result["warnings"])

    @pytest.mark.asyncio
    async def test_prepare_from_prompt_inputs_app_bundle_file_routes_to_app_inventory(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        """app_bundle_file_path targets the DISTINCT OpenX app-inventory dimension
        (targeting.app_inventory.app_bundle_id), NOT web-domain url_targeting. Bundle
        lists carry reverse-DNS bundles AND bare numeric store IDs — both survive
        verbatim and in first-seen order (a url_targeting validator would mangle them),
        and url_targeting stays empty when no web domains are supplied."""

        file_path = tmp_path / "ctv_bundles.csv"
        file_path.write_text("Bundle ID\ncom.zumobi.msnbc\n523428113\n", encoding="utf-8")

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_prompt_inputs(
            name="CTV_Bundle_Allow_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="CTV_Bundle_Allow_Package",
            demand_partner="537073256",
            buyer_ids=["393"],
            targeting={"channel": "CTV", "device_types": ["CTV"], "geo": {"includes": {"country": "US"}}},
            app_bundle_file_path=str(file_path),
            app_bundle_column="Bundle ID",
            app_bundle_match_operator="allowlist",
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["extracted_app_bundle_count"] == 2
        assert result["app_bundle_source"]["column_name"] == "Bundle ID"
        # Bundles land in app_inventory.app_bundle_id (OR-list, "==" per entry), preserving
        # first-seen order — NOT in url_targeting, which stays empty with no web domains.
        assert result["create_args_preview"]["url_targeting"] is None
        app_inventory = result["create_args_preview"]["targeting"]["app_inventory"]
        assert app_inventory["app_bundle_id"] == {
            "op": "OR",
            "val": [
                {"op": "==", "val": "com.zumobi.msnbc"},
                {"op": "==", "val": "523428113"},
            ],
        }
        assert app_inventory["app_inventory_inter_dimension_operator"] == "AND"

    @pytest.mark.asyncio
    async def test_prepare_from_prompt_inputs_allows_both_domain_and_app_bundle(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        """A single deal may carry BOTH a web-domain allowlist (url_targeting) and an
        app-bundle list (app_inventory) — they are independent OpenX dimensions, so this
        is no longer a conflict. Verified against a real Reklaim GM Display deal."""

        domain_file = tmp_path / "domains.csv"
        domain_file.write_text("domain\nexample.com\n", encoding="utf-8")
        bundle_file = tmp_path / "bundles.csv"
        bundle_file.write_text("Bundle ID\ncom.zumobi.msnbc\nB072QYQ43R\n", encoding="utf-8")

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_prompt_inputs(
            name="Domains_And_Bundles_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            demand_partner="537073256",
            buyer_ids=["393"],
            targeting={"channel": "DISPLAY", "geo": {"includes": {"country": "US"}}},
            domain_file_path=str(domain_file),
            domain_match_operator="allowlist",
            app_bundle_file_path=str(bundle_file),
            app_bundle_match_operator="allowlist",
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        # Web domain → url_targeting allowlist.
        assert result["create_args_preview"]["url_targeting"]["allowlist"] == ["example.com"]
        assert result["extracted_domain_count"] == 1
        # Bundles → app_inventory (the ASIN B072QYQ43R survives here; url_targeting drops it).
        assert result["extracted_app_bundle_count"] == 2
        app_bundle_vals = result["create_args_preview"]["targeting"]["app_inventory"]["app_bundle_id"]["val"]
        assert app_bundle_vals == [
            {"op": "==", "val": "com.zumobi.msnbc"},
            {"op": "==", "val": "B072QYQ43R"},
        ]

    @pytest.mark.asyncio
    async def test_prepare_from_prompt_inputs_rejects_app_bundle_blocklist(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        """OpenX app_inventory.app_bundle_id is an include (allowlist) set; a blocklist
        request returns a structured blocker rather than silently being treated as include."""

        bundle_file = tmp_path / "bundles.csv"
        bundle_file.write_text("Bundle ID\ncom.zumobi.msnbc\n", encoding="utf-8")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(
            return_value=httpx.Response(200, json={"data": {"optionsByPath": []}})
        )

        result = await ox_prepare_deal_from_prompt_inputs(
            name="Bundle_Blocklist_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            demand_partner="537073256",
            app_bundle_file_path=str(bundle_file),
            app_bundle_match_operator="blocklist",
        )

        assert result["ready_to_create"] is False
        assert any(b.get("code") == "app_bundle_blocklist_unsupported" for b in result["blockers"])

    @pytest.mark.asyncio
    async def test_prepare_app_bundle_xlsx_preserves_numeric_ids(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        """A 13-digit Apple App Store ID entered as a NUMBER in xlsx must extract as the
        exact integer string — not '3.2023e+12' or a trailing '.0'. That scientific-notation
        corruption is exactly what hand-pasted UI bundle lists suffer; reading via the MCP
        must be cleaner."""

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Bundles"
        worksheet.append(["App Bundle IDs"])
        worksheet.append([3202310000000])  # large integer Apple id
        worksheet.append(["com.fubotv.vix"])  # reverse-DNS string
        worksheet.append([162057])  # smaller numeric id
        worksheet.append(["B072QYQ43R"])  # Amazon ASIN
        file_path = tmp_path / "bundles.xlsx"
        workbook.save(file_path)

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_prompt_inputs(
            name="Numeric_Bundle_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            demand_partner="537073256",
            targeting={"channel": "DISPLAY", "geo": {"includes": {"country": "US"}}},
            app_bundle_file_path=str(file_path),
            app_bundle_sheet="Bundles",
            app_bundle_column="App Bundle IDs",
            app_bundle_match_operator="allowlist",
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        bundle_values = [
            entry["val"]
            for entry in result["create_args_preview"]["targeting"]["app_inventory"]["app_bundle_id"]["val"]
        ]
        assert bundle_values == ["3202310000000", "com.fubotv.vix", "162057", "B072QYQ43R"]

    @pytest.mark.asyncio
    async def test_prepare_from_brief_accepts_app_bundles_list(self, mock_openx_graphql: respx.MockRouter):
        """ox_prepare_deal_from_brief exposes a typed app_bundles list (no file needed) that
        builds the same app_inventory.app_bundle_id wire shape."""

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Brief_Bundles_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Brief_Bundles_Package",
            demand_partner="537073256",
            targeting={"channel": "DISPLAY", "geo": {"includes": {"country": "US"}}},
            app_bundles=["com.fubotv.vix", "com.fubotv.vix", "162057"],
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        # De-duplicated, first-seen order preserved.
        assert result["create_args_preview"]["targeting"]["app_inventory"]["app_bundle_id"]["val"] == [
            {"op": "==", "val": "com.fubotv.vix"},
            {"op": "==", "val": "162057"},
        ]

    @pytest.mark.asyncio
    async def test_prepare_from_prompt_inputs_blocks_when_file_supplied_without_operator(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        """When a domain file is supplied but ``domain_match_operator`` is omitted, the prepare
        path MUST hard-fail with a structured blocker. There is no implicit default — silently
        routing a blocklist file as an allowlist (or vice versa) is the exact bug this guard
        prevents."""

        file_path = tmp_path / "ambiguous.csv"
        file_path.write_text("domain\nexample.com\n", encoding="utf-8")

        def handler(request: httpx.Request) -> httpx.Response:
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_prompt_inputs(
            name="Missing_Operator_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Missing_Operator_Package",
            demand_partner="537073256",
            buyer_ids=["393"],
            targeting={"channel": "DISPLAY", "geo": ["US"]},
            domain_file_path=str(file_path),
            domain_column="domain",
        )

        assert result["success"] is True
        assert result["ready_to_create"] is False
        assert any(b["code"] == "missing_domain_match_operator" for b in result["blockers"])
        assert "domain_match_operator" in result["blocking_issues"][0]
        # No deal should be prepared when the operator is ambiguous.
        assert "create_args_preview" not in result

    @pytest.mark.asyncio
    async def test_prepare_emits_quality_flag_when_significant_share_of_rows_drop(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        """When the attached domain file is mostly non-domain values (e.g. the Reklaim/Optimum
        political-compliance mixed-inventory file where ~36% of rows are app bundle IDs), the
        trader must see this as a structured quality_flag, not just buried in a warning."""

        # 2 valid web domains + 8 obviously-bad entries (whitespace / illegal chars) → 80%
        # dropped, well over the 10% threshold. Real-world equivalent: app bundle IDs with
        # underscores or numeric Apple IDs that fail OpenX domain validation.
        file_path = tmp_path / "mixed_inventory.csv"
        file_path.write_text(
            "domain\nexample.com\nrealdomain.com\n"
            "not a domain 1\nnot a domain 2\nnot a domain 3\nnot a domain 4\n"
            "not a domain 5\nnot a domain 6\nnot a domain 7\nnot a domain 8\n",
            encoding="utf-8",
        )

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_prompt_inputs(
            name="Mixed_Inventory_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Mixed_Inventory_Package",
            demand_partner="537073256",
            buyer_ids=["393"],
            targeting={"channel": "DISPLAY", "geo": ["US"]},
            domain_file_path=str(file_path),
            domain_column="domain",
            domain_match_operator="allowlist",
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        # The structured flag must appear AND carry the dropped/valid/percent context the
        # trader needs to act on (without re-counting the warning string by hand).
        partial_flags = [f for f in result["quality_flags"] if f["flag"] == "ox_domain_extraction_partial"]
        assert len(partial_flags) == 1
        flag = partial_flags[0]
        assert flag["dropped_count"] == 8
        assert flag["valid_count"] == 2
        assert flag["dropped_percent"] == 80.0
        # The free-text warning still fires (back-compat for callers that read warnings only).
        assert any("Dropped 8 non-domain values" in w for w in result["warnings"])

    @pytest.mark.asyncio
    async def test_prepare_does_not_emit_partial_flag_for_small_drop(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        """A handful of typos in an otherwise-clean domain file should NOT trigger the
        partial-extraction quality_flag — only significant losses (>10%) warrant the alert."""

        # 9 valid web domains + 1 bogus row → ~10% dropped, exactly at the threshold (NOT over).
        file_path = tmp_path / "mostly_clean.csv"
        file_path.write_text(
            "domain\n" + "\n".join(f"example{i}.com" for i in range(9)) + "\nbad domain\n",
            encoding="utf-8",
        )

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_prompt_inputs(
            name="Mostly_Clean_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Mostly_Clean_Package",
            demand_partner="537073256",
            buyer_ids=["393"],
            targeting={"channel": "DISPLAY", "geo": ["US"]},
            domain_file_path=str(file_path),
            domain_column="domain",
            domain_match_operator="blocklist",
        )

        assert result["success"] is True
        assert not any(f["flag"] == "ox_domain_extraction_partial" for f in result["quality_flags"])
        # Free-text warning still fires for the one dropped row (back-compat).
        assert any("Dropped 1 non-domain values" in w for w in result["warnings"])

    @pytest.mark.asyncio
    async def test_prepare_from_prompt_inputs_rejects_invalid_match_operator(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        """An unknown ``domain_match_operator`` value must surface a clear failure rather
        than silently falling through to a default."""

        file_path = tmp_path / "list.csv"
        file_path.write_text("domain\nexample.com\n", encoding="utf-8")

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_prompt_inputs(
            name="Bad_Operator_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Bad_Operator_Package",
            demand_partner="537073256",
            buyer_ids=["393"],
            targeting={"channel": "DISPLAY", "geo": ["US"]},
            domain_file_path=str(file_path),
            domain_column="domain",
            domain_match_operator="exclude",  # type: ignore[arg-type]
        )

        assert result["success"] is False
        assert "domain_match_operator" in result["error"]

    @pytest.mark.asyncio
    async def test_prepare_from_prompt_inputs_reads_attached_publisher_file_for_ctv(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Publishers"
        worksheet.append(["Publisher ID"])
        worksheet.append([123])
        worksheet.append(["456"])
        worksheet.append([123])
        file_path = tmp_path / "tricoast_publishers.xlsx"
        workbook.save(file_path)

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            raise AssertionError(f"Unexpected options path: {path}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_prompt_inputs(
            name="CTV_Publisher_File_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_CTV_Targeting",
            demand_partner="TRADR / 393",
            targeting={"channel": "CTV", "device_types": ["CTV"], "geo": ["US"]},
            publisher_file_path=str(file_path),
            publisher_sheet="Publishers",
            publisher_column="Publisher ID",
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["extracted_domain_count"] == 0
        assert result["domain_source"] is None
        assert result["extracted_publisher_count"] == 2
        assert result["publisher_source"] == {
            "file_path": str(file_path),
            "sheet_name": "Publishers",
            "column_name": "Publisher ID",
            "row_count": 3,
        }
        assert result["create_args_preview"]["url_targeting"] is None
        assert result["create_args_preview"]["targeting"]["content"]["account"] == {
            "op": "INTERSECTS",
            "val": "123,456",
        }

    @pytest.mark.asyncio
    async def test_prepare_accepts_fee_partner_alias(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Fee_Alias_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR",
            buyer_ids=["393"],
            fee={"partner": "Elcano", "revenue_method": "PoM", "gross_share": "40.0"},
            targeting={"channel": "DISPLAY", "device_types": ["DESKTOP"], "geo": ["US"]},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["resolved_entities"]["fee_partner_id"] == "560610563"
        assert result["create_args_preview"]["third_party_fees_config"]["gross_share"] == "0.4"

    @pytest.mark.asyncio
    async def test_prepare_preserves_structured_targeting(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Structured_Targeting_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "device_types": ["DESKTOP", "MOBILE", "TABLET"],
                "technographic": {"language": {"op": "INTERSECTS", "val": "es"}},
                "geographic": {"includes": {"country": "US"}},
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        targeting = result["create_args_preview"]["targeting"]
        assert "rendering_context" not in targeting  # rendering_context is built by ox_create_deal, not prep
        assert targeting["channel"] == "DISPLAY"
        assert targeting["device_type"] == ["DESKTOP", "MOBILE", "TABLET"]
        assert targeting["technographic"] == {"language": {"op": "INTERSECTS", "val": "es"}}

    @pytest.mark.asyncio
    async def test_prepare_rejects_pre_resolved_state_ids(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Structured_State_Geo_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "geographic": {"includes": {"state": "3569,3674,3583,3567,3632,3635", "country": None}},
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is False
        assert any(blocker["code"] == "unvalidated_state_ids" for blocker in result["blockers"])
        assert any("Rejected pre-resolved geographic state ids" in warning for warning in result["warnings"])

    @pytest.mark.asyncio
    async def test_prepare_resolves_state_abbreviations(self, openx_api_key: str, mock_openx_graphql: respx.MockRouter):  # noqa: ARG002
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            raise AssertionError(f"Unexpected options path: {path}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="State_Targeting_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": ["HI"]},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["targeting"]["geographic"] == {
            "includes": {"state": "3595", "country": "us"}
        }

    @pytest.mark.asyncio
    async def test_prepare_routes_non_us_two_letter_country_codes_to_country(
        self, mock_openx_graphql: respx.MockRouter
    ):
        """Two-letter tokens that are valid country codes (UK, DE) but not US state
        abbreviations MUST route through the country-list wrap, not the state-resolution
        path. Before this fix, the is_states heuristic matched any 2-letter token,
        silently sending non-US codes into the US-state resolver where they failed."""

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="UK_Country_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="UK_Country_Package",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": ["UK", "DE"]},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        # MUST be wrapped as country list — NOT routed through state resolver.
        geographic = result["create_args_preview"]["targeting"]["geographic"]
        assert geographic == {"includes": {"country": "uk,de"}}
        # Defensive: ensure no state-id resolution happened.
        assert "state" not in geographic["includes"]

    @pytest.mark.asyncio
    async def test_prepare_routes_canadian_province_abbreviations_to_country_list(
        self, mock_openx_graphql: respx.MockRouter
    ):
        """Canadian province abbreviations like \"AB\" (Alberta) and \"ON\" (Ontario) are
        2-letter tokens that previously got matched by the loose is_states check and
        silently produced unresolved_state blockers (the US resolver doesn't know them).
        After the heuristic is tightened, they fall through to the country-list wrap
        instead — which is the correct behavior for a bare list of non-US tokens."""

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Canada_Province_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Canada_Province_Package",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": ["AB", "ON"]},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        # AB/ON treated as country-list tokens, not US state abbreviations.
        geographic = result["create_args_preview"]["targeting"]["geographic"]
        assert geographic == {"includes": {"country": "ab,on"}}
        assert "state" not in geographic["includes"]

    @pytest.mark.asyncio
    async def test_prepare_routes_mixed_us_state_and_non_state_to_country_list(
        self, mock_openx_graphql: respx.MockRouter
    ):
        """If even one token in the list is NOT a US state, the whole list should fall
        through to the country wrap rather than partial-resolving the US ones — mixing
        US states and non-US codes in a single geo list is itself ambiguous and the
        caller should split into separate deals (or use the structured includes shape)."""

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        # "CA" is ambiguous (California / Canada) — but with "UK" as a clear non-US token
        # alongside it, the heuristic MUST treat the whole list as countries.
        result = await ox_prepare_deal_from_brief(
            name="Mixed_Geo_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Mixed_Geo_Package",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": ["CA", "UK"]},
        )

        assert result["success"] is True
        geographic = result["create_args_preview"]["targeting"]["geographic"]
        assert geographic == {"includes": {"country": "ca,uk"}}
        assert "state" not in geographic["includes"]

    @pytest.mark.asyncio
    async def test_prepare_resolves_lowercase_state_abbreviations(self, mock_openx_graphql: respx.MockRouter):
        """The state-abbreviation match is case-insensitive — \"ca\" (lowercase) must
        still resolve as California, not get diverted to the country wrap. Pin this so
        a future refactor doesn't accidentally reintroduce case-sensitivity."""

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Lowercase_State_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Lowercase_State_Package",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": ["ca"]},
        )

        assert result["success"] is True
        # "ca" → California, resolved through the state path.
        geographic = result["create_args_preview"]["targeting"]["geographic"]
        assert "state" in geographic["includes"]
        assert geographic["includes"]["country"] == "us"

    @pytest.mark.asyncio
    @pytest.mark.usefixtures("openx_api_key")
    async def test_prepare_resolves_states_via_options_by_path(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.package.targeting.geographic.state":
                state_filter = payload.get("variables", {}).get("filter")
                if state_filter == {"state": "california*"}:
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {
                                        "id": "3588",
                                        "name": "california",
                                        "path": path,
                                        "extra": {
                                            "state": "california",
                                            "country": "united states",
                                            "type": "state",
                                            "type_id": "state-3588",
                                        },
                                    }
                                ]
                            }
                        },
                    )
                if state_filter == {"state": "texas*"}:
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {
                                        "id": "3625",
                                        "name": "texas",
                                        "path": path,
                                        "extra": {
                                            "state": "texas",
                                            "country": "united states",
                                            "type": "state",
                                            "type_id": "state-3625",
                                        },
                                    }
                                ]
                            }
                        },
                    )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            raise AssertionError(
                f"Unexpected options path/filter: {path} / {payload.get('variables', {}).get('filter')}"
            )

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="OptionsByPath_State_Targeting_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": ["CA", "TX"]},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["targeting"]["geographic"] == {
            "includes": {"state": "3588,3625", "country": "us"}
        }
        assert any("Resolved state 'CA' to verified OpenX id '3588'." in warning for warning in result["warnings"])

    @pytest.mark.asyncio
    async def test_prepare_blocks_legacy_iab_codes(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Legacy_IAB_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": ["US"], "iab_categories": ["IAB2-5", "IAB2-6"]},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is False
        assert any(blocker["code"] == "legacy_iab_codes_not_supported" for blocker in result["blockers"])
        assert any("IAB2-5" in blocker.get("details", {}).get("legacy_codes", []) for blocker in result["blockers"])
        assert any(
            blocker.get("details", {}).get("suggested_action")
            == "Use ox_list_iab_categories to discover correct OpenX names."
            for blocker in result["blockers"]
        )
        assert any(blocker.get("details", {}).get("query_hint") == "automotive" for blocker in result["blockers"])

    @pytest.mark.asyncio
    async def test_prepare_allows_audience_with_explicit_canada_country(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.package.targeting.audience.openaudience_custom":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "openaudience-123",
                                    "name": "Cars & Auto_Chrysler Enthusiasts",
                                    "path": path,
                                    "extra": {},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Audience_Canada_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "geo": {"includes": {"country": "ca"}},
                "audience": "Cars & Auto_Chrysler Enthusiasts",
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True

    @pytest.mark.asyncio
    async def test_prepare_blocks_oa_match_audience_outside_us(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.package.targeting.audience.openaudience_custom":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "openaudience-123",
                                    "name": "Cars & Auto_Chrysler Enthusiasts",
                                    "path": path,
                                    "extra": {"export_type": "oa_match"},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Audience_Canada_Restricted_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "geo": {"includes": {"country": "ca"}},
                "audience": "Cars & Auto_Chrysler Enthusiasts",
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is False
        assert any(blocker["code"] == "audience_export_type_requires_us_geo" for blocker in result["blockers"])

    @pytest.mark.asyncio
    async def test_prepare_blocks_audience_without_explicit_country(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.package.targeting.audience.openaudience_custom":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "openaudience-123",
                                    "name": "Cars & Auto_Chrysler Enthusiasts",
                                    "path": path,
                                    "extra": {},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Audience_Country_Block_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "geo": {"includes": {"state": "California"}},
                "audience": "Cars & Auto_Chrysler Enthusiasts",
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is False
        assert any(blocker["code"] == "subnational_geo_requires_country" for blocker in result["blockers"])
        assert any(blocker["code"] == "audience_requires_explicit_country" for blocker in result["blockers"])

    @pytest.mark.asyncio
    async def test_prepare_resolves_canada_excluding_alberta(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            variables = payload.get("variables", {})
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.package.targeting.geographic.state":
                assert variables.get("filter") == {"state": "alberta*", "country": "canada*"}
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "550",
                                    "name": "alberta",
                                    "path": path,
                                    "extra": {
                                        "continent": "north america",
                                        "country": "canada",
                                        "region": "prairie provinces",
                                        "state": "alberta",
                                        "type": "state",
                                        "type-id": "state-550",
                                        "type_id": "state-550",
                                    },
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            raise AssertionError(f"Unexpected options path/filter: {path} / {variables.get('filter')}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Canada_Exclude_Alberta_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "geo": {"includes": {"country": "ca"}, "excludes": {"state": "Alberta", "country": "ca"}},
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["targeting"]["geographic"] == {
            "includes": {"country": "ca"},
            "excludes": {"state": "550"},
        }
        assert any("Resolved state 'Alberta' to OpenX id '550'" in warning for warning in result["warnings"])

    @pytest.mark.asyncio
    async def test_prepare_preserves_explicit_country_exclusion_without_state(
        self, mock_openx_graphql: respx.MockRouter
    ):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            raise AssertionError(f"Unexpected options path: {path}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Country_Exclude_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "geo": {"includes": {"country": "us,ca"}, "excludes": {"country": "ca"}},
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["targeting"]["geographic"] == {
            "includes": {"country": "us,ca"},
            "excludes": {"country": "ca"},
        }

    @pytest.mark.asyncio
    async def test_prepare_adds_viewability_threshold(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            raise AssertionError(f"Unexpected options path: {path}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Viewability_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-04-15T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": {"includes": {"country": "us"}}},
            viewability_threshold=70,
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["targeting"]["viewability"] == {
            "viewability_score": {"op": ">=", "val": "0.70"}
        }
        assert any("Normalized viewability threshold" in warning for warning in result["warnings"])

    @pytest.mark.asyncio
    async def test_prepare_passes_paused_status(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            raise AssertionError(f"Unexpected options path: {path}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Paused_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-04-15T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": {"includes": {"country": "us"}}},
            status="Paused",
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["status"] == "Paused"

    @pytest.mark.asyncio
    async def test_prepare_adds_publisher_ids_to_content_account(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            raise AssertionError(f"Unexpected options path: {path}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Publisher_Targeting_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_CTV_Targeting",
            demand_partner="TRADR / 393",
            publisher_ids=["123", "456"],
            targeting={"channel": "CTV", "geo": ["US"]},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["targeting"]["content"]["account"] == {
            "op": "INTERSECTS",
            "val": "123,456",
        }

    @pytest.mark.asyncio
    async def test_prepare_passes_expected_ad_category_through_to_create_args(
        self, mock_openx_graphql: respx.MockRouter
    ):
        """When the brief flags a sensitive category (e.g. ``"Politics"``), prepare must thread
        the value through to ``create_args.expected_ad_category`` so dealCreate ships it as a
        top-level field. Without this, every political deal needs a manual UI fix to set the
        Expected Sensitive Category — exactly what surfaced on the Reklaim/Optimum batches."""

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Political_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Political_Package",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": ["US"]},
            expected_ad_category="Politics",
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["expected_ad_category"] == "Politics"

    @pytest.mark.asyncio
    async def test_prepare_omits_expected_ad_category_when_not_supplied(self, mock_openx_graphql: respx.MockRouter):
        """When ``expected_ad_category`` is omitted, ``create_args.expected_ad_category`` MUST be
        ``None`` so ``ox_create_prepared_deal`` filters it out of the dealCreate payload entirely.
        Non-political deals should not declare a sensitive category."""

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Non_Political_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Non_Political_Package",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": ["US"]},
        )

        assert result["success"] is True
        assert result["create_args_preview"]["expected_ad_category"] is None

    @pytest.mark.asyncio
    async def test_prepare_adds_excluded_publisher_ids_to_content_account(self, mock_openx_graphql: respx.MockRouter):
        """Exclude-publisher path mirrors the include path but emits NOT INTERSECTS instead
        of INTERSECTS. Wire format verified against the UI deal JSON from the Reklaim CTV
        batch: targeting.content.account = {op: "NOT INTERSECTS", val: "193155,209125"}."""

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Publisher_Exclusion_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Publisher_Exclusion_Package",
            demand_partner="TRADR / 393",
            excluded_publisher_ids=["193155", "209125"],
            targeting={"channel": "CTV", "geo": ["US"]},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["targeting"]["content"]["account"] == {
            "op": "NOT INTERSECTS",
            "val": "193155,209125",
        }

    @pytest.mark.asyncio
    async def test_prepare_blocks_when_both_publisher_lists_supplied(self, mock_openx_graphql: respx.MockRouter):
        """OpenX content.account is a single object with one op + one val. A deal that
        supplies BOTH publisher_ids (INTERSECTS) and excluded_publisher_ids (NOT INTERSECTS)
        is ambiguous — surface a structured blocker rather than silently dropping one of the
        two lists."""

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Conflicting_Publisher_Lists_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Conflicting_Publisher_Lists_Package",
            demand_partner="TRADR / 393",
            publisher_ids=["111"],
            excluded_publisher_ids=["222"],
            targeting={"channel": "DISPLAY", "geo": ["US"]},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is False
        assert any(b["code"] == "conflicting_publisher_lists" for b in result["blockers"])

    @pytest.mark.asyncio
    async def test_prepare_resolves_inventory_categories_to_metacategory_includes(
        self, mock_openx_graphql: respx.MockRouter
    ):
        """The Reklaim CTV batch needed metacategory.includes = {op: "OR", val: ["premiumctv"]}
        for the "TV by OpenX – CTV – App Bundles" inventory category. Without this resolver,
        traders had to set the category in the OpenX UI after every create. The wire shape is
        distinct from IAB categories: a list value (not comma-separated string) with op="OR"."""

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            if path == "deal.package.targeting.metacategory.includes":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "premiumctv",
                                    "name": "TV by OpenX - CTV - App Bundles",
                                    "path": path,
                                    "extra": {},
                                },
                                {
                                    "id": "premiumweb",
                                    "name": "Premium Web Inventory",
                                    "path": path,
                                    "extra": {},
                                },
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="CTV_Inventory_Categories_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="CTV_Inventory_Categories_Package",
            demand_partner="TRADR / 393",
            targeting={"channel": "CTV", "geo": ["US"]},
            inventory_categories=["TV by OpenX - CTV - App Bundles"],
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        metacategory = result["create_args_preview"]["targeting"]["metacategory"]
        # Wire shape: {op: "OR", val: ["premiumctv"]} — list, not comma-joined string.
        assert metacategory["includes"] == {"op": "OR", "val": ["premiumctv"]}
        # Setting metacategory in prepare short-circuits the create-side defaulting block,
        # so we MUST emit the policy defaults ourselves — exclude_mfa especially, since
        # every UI deal Elcano ships has it set.
        assert metacategory["exclude_mfa"] is True
        assert metacategory["excludes"] == []
        assert metacategory["keywords"] is None
        assert metacategory["inter_dimension_operator"] == "AND"

    @pytest.mark.asyncio
    async def test_prepare_blocks_unresolved_inventory_category(self, mock_openx_graphql: respx.MockRouter):
        """A trader-supplied inventory category that doesn't match any OpenX option must
        produce a structured ``unresolved_inventory_category`` blocker rather than silently
        dropping the field — silent drop is exactly how traders end up with deals that go
        live without the category they specified."""

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.package.targeting.metacategory.includes":
                # Return only categories that won't match the trader's input.
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "premiumweb", "name": "Premium Web Inventory", "path": path, "extra": {}},
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Unresolved_Inventory_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Unresolved_Inventory_Package",
            demand_partner="TRADR / 393",
            targeting={"channel": "CTV", "geo": ["US"]},
            inventory_categories=["This Category Does Not Exist In OpenX"],
        )

        assert result["success"] is True
        assert result["ready_to_create"] is False
        assert any(b["code"] == "unresolved_inventory_category" for b in result["blockers"])

    @pytest.mark.asyncio
    async def test_prepare_prefixes_bare_audience_uuid_with_openaudience(self, mock_openx_graphql: respx.MockRouter):
        """Live OpenX optionsByPath returns the bare audience UUID; dealCreate requires the
        ``openaudience-`` prefix or it returns ``invalid audience id``. The prepare path must
        normalize either shape into the prefixed wire format."""

        bare_uuid = "30fcdbfe-347a-4c24-a1bd-9884e993d432"

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.package.targeting.audience.openaudience_custom":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": bare_uuid,
                                    "name": "Reklaim > Political Affiliation > Independent",
                                    "path": path,
                                    "extra": {},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Audience_Prefix_Normalization_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Audience_Prefix_Package",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "geo": {"includes": {"country": "US"}},
                "audience": "Reklaim > Political Affiliation > Independent",
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        expected_val = f"openaudience-{bare_uuid}"
        assert result["create_args_preview"]["targeting"]["audience"]["openaudience_custom"] == {
            "op": "INTERSECTS",
            "val": expected_val,
        }
        assert result["resolved_entities"]["audience"] == {
            "op": "INTERSECTS",
            "val": expected_val,
        }

    @pytest.mark.asyncio
    async def test_prepare_does_not_double_prefix_audience_id(self, mock_openx_graphql: respx.MockRouter):
        """If optionsByPath already returns the prefixed form, the normalizer must not
        double-prefix into ``openaudience-openaudience-...``."""

        prefixed_id = "openaudience-30fcdbfe-347a-4c24-a1bd-9884e993d432"

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.deal_participants.demand_partner":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]
                        }
                    },
                )
            if path == "deal.deal_participants.buyer_ids":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "seat-map",
                                    "name": "Seats",
                                    "path": path,
                                    "extra": {"third_party_buyer_id": ["393"]},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.package.targeting.audience.openaudience_custom":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": prefixed_id,
                                    "name": "Reklaim Already Prefixed",
                                    "path": path,
                                    "extra": {},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            return httpx.Response(200, json={"data": {"optionsByPath": []}})

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_prepare_deal_from_brief(
            name="Audience_Prefixed_Already_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Audience_Prefixed_Already_Package",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "geo": {"includes": {"country": "US"}},
                "audience": "Reklaim Already Prefixed",
            },
        )

        assert result["success"] is True
        assert result["create_args_preview"]["targeting"]["audience"]["openaudience_custom"]["val"] == prefixed_id


class TestCreatePreparedAndExecute:
    @pytest.mark.asyncio
    async def test_create_prepared_deal_uses_prepared_payload(self, mock_openx_graphql: respx.MockRouter):
        captured_mutation = None

        def handler(request: httpx.Request) -> httpx.Response:
            nonlocal captured_mutation
            payload = json.loads(request.content)
            query = payload.get("query", "")
            variables = payload.get("variables", {})

            if "optionsByPath" in query:
                path = variables["path"]
                if path == "deal.deal_participants.demand_partner":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                if path == "deal.deal_participants.buyer_ids":
                    assert variables["filter"] == {"demand_partner_id": "537073256"}
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {
                                        "id": "bidswitch-seats",
                                        "name": "Bidswitch Seats",
                                        "path": path,
                                        "extra": {"third_party_buyer_id": ["393"]},
                                    }
                                ]
                            }
                        },
                    )
                if path == "deal.third_party_fees_config.partner_id":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                return httpx.Response(200, json={"data": {"optionsByPath": []}})

            if "__type" in query or "IntrospectType" in query:
                return httpx.Response(200, json=INTROSPECT_GEOGRAPHIC_ITEM_RESPONSE)
            if "dealCreate" in query:
                captured_mutation = payload
                return httpx.Response(200, json=CREATE_DEAL_SUCCESS_RESPONSE)
            if "dealById" in query:
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "dealById": {
                                "id": "deal-new-001",
                                "deal_id": "ELC-2024-NEW-001",
                                "name": "Prepared_Deal",
                                "status": "ACTIVE",
                                "currency": "USD",
                                "deal_price": 1.25,
                                "pmp_deal_type": "3",
                                "start_date": "2026-03-30T00:00:00Z",
                                "end_date": None,
                                "created_date": "2026-03-01T00:00:00Z",
                                "modified_date": "2026-03-01T00:00:00Z",
                                "deal_participants": [
                                    {"demand_partner": "537073256", "buyer_ids": ["393"], "brand_ids": []}
                                ],
                                "package": {
                                    "name": "Prepared_Package",
                                    "targeting": {
                                        "inter_dimension_operator": "AND",
                                        "rendering_context": {
                                            "op": "AND",
                                            "ad_placement": {"op": "==", "val": "BANNER"},
                                            "distribution_channel": {"op": "INTERSECTS", "val": "WEB,APP"},
                                            "device_type": {
                                                "op": "INTERSECTS",
                                                "desktop_devices": "desktop",
                                                "mobile_devices": None,
                                                "tv_devices": None,
                                            },
                                        },
                                    },
                                },
                                "third_party_fees_config": [],
                            }
                        }
                    },
                )
            raise AssertionError(f"Unexpected query: {query}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        prepared = await ox_prepare_deal_from_brief(
            name="Prepared_Deal",
            currency="USD",
            deal_price=1.25,
            start_date="2026-03-30T00:00:00Z",
            package_name="Prepared_Package",
            demand_partner="537073256",
            buyer_ids=["393"],
            targeting={"channel": "DISPLAY", "device_types": ["DESKTOP"], "geo": ["US"]},
        )

        result = await ox_create_prepared_deal(prepared["prepared_deal_id"])

        assert result["success"] is True
        assert result["prepared_deal_id"] == prepared["prepared_deal_id"]
        input_params = captured_mutation["variables"]["input"]
        assert input_params["deal_participants"][0]["demand_partner"] == "537073256"
        assert input_params["package"]["targeting"]["geographic"] == {"includes": {"country": "us"}}

    @pytest.mark.asyncio
    async def test_create_prepared_deal_ships_expected_ad_category_to_dealCreate(
        self, mock_openx_graphql: respx.MockRouter
    ):
        """End-to-end: ``expected_ad_category`` passed to prepare must appear as a top-level
        field on the dealCreate GraphQL mutation input. This is the wire-format check that
        proves the trader no longer needs the manual UI step."""
        captured_mutation: dict[str, Any] = {}

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            query = payload.get("query", "")
            variables = payload.get("variables", {})

            if "optionsByPath" in query:
                path = variables["path"]
                if path == "deal.deal_participants.demand_partner":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                if path == "deal.deal_participants.buyer_ids":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {
                                        "id": "bidswitch-seats",
                                        "name": "Bidswitch Seats",
                                        "path": path,
                                        "extra": {"third_party_buyer_id": ["393"]},
                                    }
                                ]
                            }
                        },
                    )
                if path == "deal.third_party_fees_config.partner_id":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                return httpx.Response(200, json={"data": {"optionsByPath": []}})

            if "__type" in query or "IntrospectType" in query:
                return httpx.Response(200, json=INTROSPECT_GEOGRAPHIC_ITEM_RESPONSE)
            if "dealCreate" in query:
                captured_mutation["payload"] = payload
                return httpx.Response(200, json=CREATE_DEAL_SUCCESS_RESPONSE)
            if "dealById" in query:
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "dealById": {
                                "id": "deal-new-001",
                                "deal_id": "ELC-2024-NEW-001",
                                "name": "Political_Deal",
                                "status": "ACTIVE",
                                "currency": "USD",
                                "deal_price": 0.10,
                                "pmp_deal_type": "3",
                                "start_date": "2026-03-30T00:00:00Z",
                                "end_date": None,
                                "created_date": "2026-03-01T00:00:00Z",
                                "modified_date": "2026-03-01T00:00:00Z",
                                "deal_participants": [
                                    {"demand_partner": "537073256", "buyer_ids": ["393"], "brand_ids": []}
                                ],
                                "package": {
                                    "name": "Political_Package",
                                    "targeting": {
                                        "inter_dimension_operator": "AND",
                                        "rendering_context": {
                                            "op": "AND",
                                            "ad_placement": {"op": "==", "val": "BANNER"},
                                            "distribution_channel": {"op": "INTERSECTS", "val": "WEB,APP"},
                                            "device_type": {
                                                "op": "INTERSECTS",
                                                "desktop_devices": "desktop",
                                                "mobile_devices": None,
                                                "tv_devices": None,
                                            },
                                        },
                                    },
                                },
                                "third_party_fees_config": [],
                            }
                        }
                    },
                )
            raise AssertionError(f"Unexpected query: {query}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        prepared = await ox_prepare_deal_from_brief(
            name="Political_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Political_Package",
            demand_partner="537073256",
            buyer_ids=["393"],
            targeting={"channel": "DISPLAY", "device_types": ["DESKTOP"], "geo": ["US"]},
            expected_ad_category="Politics",
        )
        result = await ox_create_prepared_deal(prepared["prepared_deal_id"])

        assert result["success"] is True
        input_params = captured_mutation["payload"]["variables"]["input"]
        assert input_params["expected_ad_category"] == "Politics"

    @pytest.mark.asyncio
    async def test_create_prepared_deal_omits_expected_ad_category_when_not_set(
        self, mock_openx_graphql: respx.MockRouter
    ):
        """When the field is omitted, dealCreate input MUST NOT carry an
        ``expected_ad_category`` key at all (rather than ``null``) — OpenX validates the enum
        and a stray null could be rejected on stricter schemas. The
        ``ox_create_prepared_deal`` shim filters ``None`` keys via dict-comprehension; this
        test pins that behavior end-to-end."""
        captured_mutation: dict[str, Any] = {}

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            query = payload.get("query", "")
            variables = payload.get("variables", {})

            if "optionsByPath" in query:
                path = variables["path"]
                if path == "deal.deal_participants.demand_partner":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                if path == "deal.deal_participants.buyer_ids":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {
                                        "id": "bidswitch-seats",
                                        "name": "Bidswitch Seats",
                                        "path": path,
                                        "extra": {"third_party_buyer_id": ["393"]},
                                    }
                                ]
                            }
                        },
                    )
                if path == "deal.third_party_fees_config.partner_id":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                return httpx.Response(200, json={"data": {"optionsByPath": []}})

            if "__type" in query or "IntrospectType" in query:
                return httpx.Response(200, json=INTROSPECT_GEOGRAPHIC_ITEM_RESPONSE)
            if "dealCreate" in query:
                captured_mutation["payload"] = payload
                return httpx.Response(200, json=CREATE_DEAL_SUCCESS_RESPONSE)
            if "dealById" in query:
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "dealById": {
                                "id": "deal-new-001",
                                "deal_id": "ELC-2024-NEW-001",
                                "name": "Non_Political_Deal",
                                "status": "ACTIVE",
                                "currency": "USD",
                                "deal_price": 0.10,
                                "pmp_deal_type": "3",
                                "start_date": "2026-03-30T00:00:00Z",
                                "end_date": None,
                                "created_date": "2026-03-01T00:00:00Z",
                                "modified_date": "2026-03-01T00:00:00Z",
                                "deal_participants": [
                                    {"demand_partner": "537073256", "buyer_ids": ["393"], "brand_ids": []}
                                ],
                                "package": {
                                    "name": "Non_Political_Package",
                                    "targeting": {
                                        "inter_dimension_operator": "AND",
                                        "rendering_context": {
                                            "op": "AND",
                                            "ad_placement": {"op": "==", "val": "BANNER"},
                                            "distribution_channel": {"op": "INTERSECTS", "val": "WEB,APP"},
                                            "device_type": {
                                                "op": "INTERSECTS",
                                                "desktop_devices": "desktop",
                                                "mobile_devices": None,
                                                "tv_devices": None,
                                            },
                                        },
                                    },
                                },
                                "third_party_fees_config": [],
                            }
                        }
                    },
                )
            raise AssertionError(f"Unexpected query: {query}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        prepared = await ox_prepare_deal_from_brief(
            name="Non_Political_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Non_Political_Package",
            demand_partner="537073256",
            buyer_ids=["393"],
            targeting={"channel": "DISPLAY", "device_types": ["DESKTOP"], "geo": ["US"]},
        )
        result = await ox_create_prepared_deal(prepared["prepared_deal_id"])

        assert result["success"] is True
        input_params = captured_mutation["payload"]["variables"]["input"]
        assert "expected_ad_category" not in input_params

    @pytest.mark.asyncio
    async def test_execute_deal_from_prompt_inputs_end_to_end(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.append(["Sites"])
        worksheet.append(["example.com"])
        file_path = tmp_path / "automotivelist.xlsx"
        workbook.save(file_path)

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            query = payload.get("query", "")
            variables = payload.get("variables", {})
            if "optionsByPath" in query:
                path = variables["path"]
                if path == "deal.deal_participants.demand_partner":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                if path == "deal.deal_participants.buyer_ids":
                    assert variables["filter"] == {"demand_partner_id": "537073256"}
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {
                                        "id": "seat-map",
                                        "name": "Seats",
                                        "path": path,
                                        "extra": {"third_party_buyer_id": ["393"]},
                                    }
                                ]
                            }
                        },
                    )
                if path == "deal.third_party_fees_config.partner_id":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [{"id": "560610563", "name": "Elcano", "path": path, "extra": {}}]
                            }
                        },
                    )
                return httpx.Response(200, json={"data": {"optionsByPath": []}})
            if "__type" in query or "IntrospectType" in query:
                return httpx.Response(200, json=INTROSPECT_GEOGRAPHIC_ITEM_RESPONSE)
            if "dealCreate" in query:
                return httpx.Response(200, json=CREATE_DEAL_SUCCESS_RESPONSE)
            if "dealById" in query:
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "dealById": {
                                "id": "deal-new-001",
                                "deal_id": "ELC-2024-NEW-001",
                                "name": "Elcano_OpenX_Test_Deal",
                                "status": "ACTIVE",
                                "currency": "USD",
                                "deal_price": 7.50,
                                "pmp_deal_type": "3",
                                "start_date": "2024-03-01T00:00:00Z",
                                "end_date": None,
                                "created_date": "2024-02-28T10:00:00Z",
                                "modified_date": "2024-02-28T10:00:00Z",
                                "deal_participants": [
                                    {"demand_partner": "537073256", "buyer_ids": ["393"], "brand_ids": []}
                                ],
                                "package": {
                                    "name": "TorqueDrive_Display_Targeting",
                                    "targeting": {
                                        "inter_dimension_operator": "AND",
                                        "rendering_context": {
                                            "op": "AND",
                                            "ad_placement": {"op": "==", "val": "BANNER"},
                                            "distribution_channel": {"op": "INTERSECTS", "val": "WEB,APP"},
                                            "device_type": {
                                                "op": "INTERSECTS",
                                                "desktop_devices": "desktop",
                                                "mobile_devices": "phone,tablet",
                                                "tv_devices": None,
                                            },
                                        },
                                    },
                                },
                                "third_party_fees_config": [],
                            }
                        }
                    },
                )
            raise AssertionError(f"Unexpected query: {query}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_execute_deal_from_prompt_inputs(
            name="Elcano_OpenX_Test_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            domain_file_path=str(file_path),
            domain_match_operator="blocklist",
            targeting={
                "channel": "DISPLAY",
                "device_types": ["DESKTOP", "MOBILE", "TABLET"],
                "geo": {"includes": {"country": "US"}},
            },
            fee={"partner_name_or_id": "560610563", "revenue_method": "PoM", "gross_share_percent": 40},
        )

        assert result["success"] is True
        assert result["preparation"]["ready_to_create"] is True
        assert result["creation"]["success"] is True
        assert result["verification"]["success"] is True
        assert result["verification_attempts"][0] == {"identifier_type": "id", "identifier_value": "deal-new-001"}

    @pytest.mark.asyncio
    async def test_execute_returns_success_when_create_succeeds_but_verification_fails(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.append(["Sites"])
        worksheet.append(["example.com"])
        file_path = tmp_path / "automotivelist.xlsx"
        workbook.save(file_path)

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            query = payload.get("query", "")
            variables = payload.get("variables", {})
            if "optionsByPath" in query:
                path = variables["path"]
                if path == "deal.deal_participants.demand_partner":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                if path == "deal.deal_participants.buyer_ids":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {
                                        "id": "seat-map",
                                        "name": "Seats",
                                        "path": path,
                                        "extra": {"third_party_buyer_id": ["393"]},
                                    }
                                ]
                            }
                        },
                    )
                if path == "deal.third_party_fees_config.partner_id":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                return httpx.Response(200, json={"data": {"optionsByPath": []}})
            if "__type" in query or "IntrospectType" in query:
                return httpx.Response(200, json=INTROSPECT_GEOGRAPHIC_ITEM_RESPONSE)
            if "dealCreate" in query:
                return httpx.Response(200, json=CREATE_DEAL_SUCCESS_RESPONSE)
            if "dealById" in query:
                return httpx.Response(401, text="Unauthorized")
            raise AssertionError(f"Unexpected query: {query}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_execute_deal_from_prompt_inputs(
            name="Elcano_OpenX_Test_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_Display_Targeting",
            demand_partner="TRADR / 393",
            domain_file_path=str(file_path),
            domain_match_operator="blocklist",
            targeting={"channel": "DISPLAY", "device_types": ["DESKTOP"], "geo": {"includes": {"country": "US"}}},
        )

        assert result["success"] is True
        assert result["creation"]["success"] is True
        assert result["verification"]["success"] is False
        assert result["verification_success"] is False
        assert result["verification_attempts"] == [
            {"identifier_type": "id", "identifier_value": "deal-new-001"},
            {"identifier_type": "deal_id", "identifier_value": "ELC-2024-NEW-001"},
        ]
        assert any("post-create verification failed" in warning.lower() for warning in result["warnings"])

    @pytest.mark.asyncio
    async def test_execute_ctv_deal_from_prompt_inputs_with_publisher_file(
        self, mock_openx_graphql: respx.MockRouter, tmp_path: Path
    ):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Publishers"
        worksheet.append(["Publisher ID"])
        worksheet.append([123])
        worksheet.append(["456"])
        file_path = tmp_path / "tricoast_publishers.xlsx"
        workbook.save(file_path)

        captured_create_input = None

        def handler(request: httpx.Request) -> httpx.Response:
            nonlocal captured_create_input

            payload = json.loads(request.content)
            query = payload.get("query", "")
            variables = payload.get("variables", {})
            if "optionsByPath" in query:
                path = variables["path"]
                if path == "deal.deal_participants.demand_partner":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                if path == "deal.deal_participants.buyer_ids":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {
                                        "id": "seat-map",
                                        "name": "Seats",
                                        "path": path,
                                        "extra": {"third_party_buyer_id": ["393"]},
                                    }
                                ]
                            }
                        },
                    )
                if path == "deal.third_party_fees_config.partner_id":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                return httpx.Response(200, json={"data": {"optionsByPath": []}})
            if "__type" in query or "IntrospectType" in query:
                return httpx.Response(200, json=INTROSPECT_GEOGRAPHIC_ITEM_RESPONSE)
            if "dealCreate" in query:
                captured_create_input = payload["variables"]["input"]
                return httpx.Response(200, json=CREATE_DEAL_SUCCESS_RESPONSE)
            if "dealById" in query:
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "dealById": {
                                "id": "deal-new-001",
                                "deal_id": "ELC-2024-NEW-001",
                                "name": "CTV Publisher Deal",
                                "status": "ACTIVE",
                                "currency": "USD",
                                "deal_price": 0.10,
                                "pmp_deal_type": "3",
                                "start_date": "2026-03-30T00:00:00Z",
                                "end_date": None,
                                "created_date": "2026-03-30T00:00:00Z",
                                "modified_date": "2026-03-30T00:00:00Z",
                                "deal_participants": [{"demand_partner": "537073256", "buyer_ids": ["393"]}],
                                "package": {
                                    "name": "TorqueDrive_CTV_Targeting",
                                    "targeting": {
                                        "inter_dimension_operator": "AND",
                                        "rendering_context": {
                                            "op": "AND",
                                            "ad_placement": {"op": "==", "val": "CTV"},
                                            "distribution_channel": {"op": "INTERSECTS", "val": "APP"},
                                            "device_type": {
                                                "op": "INTERSECTS",
                                                "desktop_devices": None,
                                                "mobile_devices": None,
                                                "tv_devices": "tv",
                                            },
                                        },
                                        "content": {"account": {"op": "INTERSECTS", "val": "123,456"}},
                                    },
                                },
                                "third_party_fees_config": [],
                            }
                        }
                    },
                )
            raise AssertionError(f"Unexpected query: {query}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_execute_deal_from_prompt_inputs(
            name="CTV_Publisher_Execute_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="TorqueDrive_CTV_Targeting",
            demand_partner="TRADR / 393",
            targeting={"channel": "CTV", "device_types": ["CTV"], "geo": ["US"]},
            publisher_file_path=str(file_path),
            publisher_sheet="Publishers",
            publisher_column="Publisher ID",
        )

        assert result["success"] is True
        assert result["preparation"]["ready_to_create"] is True
        assert result["preparation"]["extracted_publisher_count"] == 2
        assert result["creation"]["success"] is True
        assert result["verification"]["success"] is True
        assert captured_create_input is not None
        assert "url_targeting" not in captured_create_input["package"]
        assert captured_create_input["package"]["targeting"]["content"]["account"] == {
            "op": "INTERSECTS",
            "val": "123,456",
        }

    @pytest.mark.asyncio
    async def test_validate_audience_geo_reports_oa_match_us_only(self, mock_openx_graphql: respx.MockRouter):
        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            path = payload.get("variables", {}).get("path")
            if path == "deal.package.targeting.audience.openaudience_custom":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {
                                    "id": "openaudience-30fcdbfe-347a-4c24-a1bd-9884e993d432",
                                    "name": "Cars & Auto_Chrysler Enthusiasts",
                                    "path": path,
                                    "extra": {"status": "exported", "export_type": "oa_match"},
                                }
                            ]
                        }
                    },
                )
            if path == "deal.third_party_fees_config.partner_id":
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "optionsByPath": [
                                {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                            ]
                        }
                    },
                )
            raise AssertionError(f"Unexpected options path: {path}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_validate_audience_geo_compatibility(
            audience="Cars & Auto_Chrysler Enthusiasts",
            geo={"includes": {"country": "ca"}, "excludes": {"state": "Alberta", "country": "ca"}},
        )

        assert result["success"] is True
        assert result["compatible"] is False
        assert any("only in US geography" in issue for issue in result["issues"])


class TestEndToEndGoldenPath:
    """End-to-end golden test simulating the real operator prompt shape."""

    @pytest.mark.asyncio
    async def test_golden_path_display_deal_with_attachment(self, mock_openx_graphql: respx.MockRouter, tmp_path: Path):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.append(["Sites"])
        worksheet.append(["example.com"])
        worksheet.append(["autotrader.com"])
        file_path = tmp_path / "automotivelist.xlsx"
        workbook.save(file_path)

        def handler(request: httpx.Request) -> httpx.Response:
            payload = json.loads(request.content)
            query = payload.get("query", "")
            variables = payload.get("variables", {})
            if "optionsByPath" in query:
                path = variables["path"]
                if path == "deal.deal_participants.demand_partner":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                if path == "deal.deal_participants.buyer_ids":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {
                                        "id": "seat-map",
                                        "name": "Seats",
                                        "path": path,
                                        "extra": {"third_party_buyer_id": ["393"]},
                                    }
                                ]
                            }
                        },
                    )
                if path == "deal.third_party_fees_config.partner_id":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}
                                ]
                            }
                        },
                    )
                if path == "deal.package.targeting.technographic.language":
                    return httpx.Response(
                        200,
                        json={"data": {"optionsByPath": [{"id": "es", "name": "Spanish", "path": path, "extra": {}}]}},
                    )
                if path == "deal.package.targeting.audience.openaudience_custom":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {
                                        "id": "openaudience-abc123",
                                        "name": "Cars & Auto_Chrysler Enthusiasts",
                                        "path": path,
                                        "extra": {},
                                    }
                                ]
                            }
                        },
                    )
                if path == "deal.package.targeting.domain.categories_iab_v2":
                    return httpx.Response(
                        200,
                        json={
                            "data": {
                                "optionsByPath": [
                                    {
                                        "id": "18",
                                        "name": "Certified Pre-Owned Cars",
                                        "path": path,
                                        "extra": {"parent": "1"},
                                    },
                                    {"id": "25", "name": "Car Culture", "path": path, "extra": {"parent": "1"}},
                                ]
                            }
                        },
                    )
                return httpx.Response(200, json={"data": {"optionsByPath": []}})
            if "__type" in query or "IntrospectType" in query:
                return httpx.Response(200, json=INTROSPECT_GEOGRAPHIC_ITEM_RESPONSE)
            if "dealCreate" in query:
                return httpx.Response(200, json=CREATE_DEAL_SUCCESS_RESPONSE)
            if "dealById" in query:
                return httpx.Response(
                    200,
                    json={
                        "data": {
                            "dealById": {
                                "id": "deal-new-001",
                                "uid": "21b2e0e4-c0a9-fff1-8123-69534a",
                                "deal_id": "ELC-2024-NEW-001",
                                "name": "Golden_Path_Deal",
                                "status": "Active",
                                "currency": "USD",
                                "deal_price": "0.10",
                                "pmp_deal_type": "3",
                                "start_date": "2026-03-30T00:00:00Z",
                                "end_date": None,
                                "created_date": "2026-03-30T00:00:00Z",
                                "modified_date": "2026-03-30T00:00:00Z",
                                "deal_participants": [
                                    {"demand_partner": "537073256", "buyer_ids": ["393"], "brand_ids": []}
                                ],
                                "package": {
                                    "uid": "21b2e0df-c0a9-fff1-8123-69534a",
                                    "name": "Golden_Package",
                                    "targeting": {
                                        "inter_dimension_operator": "AND",
                                        "rendering_context": {
                                            "op": "AND",
                                            "ad_placement": {"op": "==", "val": "BANNER"},
                                            "distribution_channel": {"op": "INTERSECTS", "val": "WEB,APP"},
                                            "device_type": {
                                                "op": "INTERSECTS",
                                                "desktop_devices": "desktop",
                                                "mobile_devices": "phone,tablet",
                                                "tv_devices": None,
                                            },
                                        },
                                    },
                                },
                                "third_party_fees_config": [
                                    {
                                        "partner_id": "560610563",
                                        "revenue_method": "PoM",
                                        "gross_share": "0.400",
                                        "gross_cpm_cap": None,
                                    }
                                ],
                            }
                        }
                    },
                )
            raise AssertionError(f"Unexpected query: {query}")

        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=handler)

        result = await ox_execute_deal_from_prompt_inputs(
            name="Golden_Path_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            end_date="2036-03-30T00:00:00Z",
            pmp_deal_type="3",
            package_name="Golden_Package",
            demand_partner="TRADR",
            buyer_ids=["393"],
            fee={"partner": "Elcano", "revenue_method": "PoM", "gross_share": "40.0"},
            domain_file_path=str(file_path),
            domain_sheet="Sheet1",
            domain_column="Sites",
            domain_match_operator="blocklist",
            targeting={
                "channel": "DISPLAY",
                "device_types": ["DESKTOP", "MOBILE", "TABLET"],
                "geo": {"includes": {"country": "US"}},
                "language": "Spanish",
                "iab_categories": ["Certified Pre-Owned Cars", "Car Culture"],  # OpenX IAB v2 names, not legacy codes
                "audience": "Cars & Auto_Chrysler Enthusiasts",
            },
        )

        assert result["success"] is True
        assert result["phase"] == "verify"
        assert result["verification_success"] is True
        assert result["deal_url"] == "https://select.openx.com/deals/deal-new-001/details"

        prep = result["preparation"]
        assert prep["ready_to_create"] is True
        assert prep["extracted_domain_count"] == 2
        assert prep["resolved_entities"]["demand_partner"]["resolved_id"] == "537073256"
        assert prep["resolved_entities"]["fee_partner_id"] == "560610563"
        assert prep["resolved_entities"]["language"] == {"op": "INTERSECTS", "val": "es"}
        assert prep["resolved_entities"]["audience"] == {"op": "INTERSECTS", "val": "openaudience-abc123"}
        assert prep["resolved_entities"]["iab_categories"] == {"op": "INTERSECTS", "val": "18,25"}

        creation = result["creation"]
        assert creation["success"] is True
        assert creation["deal"]["deal_id"] == "ELC-2024-NEW-001"

        verification = result["verification"]
        assert verification["success"] is True
        assert verification["deal"]["status"] == "Active"
        assert verification["deal_url"] == "https://select.openx.com/deals/deal-new-001/details"


def _attribution_test_handler(request: httpx.Request) -> httpx.Response:
    """Mock handler for attribution-code auto-fee tests: resolves demand partner + Elcano fee partner."""
    payload = json.loads(request.content)
    variables = payload.get("variables", {})
    path = variables.get("path", "")
    if path == "deal.deal_participants.demand_partner":
        return httpx.Response(
            200,
            json={
                "data": {
                    "optionsByPath": [{"id": "537073292", "name": "The Trade Desk - RTB", "path": path, "extra": {}}]
                }
            },
        )
    if path == "deal.deal_participants.buyer_ids":
        return httpx.Response(
            200,
            json={
                "data": {
                    "optionsByPath": [
                        {
                            "id": "ttd-seats",
                            "name": "TTD Seats",
                            "path": path,
                            "extra": {"third_party_buyer_id": ["3233"]},
                        }
                    ]
                }
            },
        )
    if path == "deal.third_party_fees_config.partner_id":
        return httpx.Response(
            200,
            json={
                "data": {
                    "optionsByPath": [{"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}]
                }
            },
        )
    return httpx.Response(200, json={"data": {"optionsByPath": []}})


class TestElcanoDefaultCuratorFee:
    """Apply a flat 30% PoM Elcano curator fee whenever the caller passes no explicit fee.

    The deal-name attribution code (A/B/C/D00/B00) is internal Elcano rev-share accounting
    between Elcano and partners; it does NOT determine the SSP curator fee. The OpenX-side
    curator fee defaults to 30% PoM regardless of attribution code, with explicit overrides
    via the `fee=` argument when a deal needs a non-standard rate.
    """

    @pytest.mark.asyncio
    async def test_default_fee_is_30_percent_pom_for_b14_deal(self, mock_openx_graphql: respx.MockRouter):
        """The B14 deal the operator ran in production should auto-attach 30% PoM Elcano."""
        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=_attribution_test_handler)

        result = await ox_prepare_deal_from_brief(
            name="Reklaim_OpenX_Davis Elen_Toyota_NA_ACM_OLV_All_US_ELC00058_B14",
            currency="USD",
            deal_price=0.10,
            start_date="2026-05-01T00:00:00Z",
            end_date="2026-11-01T00:00:00Z",
            package_name="Reklaim_OpenX_Davis Elen_Toyota_NA_ACM_OLV_All_US_ELC00058_B14_PKG",
            demand_partner="The Trade Desk - RTB",
            buyer_ids=["3233"],
            targeting={"channel": "OLV", "geo": {"includes": {"country": "US"}}},
        )

        assert result["ready_to_create"] is True
        fee_config = result["create_args_preview"]["third_party_fees_config"]
        assert fee_config["partner_id"] == "560610563"
        assert fee_config["revenue_method"] == "PoM"
        assert fee_config["gross_share"] == "0.3"
        assert any(
            "default OpenX curator fee" in warning and "Elcano" in warning and "PoM 30%" in warning
            for warning in result["warnings"]
        )

    @pytest.mark.asyncio
    async def test_default_fee_applies_regardless_of_attribution_code(self, mock_openx_graphql: respx.MockRouter):
        """The trailing attribution code does not change the SSP curator fee — A1, D00, anything."""
        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=_attribution_test_handler)

        for trailing_token in ("A1", "B14", "D00", "B00", "Z99", "Anything"):
            result = await ox_prepare_deal_from_brief(
                name=f"Elcano_OpenX_TTD_Brand_NA_NA_NA_Display_All_US_ELC00001_{trailing_token}",
                currency="USD",
                deal_price=0.10,
                start_date="2026-05-01T00:00:00Z",
                package_name=f"{trailing_token}_Package",
                demand_partner="The Trade Desk - RTB",
                buyer_ids=["3233"],
                targeting={"channel": "DISPLAY", "geo": {"includes": {"country": "US"}}},
            )

            fee_config = result["create_args_preview"]["third_party_fees_config"]
            assert fee_config["revenue_method"] == "PoM"
            assert fee_config["gross_share"] == "0.3"

    @pytest.mark.asyncio
    async def test_explicit_fee_overrides_default(self, mock_openx_graphql: respx.MockRouter):
        """A non-standard curator rate can be set per-deal via explicit fee= without renaming."""
        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=_attribution_test_handler)

        result = await ox_prepare_deal_from_brief(
            name="Reklaim_OpenX_Davis Elen_Toyota_NA_ACM_OLV_All_US_ELC00058_B14",
            currency="USD",
            deal_price=0.10,
            start_date="2026-05-01T00:00:00Z",
            package_name="Override_Test_Package",
            demand_partner="The Trade Desk - RTB",
            buyer_ids=["3233"],
            fee={"partner_name_or_id": "Elcano", "revenue_method": "PoR", "gross_share_percent": 50},
            targeting={"channel": "OLV", "geo": {"includes": {"country": "US"}}},
        )

        fee_config = result["create_args_preview"]["third_party_fees_config"]
        assert fee_config["revenue_method"] == "PoR"
        assert fee_config["gross_share"] == "0.5"
        assert not any("default OpenX curator fee" in warning for warning in result["warnings"])


def _basic_prepare_handler(request: httpx.Request) -> httpx.Response:
    """Minimal optionsByPath handler covering demand-partner, buyer-ids, fee-partner, audience."""
    payload = json.loads(request.content)
    query = payload.get("query", "")
    if "__type" in query or "IntrospectType" in query:
        return httpx.Response(200, json=INTROSPECT_GEOGRAPHIC_ITEM_RESPONSE)
    if "optionsByPath" not in query:
        raise AssertionError(f"Unexpected query: {query}")
    path = payload.get("variables", {}).get("path")
    if path == "deal.deal_participants.demand_partner":
        return httpx.Response(
            200,
            json={
                "data": {"optionsByPath": [{"id": "537073256", "name": "Bidswitch - RTB", "path": path, "extra": {}}]}
            },
        )
    if path == "deal.deal_participants.buyer_ids":
        return httpx.Response(
            200,
            json={
                "data": {
                    "optionsByPath": [
                        {
                            "id": "seat-map",
                            "name": "Seats",
                            "path": path,
                            "extra": {"third_party_buyer_id": ["393"]},
                        }
                    ]
                }
            },
        )
    if path == "deal.third_party_fees_config.partner_id":
        return httpx.Response(
            200,
            json={
                "data": {
                    "optionsByPath": [{"id": "560610563", "name": "Elcano (fka Hyphatec)", "path": path, "extra": {}}]
                }
            },
        )
    if path == "deal.package.targeting.audience.openaudience_custom":
        return httpx.Response(
            200,
            json={
                "data": {
                    "optionsByPath": [
                        {
                            "id": "openaudience-test-1",
                            "name": "Reklaim > Political Affiliation > Democrat",
                            "path": path,
                            "extra": {"status": "exported", "export_type": "oa_match"},
                        }
                    ]
                }
            },
        )
    return httpx.Response(200, json={"data": {"optionsByPath": []}})


class TestFlatCountryListWrappedAsStructuredGeo:
    """Regression: flat country lists must be wrapped into the structured includes.country form.

    Before the fix, geo=["US"] was stored as the bare list ["us"] which OpenX silently dropped
    at create time, leaving deals with no geo filter. The fix wraps it into the structured form
    matching what ox_create_deal already produces.
    """

    @pytest.mark.asyncio
    async def test_national_us_list_becomes_structured_country_dict(self, mock_openx_graphql: respx.MockRouter):
        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=_basic_prepare_handler)

        result = await ox_prepare_deal_from_brief(
            name="Geo_National_US_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Geo_National_US_Package",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": ["US"]},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["targeting"]["geographic"] == {"includes": {"country": "us"}}

    @pytest.mark.asyncio
    async def test_string_geo_becomes_structured_country_dict(self, mock_openx_graphql: respx.MockRouter):
        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=_basic_prepare_handler)

        result = await ox_prepare_deal_from_brief(
            name="Geo_String_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Geo_String_Package",
            demand_partner="TRADR / 393",
            targeting={"channel": "DISPLAY", "geo": "US"},
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["targeting"]["geographic"] == {"includes": {"country": "us"}}

    @pytest.mark.asyncio
    async def test_wrapped_geo_satisfies_audience_country_requirement(self, mock_openx_graphql: respx.MockRouter):
        """With geo wrapped properly, audience + country combination no longer hits the
        'audience_requires_explicit_country' blocker — the deal becomes creatable end-to-end."""
        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=_basic_prepare_handler)

        result = await ox_prepare_deal_from_brief(
            name="Geo_Plus_Audience_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Geo_Plus_Audience_Package",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "geo": ["US"],
                "audience_segments": ["Reklaim > Political Affiliation > Democrat"],
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert not any(blocker["code"] == "audience_requires_explicit_country" for blocker in result["blockers"])
        # Both audience and geographic land in the prepared targeting.
        create_args = result["create_args_preview"]
        assert create_args["targeting"]["geographic"] == {"includes": {"country": "us"}}
        assert create_args["targeting"]["audience"]["openaudience_custom"] == {
            "op": "INTERSECTS",
            "val": "openaudience-test-1",
        }


class TestAudienceSegmentAliases:
    """Regression: cutlass-generated prompts often use 'audience_includes' or
    'audience_segments_include' instead of the canonical 'audience_segments' key.
    Before the fix, these were silently dropped — deals shipped with no audience targeting."""

    @pytest.mark.asyncio
    async def test_audience_includes_alias_resolves_segment(self, mock_openx_graphql: respx.MockRouter):
        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=_basic_prepare_handler)

        result = await ox_prepare_deal_from_brief(
            name="Audience_Includes_Alias_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Audience_Includes_Alias_Package",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "geo": ["US"],
                "audience_includes": ["Reklaim > Political Affiliation > Democrat"],
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["targeting"]["audience"]["openaudience_custom"] == {
            "op": "INTERSECTS",
            "val": "openaudience-test-1",
        }
        assert any(
            "Accepted 'audience_includes' as alias for 'audience_segments'" in warning for warning in result["warnings"]
        )

    @pytest.mark.asyncio
    async def test_audience_segments_include_alias_resolves_segment(self, mock_openx_graphql: respx.MockRouter):
        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=_basic_prepare_handler)

        result = await ox_prepare_deal_from_brief(
            name="Audience_Segments_Include_Alias_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Audience_Segments_Include_Alias_Package",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "geo": ["US"],
                "audience_segments_include": ["Reklaim > Political Affiliation > Democrat"],
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert result["create_args_preview"]["targeting"]["audience"]["openaudience_custom"] == {
            "op": "INTERSECTS",
            "val": "openaudience-test-1",
        }
        assert any(
            "Accepted 'audience_segments_include' as alias for 'audience_segments'" in warning
            for warning in result["warnings"]
        )

    @pytest.mark.asyncio
    async def test_canonical_audience_segments_emits_no_alias_warning(self, mock_openx_graphql: respx.MockRouter):
        mock_openx_graphql.post(OPENX_GRAPHQL_ENDPOINT).mock(side_effect=_basic_prepare_handler)

        result = await ox_prepare_deal_from_brief(
            name="Audience_Canonical_Deal",
            currency="USD",
            deal_price=0.10,
            start_date="2026-03-30T00:00:00Z",
            package_name="Audience_Canonical_Package",
            demand_partner="TRADR / 393",
            targeting={
                "channel": "DISPLAY",
                "geo": ["US"],
                "audience_segments": ["Reklaim > Political Affiliation > Democrat"],
            },
        )

        assert result["success"] is True
        assert result["ready_to_create"] is True
        assert not any("Accepted '" in warning and "as alias" in warning for warning in result["warnings"])
