"""Tests for deal_sheet_server.

These tests render real XLSX output via openpyxl and inspect the resulting
worksheet rather than mocking — the whole point of this server is to
produce a file that matches the canonical trader format, so unit tests
that don't actually open the file would miss the most important class of
bugs.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
import pytest
from deal_sheet_server import (
    DEFAULT_THEME,
    HEADER_ROW,
    SUPPORTED_SSPS,
    _extract_failure_reason,
    build_deal_sheet,
    list_deal_sheet_themes,
    validate_brief,
)


@pytest.fixture(autouse=True)
def isolate_output_dir(tmp_path, monkeypatch):
    """Send each test's generated XLSX to its own tmp dir."""
    monkeypatch.setattr("deal_sheet_server.OUTPUT_DIR", tmp_path)
    return tmp_path


def _row_values(ws, row: int, cols=range(2, 8)) -> list[object]:
    return [ws.cell(row=row, column=c).value for c in cols]


def _data_rows(ws, max_scan: int = 60) -> list[tuple[int, list[object]]]:
    """Return (row_number, values) for every non-empty row in the data area
    excluding the header at row 10."""
    out: list[tuple[int, list[object]]] = []
    for r in range(HEADER_ROW + 1, max_scan):
        vals = _row_values(ws, r)
        if any(v is not None and v != "" for v in vals):
            out.append((r, vals))
    return out


@pytest.mark.asyncio
async def test_lists_themes_includes_defaults():
    res = await list_deal_sheet_themes()
    keys = [t["key"] for t in res["themes"]]
    assert "elcano" in keys
    assert "reklaim" in keys


@pytest.mark.asyncio
async def test_builds_minimal_sheet_default_theme():
    deals = [
        {
            "partner": "Soundwave",
            "deal_name": "Elcano_OpenX_test_CTV",
            "deal_id": "OX-bef-AAA",
            "ssp": "OpenX",
            "channel": "CTV",
            "recommended_bid": "$25-$30",
        }
    ]
    res = await build_deal_sheet(deals=deals, client_name="Test Client", theme="elcano")
    assert res["success"] is True
    assert res["theme_used"] == "elcano"
    assert res["deal_count"] == 1
    assert Path(res["path"]).exists()

    wb = openpyxl.load_workbook(res["path"])
    ws = wb.active
    # Header preserved from template
    assert _row_values(ws, HEADER_ROW)[0] == "Partner"
    # First data row is the channel banner, then the deal
    rows = _data_rows(ws)
    assert rows[0][1][0] == "CTV"  # banner shows the channel name in B
    assert rows[1][1][0] == "Soundwave"  # partner cell of the data row
    assert rows[1][1][2] == "OX-bef-AAA"  # deal_id


@pytest.mark.asyncio
async def test_groups_by_channel_and_orders_known_first():
    """Deals come in mixed channels; expect output ordered CTV → OLV → Display
    regardless of input order, with banners separating each group."""
    deals = [
        {
            "partner": "P",
            "deal_name": "n1",
            "deal_id": "d1",
            "ssp": "OpenX",
            "channel": "Display",
            "recommended_bid": "$1",
        },
        {"partner": "P", "deal_name": "n2", "deal_id": "d2", "ssp": "OpenX", "channel": "OLV", "recommended_bid": "$2"},
        {"partner": "P", "deal_name": "n3", "deal_id": "d3", "ssp": "OpenX", "channel": "CTV", "recommended_bid": "$3"},
        {"partner": "P", "deal_name": "n4", "deal_id": "d4", "ssp": "OpenX", "channel": "OLV", "recommended_bid": "$2"},
    ]
    res = await build_deal_sheet(deals=deals, client_name="Sort Test", theme="elcano")
    wb = openpyxl.load_workbook(res["path"])
    ws = wb.active
    rows = _data_rows(ws)
    banners = [vals[0] for _, vals in rows if vals[1] in (None, "")]
    assert banners == ["CTV", "OLV", "Display"]


@pytest.mark.asyncio
async def test_unknown_theme_falls_back_to_default():
    deals = [
        {"partner": "X", "deal_name": "n", "deal_id": "id", "ssp": "OpenX", "channel": "CTV", "recommended_bid": "$1"}
    ]
    res = await build_deal_sheet(deals=deals, client_name="Fallback", theme="this-theme-does-not-exist")
    assert res["success"] is True
    assert res["theme_used"] == DEFAULT_THEME


@pytest.mark.asyncio
async def test_failed_deal_renders_failure_reason_in_deal_id_column():
    deals = [
        {
            "partner": "P",
            "deal_name": "ok-deal",
            "deal_id": "OX-1",
            "ssp": "OpenX",
            "channel": "CTV",
            "recommended_bid": "$5",
            "status": "created",
        },
        {
            "partner": "P",
            "deal_name": "bad-deal",
            "deal_id": "",
            "ssp": "OpenX",
            "channel": "CTV",
            "recommended_bid": "$5",
            "status": "failed",
            "failure_reason": "segment unavailable",
        },
    ]
    res = await build_deal_sheet(deals=deals, client_name="Fail Test", theme="elcano")
    wb = openpyxl.load_workbook(res["path"])
    ws = wb.active
    rows = _data_rows(ws)
    # banner, ok-deal, bad-deal
    deal_ids = [vals[2] for _, vals in rows if vals[1] not in (None, "")]
    assert deal_ids[0] == "OX-1"
    assert deal_ids[1].startswith("FAILED:") and "segment unavailable" in deal_ids[1]


@pytest.mark.asyncio
async def test_missing_optional_fields_render_dash():
    deals = [
        {
            "partner": "",  # missing
            "deal_name": "n",
            "deal_id": "id",
            "ssp": "OpenX",
            "channel": "CTV",
            # recommended_bid missing
        }
    ]
    res = await build_deal_sheet(deals=deals, client_name="Dash Test", theme="elcano")
    wb = openpyxl.load_workbook(res["path"])
    ws = wb.active
    rows = _data_rows(ws)
    deal_row = rows[1][1]  # rows[0] is the banner
    assert deal_row[0] == "—"  # partner
    assert deal_row[5] == "—"  # recommended_bid


@pytest.mark.asyncio
async def test_empty_deals_returns_error():
    res = await build_deal_sheet(deals=[], client_name="X", theme="elcano")
    assert res["success"] is False
    assert "empty" in res["error"].lower()


@pytest.mark.asyncio
async def test_missing_client_name_returns_error():
    deals = [
        {"partner": "p", "deal_name": "n", "deal_id": "id", "ssp": "openx", "channel": "CTV", "recommended_bid": "$1"}
    ]
    res = await build_deal_sheet(deals=deals, client_name="", theme="elcano")
    assert res["success"] is False


@pytest.mark.asyncio
async def test_theme_colors_applied_to_header():
    """Elcano theme must paint header background #1A0B1E; Reklaim must paint #2C31C9."""
    deals = [
        {"partner": "p", "deal_name": "n", "deal_id": "id", "ssp": "openx", "channel": "CTV", "recommended_bid": "$1"}
    ]

    elcano = await build_deal_sheet(deals=deals, client_name="E", theme="elcano")
    reklaim = await build_deal_sheet(deals=deals, client_name="R", theme="reklaim")

    e_wb = openpyxl.load_workbook(elcano["path"])
    e_color = e_wb.active.cell(row=HEADER_ROW, column=2).fill.fgColor.rgb
    r_wb = openpyxl.load_workbook(reklaim["path"])
    r_color = r_wb.active.cell(row=HEADER_ROW, column=2).fill.fgColor.rgb

    assert e_color.upper().endswith("1A0B1E")
    assert r_color.upper().endswith("2C31C9")


@pytest.mark.asyncio
async def test_logo_is_embedded():
    deals = [
        {"partner": "p", "deal_name": "n", "deal_id": "id", "ssp": "openx", "channel": "CTV", "recommended_bid": "$1"}
    ]
    res = await build_deal_sheet(deals=deals, client_name="Logo Test", theme="elcano")
    wb = openpyxl.load_workbook(res["path"])
    assert len(wb.active._images) == 1


@pytest.mark.asyncio
async def test_unknown_channel_sorts_after_known():
    deals = [
        {
            "partner": "p",
            "deal_name": "audio1",
            "deal_id": "a1",
            "ssp": "openx",
            "channel": "Audio",
            "recommended_bid": "$1",
        },
        {
            "partner": "p",
            "deal_name": "ctv1",
            "deal_id": "c1",
            "ssp": "openx",
            "channel": "CTV",
            "recommended_bid": "$1",
        },
    ]
    res = await build_deal_sheet(deals=deals, client_name="Audio Test", theme="elcano")
    wb = openpyxl.load_workbook(res["path"])
    ws = wb.active
    rows = _data_rows(ws)
    banners = [vals[0] for _, vals in rows if vals[1] in (None, "")]
    assert banners == ["CTV", "Audio"]


# ── _extract_failure_reason ────────────────────────────────────────────


@pytest.mark.parametrize(
    "payload, expected",
    [
        ({"error": "No match found for targeting key state: state"}, "No match found for targeting key state: state"),
        ({"error_message": "  bad seat  "}, "bad seat"),
        ({"message": "auth expired"}, "auth expired"),
        ({"detail": "rate limited"}, "rate limited"),
        ({"errors": ["floor below minimum"]}, "floor below minimum"),
        ({"errors": [{"message": "segment unavailable"}]}, "segment unavailable"),
        ({"errors": [{"error": "duplicate name"}, {"error": "second"}]}, "duplicate name"),
        ({"success": False}, "create returned success=false with no error detail"),
        ({}, "unknown failure"),
        (None, "unknown (no response payload captured)"),
        ("error string not dict", "unknown (no response payload captured)"),
    ],
)
def test_extract_failure_reason_handles_common_shapes(payload, expected):
    assert _extract_failure_reason(payload) == expected


def test_extract_failure_reason_prefers_top_level_over_errors_list():
    payload = {"error": "top-level wins", "errors": [{"message": "nested loses"}]}
    assert _extract_failure_reason(payload) == "top-level wins"


# ── build_deal_sheet failure-row fallbacks ─────────────────────────────


@pytest.mark.asyncio
async def test_failed_deal_uses_ssp_response_when_failure_reason_missing():
    """When the agent passes a raw `ssp_response` instead of pre-extracting
    `failure_reason`, the server extracts the reason itself. Eliminates
    per-SSP normalization burden on the agent."""
    deals = [
        {
            "partner": "p",
            "deal_name": "bad-deal",
            "deal_id": "",
            "ssp": "indexexchange",
            "channel": "CTV",
            "recommended_bid": "$5",
            "status": "failed",
            "ssp_response": {"success": False, "error": "No match found for targeting key state: state"},
        }
    ]
    res = await build_deal_sheet(deals=deals, client_name="SSP Response", theme="elcano")
    wb = openpyxl.load_workbook(res["path"])
    ws = wb.active
    rows = _data_rows(ws)
    deal_id_cell = next(vals[2] for _, vals in rows if vals[1] not in (None, ""))
    assert "FAILED" in deal_id_cell
    assert "No match found for targeting key state: state" in deal_id_cell


@pytest.mark.asyncio
async def test_failed_deal_explicit_failure_reason_wins_over_ssp_response():
    """If both are supplied, explicit failure_reason takes precedence —
    the agent had a chance to summarize and we trust that."""
    deals = [
        {
            "partner": "p",
            "deal_name": "bad-deal",
            "deal_id": "",
            "ssp": "openx",
            "channel": "CTV",
            "recommended_bid": "$5",
            "status": "failed",
            "failure_reason": "agent-curated message",
            "ssp_response": {"error": "raw API error text"},
        }
    ]
    res = await build_deal_sheet(deals=deals, client_name="Precedence", theme="elcano")
    wb = openpyxl.load_workbook(res["path"])
    ws = wb.active
    rows = _data_rows(ws)
    deal_id_cell = next(vals[2] for _, vals in rows if vals[1] not in (None, ""))
    assert "agent-curated message" in deal_id_cell
    assert "raw API error text" not in deal_id_cell


# ── unsupported-SSP guard in build_deal_sheet ──────────────────────────


@pytest.mark.asyncio
async def test_build_deal_sheet_rejects_unsupported_ssp():
    deals = [
        {
            "partner": "p",
            "deal_name": "good",
            "deal_id": "OX-1",
            "ssp": "openx",
            "channel": "CTV",
            "recommended_bid": "$5",
        },
        {
            "partner": "p",
            "deal_name": "bad",
            "deal_id": "ST-1",
            "ssp": "sharethrough",
            "channel": "CTV",
            "recommended_bid": "$5",
        },
    ]
    res = await build_deal_sheet(deals=deals, client_name="Unsupported", theme="elcano")
    assert res["success"] is False
    assert "unsupported SSP" in res["error"].lower() or "unsupported" in res["error"].lower()
    # The error must name the SSP and surface the index so the agent can fix it.
    assert any(entry["ssp"] == "sharethrough" and entry["deal_index"] == 1 for entry in res["unsupported_deals"])


@pytest.mark.parametrize(
    "trader_ssp_input",
    [
        "Index Exchange",  # human-readable form with space
        "index exchange",  # all-lower with space
        "INDEXEXCHANGE",  # all-upper, no space
        "Index-Exchange",  # punctuated
        "indexexchange",  # canonical form (control)
        "Index",  # short form alias
    ],
)
@pytest.mark.asyncio
async def test_build_deal_sheet_accepts_ssp_in_multiple_spellings(trader_ssp_input):
    """Traders type Index Exchange a dozen ways. Normalization (strip
    whitespace + punctuation, lowercase) must collapse every variant to
    the canonical token so the SSP-support gate doesn't reject deals
    just because of human-friendly spelling. Caught in the May-2026 TWC
    Beiersdorf run where `ssp: "Index Exchange"` failed and the agent
    had to retry with the lowercase token."""
    deals = [
        {
            "partner": "The Weather Company",
            "deal_name": "TWC-IX-test",
            "deal_id": "IX-test-1",
            "ssp": trader_ssp_input,
            "channel": "Display",
            "recommended_bid": "$5",
            "status": "created",
        },
    ]
    res = await build_deal_sheet(deals=deals, client_name="SSP Casing Test", theme="elcano")
    assert res["success"] is True, (
        f"input {trader_ssp_input!r} should normalize to a supported SSP, got error: {res.get('error')}"
    )


@pytest.mark.parametrize(
    "trader_ssp_input,expected_canonical",
    [
        ("Media.net", "medianet"),
        ("media.net", "medianet"),
        ("MEDIA.NET", "medianet"),
        ("Media Net", "medianet"),
    ],
)
@pytest.mark.asyncio
async def test_build_deal_sheet_normalizes_punctuated_ssps(trader_ssp_input, expected_canonical):
    """Same normalization works for SSPs with punctuation (Media.net),
    not just whitespace. `_normalize_ssp` strips all non-alphanumerics."""
    deals = [
        {
            "partner": "p",
            "deal_name": "mn-test",
            "deal_id": "MN-1",
            "ssp": trader_ssp_input,
            "channel": "Display",
            "recommended_bid": "$5",
            "status": "created",
        },
    ]
    res = await build_deal_sheet(deals=deals, client_name="Punctuation Test", theme="elcano")
    assert res["success"] is True, (
        f"{trader_ssp_input!r} should normalize to {expected_canonical!r}, got: {res.get('error')}"
    )


@pytest.mark.asyncio
async def test_validate_brief_accepts_human_friendly_ssp_spellings():
    """validate_brief's pre-audit gate uses the same normalization so the
    agent doesn't see false-positive 'unsupported SSP' errors at brief
    validation time for human-friendly trader input."""
    deals = [
        {"deal_name": "d1", "ssp": "Index Exchange", "channel": "CTV", "recommended_bid": "$5"},
        {"deal_name": "d2", "ssp": "Media.net", "channel": "OLV", "recommended_bid": "$5"},
        {"deal_name": "d3", "ssp": "OpenX", "channel": "Display", "recommended_bid": "$5"},
    ]
    res = await validate_brief(deals=deals)
    assert res["ok"] is True, f"expected human-friendly SSP names to validate, got issues: {res['issues']}"


# ── validate_brief ─────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_validate_brief_happy_path():
    deals = [
        {"partner": "Soundwave", "deal_name": "d1", "ssp": "openx", "channel": "CTV", "recommended_bid": "$27-$32"},
        {
            "partner": "Soundwave",
            "deal_name": "d2",
            "ssp": "indexexchange",
            "channel": "OLV",
            "recommended_bid": "$8-$12",
        },
    ]
    res = await validate_brief(deals=deals)
    assert res["ok"] is True
    assert res["issues"] == []
    assert res["deal_count"] == 2


@pytest.mark.asyncio
async def test_validate_brief_accepts_magnite_rejects_unknown_ssp():
    """Magnite is supported as of June 2026 (ClearLine Demand Management API);
    a genuinely unknown SSP must still be rejected with a per-deal issue."""
    deals = [
        {"partner": "p", "deal_name": "d1", "ssp": "magnite", "channel": "CTV", "recommended_bid": "$5"},
        {"partner": "p", "deal_name": "d2", "ssp": "sharethrough", "channel": "CTV", "recommended_bid": "$5"},
    ]
    res = await validate_brief(deals=deals)
    assert res["ok"] is False
    assert not any(i["deal_index"] == 0 and i["field"] == "ssp" for i in res["issues"])
    unknown_issue = next((i for i in res["issues"] if i["deal_index"] == 1 and i["field"] == "ssp"), None)
    assert unknown_issue is not None
    assert "sharethrough" in unknown_issue["message"].lower()


@pytest.mark.asyncio
async def test_validate_brief_reports_missing_required_fields():
    deals = [
        {"deal_name": "d1", "ssp": "openx", "channel": "CTV"},  # missing recommended_bid
        {"deal_name": "d2", "ssp": "openx", "recommended_bid": "$5"},  # missing channel
        {},  # missing everything
    ]
    res = await validate_brief(deals=deals)
    assert res["ok"] is False
    fields_missing = {(i["deal_index"], i["field"]) for i in res["issues"]}
    assert (0, "recommended_bid") in fields_missing
    assert (1, "channel") in fields_missing
    assert (2, "ssp") in fields_missing
    assert (2, "deal_name") in fields_missing


@pytest.mark.asyncio
async def test_validate_brief_rejects_empty_brief():
    res = await validate_brief(deals=[])
    assert res["ok"] is False
    assert res["deal_count"] == 0


@pytest.mark.asyncio
async def test_validate_brief_handles_non_dict_entry():
    res = await validate_brief(deals=["not a dict"])
    assert res["ok"] is False
    assert any("not an object" in i["message"] for i in res["issues"])


# ── missing-channel warning ────────────────────────────────────────────


@pytest.mark.asyncio
async def test_group_deals_warns_on_missing_channel(caplog):
    """Empty channel should still render (under 'Unknown') but log a warning
    so operators can spot data drift in the live log."""
    deals = [
        {
            "partner": "p",
            "deal_name": "no-channel",
            "deal_id": "OX-1",
            "ssp": "openx",
            "channel": "",  # empty
            "recommended_bid": "$5",
        }
    ]
    with caplog.at_level(logging.WARNING, logger="deal_sheet_server"):
        res = await build_deal_sheet(deals=deals, client_name="No Channel", theme="elcano")
    assert res["success"] is True
    assert any("empty/missing channel" in rec.message for rec in caplog.records)


def test_supported_ssps_exports_expected_set():
    """Sanity: the exported SUPPORTED_SSPS frozenset stays in sync with the
    multi-deal protocol's documented scope. If a new SSP is added (or an
    existing one removed) this test makes the change visible immediately."""
    assert "openx" in SUPPORTED_SSPS
    assert "indexexchange" in SUPPORTED_SSPS
    assert "pubmatic" in SUPPORTED_SSPS
    assert "xandr" in SUPPORTED_SSPS
    assert "triplelift" in SUPPORTED_SSPS
    assert "medianet" in SUPPORTED_SSPS
    assert "magnite" in SUPPORTED_SSPS  # Supported since June 2026 (ClearLine Demand Management API).


@pytest.mark.asyncio
async def test_no_trailing_empty_rows_for_short_batches():
    """The XLSX template ships with 21 rows of scaffolding sized for ~8
    deals; a 2-deal batch must not leave the unused tail rows hanging
    below the footer. The footer should be the final used row."""
    deals = [
        {
            "partner": "p",
            "deal_name": "a",
            "deal_id": "OX-1",
            "ssp": "openx",
            "channel": "CTV",
            "recommended_bid": "$5",
        },
        {
            "partner": "p",
            "deal_name": "b",
            "deal_id": "OX-2",
            "ssp": "openx",
            "channel": "CTV",
            "recommended_bid": "$5",
        },
    ]
    res = await build_deal_sheet(deals=deals, client_name="Short Batch", theme="elcano")
    wb = openpyxl.load_workbook(res["path"])
    ws = wb.active
    # Header (10) + CTV banner (11) + 2 deals (12-13) + footer (14) = max_row 14
    assert ws.max_row == 14, (
        f"expected max_row=14 for a 2-deal sheet (header+banner+2 deals+footer); "
        f"got max_row={ws.max_row} — trailing template rows weren't cleaned up"
    )


@pytest.mark.asyncio
async def test_no_trailing_rows_for_multi_channel_batches():
    """Same trailing-row cleanup must work when the batch spans multiple
    channel groups (more banners + more rows). A 3-channel batch uses
    more rows but still shouldn't leak template scaffolding below."""
    deals = [
        {
            "partner": "p",
            "deal_name": "ctv1",
            "deal_id": "OX-1",
            "ssp": "openx",
            "channel": "CTV",
            "recommended_bid": "$5",
        },
        {
            "partner": "p",
            "deal_name": "olv1",
            "deal_id": "OX-2",
            "ssp": "openx",
            "channel": "OLV",
            "recommended_bid": "$5",
        },
        {
            "partner": "p",
            "deal_name": "dsp1",
            "deal_id": "OX-3",
            "ssp": "openx",
            "channel": "Display",
            "recommended_bid": "$5",
        },
    ]
    res = await build_deal_sheet(deals=deals, client_name="Multi Channel", theme="elcano")
    wb = openpyxl.load_workbook(res["path"])
    ws = wb.active
    # Header (10) + CTV banner + ctv1 + OLV banner + olv1 + Display banner + dsp1 + footer = 17
    assert ws.max_row == 17, (
        f"expected max_row=17 for a 3-channel batch (header + 3 banners + 3 deals + footer); got max_row={ws.max_row}"
    )
