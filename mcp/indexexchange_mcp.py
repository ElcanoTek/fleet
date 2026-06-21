#!/usr/bin/env python3
"""
Index Exchange MCP Server

A Model Context Protocol (MCP) server for Index Exchange REST APIs covering:
- Authentication (login + refresh with automatic token rotation)
- Deals v3 (list DSPs, list DSP seats, create/list/get marketplace deals,
  marketplace publishers, segments, domain targeting values)
- Accounts (list account information)
- Reporting (connectivity check, create/list/run/download report specs and files)

Runs within the Victoria Terminal container environment.
"""

import asyncio
import base64
import binascii
import copy
import csv
import hashlib
import json
import logging
import os
import re
import sys
import time
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import httpx
from mcp.server.fastmcp import FastMCP
from utils.deal_id_generator import generate_external_deal_id

_HTTP_RE = re.compile(r"^https?://")
_NON_ALNUM_RE = re.compile(r"[^a-z0-9]+")
_SPACE_RE = re.compile(r"\s+")

_QUOTE_RE = re.compile(r"['\"`]")
_EXTERNAL_DEAL_ID_RE = re.compile(r"^[A-Za-z0-9\-_\.]+$")
_SUB_DOMAIN_RE = re.compile(r"^[a-z0-9._-]+$")


# Configure logging to stderr (not stdout for STDIO transport)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    stream=sys.stderr,
)
logger = logging.getLogger(__name__)

# Initialize FastMCP server
mcp = FastMCP("indexexchange_mcp")

# Constants
USER_AGENT = "victoria-terminal/1.0"
DEFAULT_BASE_URL = "https://app.indexexchange.com"
DEFAULT_TIMEOUT = 60.0
DEFAULT_DOWNLOAD_DIR = os.path.expanduser("~/Victoria/indexexchange_reports")
TOKEN_EXPIRY_BUFFER_SECONDS = 60
MAX_INLINE_DOMAIN_VALUES = 300
DIRECT_TARGET_VERIFICATION_RETRY_DELAYS_SECONDS = (2, 4, 8, 16)
DEFAULT_REPORT_FILE_POLL_INTERVAL_SECONDS = 2.0
# Real Marketplace draft reports spend 30-120s in the IX queue before the file
# is downloadable; 60s was hair-trigger and typically returned download_pending,
# prompting agents to loop on ix_list_report_files manually (burns LLM tokens
# without speeding up the IX API). 5 minutes covers every report observed in
# end-to-end testing. Tools accept poll_timeout_seconds as a per-call override.
DEFAULT_REPORT_FILE_POLL_TIMEOUT_SECONDS = 300.0

DOMAIN_PATTERN = re.compile(
    r"^(?!-)"
    r"(?:[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?\.)"
    r"+[a-zA-Z]{2,}$"
)

# App-bundle identifiers are NOT web domains: CTV/OTT inventory lists carry
# either reverse-DNS bundle IDs (com.zumobi.msnbc — these also satisfy
# DOMAIN_PATTERN) OR bare numeric store IDs (Roku channel IDs like 523428113,
# Apple/Amazon numeric app IDs). DOMAIN_PATTERN rejects the numeric form and
# any dotted bundle whose final label is numeric (com.example.app2), which is
# why a domain-only validator silently dropped the bulk of bundle lists. This
# pattern accepts bare numeric IDs and dotted identifiers with numeric labels;
# bare single-token alphanumerics (likely a wrong-column / app-name pick) are
# still rejected. Applied only in app-bundle mode (allow_app_bundle_ids=True).
APP_BUNDLE_ID_PATTERN = re.compile(r"^(?:[0-9]+|[a-z0-9][a-z0-9._-]*\.[a-z0-9._-]+)$")

IX_DSP_NAME_ALIASES: dict[str, str] = {
    "tradr": "bidswitch",
    "bidswitch": "bidswitch",
    "bidswitchrtb": "bidswitch",
    "ttd": "the trade desk",
    "dv360": "display and video 360",
}

IX_TARGETING_KEY_ALIASES: dict[str, tuple[str, ...]] = {
    "country": ("country",),
    "state": ("state", "region state", "province"),
    "device_type": ("devicetype", "device type"),
    "viewability": ("viewability",),
    "iab_categories": ("contentgenre", "content genre", "iab category", "iab categories"),
    "segment": ("segment", "audience segment", "segments"),
}

IX_TARGETING_VALUE_ALIASES: dict[str, dict[str, str]] = {
    "country": {
        "us": "USA",
        "usa": "USA",
        "unitedstates": "USA",
        "unitedstatesofamerica": "USA",
        "ca": "Canada",
        "can": "Canada",
    },
    "device_type": {
        "desktop": "Desktop",
        "mobile": "Mobile",
        "phone": "Mobile",
        "tablet": "Tablet",
        "ctv": "Connected TV",
        "connectedtv": "Connected TV",
        "connecteddevice": "Connected TV",
    },
}

IX_MARKETPLACE_REPORT_PRESETS: dict[str, dict[str, Any]] = {
    "deal_summary": {
        "description": "Daily Marketplace package activity with bid-cycle, delivery, and Marketplace fee metrics.",
        "dimensions": [
            "day",
            "deal_id",
            "deal_name",
            "partner_name",
            "dsp_name",
            "device_type",
            "inventory_channel",
            "creative_type",
        ],
        "measures": [
            "bid_request",
            "bid_error",
            "bid_timeout",
            "bid_pass",
            "bid_blocked",
            "bid_below_floor",
            "bid_in_review",
            "bid_eligible",
            "bid_chosen",
            "total_bids",
            "competing_bids",
            "impressions",
            "marketplace_media_spend",
            "marketplace_total_fee",
        ],
    },
    "deal_labels": {
        "description": "Marketplace deal label reporting for advertiser, agency, sales, and external IDs.",
        "dimensions": [
            "day",
            "deal_id",
            "deal_name",
            "advertiser",
            "agency",
            "salesperson",
            "external_reference_id",
            "dsp_id_level_1",
            "dsp_id_level_2",
            "custom",
        ],
        "measures": ["impressions", "marketplace_media_spend", "marketplace_total_fee"],
    },
    "supply_breakdown": {
        "description": "Marketplace performance by publisher, site, domain, geography, and device.",
        "dimensions": [
            "day",
            "partner_id",
            "partner_name",
            "site_id",
            "site_name",
            "domain",
            "country",
            "device_type",
            "inventory_channel",
        ],
        "measures": ["bid_request", "impressions", "marketplace_media_spend", "publisher_payment"],
    },
    "segment_performance": {
        "description": "Marketplace segment reporting including provider and billed-state dimensions.",
        "dimensions": [
            "day",
            "deal_id",
            "deal_name",
            "partner_name",
            "index_segment_id",
            "segment_id",
            "segment_name",
            "segment_owner_id",
            "segment_owner_name",
            "provider_name",
            "is_billed_segment",
            "is_included_targeting_segment",
        ],
        "measures": ["impressions", "marketplace_media_spend", "publisher_payment"],
    },
}

# Deal-type targeting defaults. Values are the create-accepted value tokens that
# show up in the Index UI for each field. Used by _ensure_deal_type_targeting_defaults
# to auto-fill device type, inventory channel, and creative-format targeting per
# the Elcano trader spec (canonical across all SSPs):
#
#   Deal type | Devices                                | Inventory   | Creative format
#   ----------|----------------------------------------|-------------|-----------------
#   display   | Personal computer + Phone + Tablet     | In-App+Web  | Banner_ANY
#   olv       | Personal computer + Phone + Tablet     | In-App+Web  | Video_ANY
#   ctv       | Connected TV + Connected device + STB  | In-App ONLY | Video_ANY
#   ott       | Phone + Tablet                         | In-App ONLY | Video_ANY
#
# Earlier revisions collapsed `display` and `olv` into one bucket and emitted
# Banner_ANY for both — that's what made the trader's OLV deals come through as
# Display in the Index UI. The split below routes OLV to Video_ANY without
# disturbing the established Display behaviour.
#
# IX device IDs (from targetingKey "Device"):
#   2 = Personal computer    3 = Connected TV       4 = Phone
#   5 = Tablet               6 = Connected device   7 = Set-top box
IX_DEVICE_VALUES_DISPLAY: tuple[str, ...] = ("2", "4", "5")  # Personal computer, Phone, Tablet
IX_DEVICE_VALUES_OLV: tuple[str, ...] = ("2", "4", "5")  # same as Display per trader spec
IX_DEVICE_VALUES_CTV: tuple[str, ...] = ("3", "6", "7")  # Connected TV, Connected device, Set-top box
IX_DEVICE_VALUES_OTT: tuple[str, ...] = ("4", "5")  # Phone, Tablet (in-app mobile video)
IX_INVENTORY_CHANNEL_VALUES_DEFAULT: tuple[str, ...] = ("App", "Site")  # In-App + Web
IX_INVENTORY_CHANNEL_VALUES_APP_ONLY: tuple[str, ...] = ("App",)  # In-App only (CTV, OTT)
IX_CREATIVE_TYPE_SIZE_VIDEO_ANY: str = "Video_ANY"  # Video (all sizes)
IX_CREATIVE_TYPE_SIZE_BANNER_ANY: str = "Banner_ANY"  # Banner (all sizes)
IX_VIDEO_ALL_LEGACY_VALUE_ID: str = "1100"  # legacy numeric targetingValueID for Video_ANY
IX_DEAL_TYPES_DISPLAY: frozenset[str] = frozenset({"display"})
IX_DEAL_TYPES_OLV: frozenset[str] = frozenset({"olv", "display/olv", "display_olv"})
IX_DEAL_TYPES_CTV: frozenset[str] = frozenset({"ctv"})
IX_DEAL_TYPES_OTT: frozenset[str] = frozenset({"ott"})
IX_DEFAULT_END_DATE_MONTHS: int = 24

# Elcano curator-margin default. Index Exchange exposes the curator margin
# via the `margin` field on the marketplace-package deal payload with
# `margin_calculation_type="P"` (Percentage of winning bid). When the caller
# omits margin_percent on the prompt-input flow we default to a flat 30%
# to match the Elcano-curated standard already shipped on OpenX/PubMatic/
# Media.net. Override per-deal by passing an explicit margin_percent.
ELCANO_DEFAULT_CURATOR_MARGIN_PERCENT: float = 30.0


# Index Exchange exposes a single login that can act on behalf of multiple
# Marketplace accounts (Reklaim, Elcano, Permutive, etc.). The names below
# resolve human-readable references to the numeric accountID expected by the
# IX REST APIs. To target a non-default account from a single MCP variant
# subprocess, set `INDEXEXCHANGE_MARKETPLACE_ACCOUNT_ID` (or the suffixed
# variant `INDEXEXCHANGE_MARKETPLACE_ACCOUNT_ID_<CLIENT>` rewritten by the
# Cutlass mcp loader) — see DEFAULT_MARKETPLACE_ACCOUNT_ID below.
IX_MARKETPLACE_ACCOUNT_IDS: dict[str, int] = {
    "reklaim": 1485234,
    "permutive": 1490424,
    "elcano": 1491166,
    "the weather company, llc": 1499155,
    "the weather company": 1499155,
    "twc": 1499155,
    "raptive": 1502939,
    "stirista": 1503605,
    "zeta global": 1507580,
    "zeta": 1507580,
}


# Default Marketplace account ID for the current MCP variant. Used when the
# caller omits `account_id` on entry-point tools like
# `ix_execute_deal_from_prompt_inputs`. Defaults to Elcano (1491166).
#
# The Cutlass mcp loader resolves variant subprocesses by rewriting suffixed
# env vars: setting `INDEXEXCHANGE_MARKETPLACE_ACCOUNT_ID_REKLAIM=1485234`
# causes the indexexchange_mcp_reklaim subprocess to see
# `INDEXEXCHANGE_MARKETPLACE_ACCOUNT_ID=1485234`. This keeps the same
# one-login-multi-account UX we ship for PubMatic.
_ELCANO_MARKETPLACE_ACCOUNT_ID: int = 1491166
try:
    DEFAULT_MARKETPLACE_ACCOUNT_ID: int = int(
        (os.environ.get("INDEXEXCHANGE_MARKETPLACE_ACCOUNT_ID") or "").strip() or _ELCANO_MARKETPLACE_ACCOUNT_ID
    )
except ValueError:
    logger.warning(
        "INDEXEXCHANGE_MARKETPLACE_ACCOUNT_ID is not an integer; falling back to Elcano default %d",
        _ELCANO_MARKETPLACE_ACCOUNT_ID,
    )
    DEFAULT_MARKETPLACE_ACCOUNT_ID = _ELCANO_MARKETPLACE_ACCOUNT_ID


# Retry safety lock for file-driven domain parity.
# Key: (account_id, external_deal_id)
# Value: (expected_domain_count, expected_domains_fingerprint_lower)
_FILE_DOMAIN_EXPECTATION_LOCKS: dict[tuple[int, str], tuple[int, str]] = {}


def _decode_jwt_exp(token: str) -> float | None:
    """Decode JWT expiry from payload without verification.

    Returns the 'exp' claim as a unix timestamp, or None if unparseable.
    """
    try:
        parts = token.split(".")
        if len(parts) < 2:
            return None
        payload_b64 = parts[1]
        # Add padding
        padding = 4 - len(payload_b64) % 4
        if padding != 4:
            payload_b64 += "=" * padding
        payload_bytes = base64.urlsafe_b64decode(payload_b64)
        payload = json.loads(payload_bytes)
        exp = payload.get("exp")
        if isinstance(exp, (int, float)):
            return float(exp)
        return None
    except (binascii.Error, json.JSONDecodeError, ValueError):
        return None
    except Exception as e:
        logger.debug("Failed to decode JWT: %s", e)
        return None


def _make_error(operation: str, status_code: int | None, message: str, details: Any = None) -> dict[str, Any]:
    """Create a normalized error dict."""
    err: dict[str, Any] = {
        "provider": "indexexchange",
        "operation": operation,
        "status_code": status_code,
        "message": message,
    }
    if details is not None:
        err["details"] = details
    return err


def _sanitize_label_value(value: str) -> str:
    sanitized = value.replace("&", " and ")
    sanitized = _QUOTE_RE.sub("", sanitized)
    return _SPACE_RE.sub(" ", sanitized).strip()


def _sanitize_labels(labels: dict[str, Any]) -> dict[str, Any]:
    sanitized_labels: dict[str, Any] = {}
    for key, value in labels.items():
        sanitized_labels[key] = _sanitize_label_value(value) if isinstance(value, str) else value
    return sanitized_labels


def _normalize_lookup_text(value: Any) -> str:
    return _NON_ALNUM_RE.sub("", str(value).strip().lower())


def _normalize_domain_candidate(value: Any) -> str | None:
    raw_value = str(value or "").strip().lower()
    if not raw_value:
        return None

    normalized_value = _HTTP_RE.sub("", raw_value)
    normalized_value = normalized_value.split("/")[0].split("?")[0].split(":")[0].strip()
    normalized_value = normalized_value.lstrip(".")
    if normalized_value.startswith("www."):
        normalized_value = normalized_value[4:]
    return normalized_value or None


def _extract_tool_error_message(result: dict[str, Any], fallback: str) -> str:
    error = result.get("error")
    if isinstance(error, dict):
        message = error.get("message")
        if isinstance(message, str) and message.strip():
            return message
    if isinstance(error, str) and error.strip():
        return error
    return fallback


def _unique_fields(fields: list[str]) -> list[str]:
    ordered_fields: list[str] = []
    seen: set[str] = set()
    for field in fields:
        normalized_field = str(field).strip()
        if not normalized_field or normalized_field in seen:
            continue
        seen.add(normalized_field)
        ordered_fields.append(normalized_field)
    return ordered_fields


def _resolve_marketplace_account_id(account: int | str) -> int:
    if isinstance(account, int):
        return account
    normalized_account = str(account).strip()
    if not normalized_account:
        raise ValueError("Marketplace account reference cannot be empty")
    if normalized_account.isdigit():
        return int(normalized_account)

    lookup_key = normalized_account.lower()
    resolved_account_id = IX_MARKETPLACE_ACCOUNT_IDS.get(lookup_key)
    if resolved_account_id is not None:
        return resolved_account_id

    raise ValueError(
        "Unknown Marketplace account name. Use one of: "
        + ", ".join(
            [
                "Reklaim",
                "Permutive",
                "Elcano",
                "The Weather Company, LLC",
                "Raptive",
                "Stirista",
                "Zeta Global",
            ]
        )
    )


def _extract_report_files(data: Any) -> list[dict[str, Any]]:
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    if isinstance(data, dict):
        for key in ("files", "reportFiles", "data"):
            value = data.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
    return []


def _extract_report_file_id(file_info: dict[str, Any]) -> str | None:
    for key in ("fileID", "fileId", "id", "reportRunID"):
        value = file_info.get(key)
        if isinstance(value, (str, int)):
            normalized_value = str(value).strip()
            if normalized_value:
                return normalized_value
    return None


def _report_file_ready(file_info: dict[str, Any]) -> bool:
    status = str(file_info.get("status", "")).strip().lower()
    if status in {"completed", "complete", "ready", "available"}:
        return True

    download_status = str(file_info.get("downloadStatus", "")).strip().lower()
    return download_status in {"new", "completed", "complete", "ready", "available"}


def _build_targeting_object(
    *, targeting_key_id: int, key_name: str, values: list[str], operator: str = "ANY_OF"
) -> dict[str, Any]:
    return {
        "targetingKeyID": targeting_key_id,
        "keyName": key_name,
        "targetingType": "standard",
        "sets": [{"operator": operator, "values": [{"value": value} for value in values]}],
    }


def _extract_domains_from_csv(
    file_path: str, column_name: str | None = None, allow_app_bundle_ids: bool = False
) -> dict[str, Any]:
    with open(file_path, newline="", encoding="utf-8-sig") as csv_file:
        rows = list(csv.reader(csv_file))

    if not rows:
        raise ValueError("CSV file is empty.")

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized_headers = {_normalize_lookup_text(header): index for index, header in enumerate(header_row) if header}

    if column_name:
        requested_column_key = _normalize_lookup_text(column_name)
        if requested_column_key not in normalized_headers:
            raise ValueError(f"Column not found in CSV: {column_name}")
        domain_column_index = normalized_headers[requested_column_key]
    else:
        preferred_columns = ["sites", "domains", "domain", "url", "urls"]
        domain_column_index = next(
            (normalized_headers[column_key] for column_key in preferred_columns if column_key in normalized_headers),
            0,
        )

    extracted_domains: list[str] = []
    invalid_values: list[str] = []
    for row in rows[1:]:
        if domain_column_index >= len(row):
            continue
        cell = row[domain_column_index]
        if allow_app_bundle_ids and isinstance(cell, float) and cell.is_integer():
            cell = int(cell)
        normalized_domain = _normalize_domain_candidate(cell)
        if normalized_domain is None:
            continue
        if DOMAIN_PATTERN.match(normalized_domain) or (
            allow_app_bundle_ids and APP_BUNDLE_ID_PATTERN.match(normalized_domain)
        ):
            extracted_domains.append(normalized_domain)
        else:
            invalid_values.append(str(cell).strip())

    return {
        "sheet_name": None,
        "column_name": header_row[domain_column_index] if header_row else None,
        "row_count": max(len(rows) - 1, 0),
        "domains": sorted(dict.fromkeys(extracted_domains)),
        "invalid_values": invalid_values,
    }


def _extract_domains_from_xlsx(
    file_path: str,
    sheet_name: str | None = None,
    column_name: str | None = None,
    allow_app_bundle_ids: bool = False,
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read Excel domain files.") from exc

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found in workbook: {sheet_name}")
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Workbook sheet is empty.")

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized_headers = {_normalize_lookup_text(header): index for index, header in enumerate(header_row) if header}

    if column_name:
        requested_column_key = _normalize_lookup_text(column_name)
        if requested_column_key not in normalized_headers:
            raise ValueError(f"Column not found in workbook: {column_name}")
        domain_column_index = normalized_headers[requested_column_key]
    else:
        preferred_columns = ["sites", "domains", "domain", "url", "urls"]
        domain_column_index = next(
            (normalized_headers[column_key] for column_key in preferred_columns if column_key in normalized_headers),
            0,
        )

    extracted_domains: list[str] = []
    invalid_values: list[str] = []
    for row in rows[1:]:
        if domain_column_index >= len(row):
            continue
        cell = row[domain_column_index]
        if allow_app_bundle_ids and isinstance(cell, float) and cell.is_integer():
            cell = int(cell)
        normalized_domain = _normalize_domain_candidate(cell)
        if normalized_domain is None:
            continue
        if DOMAIN_PATTERN.match(normalized_domain) or (
            allow_app_bundle_ids and APP_BUNDLE_ID_PATTERN.match(normalized_domain)
        ):
            extracted_domains.append(normalized_domain)
        else:
            invalid_values.append(str(cell).strip())

    return {
        "sheet_name": worksheet.title,
        "column_name": header_row[domain_column_index] if header_row else None,
        "row_count": max(len(rows) - 1, 0),
        "domains": sorted(dict.fromkeys(extracted_domains)),
        "invalid_values": invalid_values,
    }


def _extract_domains_from_file(
    file_path: str,
    sheet_name: str | None = None,
    column_name: str | None = None,
    allow_app_bundle_ids: bool = False,
) -> dict[str, Any]:
    path = Path(file_path).expanduser()
    path = (Path.cwd() / path).resolve() if not path.is_absolute() else path.resolve()

    if not path.exists() or not path.is_file():
        raise ValueError(f"Domain file does not exist or is not a file: {path}")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _extract_domains_from_csv(str(path), column_name=column_name, allow_app_bundle_ids=allow_app_bundle_ids)
    if suffix in {".xlsx", ".xlsm"}:
        return _extract_domains_from_xlsx(
            str(path), sheet_name=sheet_name, column_name=column_name, allow_app_bundle_ids=allow_app_bundle_ids
        )

    raise ValueError(f"Unsupported domain file format: {path.suffix or '<none>'}")


def _candidate_lookup_values(item: dict[str, Any], fields: tuple[str, ...]) -> list[str]:
    values: list[str] = []
    for field in fields:
        raw_value = item.get(field)
        if raw_value is None:
            continue
        text = str(raw_value).strip()
        if text:
            values.append(text)
    return values


def _resolve_unique_match(
    items: list[dict[str, Any]],
    requested_value: str,
    field_label: str,
    *,
    lookup_fields: tuple[str, ...],
    aliases: dict[str, str] | None = None,
    allow_contains_match: bool = True,
) -> dict[str, Any]:
    requested_text = str(requested_value).strip()
    normalized_requested = _normalize_lookup_text(requested_text)
    if aliases and normalized_requested in aliases:
        requested_text = aliases[normalized_requested]
        normalized_requested = _normalize_lookup_text(requested_text)

    exact_matches = [
        item
        for item in items
        if any(
            _normalize_lookup_text(candidate) == normalized_requested
            for candidate in _candidate_lookup_values(item, lookup_fields)
        )
    ]
    if len(exact_matches) == 1:
        return exact_matches[0]
    if len(exact_matches) > 1:
        raise ValueError(f"Ambiguous {field_label}: {requested_value}")

    if allow_contains_match:
        contains_matches = [
            item
            for item in items
            if any(
                normalized_requested in _normalize_lookup_text(candidate)
                or _normalize_lookup_text(candidate) in normalized_requested
                for candidate in _candidate_lookup_values(item, lookup_fields)
            )
        ]
        if len(contains_matches) == 1:
            return contains_matches[0]
        if len(contains_matches) > 1:
            raise ValueError(f"Ambiguous {field_label}: {requested_value}")

    raise LookupError(f"No match found for {field_label}: {requested_value}")


_PUBLISHER_LOOKUP_FIELDS: tuple[str, ...] = (
    "name",
    "publisherName",
    "accountName",
    "companyName",
)
_PUBLISHER_ID_LOOKUP_FIELDS: tuple[str, ...] = (
    "legacyAccountID",
    "accountID",
    "id",
)


def _find_publisher_matches(publishers: list[dict[str, Any]], requested: str) -> dict[str, Any]:
    """Return all publisher candidates matching `requested` plus the match strategy used.

    The IX publisher catalog stores names with delivery-route suffixes
    (e.g. "The Weather Company via Prebid" / "The Weather Company via
    OB"). A bare client name like "The Weather Company" typically maps
    to multiple feed variants, and the right combination per deal is a
    policy decision that belongs in the brief — not in the resolver.

    This helper surfaces ALL candidates so the caller (or operator) can
    decide. Use `_resolve_unique_match` instead when you need a hard
    single-match guarantee (e.g. seat lookup, where ambiguity is a bug).
    """
    requested_text = str(requested).strip()
    normalized = _normalize_lookup_text(requested_text)

    # 1) Exact match on any human-readable name field, or any ID field.
    exact = [
        p
        for p in publishers
        if any(_normalize_lookup_text(p.get(f)) == normalized for f in _PUBLISHER_LOOKUP_FIELDS)
        or any(_normalize_lookup_text(p.get(f)) == normalized for f in _PUBLISHER_ID_LOOKUP_FIELDS)
    ]
    if exact:
        return {"matches": exact, "strategy": "exact_name_or_id"}

    # 2) Substring on name fields.
    substring = [
        p
        for p in publishers
        if any(normalized in _normalize_lookup_text(p.get(f)) for f in _PUBLISHER_LOOKUP_FIELDS if p.get(f) is not None)
    ]
    if substring:
        return {"matches": substring, "strategy": "substring_name"}

    return {"matches": [], "strategy": "no_match"}


def _resolve_targeting_key(keys: list[dict[str, Any]], logical_name: str) -> dict[str, Any]:
    target_aliases = IX_TARGETING_KEY_ALIASES.get(logical_name, (logical_name,))
    aliases = dict.fromkeys(target_aliases, target_aliases[0])
    aliases[_normalize_lookup_text(logical_name)] = target_aliases[0]
    return _resolve_unique_match(
        keys,
        logical_name,
        f"targeting key {logical_name}",
        lookup_fields=("key", "keyName", "name", "targetingKeyID", "keyID"),
        aliases=aliases,
        allow_contains_match=True,
    )


async def _fetch_targeting_values_for_key(
    account_id: int, key_id: int, lookup_value: str | None = None
) -> list[dict[str, Any]]:
    values_result = await ix_list_targeting_values(account_id=account_id, key_id=key_id, value=lookup_value)
    if not values_result.get("success"):
        raise ValueError(
            _extract_tool_error_message(values_result, f"Failed to resolve targeting values for key {key_id}")
        )

    targeting_values = values_result.get("targeting_values")
    if not isinstance(targeting_values, list):
        raise ValueError(f"Unexpected targeting values response for key {key_id}")
    return [value for value in targeting_values if isinstance(value, dict)]


async def _resolve_targeting_value_token(
    account_id: int,
    key_id: int,
    requested_value: str,
    field_label: str,
    *,
    aliases: dict[str, str] | None = None,
) -> str:
    requested_text = str(requested_value).strip()
    normalized_requested = _normalize_lookup_text(requested_text)
    if aliases and normalized_requested in aliases:
        requested_text = aliases[normalized_requested]

    candidate_lists: list[list[dict[str, Any]]] = []
    if requested_text:
        candidate_lists.append(await _fetch_targeting_values_for_key(account_id, key_id, requested_text))
    candidate_lists.append(await _fetch_targeting_values_for_key(account_id, key_id))

    last_error: Exception | None = None
    for candidates in candidate_lists:
        if not candidates:
            continue
        if len(candidates) == 1:
            sole_candidate = candidates[0]
            value_token = sole_candidate.get("value")
            if isinstance(value_token, str) and value_token.strip():
                return value_token.strip()
            value_id = sole_candidate.get("targetingValueID")
            if value_id is None:
                value_id = sole_candidate.get("valueID")
            if value_id is not None:
                return str(value_id)
        normalized_requested = _normalize_lookup_text(requested_text)
        same_name_matches = [
            candidate
            for candidate in candidates
            if _normalize_lookup_text(
                candidate.get("name") or candidate.get("label") or candidate.get("displayName") or ""
            )
            == normalized_requested
        ]
        if len(same_name_matches) > 1:
            first_match = same_name_matches[0]
            value_token = first_match.get("value")
            if isinstance(value_token, str) and value_token.strip():
                return value_token.strip()
            value_id = first_match.get("targetingValueID")
            if value_id is None:
                value_id = first_match.get("valueID")
            if value_id is not None:
                return str(value_id)
        try:
            matched = _resolve_unique_match(
                candidates,
                requested_text,
                field_label,
                lookup_fields=("value", "name", "label", "displayName", "targetingValueID", "valueID"),
                aliases=aliases,
                allow_contains_match=True,
            )
            value_token = matched.get("value")
            if isinstance(value_token, str) and value_token.strip():
                return value_token.strip()

            value_id = matched.get("targetingValueID")
            if value_id is None:
                value_id = matched.get("valueID")
            if value_id is not None:
                return str(value_id)
        except (LookupError, ValueError) as exc:
            last_error = exc

    if last_error is not None:
        raise ValueError(str(last_error))
    raise ValueError(f"No targeting value found for {field_label}: {requested_value}")


def _dedupe_preserving_order(values: list[Any]) -> list[Any]:
    seen: set[Any] = set()
    deduped: list[Any] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        deduped.append(value)
    return deduped


def _parse_dwp_reference(reference: str) -> tuple[str | None, str]:
    normalized_reference = str(reference).strip()
    if not normalized_reference:
        raise ValueError("Deals with Publishers reference cannot be empty")

    if "•" in normalized_reference:
        left, right = [part.strip() for part in normalized_reference.split("•", 1)]
        if right:
            return (left or None), right

    return None, normalized_reference


async def _resolve_dwp_internal_deal_ids(
    account_id: int,
    references: list[str],
) -> tuple[list[int], list[dict[str, Any]], list[str]]:
    resolved_ids: list[int] = []
    resolution_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    allocated_ids_by_name: dict[str, set[int]] = {}

    for reference in references:
        external_deal_id_hint, name_hint = _parse_dwp_reference(reference)
        deals_result = await ix_list_deals_v3(
            class_ids=[5],
            targeted_marketplace_account_ids=[account_id],
            search=name_hint,
            page_size=50,
        )
        if not deals_result.get("success"):
            raise ValueError(
                _extract_tool_error_message(deals_result, f"Failed to search Deals with Publishers for {reference}")
            )

        deals = [deal for deal in deals_result.get("deals", []) if isinstance(deal, dict)]
        exact_name_matches = [
            deal for deal in deals if _normalize_lookup_text(deal.get("name")) == _normalize_lookup_text(name_hint)
        ]
        if external_deal_id_hint:
            exact_name_matches = [
                deal
                for deal in exact_name_matches
                if _normalize_lookup_text(deal.get("externalDealID")) == _normalize_lookup_text(external_deal_id_hint)
            ]

        if not exact_name_matches:
            raise ValueError(f"Could not resolve Deals with Publishers reference: {reference}")

        sorted_matches = sorted(
            exact_name_matches,
            key=lambda deal: int(deal.get("internalDealID", 0)) if str(deal.get("internalDealID", "")).isdigit() else 0,
        )
        allocated_for_name = allocated_ids_by_name.setdefault(_normalize_lookup_text(name_hint), set())
        selected_match = next(
            (
                deal
                for deal in sorted_matches
                if isinstance(deal.get("internalDealID"), int) and deal["internalDealID"] not in allocated_for_name
            ),
            None,
        )

        if selected_match is None:
            if len(sorted_matches) == 1:
                selected_match = sorted_matches[0]
                warnings.append(
                    f"Reused Deals with Publishers internalDealID {selected_match.get('internalDealID')} for repeated reference '{reference}'."
                )
            else:
                raise ValueError(
                    f"More repeated Deals with Publishers references were supplied than distinct live matches for '{reference}'."
                )

        internal_deal_id = selected_match.get("internalDealID")
        if not isinstance(internal_deal_id, int):
            raise ValueError(f"Resolved Deals with Publishers match is missing internalDealID: {reference}")

        allocated_for_name.add(internal_deal_id)
        resolved_ids.append(internal_deal_id)
        resolution_rows.append(
            {
                "reference": reference,
                "name": selected_match.get("name"),
                "external_deal_id": selected_match.get("externalDealID"),
                "internal_deal_id": internal_deal_id,
            }
        )

    return resolved_ids, resolution_rows, warnings


DIRECT_TARGET_VISIBILITY_FAILURE_FLAGS = {
    "internal_deal_id_visibility_failed",
    # `publisher_id_visibility_pending` (info-severity, renamed from
    # `publisher_id_visibility_failed`) is no longer treated as a strict
    # failure — it represents the IX read-back caching lag, not a write
    # error. Kept here as a comment so anyone re-adding it knows why.
}


def _quality_flags_require_strict_failure(
    quality_flags: list[dict[str, Any]],
    *,
    expected_domain_count: int | None,
    expected_domains_fingerprint: str | None,
) -> bool:
    if not quality_flags:
        return False

    for flag in quality_flags:
        # Informational flags (e.g. read-back caching lag) never require
        # strict failure — they describe expected behavior, not write errors.
        if flag.get("severity") == "info":
            continue
        flag_name = str(flag.get("flag", "")).strip()
        if flag_name in DIRECT_TARGET_VISIBILITY_FAILURE_FLAGS:
            continue
        return True

    return expected_domain_count is not None or expected_domains_fingerprint is not None


def _build_marketplace_deal_payload(
    *,
    account_id: int,
    name: str,
    external_deal_id: str,
    start_date: str,
    floor: float,
    dsp_id: int,
    end_date: str,
    auction_type: str = "first",
    open_market: bool | None = None,
    seat_ids: list[str] | None = None,
    publisher_ids: list[int] | None = None,
    internal_deal_ids: list[int] | None = None,
    margin: float | None = None,
    margin_calculation_type: str | None = None,
    targeting: list[dict] | None = None,
    labels: dict | None = None,
) -> dict:
    """Build the v3 POST /api/deals/v3/deals payload for a Marketplace Package (classID=4).

    This function is pure Python with no network I/O and is safe to unit-test directly.
    All validation that can be done without a network call is performed here and raises
    ValueError on failure so callers receive a clean error message.

    Args:
        account_id: The marketplace account ID (integer shown in the IX UI top-left).
        name: Deal name, 1-255 characters.
        external_deal_id: Unique external deal ID, 3-64 characters. May contain letters,
            numbers, dashes (-), underscores (_), and periods (.). Cannot start with "0"
            or contain spaces.
        start_date: Deal start date in YYYY-MM-DD format.
        floor: Minimum bid price. Must be >= 0.10 for Marketplace Packages (classID=4).
        dsp_id: DSP integer ID from GET /api/deals/v1/dsps?validForClassID=4.
        end_date: Required deal end date in YYYY-MM-DD format.
        auction_type: "first" (default) or "fixed".
        open_market: Whether the deal competes with open market bids. When omitted (None,
            the default) the field is not sent, so accounts that exclude it won't error.
        seat_ids: Optional list of DSP extended seat ID strings. Required for The Trade
            Desk, Google DV360, Xandr, and Quantcast.
        publisher_ids: Optional list of marketplace publisher legacyAccountIDs to target
            directly inside the Marketplace Package. Use
            ix_list_marketplace_publishers(account_id) to resolve them.
        internal_deal_ids: Optional list of classID=5 internalDealIDs to target inside
            the Marketplace Package. Use ix_list_deals_v3(..., class_ids=[5],
            targeted_marketplace_account_ids=[account_id]) to resolve them.
        margin: Optional deal-specific fee percentage (e.g. 5.5 for 5.5%).
        margin_calculation_type: "P" (percentage) or "C" (CPM), required when margin set.
        targeting: Optional list of v3 targeting objects. Each object must have the shape:
            {
              "targetingKeyID": 9,         # required numeric targeting key ID
              "keyName": "Country",          # case-insensitive targeting key name
              "targetingType": "standard",   # "standard" or "custom"
              "sets": [
                {
                  "operator": "ANY_OF",      # "ANY_OF" (inclusion) or "NONE_OF" (exclusion)
                  "values": [
                    {"value": "USA"}         # create payload value token as string
                  ]
                }
              ]
            }
            Values in create payload must be strings accepted by the create endpoint.
            For standard keys, this is typically the returned "value" token from
            ix_list_targeting_values. Numeric targetingValueID strings are accepted as
            compatibility input for standard keys and translated automatically at
            create time when possible.

            IMPORTANT: For Domain/AppBundle key 120, pass domain/app-bundle strings in
            create payload values. Do not pass numeric value IDs from
            ix_create_domain_targeting_values.

            Known targetingKeyID references:
              - 3: DeviceType
              - 8: Viewability
              - 9: Country
              - 10: creativeTypeSize (use for OLV/Video: value "Video (all sizes)")
              - 11: contentGenre (use for IAB Categories/Entertainment)
              - 120: Domain and app bundle (for package deals)

            Audience Segment targeting:
              Segments from ix_list_segments use a SEPARATE targeting mechanism.
              Use the segment's targetingKeyID and keyName as returned by
              ix_list_targeting_keys (commonly key "Segment" or similar).
              Pass the segment ID as a string value. Use operator "ANY_OF" for
              segment inclusion and "NONE_OF" for segment exclusion.
              Example segment inclusion:
                {"targetingKeyID": <segment_key_id>, "keyName": "Segment",
                 "targetingType": "standard",
                 "sets": [{"operator": "ANY_OF",
                           "values": [{"value": "280"}, {"value": "3007"}]}]}
              Example segment exclusion:
                {"targetingKeyID": <segment_key_id>, "keyName": "Segment",
                 "targetingType": "standard",
                 "sets": [{"operator": "NONE_OF",
                           "values": [{"value": "308129"}]}]}
        labels: Optional dict with any of: advertiser, agency, custom,
            externalReferenceID, salesperson (all strings or null). Use "custom"
            for custom metadata labels needed for downstream reporting.

    Returns:
        A dict ready to be JSON-serialised and sent as the request body.

    Raises:
        ValueError: If any validation constraint is violated.
    """
    # --- name ---
    if not name or not (1 <= len(name) <= 255):
        raise ValueError(f"name must be 1\u2013255 characters, got {len(name)!r}")

    # --- externalDealID ---
    if not (3 <= len(external_deal_id) <= 64):
        raise ValueError(f"externalDealID must be 3\u201364 characters, got {len(external_deal_id)!r}")
    if " " in external_deal_id:
        raise ValueError("externalDealID cannot contain spaces")
    if external_deal_id.startswith("0"):
        raise ValueError("externalDealID cannot start with '0'")
    if not _EXTERNAL_DEAL_ID_RE.match(external_deal_id):
        raise ValueError("externalDealID may only contain letters, numbers, dashes, underscores, and periods")

    # --- floor ---
    if floor < 0.10:
        raise ValueError(f"floor must be >= 0.10 for Marketplace Packages (classID=4), got {floor}")
    if floor > _DEAL_FLOOR_MAXIMUM:
        raise ValueError(f"floor must be <= {_DEAL_FLOOR_MAXIMUM}, got {floor}")

    # --- end_date ---
    if not end_date:
        raise ValueError("end_date is required and must be a YYYY-MM-DD string")

    # --- auction_type ---
    if auction_type not in ("first", "fixed"):
        raise ValueError(f"auction_type must be 'first' or 'fixed', got {auction_type!r}")

    # --- margin ---
    if margin is not None and margin_calculation_type not in ("P", "C"):
        raise ValueError(
            f"margin_calculation_type must be 'P' or 'C' when margin is provided, got {margin_calculation_type!r}"
        )

    # --- publisher_ids / internal_deal_ids mutual exclusivity ---
    if publisher_ids is not None and internal_deal_ids is not None:
        raise ValueError(
            "publisher_ids and internal_deal_ids are mutually exclusive; target publishers or deals, not both"
        )

    # --- publisher_ids ---
    if publisher_ids is not None:
        if not publisher_ids:
            raise ValueError("publisher_ids must contain at least one publisher ID when provided")
        for i, publisher_id in enumerate(publisher_ids):
            if not isinstance(publisher_id, int) or isinstance(publisher_id, bool):
                raise ValueError(
                    f"publisher_ids[{i}] must be an integer legacyAccountID, got {type(publisher_id).__name__}"
                )

    # --- internal_deal_ids ---
    if internal_deal_ids is not None:
        if not internal_deal_ids:
            raise ValueError("internal_deal_ids must contain at least one internal deal ID when provided")
        for i, deal_id in enumerate(internal_deal_ids):
            if not isinstance(deal_id, int) or isinstance(deal_id, bool):
                raise ValueError(
                    f"internal_deal_ids[{i}] must be an integer internalDealID, got {type(deal_id).__name__}"
                )

    # --- targeting ---
    if targeting is not None:
        for i, t in enumerate(targeting):
            if "keyName" not in t:
                raise ValueError(f"targeting[{i}] missing required field 'keyName'")
            # Segment targeting (`keyName: "segmentid"`) is special: the IX
            # API accepts it without a targetingKeyID, and the segment key
            # isn't returned by `ix_list_targeting_keys` (only the
            # inclusion-only `im_segments` is). Allow the field to be
            # omitted for segment targeting; require it for everything else.
            keyname_lower = str(t.get("keyName", "")).strip().lower()
            is_segment_targeting = keyname_lower == "segmentid"
            if "targetingKeyID" not in t and not is_segment_targeting:
                raise ValueError(f"targeting[{i}] missing required field 'targetingKeyID'")
            if "targetingKeyID" in t and (
                not isinstance(t["targetingKeyID"], int) or isinstance(t["targetingKeyID"], bool)
            ):
                raise ValueError(
                    f"targeting[{i}].targetingKeyID must be an integer, got {type(t['targetingKeyID']).__name__}"
                )
            if t.get("targetingType") not in ("standard", "custom"):
                raise ValueError(
                    f"targeting[{i}].targetingType must be 'standard' or 'custom', got {t.get('targetingType')!r}"
                )
            sets = t.get("sets")
            if not sets:
                raise ValueError(f"targeting[{i}].sets must have at least one entry")
            for j, s in enumerate(sets):
                if s.get("operator") not in ("ANY_OF", "NONE_OF"):
                    raise ValueError(
                        f"targeting[{i}].sets[{j}].operator must be 'ANY_OF' or 'NONE_OF', got {s.get('operator')!r}"
                    )
                values = s.get("values")
                if not values:
                    raise ValueError(f"targeting[{i}].sets[{j}].values must have at least one entry")
                for k, v in enumerate(values):
                    if "value" not in v:
                        raise ValueError(f"targeting[{i}].sets[{j}].values[{k}] missing 'value'")
                    if not isinstance(v["value"], str):
                        raise ValueError(
                            f"targeting[{i}].sets[{j}].values[{k}].value must be a "
                            f"string (even for numeric IDs), got {type(v['value']).__name__}"
                        )

    targeting_payload = copy.deepcopy(targeting) if targeting is not None else None
    # Strip targetingKeyID from segment targeting objects — the API stores
    # `keyName: "segmentid"` without an ID and the agent / callers may pass
    # a placeholder integer just to satisfy the older validator. Keep the
    # request lean.
    if targeting_payload is not None:
        for target in targeting_payload:
            if str(target.get("keyName", "")).strip().lower() == "segmentid":
                target.pop("targetingKeyID", None)
    if internal_deal_ids is not None:
        internal_deal_targeting = {
            "targetingKeyID": 701,
            "keyName": "internaldealid",
            "targetingType": "standard",
            "sets": [{"operator": "ANY_OF", "values": [{"value": str(deal_id)} for deal_id in internal_deal_ids]}],
        }
        if targeting_payload is None:
            targeting_payload = [internal_deal_targeting]
        else:
            targeting_payload = [
                target
                for target in targeting_payload
                if str(target.get("keyName", "")).strip().lower() != "internaldealid"
            ]
            targeting_payload.insert(0, internal_deal_targeting)

    # --- Build payload ---
    marketplace_configurations: dict = {"dspID": dsp_id}
    if seat_ids is not None:
        marketplace_configurations["seatIDs"] = seat_ids
    if margin is not None:
        marketplace_configurations["margin"] = margin
    if margin_calculation_type is not None:
        marketplace_configurations["marginCalculationType"] = margin_calculation_type

    payload: dict = {
        "account": {"accountID": account_id},
        "classID": 4,
        "name": name,
        "externalDealID": external_deal_id,
        "startDate": start_date,
        "floor": floor,
        "auctionType": auction_type,
        "marketplaceConfigurations": marketplace_configurations,
    }

    if open_market is not None:
        payload["openMarket"] = open_market
    payload["endDate"] = end_date
    if publisher_ids is not None:
        payload["publisherIDs"] = publisher_ids
    if targeting_payload is not None:
        payload["targeting"] = targeting_payload
    if labels is not None:
        payload["labels"] = _sanitize_labels(labels)

    return payload


# Per-class floor minimums from the Update deal settings docs: Direct Deal
# (1), Inventory Package (3), and Deal with Marketplaces (5) accept 0.01;
# Marketplace Package (4) requires 0.10. 99999.99 is the documented maximum
# for every class.
_DEAL_CLASS_FLOOR_MINIMUM: dict[int, float] = {1: 0.01, 3: 0.01, 4: 0.10, 5: 0.01}
_DEAL_FLOOR_MAXIMUM = 99999.99

_DEAL_UPDATE_STATUS_VALUES = ("A", "P", "active", "paused")

_DATE_ONLY_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _validate_update_targeting_shape(targeting: list[dict]) -> None:
    """Shape-check a PATCH targeting array. Raises ValueError on problems.

    Update targeting uses the same object shape the GET response returns
    ({keyName, targetingType, sets:[{operator, values:[{value, label?}]}]}),
    so values read back from ix_get_deal_settings can be edited and sent
    back verbatim. No name->token normalization is applied here — the PATCH
    body must already carry canonical values.
    """
    for i, t in enumerate(targeting):
        if not isinstance(t, dict):
            raise ValueError(f"targeting[{i}] must be an object, got {type(t).__name__}")
        if not str(t.get("keyName", "")).strip():
            raise ValueError(f"targeting[{i}] missing required field 'keyName'")
        if t.get("targetingType") not in ("standard", "custom"):
            raise ValueError(f"targeting[{i}].targetingType must be 'standard' or 'custom'")
        sets = t.get("sets")
        if not isinstance(sets, list) or not sets:
            raise ValueError(f"targeting[{i}].sets must be a non-empty list")
        for j, s in enumerate(sets):
            if not isinstance(s, dict) or s.get("operator") not in ("ANY_OF", "NONE_OF"):
                raise ValueError(f"targeting[{i}].sets[{j}].operator must be 'ANY_OF' or 'NONE_OF'")
            values = s.get("values")
            if not isinstance(values, list) or not values:
                raise ValueError(f"targeting[{i}].sets[{j}].values must be a non-empty list")
            for k, v in enumerate(values):
                if not isinstance(v, dict) or not str(v.get("value", "")).strip():
                    raise ValueError(f"targeting[{i}].sets[{j}].values[{k}] must be an object with a 'value' string")


def _build_deal_update_payload(
    *,
    current_deal: dict,
    name: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    floor: float | None = None,
    auction_type: str | None = None,
    status: str | None = None,
    open_market: bool | None = None,
    labels: dict | None = None,
    targeting: list[dict] | None = None,
    direct_configurations: dict | None = None,
    ivp_configurations: dict | None = None,
    marketplace_configurations: dict | None = None,
    marketplace_participant_configurations: dict | None = None,
) -> dict:
    """Build the PATCH /api/deals/v3/deals/{internalDealID} body.

    Pure Python with no network I/O — safe to unit-test directly. Only the
    fields the caller supplied are included (partial update); validation that
    needs the deal's current state (class-aware floor minimum, date ordering
    against the unchanged date) reads it from `current_deal`, the object the
    caller just fetched for the ETag. Raises ValueError on validation failure.
    """
    payload: dict[str, Any] = {}

    if name is not None:
        if not name or len(name) > 255:
            raise ValueError(f"name must be 1–255 characters, got length {len(name)}")
        payload["name"] = name

    if start_date is not None:
        if not _DATE_ONLY_RE.match(start_date):
            raise ValueError(f"start_date must be YYYY-MM-DD, got {start_date!r}")
        payload["startDate"] = start_date
    if end_date is not None:
        if not _DATE_ONLY_RE.match(end_date):
            raise ValueError(f"end_date must be YYYY-MM-DD, got {end_date!r}")
        payload["endDate"] = end_date
    # endDate cannot be before startDate — compare against the value being
    # set, falling back to the deal's current value for whichever side this
    # PATCH leaves unchanged.
    effective_start = start_date if start_date is not None else current_deal.get("startDate")
    effective_end = end_date if end_date is not None else current_deal.get("endDate")
    dates_touched = start_date is not None or end_date is not None
    if dates_touched and effective_start and effective_end and str(effective_end) < str(effective_start):
        raise ValueError(f"endDate {effective_end!r} cannot be before startDate {effective_start!r}")

    if floor is not None:
        class_id = current_deal.get("classID")
        minimum = _DEAL_CLASS_FLOOR_MINIMUM.get(class_id, 0.01) if isinstance(class_id, int) else 0.01
        if floor < minimum:
            raise ValueError(f"floor must be >= {minimum} for classID={class_id}, got {floor}")
        if floor > _DEAL_FLOOR_MAXIMUM:
            raise ValueError(f"floor must be <= {_DEAL_FLOOR_MAXIMUM}, got {floor}")
        payload["floor"] = floor

    if auction_type is not None:
        if auction_type not in ("first", "fixed"):
            raise ValueError(f"auction_type must be 'first' or 'fixed', got {auction_type!r}")
        payload["auctionType"] = auction_type

    if status is not None:
        if status not in _DEAL_UPDATE_STATUS_VALUES:
            raise ValueError(f"status must be one of {_DEAL_UPDATE_STATUS_VALUES}, got {status!r}")
        payload["status"] = status

    if open_market is not None:
        payload["openMarket"] = open_market

    if labels is not None:
        payload["labels"] = _sanitize_labels(labels)

    if targeting is not None:
        _validate_update_targeting_shape(targeting)
        payload["targeting"] = targeting

    # Per-class configuration objects pass through as-is; the API ignores the
    # ones that don't match the deal's classID, but sending a mismatched one
    # is almost always a caller bug — flag it when we know the class.
    class_id = current_deal.get("classID")
    for arg_value, key, applies_to_class in (
        (direct_configurations, "directConfigurations", 1),
        (ivp_configurations, "ivpConfigurations", 3),
        (marketplace_configurations, "marketplaceConfigurations", 4),
        (marketplace_participant_configurations, "marketplaceParticipantConfigurations", 5),
    ):
        if arg_value is None:
            continue
        if isinstance(class_id, int) and class_id != applies_to_class:
            raise ValueError(f"{key} only applies to classID={applies_to_class} deals; this deal is classID={class_id}")
        payload[key] = arg_value

    if not payload:
        raise ValueError("No update fields provided — pass at least one field to change.")

    return payload


async def _normalize_targeting_for_create(
    client: "IndexExchangeClient",
    account_id: int,
    targeting: list[dict] | None,
    allow_partial_targeting: bool = False,
) -> tuple[list[dict] | None, dict[str, Any] | None]:
    """Normalize targeting payload for IX create endpoint semantics.

    - Canonicalize keyName from targeting key metadata when available.
    - For standard keys (except Domain/AppBundle), translate numeric value IDs to
      create-accepted value tokens.
    - For Domain/AppBundle key (120), pass literal domains/app bundles as provided.
    """
    if targeting is None:
        return None, None

    domain_stats: dict[str, Any] = {
        "source_domain_rows": 0,
        "normalized_unique_domains": 0,
        "domain_values_in_create_payload": 0,
        "invalid_count": 0,
        "invalid_samples": [],
        "duplicate_count": 0,
    }

    def normalize_domain_literal(raw_value: str) -> tuple[str | None, str | None]:
        value = raw_value.strip().lower()
        if not value:
            return None, "empty"
        if value in {"domain", "domains", "domain and app bundle", "domain/app bundle", "site", "url"}:
            return None, "header"
        if value.startswith("http://") or value.startswith("https://"):
            return None, "protocol"
        if any(ch in value for ch in ["/", "?", "#", " ", "\t", "\r", "\n"]):
            return None, "path_or_whitespace"
        if ":" in value:
            return None, "port_or_invalid"
        if value.startswith(".") or value.endswith("."):
            return None, "dot_boundary"
        if value.isdigit():
            return value, None
        if "." not in value:
            return None, "missing_dot"
        if not _SUB_DOMAIN_RE.match(value):
            return None, "invalid_chars"
        return value, None

    publisher_account_id = await _resolve_targeting_publisher_account_id(client, account_id)
    keys_resp = await client.request(
        "GET",
        "/api/supply-configuration/v1/inventory-groups/targets",
        params={"publisherAccountID": publisher_account_id},
    )
    keys = keys_resp if isinstance(keys_resp, list) else keys_resp.get("data", keys_resp)

    key_name_by_id: dict[int, str] = {}
    if isinstance(keys, list):
        for key in keys:
            if not isinstance(key, dict):
                continue
            key_id = key.get("targetingKeyID")
            if not isinstance(key_id, int):
                continue
            name = key.get("key") or key.get("keyName")
            if isinstance(name, str) and name:
                key_name_by_id[key_id] = name

    normalized_targeting = copy.deepcopy(targeting)
    numeric_values_by_key: dict[int, set[str]] = {}

    for targeting_obj in normalized_targeting:
        key_id = targeting_obj.get("targetingKeyID")
        if not isinstance(key_id, int):
            continue

        canonical_name = key_name_by_id.get(key_id)
        if canonical_name:
            targeting_obj["keyName"] = canonical_name
        key_name = str(canonical_name or targeting_obj.get("keyName", ""))
        is_segment_key = "segment" in key_name.lower()

        sets = targeting_obj.get("sets")
        if not isinstance(sets, list):
            continue

        for set_obj in sets:
            values = set_obj.get("values")
            if not isinstance(values, list):
                continue

            for value_obj in values:
                if not isinstance(value_obj, dict):
                    continue
                raw_value = value_obj.get("value")
                if not isinstance(raw_value, str):
                    continue
                value_str = raw_value.strip()

                if key_id == 120:
                    domain_stats["source_domain_rows"] += 1

                if key_id != 120 and not is_segment_key and value_str.isdigit():
                    numeric_values_by_key.setdefault(key_id, set()).add(value_str)

    domain_values_seen: set[str] = set()
    for targeting_obj in normalized_targeting:
        if targeting_obj.get("targetingKeyID") != 120:
            continue
        sets = targeting_obj.get("sets")
        if not isinstance(sets, list):
            continue
        for set_obj in sets:
            values = set_obj.get("values")
            if not isinstance(values, list):
                continue
            normalized_values: list[dict[str, str]] = []
            for value_obj in values:
                if not isinstance(value_obj, dict):
                    continue
                raw_value = value_obj.get("value")
                if not isinstance(raw_value, str):
                    continue
                normalized_value, invalid_reason = normalize_domain_literal(raw_value)
                if normalized_value is None:
                    domain_stats["invalid_count"] += 1
                    samples = domain_stats["invalid_samples"]
                    if len(samples) < 10:
                        samples.append({"value": raw_value, "reason": invalid_reason})
                    continue
                if normalized_value in domain_values_seen:
                    domain_stats["duplicate_count"] += 1
                    continue
                domain_values_seen.add(normalized_value)
                normalized_values.append({"value": normalized_value})
            set_obj["values"] = normalized_values

    domain_stats["normalized_unique_domains"] = len(domain_values_seen)
    domain_stats["domain_values_in_create_payload"] = len(domain_values_seen)

    if domain_stats["source_domain_rows"] > 0:
        if domain_stats["normalized_unique_domains"] == 0:
            raise ValueError(
                "Domain/AppBundle targeting (key 120) has no valid values after normalization. "
                "Provide valid literal domains/app bundles."
            )
        if domain_stats["invalid_count"] > 0 and not allow_partial_targeting:
            raise ValueError(
                "Domain/AppBundle targeting includes invalid entries and strict mode is enabled. "
                f"source_rows={domain_stats['source_domain_rows']}, invalid_count={domain_stats['invalid_count']}, "
                f"normalized_unique_domains={domain_stats['normalized_unique_domains']}. "
                "Set allow_partial_targeting=true to proceed with valid subset."
            )

    for key_id, numeric_values in numeric_values_by_key.items():
        values_resp = await client.request(
            "GET",
            f"/api/supply-configuration/v1/inventory-groups/targets/{key_id}/values",
            params={"publisherAccountID": publisher_account_id},
        )
        values = values_resp if isinstance(values_resp, list) else values_resp.get("data", values_resp)

        id_to_value: dict[str, str] = {}
        token_values: set[str] = set()
        if isinstance(values, list):
            for item in values:
                if not isinstance(item, dict):
                    continue
                value_id = item.get("targetingValueID")
                if value_id is None:
                    value_id = item.get("valueID")
                value_token = item.get("value")
                if value_id is None or not isinstance(value_token, str):
                    continue
                id_to_value[str(value_id)] = value_token
                token_values.add(value_token.strip())

        missing_ids = sorted(v for v in numeric_values if v not in id_to_value and v not in token_values)
        if missing_ids:
            raise ValueError(
                "Unable to translate targeting value IDs for create payload "
                f"(key {key_id}): missing IDs {missing_ids}. "
                "Pass create-accepted value tokens instead of IDs."
            )

        for targeting_obj in normalized_targeting:
            if targeting_obj.get("targetingKeyID") != key_id:
                continue
            sets = targeting_obj.get("sets")
            if not isinstance(sets, list):
                continue
            for set_obj in sets:
                values = set_obj.get("values")
                if not isinstance(values, list):
                    continue
                for value_obj in values:
                    if not isinstance(value_obj, dict):
                        continue
                    raw_value = value_obj.get("value")
                    if isinstance(raw_value, str):
                        trimmed = raw_value.strip()
                        if trimmed in id_to_value and trimmed not in token_values:
                            value_obj["value"] = id_to_value[trimmed]

    return normalized_targeting, domain_stats


def _collect_values_for_key_name(targeting: Any, expected_key_name: str) -> list[str]:
    if not isinstance(targeting, list):
        return []

    normalized_expected = expected_key_name.strip().lower()
    values: list[str] = []
    for targeting_obj in targeting:
        if not isinstance(targeting_obj, dict):
            continue
        key_name = str(targeting_obj.get("keyName", "")).strip().lower()
        if key_name != normalized_expected:
            continue
        sets = targeting_obj.get("sets")
        if not isinstance(sets, list):
            continue
        for set_obj in sets:
            if not isinstance(set_obj, dict):
                continue
            set_values = set_obj.get("values")
            if not isinstance(set_values, list):
                continue
            for value_obj in set_values:
                if not isinstance(value_obj, dict):
                    continue
                raw_value = value_obj.get("value")
                if isinstance(raw_value, str) and raw_value not in values:
                    values.append(raw_value)
    return values


def _collect_direct_ids_from_deal(
    deal: Any,
    *,
    explicit_fields: tuple[str, ...],
    list_fields: tuple[str, ...],
    list_item_keys: tuple[str, ...],
    targeting_key_name: str,
) -> list[str]:
    if not isinstance(deal, dict):
        return []

    values: list[str] = []

    def append_unique(raw: Any) -> None:
        if isinstance(raw, (str, int)) and not isinstance(raw, bool):
            normalized = str(raw)
            if normalized not in values:
                values.append(normalized)

    for field_name in explicit_fields:
        raw_field_value = deal.get(field_name)
        if isinstance(raw_field_value, list):
            for item in raw_field_value:
                append_unique(item)

    marketplace_configurations = deal.get("marketplaceConfigurations")
    if isinstance(marketplace_configurations, dict):
        for field_name in explicit_fields:
            raw_field_value = marketplace_configurations.get(field_name)
            if isinstance(raw_field_value, list):
                for item in raw_field_value:
                    append_unique(item)

    for list_field in list_fields:
        raw_list = deal.get(list_field)
        if not isinstance(raw_list, list):
            continue
        for item in raw_list:
            if not isinstance(item, dict):
                append_unique(item)
                continue
            for key in list_item_keys:
                if key in item:
                    append_unique(item.get(key))
                    break

    targeting_values = _collect_values_for_key_name(deal.get("targeting"), targeting_key_name)
    for targeting_value in targeting_values:
        append_unique(targeting_value)

    return values


async def _fetch_deal_with_direct_target_retries(
    client: "IndexExchangeClient",
    *,
    internal_deal_id: int | str,
    expect_internal_deal_ids: bool,
    expect_publisher_ids: bool,
) -> Any:
    persisted_deal = await client.request("GET", f"/api/deals/v3/deals/{internal_deal_id}")

    for delay_seconds in DIRECT_TARGET_VERIFICATION_RETRY_DELAYS_SECONDS:
        has_internal_ids = bool(
            _collect_direct_ids_from_deal(
                persisted_deal,
                explicit_fields=("internalDealIDs",),
                list_fields=("internalDeals", "deals"),
                list_item_keys=("internalDealID", "dealID", "id"),
                targeting_key_name="internaldealid",
            )
        )
        has_publisher_ids = bool(
            _collect_direct_ids_from_deal(
                persisted_deal,
                explicit_fields=("publisherIDs",),
                list_fields=("publishers",),
                list_item_keys=("legacyAccountID", "publisherID", "accountID", "id"),
                targeting_key_name="publisherid",
            )
        )

        if (not expect_internal_deal_ids or has_internal_ids) and (not expect_publisher_ids or has_publisher_ids):
            break

        logger.info(
            "Retrying direct target verification for deal %s after %ss",
            internal_deal_id,
            delay_seconds,
        )
        await asyncio.sleep(delay_seconds)
        persisted_deal = await client.request("GET", f"/api/deals/v3/deals/{internal_deal_id}")

    return persisted_deal


def _extract_inventory_group_targeting_values(
    inventory_group: dict[str, Any],
    *,
    key_name: str,
) -> list[str]:
    targeting = inventory_group.get("targeting")
    if not isinstance(targeting, list):
        return []

    normalized_key_name = key_name.strip().lower()
    values: list[str] = []
    for targeting_entry in targeting:
        if not isinstance(targeting_entry, dict):
            continue
        entry_key_name = str(targeting_entry.get("key") or targeting_entry.get("keyName") or "").strip().lower()
        if entry_key_name != normalized_key_name:
            continue
        entry_values = targeting_entry.get("values")
        if not isinstance(entry_values, list):
            continue
        for value_entry in entry_values:
            if not isinstance(value_entry, dict):
                continue
            raw_value = value_entry.get("value")
            if isinstance(raw_value, str) and raw_value not in values:
                values.append(raw_value)
    return values


def _collect_domain_values_from_inventory_group(inventory_group: dict[str, Any]) -> list[str]:
    return _extract_inventory_group_targeting_values(inventory_group, key_name="domain")


def _collect_values_for_key(targeting: Any, expected_key_name: str) -> list[str]:
    if not isinstance(targeting, list):
        return []

    normalized_expected = expected_key_name.strip().lower()
    values: list[str] = []
    for targeting_obj in targeting:
        if not isinstance(targeting_obj, dict):
            continue
        key_name = str(targeting_obj.get("keyName") or targeting_obj.get("key") or "").strip().lower()
        if key_name != normalized_expected:
            continue
        sets = targeting_obj.get("sets")
        if isinstance(sets, list):
            for set_obj in sets:
                if not isinstance(set_obj, dict):
                    continue
                set_values = set_obj.get("values")
                if not isinstance(set_values, list):
                    continue
                for value_obj in set_values:
                    if not isinstance(value_obj, dict):
                        continue
                    raw_value = value_obj.get("value")
                    if isinstance(raw_value, str) and raw_value not in values:
                        values.append(raw_value)
            continue
        direct_values = targeting_obj.get("values")
        if isinstance(direct_values, list):
            for value_obj in direct_values:
                if not isinstance(value_obj, dict):
                    continue
                raw_value = value_obj.get("value")
                if isinstance(raw_value, str) and raw_value not in values:
                    values.append(raw_value)
    return values


async def _find_deal_specific_inventory_group(
    client: "IndexExchangeClient",
    *,
    account_id: int,
    external_deal_id: str,
    dsp_id: int,
) -> dict[str, Any] | None:
    publisher_account_id = await _resolve_targeting_publisher_account_id(client, account_id)
    search_candidates = [
        external_deal_id,
        f"ExternalDealID_{external_deal_id}",
        f"ExternalDealID_{external_deal_id}_DspID_{dsp_id}",
    ]

    last_response: dict[str, Any] | None = None
    for search_value in search_candidates:
        response = await client.request(
            "GET",
            "/api/supply-configuration/v1/inventory-groups",
            params={
                "publisherAccountID": publisher_account_id,
                "pageSize": 50,
                "pageOffset": 0,
                "search": search_value,
            },
        )
        if isinstance(response, dict):
            last_response = response
        inventory_groups = response.get("inventoryGroups") if isinstance(response, dict) else None
        if not isinstance(inventory_groups, list):
            continue
        for inventory_group in inventory_groups:
            if not isinstance(inventory_group, dict):
                continue
            name = str(inventory_group.get("name") or "")
            group_type = str(inventory_group.get("type") or "")
            if external_deal_id in name and group_type.lower() == "deal specific":
                return inventory_group

    if (
        last_response
        and isinstance(last_response.get("inventoryGroups"), list)
        and len(last_response["inventoryGroups"]) == 1
    ):
        inventory_group = last_response["inventoryGroups"][0]
        if isinstance(inventory_group, dict):
            return inventory_group

    return None


def _has_value_for_key_name(targeting: Any, expected_key_name: str, expected_value: str) -> bool:
    normalized_value = expected_value.strip().lower()
    return any(
        value.strip().lower() == normalized_value
        for value in _collect_values_for_key_name(targeting, expected_key_name)
    )


def _default_rolling_end_date(start_date: str, months: int = IX_DEFAULT_END_DATE_MONTHS) -> str:
    """Return a YYYY-MM-DD date `months` calendar months after start_date.

    Used to fill in deal end_date defaults when the caller does not specify one.
    Clamps day-of-month when the target month is shorter (e.g. Aug 31 + 6mo -> Feb 28/29).
    The default-applied month count is `IX_DEFAULT_END_DATE_MONTHS` (currently 24).
    """
    start = date.fromisoformat(start_date)
    total_month = start.month + months
    end_year = start.year + (total_month - 1) // 12
    end_month = ((total_month - 1) % 12) + 1
    # Clamp day to last valid day of end_month.
    days_in_month = [
        31,
        29 if end_year % 4 == 0 and (end_year % 100 != 0 or end_year % 400 == 0) else 28,
        31,
        30,
        31,
        30,
        31,
        31,
        30,
        31,
        30,
        31,
    ][end_month - 1]
    end_day = min(start.day, days_in_month)
    return date(end_year, end_month, end_day).isoformat()


def _normalize_deal_type(deal_type: str | None) -> str | None:
    """Return canonical "display", "olv", "ctv", "ott", or None for an input deal_type hint."""
    if deal_type is None:
        return None
    normalized = deal_type.strip().lower()
    if not normalized:
        return None
    if normalized in IX_DEAL_TYPES_CTV:
        return "ctv"
    if normalized in IX_DEAL_TYPES_OTT:
        return "ott"
    if normalized in IX_DEAL_TYPES_OLV:
        return "olv"
    if normalized in IX_DEAL_TYPES_DISPLAY:
        return "display"
    return None


async def _ensure_deal_type_targeting_defaults(
    client: "IndexExchangeClient",
    account_id: int,
    targeting: list[dict] | None,
    *,
    deal_type: str | None = None,
) -> list[dict] | None:
    """Auto-fill Index-UI defaults per canonical channel.

    Branches on the explicit `deal_type` first (preferred — the brief and the
    multi-deal protocol both pass it); when omitted, falls back to inferring
    from existing device-type targeting values (legacy path).

    Canonical defaults (trader spec):
      - display: devices=PC+Phone+Tablet, inventory=In-App+Web, creative=Banner_ANY
      - olv:     devices=PC+Phone+Tablet, inventory=In-App+Web, creative=Video_ANY
      - ctv:     devices=CTV+Connected device+Set-top box, inventory=In-App only,
                 creative=Video_ANY
      - ott:     devices=Phone+Tablet, inventory=In-App only, creative=Video_ANY

    Existing user-provided targeting values for a given key are left alone when
    they already cover the canonical set; only missing pieces are filled in.
    """
    normalized_deal_type = _normalize_deal_type(deal_type)

    if targeting is None and normalized_deal_type is None:
        return None

    resolved_targeting: list[dict] = copy.deepcopy(targeting) if targeting is not None else []

    # Legacy inference path: when deal_type isn't passed, look at existing
    # device-type tokens to guess whether this is CTV or Display. OLV/OTT
    # cannot be inferred from devices alone (OLV shares Display's device set,
    # OTT shares Phone+Tablet with mobile-Display) — these require an
    # explicit deal_type to route to the Video creative format.
    if normalized_deal_type is None:
        device_values = _collect_values_for_key_name(resolved_targeting, "devicetype")
        has_ctv_signal = any(value in set(IX_DEVICE_VALUES_CTV) for value in device_values)
        has_display_signal = any(value in set(IX_DEVICE_VALUES_DISPLAY) for value in device_values)
        if has_ctv_signal and has_display_signal:
            # Mixed signal — preserves prior behaviour treating ambiguous
            # mixes as CTV.
            normalized_deal_type = "ctv"
        elif has_ctv_signal:
            normalized_deal_type = "ctv"
        elif has_display_signal:
            normalized_deal_type = "display"
        else:
            return resolved_targeting if targeting is not None else None

    targeting_keys_resp = await client.request(
        "GET",
        "/api/supply-configuration/v1/inventory-groups/targets",
        params={"publisherAccountID": await _resolve_targeting_publisher_account_id(client, account_id)},
    )
    targeting_keys = (
        targeting_keys_resp
        if isinstance(targeting_keys_resp, list)
        else targeting_keys_resp.get("data", targeting_keys_resp)
    )

    key_name_to_id: dict[str, int] = {}
    if isinstance(targeting_keys, list):
        for key in targeting_keys:
            if not isinstance(key, dict):
                continue
            key_id = key.get("targetingKeyID")
            key_name = key.get("key") or key.get("keyName")
            if isinstance(key_id, int) and isinstance(key_name, str) and key_name.strip():
                key_name_to_id[key_name.strip().lower()] = key_id

    device_key_id = key_name_to_id.get("devicetype", 3)
    creative_key_id = key_name_to_id.get("creativetypesize", 10)
    inventory_key_id = key_name_to_id.get("inventorychannel", 272)

    canonical_inventory_values: tuple[str, ...]
    if normalized_deal_type == "ctv":
        canonical_device_values = list(IX_DEVICE_VALUES_CTV)
        canonical_creative_token = IX_CREATIVE_TYPE_SIZE_VIDEO_ANY
        canonical_inventory_values = IX_INVENTORY_CHANNEL_VALUES_APP_ONLY
        legacy_creative_tokens = {IX_VIDEO_ALL_LEGACY_VALUE_ID}
    elif normalized_deal_type == "ott":
        canonical_device_values = list(IX_DEVICE_VALUES_OTT)
        canonical_creative_token = IX_CREATIVE_TYPE_SIZE_VIDEO_ANY
        canonical_inventory_values = IX_INVENTORY_CHANNEL_VALUES_APP_ONLY
        legacy_creative_tokens = {IX_VIDEO_ALL_LEGACY_VALUE_ID}
    elif normalized_deal_type == "olv":
        canonical_device_values = list(IX_DEVICE_VALUES_OLV)
        canonical_creative_token = IX_CREATIVE_TYPE_SIZE_VIDEO_ANY
        canonical_inventory_values = IX_INVENTORY_CHANNEL_VALUES_DEFAULT
        legacy_creative_tokens = {IX_VIDEO_ALL_LEGACY_VALUE_ID}
    else:  # display
        canonical_device_values = list(IX_DEVICE_VALUES_DISPLAY)
        canonical_creative_token = IX_CREATIVE_TYPE_SIZE_BANNER_ANY
        canonical_inventory_values = IX_INVENTORY_CHANNEL_VALUES_DEFAULT
        legacy_creative_tokens = set()

    canonical_device_value_objs = [{"value": value} for value in canonical_device_values]
    device_target_found = False
    for targeting_obj in resolved_targeting:
        if not isinstance(targeting_obj, dict):
            continue
        key_name = str(targeting_obj.get("keyName", "")).strip().lower()
        key_id = targeting_obj.get("targetingKeyID")
        if key_name != "devicetype" and key_id != device_key_id:
            continue
        sets = targeting_obj.get("sets")
        if not isinstance(sets, list) or not sets:
            continue
        for set_obj in sets:
            if not isinstance(set_obj, dict):
                continue
            if str(set_obj.get("operator", "")).upper() != "ANY_OF":
                continue
            set_obj["values"] = copy.deepcopy(canonical_device_value_objs)
            device_target_found = True
            break
        targeting_obj["targetingKeyID"] = device_key_id
        targeting_obj["keyName"] = "DeviceType"
        if device_target_found:
            break

    if not device_target_found:
        resolved_targeting.append(
            {
                "targetingKeyID": device_key_id,
                "keyName": "DeviceType",
                "targetingType": "standard",
                "sets": [{"operator": "ANY_OF", "values": copy.deepcopy(canonical_device_value_objs)}],
            }
        )

    if not _has_value_for_key_name(resolved_targeting, "creativetypesize", canonical_creative_token) and not any(
        _has_value_for_key_name(resolved_targeting, "creativetypesize", token) for token in legacy_creative_tokens
    ):
        resolved_targeting.append(
            {
                "targetingKeyID": creative_key_id,
                "keyName": "creativeTypeSize",
                "targetingType": "standard",
                "sets": [{"operator": "ANY_OF", "values": [{"value": canonical_creative_token}]}],
            }
        )

    if canonical_inventory_values:
        existing_inventory_values = {
            value.strip() for value in _collect_values_for_key_name(resolved_targeting, "inventorychannel")
        }
        missing_inventory_values = [
            value for value in canonical_inventory_values if value not in existing_inventory_values
        ]
        if missing_inventory_values:
            inventory_target_found = False
            for targeting_obj in resolved_targeting:
                if not isinstance(targeting_obj, dict):
                    continue
                key_name = str(targeting_obj.get("keyName", "")).strip().lower()
                key_id = targeting_obj.get("targetingKeyID")
                if key_name != "inventorychannel" and key_id != inventory_key_id:
                    continue
                sets = targeting_obj.get("sets")
                if not isinstance(sets, list) or not sets:
                    continue
                for set_obj in sets:
                    if not isinstance(set_obj, dict):
                        continue
                    if str(set_obj.get("operator", "")).upper() != "ANY_OF":
                        continue
                    existing_values = [value for value in set_obj.get("values", []) if isinstance(value, dict)]
                    existing_tokens = {
                        str(value.get("value", "")).strip()
                        for value in existing_values
                        if isinstance(value.get("value"), str)
                    }
                    for missing in missing_inventory_values:
                        if missing not in existing_tokens:
                            existing_values.append({"value": missing})
                            existing_tokens.add(missing)
                    set_obj["values"] = existing_values
                    inventory_target_found = True
                    break
                targeting_obj["targetingKeyID"] = inventory_key_id
                targeting_obj["keyName"] = "inventoryChannel"
                if inventory_target_found:
                    break
            if not inventory_target_found:
                resolved_targeting.append(
                    {
                        "targetingKeyID": inventory_key_id,
                        "keyName": "inventoryChannel",
                        "targetingType": "standard",
                        "sets": [
                            {
                                "operator": "ANY_OF",
                                "values": [{"value": value} for value in canonical_inventory_values],
                            }
                        ],
                    }
                )

    return resolved_targeting


def _count_domain_values_in_targeting(targeting: Any) -> int:
    if not isinstance(targeting, list):
        return 0
    total = 0
    for targeting_obj in targeting:
        if not isinstance(targeting_obj, dict):
            continue
        key_name = str(targeting_obj.get("keyName", "")).strip().lower()
        key_id = targeting_obj.get("targetingKeyID")
        is_domain_key = key_id == 120 or key_name in {"domain", "domain and app bundle", "domain_and_app_bundle"}
        if not is_domain_key:
            continue
        sets = targeting_obj.get("sets")
        if not isinstance(sets, list):
            continue
        for set_obj in sets:
            if not isinstance(set_obj, dict):
                continue
            values = set_obj.get("values")
            if isinstance(values, list):
                total += len(values)
    return total


def _has_domain_targeting(targeting: Any) -> bool:
    if not isinstance(targeting, list):
        return False
    for targeting_obj in targeting:
        if not isinstance(targeting_obj, dict):
            continue
        key_name = str(targeting_obj.get("keyName", "")).strip().lower()
        key_id = targeting_obj.get("targetingKeyID")
        if key_id == 120 or key_name in {"domain", "domain and app bundle", "domain_and_app_bundle"}:
            return True
    return False


def _collect_domain_values_in_targeting(targeting: Any) -> list[str]:
    if not isinstance(targeting, list):
        return []
    values_out: list[str] = []
    for targeting_obj in targeting:
        if not isinstance(targeting_obj, dict):
            continue
        key_name = str(targeting_obj.get("keyName", "")).strip().lower()
        key_id = targeting_obj.get("targetingKeyID")
        is_domain_key = key_id == 120 or key_name in {"domain", "domain and app bundle", "domain_and_app_bundle"}
        if not is_domain_key:
            continue
        sets = targeting_obj.get("sets")
        if not isinstance(sets, list):
            continue
        for set_obj in sets:
            if not isinstance(set_obj, dict):
                continue
            set_values = set_obj.get("values")
            if not isinstance(set_values, list):
                continue
            for value_obj in set_values:
                if not isinstance(value_obj, dict):
                    continue
                value = value_obj.get("value")
                if isinstance(value, str):
                    trimmed = value.strip()
                    if trimmed:
                        values_out.append(trimmed)
    return values_out


def _domain_values_fingerprint(values: list[str]) -> str:
    normalized = sorted({value.strip().lower() for value in values if isinstance(value, str) and value.strip()})
    payload = json.dumps(normalized, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _clear_file_domain_expectation_locks_for_tests() -> None:
    """Test-only helper to reset file-driven expectation locks between tests."""
    _FILE_DOMAIN_EXPECTATION_LOCKS.clear()


_DOMAIN_OPERATOR_ALIASES = {
    "allowlist": "ANY_OF",
    "allow": "ANY_OF",
    "allow_list": "ANY_OF",
    "include": "ANY_OF",
    "inclusion": "ANY_OF",
    "any_of": "ANY_OF",
    "blocklist": "NONE_OF",
    "block": "NONE_OF",
    "block_list": "NONE_OF",
    "exclude": "NONE_OF",
    "exclusion": "NONE_OF",
    "none_of": "NONE_OF",
}


def _normalize_domain_operator(value: str) -> str:
    """Map a human-readable domain/app-bundle match operator to the IX wire value.

    Accepts the brief-level vocabulary (``allowlist`` / ``blocklist``), common
    synonyms (``include`` / ``exclude``), or the raw wire values (``ANY_OF`` /
    ``NONE_OF``). Returns ``"ANY_OF"`` (allowlist/inclusion) or ``"NONE_OF"``
    (blocklist/exclusion).
    """
    if value is None:
        raise ValueError("domain_match_operator must not be None")
    normalized = _DOMAIN_OPERATOR_ALIASES.get(str(value).strip().lower())
    if normalized is None:
        raise ValueError(
            f"domain_match_operator must be 'allowlist'/'blocklist' (or 'ANY_OF'/'NONE_OF'), got {value!r}"
        )
    return normalized


def _build_domain_targeting_object(domains: list[str], operator: str = "NONE_OF") -> dict[str, Any]:
    """Build a key-120 targeting object for domain/app-bundle values.

    Args:
        domains: List of domain literal strings.
        operator: "ANY_OF" (include only these domains) or "NONE_OF" (exclude
            these domains).  Defaults to "NONE_OF" because the most common
            use-case for file-driven domain lists is a blocklist/exclusion list.
    """
    if operator not in ("ANY_OF", "NONE_OF"):
        raise ValueError(f"operator must be 'ANY_OF' or 'NONE_OF', got {operator!r}")
    return {
        "targetingKeyID": 120,
        "keyName": "Domain",
        "targetingType": "standard",
        "sets": [{"operator": operator, "values": [{"value": d} for d in domains]}],
    }


def _load_domains_from_source(domain_source: Any) -> tuple[list[str], dict[str, Any]]:
    """Load raw domain values from a server-side source.

    Supported shapes:
      - "/path/to/file.csv"
      - {"file_path": "/path/to/file.csv"}
      - {"type": "file_path", "file_path": "/path/to/file.csv"}

    domain_list_id is intentionally not implemented yet; callers should use file_path.
    """
    source_type = "file_path"
    file_path: str | None = None
    sheet_name: str | None = None
    column_name: str | None = None
    domain_list_id: str | None = None
    allow_app_bundle_ids = False

    if isinstance(domain_source, str):
        file_path = domain_source
    elif isinstance(domain_source, dict):
        source_type = str(domain_source.get("type", "file_path")).strip().lower()
        file_path_raw = domain_source.get("file_path")
        if isinstance(file_path_raw, str) and file_path_raw.strip():
            file_path = file_path_raw.strip()
        sheet_name_raw = domain_source.get("sheet_name")
        if isinstance(sheet_name_raw, str) and sheet_name_raw.strip():
            sheet_name = sheet_name_raw.strip()
        column_name_raw = domain_source.get("column_name")
        if isinstance(column_name_raw, str) and column_name_raw.strip():
            column_name = column_name_raw.strip()
        domain_list_id_raw = domain_source.get("domain_list_id")
        if isinstance(domain_list_id_raw, str) and domain_list_id_raw.strip():
            domain_list_id = domain_list_id_raw.strip()
        allow_app_bundle_ids = bool(domain_source.get("allow_app_bundle_ids", False))
    else:
        raise ValueError("domain_source must be a string path or an object containing file_path")

    if file_path and domain_list_id:
        raise ValueError("domain_source cannot include both file_path and domain_list_id")

    if domain_list_id:
        raise ValueError("domain_source.domain_list_id is not implemented yet; use domain_source.file_path for now")

    if source_type != "file_path":
        raise ValueError(f"Unsupported domain_source type {source_type!r}; use 'file_path'")

    if not file_path:
        raise ValueError("domain_source.file_path is required")

    path = Path(file_path).expanduser()
    path = (Path.cwd() / path).resolve() if not path.is_absolute() else path.resolve()

    if not path.exists() or not path.is_file():
        raise ValueError(f"domain_source.file_path does not exist or is not a file: {path}")

    file_result = _extract_domains_from_file(
        str(path), sheet_name=sheet_name, column_name=column_name, allow_app_bundle_ids=allow_app_bundle_ids
    )
    raw_domains = file_result["domains"]
    if not raw_domains:
        raise ValueError(f"domain_source.file_path has no readable rows: {path}")

    metadata = {
        "domain_source_type": "file_path",
        "domain_source_file_path": str(path),
        "domain_source_rows_loaded": len(raw_domains),
        "domain_source_header_rows_removed": 1 if file_result.get("column_name") else 0,
    }
    if file_result.get("sheet_name"):
        metadata["domain_source_sheet_name"] = file_result["sheet_name"]
    if file_result.get("column_name"):
        metadata["domain_source_column_name"] = file_result["column_name"]
    if file_result.get("invalid_values"):
        metadata["domain_source_invalid_rows_dropped"] = len(file_result["invalid_values"])
    return raw_domains, metadata


class IndexExchangeClient:
    """Client for interacting with the Index Exchange REST APIs.

    Supports two auth modes:
    A) Service account: INDEXEXCHANGE_SERVICE_ID + INDEXEXCHANGE_SERVICE_SECRET
    B) User account: INDEXEXCHANGE_USERNAME + INDEXEXCHANGE_PASSWORD

    Service account credentials are preferred when both are present.
    """

    def __init__(self) -> None:
        self.base_url = os.environ.get("INDEXEXCHANGE_BASE_URL", DEFAULT_BASE_URL).rstrip("/")
        timeout_str = os.environ.get("INDEXEXCHANGE_TIMEOUT_SECONDS", "")
        try:
            self._timeout = float(timeout_str) if timeout_str else DEFAULT_TIMEOUT
        except ValueError:
            self._timeout = DEFAULT_TIMEOUT

        # Always use the default — the INDEXEXCHANGE_DOWNLOAD_DIR override was
        # routinely set to a developer's MacOS path in .bashrc (e.g.
        # /Users/...) which fails on Linux deployments. The default works
        # everywhere; callers who really need a different dir can pass
        # output_dir per-call.
        self._download_dir = DEFAULT_DOWNLOAD_DIR

        # Credentials
        self._service_id = os.environ.get("INDEXEXCHANGE_SERVICE_ID", "")
        self._service_secret = os.environ.get("INDEXEXCHANGE_SERVICE_SECRET", "")
        self._username = os.environ.get("INDEXEXCHANGE_USERNAME", "")
        self._password = os.environ.get("INDEXEXCHANGE_PASSWORD", "")

        # Token state
        self._access_token: str = ""
        self._refresh_token: str = ""
        self._token_expiry: float = 0.0  # unix timestamp

        self._http_client: httpx.AsyncClient | None = None

    @property
    def auth_mode(self) -> str | None:
        """Return the configured auth mode or None if not configured."""
        if self._service_id and self._service_secret:
            return "service_account"
        if self._username and self._password:
            return "user_account"
        return None

    def _is_configured(self) -> bool:
        return self.auth_mode is not None

    async def _get_http_client(self) -> httpx.AsyncClient:
        if self._http_client is None:
            self._http_client = httpx.AsyncClient(timeout=self._timeout, follow_redirects=True)
        return self._http_client

    def _token_is_valid(self) -> bool:
        """Check if the cached access token is still valid (with safety buffer)."""
        if not self._access_token:
            return False
        return time.time() < (self._token_expiry - TOKEN_EXPIRY_BUFFER_SECONDS)

    async def _login(self) -> None:
        """Authenticate with Index Exchange and obtain access + refresh tokens."""
        if not self._is_configured():
            raise ValueError(
                "Index Exchange not configured. Set INDEXEXCHANGE_SERVICE_ID + "
                "INDEXEXCHANGE_SERVICE_SECRET, or INDEXEXCHANGE_USERNAME + "
                "INDEXEXCHANGE_PASSWORD environment variables."
            )

        client = await self._get_http_client()
        url = f"{self.base_url}/api/authentication/v1/login"

        if self.auth_mode == "service_account":
            body = {
                "username": self._service_id,
                "password": self._service_secret,
            }
            logger.info("Logging in to Index Exchange (service account)")
        else:
            body = {
                "username": self._username,
                "password": self._password,
            }
            logger.info("Logging in to Index Exchange (user account)")

        try:
            response = await client.post(
                url,
                json=body,
                headers={
                    "Content-Type": "application/json",
                    "User-Agent": USER_AGENT,
                },
            )
            response.raise_for_status()
            data = response.json()

            # Handle nested response: loginResponse.authResponse.access_token
            auth_response = data
            if "loginResponse" in data:
                auth_response = data["loginResponse"].get("authResponse", {})

            self._access_token = auth_response.get("access_token", "")
            self._refresh_token = auth_response.get("refresh_token", "")

            if not self._access_token:
                raise ValueError("Login succeeded but no access_token in response")

            # Parse JWT exp
            exp = _decode_jwt_exp(self._access_token)
            if exp:
                self._token_expiry = exp
            else:
                # Fallback: assume 10 min from now
                self._token_expiry = time.time() + 600

            logger.info("Successfully obtained Index Exchange access token")

        except httpx.HTTPStatusError as e:
            logger.error("Login failed with HTTP status %d", e.response.status_code)
            raise ValueError(f"Index Exchange login failed: HTTP {e.response.status_code}") from e
        except ValueError:
            raise
        except Exception as e:
            logger.error("Login failed: %s", type(e).__name__)
            raise ValueError(f"Index Exchange login failed: {type(e).__name__}") from e

    async def _refresh(self) -> bool:
        """Attempt to refresh the access token. Returns True on success."""
        if not self._refresh_token:
            return False

        client = await self._get_http_client()
        url = f"{self.base_url}/api/authentication/v1/refresh"

        try:
            response = await client.post(
                url,
                json={"refreshToken": self._refresh_token},
                headers={
                    "Content-Type": "application/json",
                    "User-Agent": USER_AGENT,
                },
            )
            if response.status_code == 401:
                logger.warning("Refresh token rejected (401), will re-login")
                return False
            response.raise_for_status()
            data = response.json()

            self._access_token = data.get("access_token", "")
            # Some APIs return a new refresh token too
            if data.get("refresh_token"):
                self._refresh_token = data["refresh_token"]

            if not self._access_token:
                return False

            exp = _decode_jwt_exp(self._access_token)
            if exp:
                self._token_expiry = exp
            else:
                self._token_expiry = time.time() + 600

            logger.info("Successfully refreshed Index Exchange access token")
            return True

        except Exception:
            logger.warning("Token refresh failed, will re-login")
            return False

    async def _ensure_access_token(self) -> str:
        """Ensure we have a valid access token, refreshing or logging in as needed."""
        if self._token_is_valid():
            return self._access_token

        # Try refresh first if we have a refresh token
        if self._refresh_token and await self._refresh():
            return self._access_token

        # Fall back to login
        await self._login()
        return self._access_token

    async def request(
        self,
        method: str,
        path: str,
        *,
        params: dict[str, Any] | None = None,
        json_body: dict[str, Any] | list[Any] | None = None,
        headers: dict[str, str] | None = None,
        accept: str = "application/json",
        raw_response: bool = False,
    ) -> Any:
        """Execute an authenticated HTTP request against the IX API.

        Retries once on 401 by forcing re-auth.
        """
        token = await self._ensure_access_token()
        client = await self._get_http_client()
        url = f"{self.base_url}{path}"

        req_headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "Accept": accept,
            "User-Agent": USER_AGENT,
        }
        if headers:
            req_headers.update(headers)

        logger.info("Executing %s request to Index Exchange: %s", method.upper(), path)

        for attempt in range(2):
            try:
                response = await client.request(
                    method.upper(),
                    url,
                    headers=req_headers,
                    params=params,
                    json=json_body,
                )

                if response.status_code == 401 and attempt == 0:
                    logger.warning("Received 401, forcing re-auth and retrying")
                    # Force full re-login
                    self._access_token = ""
                    self._refresh_token = ""
                    self._token_expiry = 0.0
                    token = await self._ensure_access_token()
                    req_headers["Authorization"] = f"Bearer {token}"
                    continue

                response.raise_for_status()

                if raw_response:
                    return response

                return response.json()

            except httpx.HTTPStatusError as e:
                body_text = e.response.text[:500] if e.response.text else ""
                logger.error(
                    "HTTP error %d on %s %s",
                    e.response.status_code,
                    method.upper(),
                    path,
                )
                raise ValueError(f"HTTP {e.response.status_code}: {body_text}") from e
            except ValueError:
                raise
            except Exception as e:
                logger.error("Request error on %s %s: %s", method.upper(), path, type(e).__name__)
                raise

        # Should not reach here
        raise ValueError("Request failed after retry")  # pragma: no cover

    async def download_file(
        self, path: str, directory: Path, filename_base: str, filename_hint: str | None = None
    ) -> tuple[Path, int, str, str]:
        """Download a file from the given path to the destination directory.

        Automatically determines extension based on headers or hint.

        Returns:
            Tuple of (file_path, file_size, sha256_hash, content_type)
        """
        token = await self._ensure_access_token()
        client = await self._get_http_client()
        url = f"{self.base_url}{path}"

        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/octet-stream",
            "User-Agent": USER_AGENT,
        }

        logger.info("Downloading file from Index Exchange: %s", path)

        for attempt in range(2):
            try:
                async with client.stream("GET", url, headers=headers) as response:
                    if response.status_code == 401 and attempt == 0:
                        logger.warning("Received 401 during download, forcing re-auth and retrying")
                        self._access_token = ""
                        self._refresh_token = ""
                        self._token_expiry = 0.0
                        token = await self._ensure_access_token()
                        headers["Authorization"] = f"Bearer {token}"
                        continue

                    response.raise_for_status()

                    content_type = response.headers.get("content-type", "")
                    content_disp = response.headers.get("content-disposition", "")

                    ext = ".bin"
                    if filename_hint:
                        suffixes = Path(filename_hint).suffixes
                        if suffixes:
                            ext = "".join(suffixes)
                    elif "csv" in content_type or "csv" in content_disp:
                        ext = ".csv"
                    elif "gzip" in content_type or content_disp.endswith(".gz"):
                        ext = ".csv.gz"
                    elif "zip" in content_type or content_disp.endswith(".zip"):
                        ext = ".csv.zip"

                    destination = directory / f"{filename_base}{ext}"

                    sha256 = hashlib.sha256()
                    bytes_downloaded = 0

                    with destination.open("wb") as f:
                        async for chunk in response.aiter_bytes():
                            f.write(chunk)
                            sha256.update(chunk)
                            bytes_downloaded += len(chunk)

                    return destination, bytes_downloaded, sha256.hexdigest(), content_type

            except httpx.HTTPStatusError as e:
                # Can't read response text easily in stream mode if already consumed,
                # but if status error raised immediately, we might not have consumed it.
                # However, stream context usually doesn't load body automatically on error unless we read it.
                # For simplicity, we just log status code.
                logger.error(
                    "HTTP error %d on GET %s",
                    e.response.status_code,
                    path,
                )
                raise ValueError(f"HTTP {e.response.status_code}") from e
            except ValueError:
                raise
            except Exception as e:
                logger.error("Download error on GET %s: %s", path, type(e).__name__)
                raise

        raise ValueError("Download failed after retry")


# Global client instance
_ix_client: IndexExchangeClient | None = None


def get_ix_client() -> IndexExchangeClient:
    """Get or create the Index Exchange client singleton."""
    global _ix_client
    if _ix_client is None:
        _ix_client = IndexExchangeClient()
    return _ix_client


async def _resolve_targeting_publisher_account_id(client: IndexExchangeClient, account_id: int) -> int:
    """Resolve publisher account scope for targeting discovery endpoints.

    Marketplace accounts must use marketplace.legacyMarketplaceID as
    publisherAccountID for inventory-group targeting endpoints.

    Note: IX UI uses the query parameter name "publisherAccountID" even in
    marketplace flows where the value is a legacyMarketplaceID.
    """
    accounts_resp = await client.request(
        "GET",
        "/api/accounts/v2/accounts/",
        params={"accountIDs": account_id},
    )

    accounts: Any
    if isinstance(accounts_resp, list):
        accounts = accounts_resp
    elif isinstance(accounts_resp, dict):
        if isinstance(accounts_resp.get("data"), list):
            accounts = accounts_resp["data"]
        elif isinstance(accounts_resp.get("accounts"), list):
            accounts = accounts_resp["accounts"]
        else:
            accounts = accounts_resp
    else:
        accounts = accounts_resp

    if not isinstance(accounts, list):
        return account_id

    account = next((a for a in accounts if isinstance(a, dict) and a.get("accountID") == account_id), None)
    if not account:
        return account_id

    if account.get("accountType") != "marketplace":
        return account_id

    marketplace = account.get("marketplace")
    if not isinstance(marketplace, dict):
        return account_id

    legacy_marketplace_id = marketplace.get("legacyMarketplaceID")
    if isinstance(legacy_marketplace_id, int):
        return legacy_marketplace_id
    if isinstance(legacy_marketplace_id, str) and legacy_marketplace_id.isdigit():
        return int(legacy_marketplace_id)
    return account_id


# =============================================================================
# MCP Tools — Auth
# =============================================================================


@mcp.tool()
async def ix_auth_status() -> dict[str, Any]:
    """Check Index Exchange authentication status.

    Returns which auth mode is configured, whether a token is cached,
    and the token expiry timestamp. Never returns raw tokens.

    Returns:
        Dictionary with auth_mode, token_cached, token_expiry, configured.
    """
    client = get_ix_client()
    return {
        "configured": client._is_configured(),
        "auth_mode": client.auth_mode,
        "token_cached": bool(client._access_token),
        "token_expiry": client._token_expiry if client._access_token else None,
        "token_valid": client._token_is_valid(),
    }


@mcp.tool()
async def ix_auth_login(force: bool = False) -> dict[str, Any]:
    """Trigger Index Exchange login and return token expiry.

    Args:
        force: If True, forces a fresh login even if a valid token exists.

    Returns:
        Dictionary with success, auth_mode, token_expiry.
    """
    try:
        client = get_ix_client()
        if force:
            client._access_token = ""
            client._refresh_token = ""
            client._token_expiry = 0.0
        await client._ensure_access_token()
        return {
            "success": True,
            "auth_mode": client.auth_mode,
            "token_expiry": client._token_expiry,
        }
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("login", None, str(e)),
        }


# =============================================================================
# MCP Tools — Deals
# =============================================================================


@mcp.tool()
async def ix_list_dsps(valid_for_class_id: int | None = None) -> dict[str, Any]:
    """List available DSPs (Demand Side Platforms) from Index Exchange.

    Args:
        valid_for_class_id: Optional class ID to filter DSPs.

    Returns:
        Dictionary with success and dsps list, or error.
    """
    try:
        client = get_ix_client()
        params: dict[str, Any] = {}
        if valid_for_class_id is not None:
            params["validForClassID"] = valid_for_class_id
        data = await client.request("GET", "/api/deals/v1/dsps", params=params or None)
        return {"success": True, "dsps": data}
    except Exception as e:
        return {"success": False, "error": _make_error("list_dsps", None, str(e))}


@mcp.tool()
async def ix_list_dsp_seats(
    dsp_id: int,
    name_like: str | None = None,
    seat_id_like: str | None = None,
) -> dict[str, Any]:
    """List seats for a specific DSP.

    Major DSPs can return tens of thousands of seats (e.g. Amazon's
    catalog is ~1MB of JSON). Pass `name_like` or `seat_id_like` to
    filter server-side before returning — both substring match,
    case-insensitive.

    Args:
        dsp_id: The DSP identifier.
        name_like: Substring filter against the seat's `name` /
            `seatName` field. Use when the caller knows the
            human-readable seat name (e.g. "GroupM - Xaxis").
        seat_id_like: Substring filter against `seatID` and
            `extendedSeatID`. Use when the caller knows the numeric
            seatID (e.g. "5030037") or extendedSeatID (e.g.
            "AMZATKFYXZ39AR77").

    Returns:
        Dictionary with success and seats list, or error.
    """
    try:
        client = get_ix_client()
        data = await client.request("GET", f"/api/deals/v1/dsps/{dsp_id}/seats")
        # The API returns either a bare list or {"seats": [...]}. Normalize
        # to the inner list for filtering, then preserve the original
        # envelope shape on the way out.
        if isinstance(data, dict) and isinstance(data.get("seats"), list):
            seats = data["seats"]
        elif isinstance(data, list):
            seats = data
        else:
            seats = []

        def _matches(seat: dict[str, Any]) -> bool:
            if name_like:
                needle = _normalize_lookup_text(name_like)
                names = [seat.get("name"), seat.get("seatName")]
                if not any(needle in _normalize_lookup_text(n) for n in names if n is not None):
                    return False
            if seat_id_like:
                needle = _normalize_lookup_text(seat_id_like)
                ids = [seat.get("seatID"), seat.get("extendedSeatID")]
                if not any(needle in _normalize_lookup_text(i) for i in ids if i is not None):
                    return False
            return True

        if name_like or seat_id_like:
            seats = [s for s in seats if isinstance(s, dict) and _matches(s)]

        # Preserve the {"seats": {"seats": [...]}} shape the IX API returns
        # for some accounts, so callers that destructure either form keep
        # working. The simple list form is the historical default.
        return {"success": True, "seats": seats}
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("list_dsp_seats", None, str(e)),
        }


@mcp.tool()
async def ix_create_marketplace_deal(
    account_id: int,
    name: str,
    start_date: str,
    floor: float,
    dsp_id: int,
    end_date: str | None = None,
    external_deal_id: str | None = None,
    auction_type: str = "first",
    open_market: bool | None = None,
    seat_ids: list[str] | None = None,
    publisher_ids: list[int] | None = None,
    internal_deal_ids: list[int] | None = None,
    margin: float | None = None,
    margin_calculation_type: str | None = None,
    targeting: list[dict] | None = None,
    labels: dict | None = None,
    domain_source: dict | str | None = None,
    domain_operator: str = "NONE_OF",
    allow_partial_targeting: bool = False,
    expected_domain_count: int | None = None,
    expected_domains_fingerprint: str | None = None,
    deal_type: str | None = None,
) -> dict:
    """Create a Marketplace Package deal (classID=4) via POST /api/deals/v3/deals.

    This tool is for Marketplace Partner accounts only. It bundles inventory from
    multiple publishers, applies targeting, and prices the package for a single buyer DSP.

    Typical workflow:
      1. ix_list_dsps(valid_for_class_id=4)            -> pick dsp_id
      2. ix_list_dsp_seats(dsp_id)                     -> pick seat_ids (required for TTD,
                                                          DV360, Xandr, Quantcast)
      3. ix_list_marketplace_publishers(account_id)    -> pick publisher legacyAccountIDs
      4. ix_list_targeting_keys(account_id)            -> discover targeting key IDs
         (including the segment targeting key)
      5. ix_list_segments(account_id)                  -> pick segment IDs for
         inclusion/exclusion targeting
      6. ix_create_domain_targeting_values(domains, account_id) -> validate/prepare domain list if needed
      7. Call this tool with ALL resolved targeting values:
         - Standard targeting (DeviceType, Country, etc.)
         - Segment targeting (include AND exclude) using the segment key from step 4
         - Domain targeting (inline literals for key 120, or use domain_source)

    Args:
        account_id: The marketplace account ID (integer in the IX UI top-left corner).
        name: Deal name, 1-255 characters.
        external_deal_id: Optional globally unique external deal ID, 3-64 characters.
            Allowed characters: letters, numbers, dashes (-), underscores (_), periods (.).
            Cannot start with "0" or contain spaces. If omitted, a compliant ID is
            auto-generated (example format: "IX17429061325300000").
        start_date: Deal start date in YYYY-MM-DD format. Example: "2026-03-01".
        floor: Minimum bid price. Must be >= 0.10 for Marketplace Packages.
        dsp_id: DSP integer ID obtained from ix_list_dsps(valid_for_class_id=4).
        end_date: Deal end date in YYYY-MM-DD format. Must be >= start_date. When
            omitted, defaults to a rolling end date 24 months (2 years) after start_date.
        auction_type: "first" (first-price, default) or
            "fixed" (winning bidder pays the floor price regardless of bid).
        open_market: Whether the deal competes with open market bids. When omitted
            (default), the field is excluded from the request body entirely—required
            for accounts where openMarket is an excluded/unsupported field.
        seat_ids: Optional list of DSP extended seat ID strings from ix_list_dsp_seats.
            Required for: The Trade Desk, Google Display & Video 360, Xandr, Quantcast.
        publisher_ids: Optional list of marketplace publisher legacyAccountIDs to target
            directly inside this Marketplace Package. Resolve them via
            ix_list_marketplace_publishers(account_id). If publisher_ids is provided,
            it targets publishers directly instead of relying on a targeting-key
            workaround like publisherid in the targeting list.
        internal_deal_ids: Optional list of classID=5 internalDealIDs to target inside
            this Marketplace Package. Resolve them via ix_list_deals_v3 with
            class_ids=[5] and targeted_marketplace_account_ids=[account_id]. If
            internal_deal_ids is provided, it targets deals directly instead of relying
            on a targeting-key workaround like InboundDealID.
        margin: Optional deal-specific Marketplace Owner fee. E.g. 5.5 for 5.5%.
            If omitted, the account default margin is used.
        margin_calculation_type: Required when margin is provided. "P" = percentage of
            winning bid, "C" = flat CPM fee.
        targeting: Optional list of v3 targeting objects. At least one of targeting,
            publisher_ids, internal_deal_ids, or segments must
            be specified to create a valid Marketplace Package.

            IMPORTANT: publisher_ids and internal_deal_ids are mutually exclusive per
            the IX Marketplace Package workflow. If targeting deals, do not also target
            publishers. If targeting publishers, do not also target deals.

            Each targeting object shape:
            {
              "targetingKeyID": 9,         # Required integer targeting key ID
              "keyName": "Country",          # See Targeting Keys Reference in API docs
              "targetingType": "standard",   # "standard" or "custom"
              "sets": [
                {
                  "operator": "ANY_OF",      # "ANY_OF" = include, "NONE_OF" = exclude
                  "values": [
                    {"value": "USA"},        # create-accepted value token string
                    {"value": "456"}
                  ]
                }
              ]
            }
            Values must be create-accepted tokens represented as strings.
            Numeric targetingValueID strings are accepted as compatibility input for
            standard keys and translated to create tokens automatically when possible.
            For Domain/AppBundle key 120, pass domain/app-bundle literals (not IDs).

            Known targetingKeyID references:
              - 3: DeviceType
              - 8: Viewability
              - 9: Country
              - 10: creativeTypeSize (use for OLV/Video: value "Video (all sizes)")
              - 11: contentGenre (use for IAB Categories/Entertainment)
              - 120: Domain and app bundle (for package deals)

            Audience Segment targeting:
              Segments from ix_list_segments use a SEPARATE targeting mechanism.
              Use the segment's targetingKeyID and keyName as returned by
              ix_list_targeting_keys (commonly key "Segment" or similar).
              Pass the segment ID as a string value. Use operator "ANY_OF" for
              segment inclusion and "NONE_OF" for segment exclusion.
              Example segment inclusion:
                {"targetingKeyID": <segment_key_id>, "keyName": "Segment",
                 "targetingType": "standard",
                 "sets": [{"operator": "ANY_OF",
                           "values": [{"value": "280"}, {"value": "3007"}]}]}
              Example segment exclusion:
                {"targetingKeyID": <segment_key_id>, "keyName": "Segment",
                 "targetingType": "standard",
                 "sets": [{"operator": "NONE_OF",
                           "values": [{"value": "308129"}]}]}

            NOTE: Exact Deal with Marketplaces / Deals with Publishers targeting should
            prefer the dedicated internal_deal_ids argument when available rather than
            attempting to encode those deals through a targeting key like InboundDealID.

            NOTE: Exact marketplace publisher targeting should prefer the dedicated
            publisher_ids argument when available rather than attempting to encode those
            publishers through a targeting key like publisherid in the targeting list.

        labels: Optional reporting labels dict. Accepted keys (all optional, string|null):
            advertiser, agency, custom, externalReferenceID, salesperson.
            The "custom" field can store custom metadata labels.
        domain_source: Optional server-side domain source for key 120 targeting, used to avoid
            oversized inline payloads. Supported values:
              - "/absolute/or/relative/path/to/domains.csv"
              - {"file_path": "/path/to/domains.csv"}
              - {"type": "file_path", "file_path": "/path/to/domains.csv"}

            When provided, the server reads the file and injects key 120 values into
            create payload internally. Do not include key 120 in targeting at the same time.
        domain_operator: Operator for file-driven domain targeting. "NONE_OF" (default,
            exclude/blocklist) or "ANY_OF" (include/allowlist). Only used when
            domain_source is provided. For inline targeting, set the operator directly
            in the targeting object.
        allow_partial_targeting: Default false (strict mode). When false, create is
            blocked if domain normalization finds invalid key 120 values. When true,
            invalid domain entries are dropped and create proceeds with valid subset.
        expected_domain_count: Optional hard guardrail for file-driven workflows.
            If provided, create is blocked unless the submitted key 120 domain count
            exactly matches this value.
        expected_domains_fingerprint: Optional hard guardrail for file-driven workflows.
            SHA-256 fingerprint for normalized domain literals. If provided, create is
            blocked unless the submitted domain fingerprint exactly matches this value.
        deal_type: Optional hint to force deal-type targeting defaults. Accepted values:
            "display", "olv" (treated as display/OLV), or "ctv". When set, the MCP
            auto-fills device types, inventory channels (display/OLV), and creative
            format (Banner (all sizes) for display/OLV, Video (all sizes) for CTV)
            even if the caller did not specify them. Inferred from device targeting
            when not set.

    Returns:
        On success: {
            "success": True,
            "deal": <API response dict>,
            "internal_deal_id": int,
            "deal_url": str | None,
        }
        On failure: {"success": False, "error": <error dict>}
    """
    try:
        if external_deal_id is None:
            external_deal_id = generate_external_deal_id()

        if not end_date:
            end_date = _default_rolling_end_date(start_date)

        if publisher_ids is not None and internal_deal_ids is not None:
            return {
                "success": False,
                "error": _make_error(
                    "create_marketplace_deal",
                    None,
                    "publisher_ids and internal_deal_ids are mutually exclusive; target publishers or deals, not both.",
                ),
            }

        normalized_input_targeting = copy.deepcopy(targeting) if targeting is not None else None

        client = get_ix_client()

        direct_targeting_metadata: dict[str, Any] = {}
        if internal_deal_ids is not None:
            direct_targeting_metadata["internal_deal_ids"] = [str(value) for value in internal_deal_ids]
        if publisher_ids is not None:
            direct_targeting_metadata["publisher_ids"] = [str(value) for value in publisher_ids]

        has_inline_domain_targeting = _has_domain_targeting(normalized_input_targeting)

        if domain_source is not None and has_inline_domain_targeting:
            return {
                "success": False,
                "error": _make_error(
                    "create_marketplace_deal",
                    None,
                    "domain_source cannot be used together with inline key 120 targeting. Use one mode only.",
                ),
            }

        if has_inline_domain_targeting:
            inline_domain_values = _collect_domain_values_in_targeting(normalized_input_targeting)
            inline_domain_count = len(inline_domain_values)
            if inline_domain_count > MAX_INLINE_DOMAIN_VALUES:
                return {
                    "success": False,
                    "error": _make_error(
                        "create_marketplace_deal",
                        None,
                        "Inline domain targeting payload is too large for reliable transport. "
                        f"inline_domain_count={inline_domain_count}, max_inline_domain_values={MAX_INLINE_DOMAIN_VALUES}. "
                        "Use domain_source.file_path instead.",
                    ),
                    "domain_diagnostics": {
                        "inline_domain_values": inline_domain_count,
                        "max_inline_domain_values": MAX_INLINE_DOMAIN_VALUES,
                    },
                }

        domain_source_metadata: dict[str, Any] = {}
        if domain_source is not None:
            source_domains, domain_source_metadata = _load_domains_from_source(domain_source)
            targeting_list = normalized_input_targeting or []
            targeting_list.append(_build_domain_targeting_object(source_domains, operator=domain_operator))
            normalized_input_targeting = targeting_list

        normalized_input_targeting = await _ensure_deal_type_targeting_defaults(
            client, account_id, normalized_input_targeting, deal_type=deal_type
        )

        normalized_targeting, domain_stats = await _normalize_targeting_for_create(
            client, account_id, normalized_input_targeting, allow_partial_targeting=allow_partial_targeting
        )
        if domain_stats is not None and domain_source_metadata:
            domain_stats.update(domain_source_metadata)

        has_domain_targeting = _has_domain_targeting(normalized_targeting)
        submitted_domain_values = _collect_domain_values_in_targeting(normalized_targeting)
        submitted_domain_count = len(submitted_domain_values)
        submitted_domains_fingerprint = _domain_values_fingerprint(submitted_domain_values)

        if has_domain_targeting and ((expected_domain_count is None) != (expected_domains_fingerprint is None)):
            return {
                "success": False,
                "error": _make_error(
                    "create_marketplace_deal",
                    None,
                    "expected_domain_count and expected_domains_fingerprint must be provided together.",
                ),
                "domain_diagnostics": {
                    **(domain_stats or {}),
                    "expected_domain_count": expected_domain_count,
                    "expected_domains_fingerprint": expected_domains_fingerprint,
                    "submitted_domain_values": submitted_domain_count,
                    "submitted_domains_fingerprint": submitted_domains_fingerprint,
                    "submitted_domain_sample": submitted_domain_values[:10],
                },
            }

        retry_lock_key = (account_id, external_deal_id)
        if has_domain_targeting and expected_domain_count is not None and expected_domains_fingerprint is not None:
            normalized_expected_fingerprint = expected_domains_fingerprint.strip().lower()
            locked_expected = _FILE_DOMAIN_EXPECTATION_LOCKS.get(retry_lock_key)
            if locked_expected is None:
                _FILE_DOMAIN_EXPECTATION_LOCKS[retry_lock_key] = (
                    expected_domain_count,
                    normalized_expected_fingerprint,
                )
            elif locked_expected != (expected_domain_count, normalized_expected_fingerprint):
                locked_count, locked_fingerprint = locked_expected
                return {
                    "success": False,
                    "error": _make_error(
                        "create_marketplace_deal",
                        None,
                        "Expected domain parity values changed across retries for this deal key; refusing to proceed.",
                    ),
                    "domain_diagnostics": {
                        **(domain_stats or {}),
                        "deal_parity_lock_key": {
                            "account_id": account_id,
                            "external_deal_id": external_deal_id,
                        },
                        "locked_expected_domain_count": locked_count,
                        "locked_expected_domains_fingerprint": locked_fingerprint,
                        "expected_domain_count": expected_domain_count,
                        "expected_domains_fingerprint": expected_domains_fingerprint,
                        "submitted_domain_values": submitted_domain_count,
                        "submitted_domains_fingerprint": submitted_domains_fingerprint,
                        "submitted_domain_sample": submitted_domain_values[:10],
                    },
                }

        if expected_domain_count is not None and submitted_domain_count != expected_domain_count:
            return {
                "success": False,
                "error": _make_error(
                    "create_marketplace_deal",
                    None,
                    "Domain count mismatch for file-driven targeting."
                    f" expected_domain_count={expected_domain_count}, submitted_domain_count={submitted_domain_count}",
                ),
                "domain_diagnostics": {
                    **(domain_stats or {}),
                    "deal_parity_lock_key": {
                        "account_id": account_id,
                        "external_deal_id": external_deal_id,
                    },
                    "expected_domain_count": expected_domain_count,
                    "expected_domains_fingerprint": expected_domains_fingerprint,
                    "submitted_domain_values": submitted_domain_count,
                    "submitted_domains_fingerprint": submitted_domains_fingerprint,
                    "submitted_domain_sample": submitted_domain_values[:10],
                },
            }

        if (
            expected_domains_fingerprint is not None
            and submitted_domains_fingerprint.lower() != expected_domains_fingerprint.strip().lower()
        ):
            return {
                "success": False,
                "error": _make_error(
                    "create_marketplace_deal",
                    None,
                    "Domain fingerprint mismatch for file-driven targeting."
                    f" expected_domains_fingerprint={expected_domains_fingerprint},"
                    f" submitted_domains_fingerprint={submitted_domains_fingerprint}",
                ),
                "domain_diagnostics": {
                    **(domain_stats or {}),
                    "deal_parity_lock_key": {
                        "account_id": account_id,
                        "external_deal_id": external_deal_id,
                    },
                    "expected_domain_count": expected_domain_count,
                    "expected_domains_fingerprint": expected_domains_fingerprint,
                    "submitted_domain_values": submitted_domain_count,
                    "submitted_domains_fingerprint": submitted_domains_fingerprint,
                    "submitted_domain_sample": submitted_domain_values[:10],
                },
            }

        payload = _build_marketplace_deal_payload(
            account_id=account_id,
            name=name,
            external_deal_id=external_deal_id,
            start_date=start_date,
            floor=floor,
            dsp_id=dsp_id,
            end_date=end_date,
            auction_type=auction_type,
            open_market=open_market,
            seat_ids=seat_ids,
            publisher_ids=publisher_ids,
            internal_deal_ids=internal_deal_ids,
            margin=margin,
            margin_calculation_type=margin_calculation_type,
            targeting=normalized_targeting,
            labels=labels,
        )
    except ValueError as e:
        return {"success": False, "error": _make_error("create_marketplace_deal", None, str(e))}

    logger.debug(
        "Creating v3 Marketplace Package deal: name=%r externalDealID=%r floor=%s dspID=%s",
        name,
        external_deal_id,
        floor,
        dsp_id,
    )

    try:
        data = await client.request("POST", "/api/deals/v3/deals", json_body=payload)
        internal_deal_id = data.get("internalDealID")
        deal_url = None
        if isinstance(internal_deal_id, int) or (isinstance(internal_deal_id, str) and internal_deal_id.isdigit()):
            deal_url = f"{client.base_url}/deals/{internal_deal_id}/show?account_id={account_id}"

        submitted_domain_count = _count_domain_values_in_targeting(normalized_targeting)
        persisted_domain_count = None
        verification = None
        quality_flags: list[dict[str, Any]] = []
        persisted_domains_fingerprint = None

        has_direct_targeting = bool(direct_targeting_metadata.get("internal_deal_ids")) or bool(
            direct_targeting_metadata.get("publisher_ids")
        )
        needs_post_create_verification = submitted_domain_count > 0 or has_direct_targeting

        if internal_deal_id is not None and needs_post_create_verification:
            try:
                persisted_deal = await _fetch_deal_with_direct_target_retries(
                    client,
                    internal_deal_id=internal_deal_id,
                    expect_internal_deal_ids=bool(direct_targeting_metadata.get("internal_deal_ids")),
                    expect_publisher_ids=bool(direct_targeting_metadata.get("publisher_ids")),
                )
                verification = {}

                persisted_targeting_source = persisted_deal.get("targeting")
                inventory_group = None
                inventory_group_targeting = None

                if submitted_domain_count > 0:
                    try:
                        inventory_group = await _find_deal_specific_inventory_group(
                            client,
                            account_id=account_id,
                            external_deal_id=external_deal_id,
                            dsp_id=dsp_id,
                        )
                    except Exception as inventory_group_err:
                        logger.info(
                            "Inventory group lookup unavailable for domain verification: %s", inventory_group_err
                        )
                        inventory_group = None
                    inventory_group_targeting = (
                        inventory_group.get("targeting") if isinstance(inventory_group, dict) else None
                    )
                    if isinstance(inventory_group, dict):
                        verification["inventory_group_id"] = inventory_group.get("inventoryGroupID")
                        verification["inventory_group_name"] = inventory_group.get("name")
                    if isinstance(inventory_group_targeting, list) and not _collect_domain_values_in_targeting(
                        persisted_deal.get("targeting")
                    ):
                        persisted_targeting_source = inventory_group_targeting

                if submitted_domain_count > 0:
                    if isinstance(inventory_group, dict):
                        persisted_domain_values = _collect_domain_values_from_inventory_group(inventory_group)
                        persisted_domain_count = len(persisted_domain_values)
                    else:
                        persisted_domain_count = _count_domain_values_in_targeting(persisted_targeting_source)
                        persisted_domain_values = _collect_domain_values_in_targeting(persisted_targeting_source)
                    persisted_domains_fingerprint = _domain_values_fingerprint(persisted_domain_values)
                    parity_ok = persisted_domain_count == submitted_domain_count
                    fingerprint_parity_ok = persisted_domains_fingerprint == submitted_domains_fingerprint
                    verification.update(
                        {
                            "domain_count_parity": parity_ok,
                            "domain_fingerprint_parity": fingerprint_parity_ok,
                            "submitted_domain_values": submitted_domain_count,
                            "persisted_domain_values": persisted_domain_count,
                            "submitted_domains_fingerprint": submitted_domains_fingerprint,
                            "persisted_domains_fingerprint": persisted_domains_fingerprint,
                        }
                    )
                    if not parity_ok:
                        quality_flags.append(
                            {
                                "flag": "domain_count_parity_failed",
                                "impact": "Persisted Domain/AppBundle targeting count differs from submitted payload",
                                "submitted_domain_values": submitted_domain_count,
                                "persisted_domain_values": persisted_domain_count,
                            }
                        )
                    if not fingerprint_parity_ok:
                        quality_flags.append(
                            {
                                "flag": "domain_fingerprint_parity_failed",
                                "impact": "Persisted Domain/AppBundle targeting fingerprint differs from submitted payload",
                                "submitted_domains_fingerprint": submitted_domains_fingerprint,
                                "persisted_domains_fingerprint": persisted_domains_fingerprint,
                            }
                        )

                persisted_internal_ids = _collect_direct_ids_from_deal(
                    persisted_deal,
                    explicit_fields=("internalDealIDs",),
                    list_fields=("internalDeals", "deals"),
                    list_item_keys=("internalDealID", "dealID", "id"),
                    targeting_key_name="internaldealid",
                )
                if "internal_deal_ids" in direct_targeting_metadata:
                    submitted_internal_ids = sorted(direct_targeting_metadata["internal_deal_ids"] or [])
                    persisted_internal_ids_sorted = sorted(persisted_internal_ids)
                    verification["internal_deal_ids_submitted"] = submitted_internal_ids
                    verification["internal_deal_ids_persisted"] = persisted_internal_ids_sorted
                    verification["internal_deal_ids_visible"] = bool(persisted_internal_ids_sorted)
                    if persisted_internal_ids_sorted:
                        internal_ids_ok = persisted_internal_ids_sorted == submitted_internal_ids
                        verification["internal_deal_ids_parity"] = internal_ids_ok
                    else:
                        verification["internal_deal_ids_parity"] = False
                        quality_flags.append(
                            {
                                "flag": "internal_deal_id_visibility_failed",
                                "impact": "Persisted internaldealid targeting was not visible after post-create verification retries",
                                "submitted_internal_deal_ids": submitted_internal_ids,
                            }
                        )
                    if persisted_internal_ids_sorted and not verification["internal_deal_ids_parity"]:
                        quality_flags.append(
                            {
                                "flag": "internal_deal_id_parity_failed",
                                "impact": "Persisted internaldealid targeting differs from submitted internal_deal_ids",
                                "submitted_internal_deal_ids": submitted_internal_ids,
                                "persisted_internal_deal_ids": persisted_internal_ids_sorted,
                            }
                        )

                persisted_publisher_ids = _collect_direct_ids_from_deal(
                    persisted_deal,
                    explicit_fields=("publisherIDs",),
                    list_fields=("publishers",),
                    list_item_keys=("legacyAccountID", "publisherID", "accountID", "id"),
                    targeting_key_name="publisherid",
                )
                if "publisher_ids" in direct_targeting_metadata:
                    submitted_publisher_ids = sorted(direct_targeting_metadata["publisher_ids"] or [])
                    persisted_publisher_ids_sorted = sorted(persisted_publisher_ids)
                    verification["publisher_ids_submitted"] = submitted_publisher_ids
                    verification["publisher_ids_persisted"] = persisted_publisher_ids_sorted
                    verification["publisher_ids_visible"] = bool(persisted_publisher_ids_sorted)
                    if persisted_publisher_ids_sorted:
                        publisher_ids_ok = persisted_publisher_ids_sorted == submitted_publisher_ids
                        verification["publisher_ids_parity"] = publisher_ids_ok
                    else:
                        verification["publisher_ids_parity"] = False
                        # IX caches publisher targeting separately from the deal
                        # body; it often takes longer than our verification
                        # retry window (~30s) to appear on the GET. The deal
                        # itself is created correctly — this is a read-visibility
                        # lag, not a write failure. Tag as `info` so downstream
                        # renderers and the agent's summary don't surface it as
                        # an error.
                        quality_flags.append(
                            {
                                "flag": "publisher_id_visibility_pending",
                                "severity": "info",
                                "impact": (
                                    "Publisher targeting was submitted to the create endpoint but is not yet "
                                    "visible on the read-back. IX caches publisher targeting separately; the "
                                    "deal is created correctly. Confirm in the IX UI after ~1 minute if needed."
                                ),
                                "submitted_publisher_ids": submitted_publisher_ids,
                            }
                        )
                    if persisted_publisher_ids_sorted and not verification["publisher_ids_parity"]:
                        quality_flags.append(
                            {
                                "flag": "publisher_id_parity_failed",
                                "impact": "Persisted publisherid targeting differs from submitted publisher_ids",
                                "submitted_publisher_ids": submitted_publisher_ids,
                                "persisted_publisher_ids": persisted_publisher_ids_sorted,
                            }
                        )

                persisted_targeting_for_checks = persisted_deal.get("targeting")
                if not isinstance(persisted_targeting_for_checks, list) and isinstance(inventory_group_targeting, list):
                    persisted_targeting_for_checks = inventory_group_targeting

                if isinstance(persisted_targeting_for_checks, list):
                    submitted_targeting_checks = {
                        "devicetype": sorted(_collect_values_for_key(normalized_targeting, "devicetype")),
                        "country": sorted(_collect_values_for_key(normalized_targeting, "country")),
                        "creativetypesize": sorted(_collect_values_for_key(normalized_targeting, "creativetypesize")),
                    }
                    if submitted_domain_count > 0 and not any(submitted_targeting_checks.values()):
                        submitted_targeting_checks = {}
                    persisted_targeting_checks = {
                        key_name: sorted(_collect_values_for_key(persisted_targeting_for_checks, key_name))
                        for key_name in submitted_targeting_checks
                    }
                    verification["persisted_targeting_source"] = (
                        "inventory_group" if persisted_targeting_for_checks == inventory_group_targeting else "deal"
                    )
                    verification["targeting_checks"] = {
                        key_name: {
                            "submitted": submitted_targeting_checks[key_name],
                            "persisted": persisted_targeting_checks[key_name],
                            "parity": submitted_targeting_checks[key_name] == persisted_targeting_checks[key_name],
                        }
                        for key_name in submitted_targeting_checks
                    }
                    for key_name, targeting_check in verification["targeting_checks"].items():
                        if not targeting_check["parity"]:
                            quality_flags.append(
                                {
                                    "flag": f"{key_name}_parity_failed",
                                    "impact": f"Persisted {key_name} targeting differs from submitted targeting.",
                                    "submitted_values": targeting_check["submitted"],
                                    "persisted_values": targeting_check["persisted"],
                                }
                            )

                if submitted_domain_count > 0 and isinstance(inventory_group, dict):
                    verification["domain_persisted_targeting_source"] = "inventory_group"
            except Exception as verify_err:
                quality_flags.append(
                    {
                        "flag": "post_create_verification_failed",
                        "impact": str(verify_err),
                    }
                )

        if domain_stats is not None:
            domain_stats["expected_domain_count"] = expected_domain_count
            domain_stats["expected_domains_fingerprint"] = expected_domains_fingerprint
            domain_stats["submitted_domain_values"] = submitted_domain_count
            domain_stats["submitted_domains_fingerprint"] = submitted_domains_fingerprint
            if persisted_domain_count is not None:
                domain_stats["persisted_domain_values"] = persisted_domain_count
            if persisted_domains_fingerprint is not None:
                domain_stats["persisted_domains_fingerprint"] = persisted_domains_fingerprint

        enforce_strict_post_create_parity = (
            not allow_partial_targeting or expected_domain_count is not None or expected_domains_fingerprint is not None
        )

        # Info-severity flags are informational (e.g. read-back caching lag)
        # and must not flip verification_success to False — they're not a
        # signal that anything is wrong with the persisted deal.
        verification_success = not any(flag.get("severity", "warning") != "info" for flag in quality_flags)
        strict_failure_required = _quality_flags_require_strict_failure(
            quality_flags,
            expected_domain_count=expected_domain_count,
            expected_domains_fingerprint=expected_domains_fingerprint,
        )

        if quality_flags and enforce_strict_post_create_parity and strict_failure_required:
            return {
                "success": False,
                "error": _make_error(
                    "create_marketplace_deal",
                    None,
                    "Deal created but post-create verification failed in strict mode.",
                    details={"quality_flags": quality_flags},
                ),
                "deal": data,
                "internal_deal_id": internal_deal_id,
                "deal_url": deal_url,
                "domain_diagnostics": domain_stats,
                "verification": verification,
                "verification_success": verification_success,
                "quality_flags": quality_flags,
            }

        return {
            "success": True,
            "deal": data,
            "internal_deal_id": internal_deal_id,
            "deal_url": deal_url,
            "domain_diagnostics": domain_stats,
            "verification": verification,
            "verification_success": verification_success,
            "quality_flags": quality_flags,
            "warnings": [flag.get("impact") for flag in quality_flags if isinstance(flag.get("impact"), str)],
        }
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("create_marketplace_deal", None, str(e)),
        }


def _resolve_segment_with_prefix_tolerance(
    segments: list[dict[str, Any]],
    segment_name: str,
) -> dict[str, Any]:
    """Match a segment name against the catalog, tolerating publisher prefixes.

    Trader prompts often namespace segments under the data partner's name
    (e.g. ``"The Weather Company > Weather Targeting > X"``) but the IX
    catalog stores them without that prefix (``"Weather Targeting > X"``).
    Try the full name first, then progressively strip leading ``" > "``
    components and retry. Raises LookupError on no unique match — same
    contract as ``_resolve_unique_match``.
    """
    candidates: list[str] = [segment_name]
    parts = [p.strip() for p in segment_name.split(" > ") if p.strip()]
    # Add each progressively-shorter suffix as a candidate (skip the bare
    # last token unless the full name is a single token — too generic
    # otherwise, e.g. "Sunny" matches dozens of catalog entries).
    for i in range(1, len(parts)):
        candidates.append(" > ".join(parts[i:]))

    last_error: Exception | None = None
    seen: set[str] = set()
    for candidate in candidates:
        if not candidate or candidate in seen:
            continue
        seen.add(candidate)
        try:
            return _resolve_unique_match(
                segments,
                candidate,
                "segment",
                lookup_fields=("externalSegmentName", "name", "segmentName", "label", "id", "segmentID"),
                allow_contains_match=True,
            )
        except (LookupError, ValueError) as exc:
            last_error = exc
            continue
    raise last_error if last_error is not None else LookupError(f"No match found for segment: {segment_name}")


@mcp.tool()
async def ix_execute_deal_from_prompt_inputs(
    name: str,
    start_date: str,
    floor: float,
    dsp_name: str,
    account_id: int | str | None = None,
    end_date: str | None = None,
    external_deal_id: str | None = None,
    seat_name: str | None = None,
    publisher_names: list[str] | None = None,
    publisher_ids: list[int] | None = None,
    deals_with_publishers: list[str] | None = None,
    segment_names: list[str] | None = None,
    excluded_segment_names: list[str] | None = None,
    domain_file_path: str | None = None,
    domain_sheet: str | None = None,
    domain_column: str | None = None,
    domain_match_operator: str = "blocklist",
    app_bundle_file_path: str | None = None,
    app_bundle_sheet: str | None = None,
    app_bundle_column: str | None = None,
    app_bundle_match_operator: str | None = None,
    geo_countries: list[str] | None = None,
    geo_states: list[str] | None = None,
    dma_codes: list[str] | None = None,
    device_types: list[str] | None = None,
    iab_categories: list[str] | None = None,
    viewability_threshold: int | None = None,
    margin_percent: float | None = None,
    auction_type: str = "first",
    open_market: bool | None = None,
    labels: dict | None = None,
    deal_type: str | None = None,
) -> dict[str, Any]:
    """Resolve human-readable IX deal inputs and create a marketplace deal in one call.

    Deal-type defaults
    ------------------
    Pass `deal_type` to apply the Index UI's canonical General-settings defaults
    even when the prompt omits them. All four canonical channels are supported:

    - `deal_type="display"` → devices Personal computer + Phone + Tablet,
      inventory In-App + Web, creative Banner (all sizes).
    - `deal_type="olv"` → same devices/inventory as display, but creative
      Video (all sizes). OLV is NOT a Banner variant — it is web+app video.
    - `deal_type="ctv"` → devices Connected TV + Connected device + Set-top
      box, inventory In-App only, creative Video (all sizes).
    - `deal_type="ott"` → devices Phone + Tablet (mobile-app video, NOT
      CTV devices), inventory In-App only, creative Video (all sizes).

    When `deal_type` is omitted the type is inferred from any `device_types`
    passed in (e.g. "CTV" triggers CTV defaults) — but OLV and OTT cannot be
    inferred from devices alone (they share device sets with display/mobile),
    so pass `deal_type` explicitly for those two.

    When `end_date` is omitted it defaults to a rolling end date 24 months
    (2 years) after `start_date`.

    Segment targeting
    -----------------
    `segment_names` resolves to ANY_OF (inclusion); `excluded_segment_names`
    resolves to NONE_OF (exclusion) on the same `keyName: "segmentid"`
    targeting object. Both lists tolerate publisher prefixes (e.g.
    `"The Weather Company > Weather Targeting > X"` matches a catalog
    entry stored as `"Weather Targeting > X"`). IX requires segments at
    create time — they cannot be added later via the API.

    Domain / app-bundle targeting
    -----------------------------
    Web-domain lists and CTV/app-bundle lists both land on the same key-120
    `Domain` targeting object, but use DIFFERENT parameters so the values are
    validated correctly. A deal supplies one or the other (passing both raises):

    - Web domains (Display/OLV): `domain_file_path` (+ optional `domain_sheet` /
      `domain_column`), operator via `domain_match_operator`.
    - App bundles (CTV/OTT): `app_bundle_file_path` (+ optional
      `app_bundle_sheet` / `app_bundle_column`), operator via
      `app_bundle_match_operator`. App-bundle mode accepts both reverse-DNS
      bundle IDs (`com.zumobi.msnbc`) AND bare numeric store IDs (Roku/Apple/
      Amazon, e.g. `523428113`) — the domain path rejects the latter, so a
      bundle list MUST go through `app_bundle_file_path` to avoid silently
      dropping numeric IDs.

    Both operators take `"allowlist"` (aliases: `allow`, `include`, `ANY_OF`) →
    serve ONLY the listed values (inclusion), or `"blocklist"` (aliases:
    `block`, `exclude`, `NONE_OF`, the default) → serve everywhere EXCEPT the
    listed values (exclusion). The resolved wire operator and the count of
    dropped values are echoed back as warnings whenever a file is loaded, so the
    caller can verify intent against the persisted deal.

    Geo targeting
    -------------
    `geo_countries` and `geo_states` build country / region targeting.
    `dma_codes` builds Designated Market Area targeting on IX's `ZipCode`
    key (#781 — the same key also accepts ZIP/postal codes; pass the
    Nielsen DMA number such as "602" for Chicago).

    Seat handling
    -------------
    `seat_name` accepts the human-readable seat name (e.g. `"GroupM -
    Xaxis"`), the numeric `seatID` string (e.g. `"5030037"`), or the
    `extendedSeatID` string (e.g. `"AMZATKFYXZ39AR77"`). The MCP resolves
    the input against the DSP's seat catalog and forwards the
    `extendedSeatID` to the IX create endpoint — which is what the API
    requires regardless of which form the caller supplied.

    External deal ID
    ----------------
    `external_deal_id` sets the top-level `externalDealID` on the created
    deal (3-64 chars; letters, numbers, dash, underscore, period). Use
    this when the deal needs to round-trip back to an upstream order ID
    or trafficking system. Omit to let IX auto-generate one.

    Account selection
    -----------------
    `account_id` accepts a numeric Marketplace account ID, a known account
    name (e.g. `"Reklaim"`, `"Permutive"`, `"Zeta Global"`), or `None`. When
    omitted, the variant default `DEFAULT_MARKETPLACE_ACCOUNT_ID` is used —
    set `INDEXEXCHANGE_MARKETPLACE_ACCOUNT_ID` in the environment (or the
    suffixed variant `INDEXEXCHANGE_MARKETPLACE_ACCOUNT_ID_<CLIENT>` for
    the mcp loader) to switch the default per subprocess.
    """
    logger.info("ix_execute_deal_from_prompt_inputs called with name: %s file: %s", name, domain_file_path)

    warnings: list[str] = []
    pre_create_quality_flags: list[dict[str, Any]] = []

    if account_id is None:
        resolved_account_id = DEFAULT_MARKETPLACE_ACCOUNT_ID
    else:
        try:
            resolved_account_id = _resolve_marketplace_account_id(account_id)
        except ValueError as exc:
            return {
                "success": False,
                "error": _make_error("execute_deal_from_prompt_inputs", None, str(exc)),
            }
    account_id = resolved_account_id

    # Curator-margin default: when caller omits margin_percent, apply the
    # standard 30% Elcano Marketplace Owner fee (Percentage of winning bid).
    # Mirrors the OpenX/PubMatic/Media.net Elcano default-margin behaviour
    # so all four curator-margin-capable SSPs are aligned.
    if margin_percent is None:
        margin_percent = ELCANO_DEFAULT_CURATOR_MARGIN_PERCENT
        warnings.append(
            f"Applied default Elcano curator margin: {ELCANO_DEFAULT_CURATOR_MARGIN_PERCENT:g}% "
            "(Percentage of winning bid). Pass margin_percent= to override."
        )
        pre_create_quality_flags.append(
            {
                "flag": "ix_default_curator_margin_applied",
                "impact": (
                    f"Auto-applied flat {ELCANO_DEFAULT_CURATOR_MARGIN_PERCENT:g}% Marketplace Owner fee for Elcano."
                ),
                "margin_percent": ELCANO_DEFAULT_CURATOR_MARGIN_PERCENT,
                "margin_calculation_type": "P",
            }
        )

    if not end_date:
        end_date = _default_rolling_end_date(start_date)

    try:
        dsps_result = await ix_list_dsps(valid_for_class_id=4)
        if not dsps_result.get("success"):
            raise ValueError(_extract_tool_error_message(dsps_result, "Failed to load DSPs"))
        dsps = [dsp for dsp in dsps_result.get("dsps", []) if isinstance(dsp, dict)]
        dsp = _resolve_unique_match(
            dsps,
            dsp_name,
            "DSP",
            lookup_fields=("name", "dspName", "dspID", "id"),
            aliases=IX_DSP_NAME_ALIASES,
            allow_contains_match=True,
        )
        dsp_id = dsp.get("dspID") or dsp.get("id")
        if not isinstance(dsp_id, int):
            raise ValueError(f"Resolved DSP is missing dspID: {dsp_name}")

        resolved_seat_ids: list[str] | None = None
        if seat_name:
            seat_reference = seat_name.strip()
            if "/" in seat_reference:
                _, right = [part.strip() for part in seat_reference.split("/", 1)]
                if right:
                    seat_reference = right
            # Always resolve through ix_list_dsp_seats. The IX create API
            # requires the `extendedSeatID` string (e.g. "AMZATKFYXZ39AR77"),
            # not the numeric `seatID` — so even when the caller passes a
            # pure-digit seatID we have to look it up and forward the
            # extendedSeatID. The seats catalog stores both fields and the
            # lookup matches either one (or the human-readable `name` field).
            seats_result = await ix_list_dsp_seats(dsp_id=dsp_id)
            if not seats_result.get("success"):
                raise ValueError(_extract_tool_error_message(seats_result, f"Failed to load seats for DSP {dsp_id}"))
            seats_payload = seats_result.get("seats", [])
            if isinstance(seats_payload, dict):
                seats_payload = seats_payload.get("seats", [])
            seats = [seat for seat in seats_payload if isinstance(seat, dict)]
            # For pure-digit input, restrict to exact-match on seatID/extendedSeatID
            # to avoid contains-match collisions with substring-similar IDs
            # (e.g. "100" inside "1001", "10000"). For non-digit input, fall back
            # to the standard fuzzy match path.
            if seat_reference.isdigit():
                seat = _resolve_unique_match(
                    seats,
                    seat_reference,
                    "seat",
                    lookup_fields=("seatID", "extendedSeatID", "id"),
                    allow_contains_match=False,
                )
            else:
                seat = _resolve_unique_match(
                    seats,
                    seat_reference,
                    "seat",
                    lookup_fields=("extendedSeatID", "seatID", "seatName", "name", "id"),
                    allow_contains_match=True,
                )
            seat_id = seat.get("extendedSeatID")
            if seat_id is None:
                seat_id = seat.get("seatID")
            if seat_id is None:
                seat_id = seat.get("id")
            if seat_id is None:
                raise ValueError(f"Resolved seat is missing a usable ID: {seat_name}")
            resolved_seat_ids = [str(seat_id)]

        resolved_publisher_ids = [
            publisher_id for publisher_id in (publisher_ids or []) if isinstance(publisher_id, int)
        ]
        if publisher_names:
            publishers_result = await ix_list_marketplace_publishers(marketplace_account_id=account_id)
            if not publishers_result.get("success"):
                raise ValueError(_extract_tool_error_message(publishers_result, "Failed to load publishers"))
            publishers = [
                publisher for publisher in publishers_result.get("publishers", []) if isinstance(publisher, dict)
            ]
            for publisher_name in publisher_names:
                exact_or_unique = _find_publisher_matches(publishers, publisher_name)
                if len(exact_or_unique["matches"]) == 1:
                    publisher = exact_or_unique["matches"][0]
                    publisher_id = publisher.get("legacyAccountID")
                    if not isinstance(publisher_id, int):
                        publisher_id = publisher.get("accountID")
                    if not isinstance(publisher_id, int):
                        publisher_id = publisher.get("id")
                    if isinstance(publisher_id, int):
                        resolved_publisher_ids.append(publisher_id)
                    continue

                # Zero or many matches — refuse to guess. Emit a structured
                # quality flag listing the candidates and instructing the
                # caller to re-issue with explicit publisher_ids. Policy
                # (which feed variants belong on this deal) stays with the
                # brief, not the tool. See conversation: TWC has 2-3 feed
                # variants and the right combination is account-specific.
                pre_create_quality_flags.append(
                    {
                        "flag": "ix_publisher_resolution_ambiguous"
                        if exact_or_unique["matches"]
                        else "ix_publisher_resolution_not_found",
                        "impact": (
                            f"Could not uniquely resolve publisher_name='{publisher_name}': "
                            f"{len(exact_or_unique['matches'])} candidates. Re-call with explicit "
                            "publisher_ids=[...] to pin the feed variants."
                        ),
                        "requested_name": publisher_name,
                        "candidates": [
                            {
                                "legacyAccountID": cand.get("legacyAccountID"),
                                "accountID": cand.get("accountID"),
                                "name": cand.get("name"),
                                "accountStatus": cand.get("accountStatus"),
                                "currency": cand.get("currency"),
                                "isOptedOutOfMarketplaces": cand.get("isOptedOutOfMarketplaces", False),
                            }
                            for cand in exact_or_unique["matches"][:10]
                        ],
                        "match_strategy": exact_or_unique["strategy"],
                    }
                )
                warnings.append(
                    f"Publisher '{publisher_name}' resolution returned "
                    f"{len(exact_or_unique['matches'])} candidates; pass publisher_ids=[...] to disambiguate."
                )

        resolved_publisher_ids = _dedupe_preserving_order(resolved_publisher_ids)

        resolved_internal_deal_ids: list[int] | None = None
        if deals_with_publishers:
            resolved_internal_deal_ids, resolution_rows, dwp_warnings = await _resolve_dwp_internal_deal_ids(
                account_id,
                deals_with_publishers,
            )
            warnings.extend(dwp_warnings)
            if len(resolution_rows) != len(deals_with_publishers):
                raise ValueError("Deals with Publishers resolution did not return a one-to-one mapping.")

        targeting_keys_result = await ix_list_targeting_keys(account_id=account_id)
        if not targeting_keys_result.get("success"):
            raise ValueError(_extract_tool_error_message(targeting_keys_result, "Failed to load targeting keys"))
        targeting_keys = [key for key in targeting_keys_result.get("targeting_keys", []) if isinstance(key, dict)]

        targeting: list[dict[str, Any]] = []

        if geo_countries:
            country_key = _resolve_targeting_key(targeting_keys, "country")
            country_values = [
                await _resolve_targeting_value_token(
                    account_id,
                    int(country_key.get("targetingKeyID") or country_key.get("keyID")),
                    country,
                    "country",
                    aliases=IX_TARGETING_VALUE_ALIASES["country"],
                )
                for country in geo_countries
            ]
            targeting.append(
                _build_targeting_object(
                    targeting_key_id=int(country_key.get("targetingKeyID") or country_key.get("keyID")),
                    key_name=str(country_key.get("key") or country_key.get("keyName") or "Country"),
                    values=_dedupe_preserving_order(country_values),
                )
            )

        if geo_states:
            state_key = _resolve_targeting_key(targeting_keys, "state")
            state_values = [
                await _resolve_targeting_value_token(
                    account_id,
                    int(state_key.get("targetingKeyID") or state_key.get("keyID")),
                    state,
                    "state",
                )
                for state in geo_states
            ]
            targeting.append(
                _build_targeting_object(
                    targeting_key_id=int(state_key.get("targetingKeyID") or state_key.get("keyID")),
                    key_name=str(state_key.get("key") or state_key.get("keyName") or "State"),
                    values=_dedupe_preserving_order(state_values),
                )
            )

        if dma_codes:
            # IX's `ZipCode` key (#781) accepts both ZIP/postal codes AND
            # Nielsen DMA codes (e.g. "602" for Chicago). The catalog
            # description says "Zip and Post Code" but the API stores DMA
            # entries under the same key. Trader-confirmed via existing
            # deals that use value="602" with keyName="zipcode".
            normalized_dma_codes = _dedupe_preserving_order(
                [str(code).strip() for code in dma_codes if str(code).strip()]
            )
            if normalized_dma_codes:
                targeting.append(
                    _build_targeting_object(
                        targeting_key_id=781,
                        key_name="ZipCode",
                        values=normalized_dma_codes,
                    )
                )

        if device_types:
            device_key = _resolve_targeting_key(targeting_keys, "device_type")
            device_values = [
                await _resolve_targeting_value_token(
                    account_id,
                    int(device_key.get("targetingKeyID") or device_key.get("keyID")),
                    device_type,
                    "device type",
                    aliases=IX_TARGETING_VALUE_ALIASES["device_type"],
                )
                for device_type in device_types
            ]
            targeting.append(
                _build_targeting_object(
                    targeting_key_id=int(device_key.get("targetingKeyID") or device_key.get("keyID")),
                    key_name=str(device_key.get("key") or device_key.get("keyName") or "DeviceType"),
                    values=_dedupe_preserving_order(device_values),
                )
            )

        if viewability_threshold is not None:
            viewability_key = _resolve_targeting_key(targeting_keys, "viewability")
            threshold_value = format(float(viewability_threshold) / 100.0, ".2f")
            viewability_value = await _resolve_targeting_value_token(
                account_id,
                int(viewability_key.get("targetingKeyID") or viewability_key.get("keyID")),
                threshold_value,
                "viewability threshold",
            )
            targeting.append(
                _build_targeting_object(
                    targeting_key_id=int(viewability_key.get("targetingKeyID") or viewability_key.get("keyID")),
                    key_name=str(viewability_key.get("key") or viewability_key.get("keyName") or "Viewability"),
                    values=[viewability_value],
                )
            )

        if iab_categories:
            iab_key = _resolve_targeting_key(targeting_keys, "iab_categories")
            iab_values = [
                await _resolve_targeting_value_token(
                    account_id,
                    int(iab_key.get("targetingKeyID") or iab_key.get("keyID")),
                    category,
                    "IAB category",
                )
                for category in iab_categories
            ]
            targeting.append(
                _build_targeting_object(
                    targeting_key_id=int(iab_key.get("targetingKeyID") or iab_key.get("keyID")),
                    key_name=str(iab_key.get("key") or iab_key.get("keyName") or "contentGenre"),
                    values=_dedupe_preserving_order(iab_values),
                )
            )

        # Segment targeting: includes (ANY_OF) + excludes (NONE_OF). Both
        # land under a single `keyName: "segmentid"` targeting object — the
        # API rejects the `im_segments` key returned by ix_list_targeting_keys
        # for exclusion, and the actual segment-targeting key isn't in the
        # public list. Fetch the catalog once and resolve both lists against
        # it. Names tolerate publisher prefixes (e.g. "The Weather Company > X"
        # → matches "X" in the catalog).
        resolved_segment_include_ids: list[str] = []
        resolved_segment_exclude_ids: list[str] = []
        if segment_names or excluded_segment_names:
            segments_result = await ix_list_segments(account_id=account_id)
            if not segments_result.get("success"):
                raise ValueError(_extract_tool_error_message(segments_result, "Failed to load segments"))
            segments = [segment for segment in segments_result.get("segments", []) if isinstance(segment, dict)]

            def _resolve_segment_id(segment_name: str) -> str:
                segment = _resolve_segment_with_prefix_tolerance(segments, segment_name)
                segment_id = segment.get("id")
                if segment_id is None:
                    segment_id = segment.get("segmentID")
                if segment_id is None:
                    raise ValueError(f"Resolved segment is missing an ID: {segment_name}")
                return str(segment_id)

            for segment_name in segment_names or []:
                resolved_segment_include_ids.append(_resolve_segment_id(segment_name))
            for segment_name in excluded_segment_names or []:
                resolved_segment_exclude_ids.append(_resolve_segment_id(segment_name))

            sets: list[dict[str, Any]] = []
            if resolved_segment_include_ids:
                sets.append(
                    {
                        "operator": "ANY_OF",
                        "values": [{"value": sid} for sid in _dedupe_preserving_order(resolved_segment_include_ids)],
                    }
                )
            if resolved_segment_exclude_ids:
                sets.append(
                    {
                        "operator": "NONE_OF",
                        "values": [{"value": sid} for sid in _dedupe_preserving_order(resolved_segment_exclude_ids)],
                    }
                )
            if sets:
                targeting.append(
                    {
                        "keyName": "segmentid",
                        "targetingType": "standard",
                        "sets": sets,
                    }
                )

        # Domain and app-bundle lists land on the same key-120 Domain targeting
        # object. A deal supplies one or the other; app-bundle mode additionally
        # accepts numeric store IDs (Roku/Apple/Amazon) that a web-domain
        # validator would otherwise drop.
        if domain_file_path and app_bundle_file_path:
            raise ValueError(
                "Provide either domain_file_path or app_bundle_file_path, not both — they target the same key-120 list."
            )

        domain_source = None
        resolved_domain_operator = "NONE_OF"
        list_file_path = app_bundle_file_path or domain_file_path
        if list_file_path:
            is_app_bundle = app_bundle_file_path is not None
            if is_app_bundle:
                list_sheet = app_bundle_sheet
                list_column = app_bundle_column
                list_operator_raw = app_bundle_match_operator if app_bundle_match_operator is not None else "blocklist"
                list_kind = "app-bundle"
                operator_field = "app_bundle_match_operator"
            else:
                list_sheet = domain_sheet
                list_column = domain_column
                list_operator_raw = domain_match_operator
                list_kind = "domain"
                operator_field = "domain_match_operator"

            resolved_domain_operator = _normalize_domain_operator(list_operator_raw)
            file_result = _extract_domains_from_file(
                list_file_path,
                sheet_name=list_sheet,
                column_name=list_column,
                allow_app_bundle_ids=is_app_bundle,
            )
            if not file_result["domains"]:
                raise ValueError(f"No valid {list_kind} values found in file: {list_file_path}")
            if file_result["invalid_values"]:
                warnings.append(
                    f"Dropped {len(file_result['invalid_values'])} non-{list_kind} values from {list_file_path}."
                )
            warnings.append(
                f"Applied {list_kind} targeting operator {resolved_domain_operator} "
                f"({operator_field}={list_operator_raw!r}) to "
                f"{len(file_result['domains'])} values from {list_file_path}."
            )
            domain_source = {
                "file_path": list_file_path,
                "sheet_name": list_sheet,
                "column_name": list_column,
                "allow_app_bundle_ids": is_app_bundle,
            }

        create_result = await ix_create_marketplace_deal(
            account_id=account_id,
            name=name,
            start_date=start_date,
            floor=floor,
            dsp_id=dsp_id,
            end_date=end_date,
            external_deal_id=external_deal_id,
            auction_type=auction_type,
            open_market=open_market,
            seat_ids=resolved_seat_ids,
            publisher_ids=resolved_publisher_ids or None,
            internal_deal_ids=resolved_internal_deal_ids,
            margin=margin_percent,
            margin_calculation_type="P" if margin_percent is not None else None,
            targeting=targeting or None,
            labels=labels,
            domain_source=domain_source,
            domain_operator=resolved_domain_operator,
            deal_type=deal_type,
        )
        create_warnings = create_result.get("warnings")
        if isinstance(create_warnings, list):
            warnings.extend(str(warning) for warning in create_warnings)
        create_quality_flags = create_result.get("quality_flags") or []
        merged_quality_flags = pre_create_quality_flags + (
            list(create_quality_flags) if isinstance(create_quality_flags, list) else []
        )
        if create_result.get("success"):
            return {
                "success": True,
                "deal_url": create_result.get("deal_url"),
                "deal": create_result.get("deal"),
                "warnings": _dedupe_preserving_order(warnings),
                "verification": create_result.get("verification"),
                "verification_success": create_result.get("verification_success"),
                "quality_flags": merged_quality_flags,
                "error": None,
            }

        return {
            "success": False,
            "deal_url": create_result.get("deal_url"),
            "deal": create_result.get("deal"),
            "warnings": _dedupe_preserving_order(warnings),
            "verification": create_result.get("verification"),
            "verification_success": create_result.get("verification_success"),
            "quality_flags": merged_quality_flags,
            "error": _extract_tool_error_message(create_result, "Failed to create marketplace deal"),
        }
    except Exception as e:
        logger.error("ix_execute_deal_from_prompt_inputs failed: %s", e)
        return {
            "success": False,
            "deal_url": None,
            "deal": None,
            "warnings": _dedupe_preserving_order(warnings),
            "verification": None,
            "verification_success": False,
            "quality_flags": list(pre_create_quality_flags),
            "error": str(e),
        }


@mcp.tool()
async def ix_list_deals_v3(
    account_ids: list[int] | None = None,
    auction_type: str | None = None,
    bidding_strategy: list[str] | None = None,
    class_ids: list[int] | None = None,
    dsp_ids: list[int] | None = None,
    internal_deal_ids: list[int] | None = None,
    modified_date_since: str | None = None,
    page_offset: int = 0,
    page_size: int = 100,
    search: str | None = None,
    status: str | None = None,
    sort_field: str | None = None,
    sort_order: str | None = None,
    targeted_marketplace_account_ids: list[int] | None = None,
) -> dict:
    """List deals via GET /api/deals/v3/deals.

    Use class_ids=[5] with targeted_marketplace_account_ids=[your_account_id] to find
    'Deal with Marketplaces' deals that can be targeted inside a Marketplace Package.

    Args:
        account_ids: Filter by account IDs.
        auction_type: Filter by auction type: "first" or "fixed".
        bidding_strategy: Filter by bidding strategy for classID 1/5 deals:
            "standard", "preferredPrice", "priorityBidding",
            "programmaticGuaranteed", or "other".
        class_ids: Filter by deal class: 1 (Direct Deal), 3 (Inventory Package),
            4 (Marketplace Package), 5 (Deal with Marketplaces).
        dsp_ids: Filter by DSP IDs.
        internal_deal_ids: Filter by internal deal IDs.
        modified_date_since: Return deals modified after this ISO-8601 datetime string.
        page_offset: Pagination offset (default 0).
        page_size: Results per page, max 2000 (default 100).
        search: Free-text search string (matches internal ID, external ID, or name).
        status: Filter by status: "active", "paused", "expired", or "auto-paused".
        sort_field: Field name to sort by.
        sort_order: "asc" or "desc".
        targeted_marketplace_account_ids: Filter by targeted Marketplace Partner account IDs.

    Returns:
        {"success": True, "deals": [...], "total": int} or {"success": False, "error": ...}
    """
    if page_size > 2000:
        return {
            "success": False,
            "error": _make_error("list_deals_v3", None, "page_size cannot exceed 2000"),
        }
    if status and status not in ("active", "paused", "expired", "auto-paused"):
        return {
            "success": False,
            "error": _make_error(
                "list_deals_v3",
                None,
                f"status must be one of: active, paused, expired, auto-paused; got {status!r}",
            ),
        }
    if auction_type and auction_type not in ("first", "fixed"):
        return {
            "success": False,
            "error": _make_error(
                "list_deals_v3", None, f"auction_type must be 'first' or 'fixed'; got {auction_type!r}"
            ),
        }

    params: dict = {"pageOffset": page_offset, "pageSize": page_size}
    if account_ids:
        params["accountIDs"] = ",".join(str(x) for x in account_ids)
    if auction_type:
        params["auctionType"] = auction_type
    if bidding_strategy:
        params["biddingStrategy"] = ",".join(bidding_strategy)
    if class_ids:
        params["classIDs"] = ",".join(str(x) for x in class_ids)
    if dsp_ids:
        params["dspIDs"] = ",".join(str(x) for x in dsp_ids)
    if internal_deal_ids:
        params["internalDealIDs"] = ",".join(str(x) for x in internal_deal_ids)
    if modified_date_since:
        params["modifiedDateSince"] = modified_date_since
    if search:
        params["search"] = search
    if status:
        params["status"] = status
    if sort_field:
        params["sortField"] = sort_field
    if sort_order:
        params["sortOrder"] = sort_order
    if targeted_marketplace_account_ids:
        params["targetedMarketplaceAccountIDs"] = ",".join(str(x) for x in targeted_marketplace_account_ids)

    try:
        client = get_ix_client()
        data = await client.request("GET", "/api/deals/v3/deals", params=params)
        deals = data if isinstance(data, list) else data.get("deals", data)
        # The documented response field is totalCount; "total" is kept as a
        # fallback for older fixtures/responses that used it.
        total = len(deals) if isinstance(data, list) else data.get("totalCount", data.get("total", len(deals)))
        return {"success": True, "deals": deals, "total": total}
    except Exception as e:
        return {"success": False, "error": _make_error("list_deals_v3", None, str(e))}


@mcp.tool()
async def ix_get_deal_settings(internal_deal_id: int) -> dict:
    """Retrieve full settings for a single deal via GET /api/deals/v3/deals/{internalDealID}.

    Returns all deal properties including seatIDs and the full targeting array
    (not available in the list endpoint), plus the deal's `etag` — the entity
    tag IX requires in the If-Match header of any update. ix_update_deal
    fetches a fresh etag itself, so you only need this value when you want to
    pin an update to a specific version you have inspected.

    Args:
        internal_deal_id: The integer internal deal ID returned by ix_list_deals_v3
            or ix_create_marketplace_deal.

    Returns:
        {"success": True, "deal": <deal object>, "etag": <str|None>}
        or {"success": False, "error": ...}
    """
    try:
        client = get_ix_client()
        response = await client.request("GET", f"/api/deals/v3/deals/{internal_deal_id}", raw_response=True)
        return {
            "success": True,
            "deal": response.json(),
            "etag": response.headers.get("etag"),
        }
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("get_deal_settings", None, str(e)),
        }


@mcp.tool()
async def ix_update_deal(
    internal_deal_id: int,
    name: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    floor: float | None = None,
    auction_type: str | None = None,
    status: str | None = None,
    open_market: bool | None = None,
    labels: dict | None = None,
    targeting: list[dict] | None = None,
    direct_configurations: dict | None = None,
    ivp_configurations: dict | None = None,
    marketplace_configurations: dict | None = None,
    marketplace_participant_configurations: dict | None = None,
    etag: str | None = None,
) -> dict:
    """Update a deal via PATCH /api/deals/v3/deals/{internalDealID}.

    Partial update: only the arguments you pass are sent; everything else is
    left untouched. Works for any deal class — Direct Deal (1), Inventory
    Package (3), Marketplace Package (4), or Deal with Marketplaces (5).

    ETag concurrency: IX requires an If-Match header carrying the deal's
    current entity tag. This tool fetches the deal first (GET deal settings)
    to obtain the etag and the current state used for validation, then sends
    the PATCH. If IX reports a version conflict and you did NOT pin an `etag`,
    the tool re-fetches once and retries; a caller-pinned `etag` is never
    silently replaced — the conflict is returned instead.

    Args:
        internal_deal_id: Integer internal deal ID (from ix_list_deals_v3 /
            ix_create_marketplace_deal — NOT the external "IX17..." ID).
        name: New deal name, 1-255 characters.
        start_date: New start date, YYYY-MM-DD (UTC).
        end_date: New end date, YYYY-MM-DD (UTC). Cannot end up before the
            deal's (possibly unchanged) start date.
        floor: New floor price. Minimum 0.10 for Marketplace Packages
            (classID=4), 0.01 for other classes; maximum 99999.99. The class
            is read from the fetched deal.
        auction_type: "first" (highest bid wins at bid price) or "fixed"
            (winner pays the floor).
        status: "A"/"active" or "P"/"paused".
        open_market: Whether the deal competes with open market bids.
        labels: Reporting labels dict (advertiser, agency, custom,
            externalReferenceID, salesperson — all optional strings).
        targeting: FULL REPLACEMENT of the deal's targeting array — IX
            replaces the whole configuration with what you send, so always
            start from the current array returned by ix_get_deal_settings (or
            this tool's pre-fetch), edit it, and pass the COMPLETE result.
            Sending only the changed key silently drops every other key.
            Shape per object: {"keyName": ..., "targetingType":
            "standard"|"custom", "sets": [{"operator": "ANY_OF"|"NONE_OF",
            "values": [{"value": "...", "label": "..."}]}]}. Values must be
            canonical (as returned by GET) — no name resolution is applied.
        direct_configurations: classID=1 settings (applyBlocks,
            postAuctionDiscount, priority, ...). Rejected if the deal is a
            different class.
        ivp_configurations: classID=3 settings. Rejected on other classes.
        marketplace_configurations: classID=4 settings. Rejected on other classes.
        marketplace_participant_configurations: classID=5 settings. Rejected
            on other classes.
        etag: Optional entity tag from a prior ix_get_deal_settings call. Pin
            this when the update must only apply to the exact version you
            inspected (e.g. after editing its targeting array).

    Returns:
        On success: {"success": True, "deal": <updated deal>, "updated_fields":
            [...], "internal_deal_id": int, "etag_used": str}.
        On failure: {"success": False, "error": <error dict>} — including a
            structured etag_conflict error when a pinned etag is stale.
    """
    try:
        client = get_ix_client()

        # Always fetch current state: validation needs classID + current
        # dates, and the PATCH needs a current etag when none was pinned.
        current_response = await client.request("GET", f"/api/deals/v3/deals/{internal_deal_id}", raw_response=True)
        current_deal = current_response.json()
        current_etag = current_response.headers.get("etag")

        caller_pinned_etag = etag is not None
        effective_etag = etag if caller_pinned_etag else current_etag
        if not effective_etag:
            return {
                "success": False,
                "error": _make_error(
                    "update_deal",
                    None,
                    "IX did not return an etag for this deal and none was supplied — "
                    "cannot build the required If-Match header.",
                ),
            }

        payload = _build_deal_update_payload(
            current_deal=current_deal,
            name=name,
            start_date=start_date,
            end_date=end_date,
            floor=floor,
            auction_type=auction_type,
            status=status,
            open_market=open_market,
            labels=labels,
            targeting=targeting,
            direct_configurations=direct_configurations,
            ivp_configurations=ivp_configurations,
            marketplace_configurations=marketplace_configurations,
            marketplace_participant_configurations=marketplace_participant_configurations,
        )

        async def _patch(if_match: str) -> Any:
            return await client.request(
                "PATCH",
                f"/api/deals/v3/deals/{internal_deal_id}",
                json_body=payload,
                headers={"If-Match": if_match},
            )

        try:
            data = await _patch(effective_etag)
        except ValueError as e:
            message = str(e)
            is_etag_conflict = message.startswith(("HTTP 412", "HTTP 409", "HTTP 428"))
            if is_etag_conflict and not caller_pinned_etag:
                # Someone changed the deal between our GET and PATCH. We never
                # pinned a version, so re-fetch the etag once and retry.
                logger.warning("IX etag conflict on deal %d; re-fetching and retrying once", internal_deal_id)
                retry_response = await client.request(
                    "GET", f"/api/deals/v3/deals/{internal_deal_id}", raw_response=True
                )
                retry_etag = retry_response.headers.get("etag")
                if not retry_etag:
                    raise
                effective_etag = retry_etag
                data = await _patch(retry_etag)
            elif is_etag_conflict:
                return {
                    "success": False,
                    "error": _make_error(
                        "update_deal",
                        None,
                        f"etag_conflict: the pinned etag is stale ({message}). The deal changed since you "
                        "read it — re-run ix_get_deal_settings, review the current state, and retry.",
                    ),
                }
            else:
                raise

        return {
            "success": True,
            "deal": data,
            "updated_fields": sorted(payload.keys()),
            "internal_deal_id": internal_deal_id,
            "etag_used": effective_etag,
        }
    except Exception as e:
        return {"success": False, "error": _make_error("update_deal", None, str(e))}


@mcp.tool()
async def ix_list_marketplace_publishers(
    marketplace_account_id: int,
    name_like: str | None = None,
) -> dict:
    """List publishers available to a Marketplace Partner account.

    Calls GET /api/accounts/v1/marketplaces/{accountID}/publishers.

    The returned publishers[].legacyAccountID values are the publisherIDs used when
    building publisher-level targeting for a Marketplace Package (classID=4).

    Large accounts return 5,000+ publisher entries (~1MB) — pass
    `name_like` to filter server-side on `name` / `publisherName` /
    `accountName` / `companyName`. Substring match, case-insensitive.

    Args:
        marketplace_account_id: The Marketplace Partner account ID (integer shown in
            the IX UI top-left corner, same value passed as account_id when creating
            deals).
        name_like: Optional substring filter against publisher name fields.

    Returns:
        {"success": True, "publishers": [...]} or {"success": False, "error": ...}
    """
    try:
        client = get_ix_client()
        data = await client.request(
            "GET",
            f"/api/accounts/v1/marketplaces/{marketplace_account_id}/publishers",
        )
        publishers = data if isinstance(data, list) else data.get("publishers", data)
        if name_like and isinstance(publishers, list):
            needle = _normalize_lookup_text(name_like)

            def _matches(pub: dict[str, Any]) -> bool:
                for field in ("name", "publisherName", "accountName", "companyName"):
                    candidate = pub.get(field)
                    if candidate is not None and needle in _normalize_lookup_text(candidate):
                        return True
                return False

            publishers = [p for p in publishers if isinstance(p, dict) and _matches(p)]
        return {"success": True, "publishers": publishers}
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("list_marketplace_publishers", None, str(e)),
        }


@mcp.tool()
async def ix_find_publisher_by_name(
    marketplace_account_id: int,
    name: str,
) -> dict:
    """Look up publishers by name and return all candidates with disambiguation context.

    Use this when a brief carries a canonical client name (e.g. "The
    Weather Company") that may map to multiple delivery-route feeds in
    IX's catalog ("via Prebid" / "via OB" / "via Adapter"). The tool
    returns ALL candidates — it refuses to pick one for you, because
    the right combination is policy that belongs in the brief.

    Match strategy:
      1. Exact match (case- and punctuation-insensitive) on any name field
         or any ID field. If anything matches exactly, only the exact
         matches are returned.
      2. If no exact match, substring match against name fields.

    Args:
        marketplace_account_id: The Marketplace Partner account ID.
        name: The publisher name (or partial name) to look up.

    Returns:
        {
            "success": True,
            "match_count": <int>,
            "strategy": "exact_name_or_id" | "substring_name" | "no_match",
            "candidates": [
                {
                    "legacyAccountID": <int>,   # use this as publisher_id in deals
                    "accountID": <int>,
                    "name": <str>,
                    "accountStatus": "A" | ...,
                    "currency": "USD" | ...,
                    "isOptedOutOfMarketplaces": <bool>,
                },
                ...
            ],
        }
    """
    try:
        listing = await ix_list_marketplace_publishers(
            marketplace_account_id=marketplace_account_id,
            name_like=name,
        )
        if not listing.get("success"):
            return {
                "success": False,
                "error": listing.get("error")
                or _make_error("find_publisher_by_name", None, "Failed to load publishers"),
            }
        publishers = [p for p in listing.get("publishers", []) if isinstance(p, dict)]
        result = _find_publisher_matches(publishers, name)
        matches = result["matches"]
        return {
            "success": True,
            "match_count": len(matches),
            "strategy": result["strategy"],
            "candidates": [
                {
                    "legacyAccountID": cand.get("legacyAccountID"),
                    "accountID": cand.get("accountID"),
                    "name": cand.get("name"),
                    "accountStatus": cand.get("accountStatus"),
                    "currency": cand.get("currency"),
                    "isOptedOutOfMarketplaces": cand.get("isOptedOutOfMarketplaces", False),
                }
                for cand in matches
            ],
        }
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("find_publisher_by_name", None, str(e)),
        }


@mcp.tool()
async def ix_list_segments(account_id: int, name_like: str | None = None) -> dict:
    """List audience segments available for deal targeting.

    Calls GET /api/segments/v2/segments?accountID={account_id}.

    Production catalogs can return tens of thousands of segments
    (90k+ on TWC accounts) — pass `name_like` to filter client-side
    on common name fields (`externalSegmentName`, `name`, `segmentName`,
    `label`) before returning. Substring match, case-insensitive.

    The returned segment IDs can be used in deal targeting objects. For
    segment targeting, use `keyName: "segmentid"` directly — `targetingKeyID`
    is optional/ignored by the API for segment targeting (the `im_segments`
    key returned by `ix_list_targeting_keys` only supports inclusion, not
    exclusion). The high-level `ix_execute_deal_from_prompt_inputs` tool
    handles this automatically.

    Example targeting object for segment inclusion + exclusion (single
    targeting object with two sets — the API stores them under one
    `keyName: "segmentid"` entry):
      {"keyName": "segmentid",
       "targetingType": "standard",
       "sets": [
         {"operator": "ANY_OF",  "values": [{"value": "280"}, {"value": "3007"}]},
         {"operator": "NONE_OF", "values": [{"value": "308129"}]}
       ]}

    IMPORTANT: Segment targeting MUST be included in the targeting payload
    passed to ix_create_marketplace_deal. Segments cannot be applied after
    deal creation via the API; omitting them requires manual UI configuration.

    Args:
        account_id: The marketplace account ID.
        name_like: Optional case-insensitive substring filter applied
            client-side to the segment name fields. Returns the full
            catalog when omitted.

    Returns:
        {"success": True, "segments": [...]} or {"success": False, "error": ...}
    """
    try:
        client = get_ix_client()
        data = await client.request(
            "GET",
            "/api/segments/v2/segments",
            params={"accountID": account_id},
        )
        segments = data if isinstance(data, list) else data.get("segments", data)
        if name_like and isinstance(segments, list):
            needle = _normalize_lookup_text(name_like)
            if needle:
                lookup_fields = ("externalSegmentName", "name", "segmentName", "label")
                segments = [
                    seg
                    for seg in segments
                    if isinstance(seg, dict)
                    and any(needle in _normalize_lookup_text(seg.get(field)) for field in lookup_fields)
                ]
        return {"success": True, "segments": segments}
    except Exception as e:
        return {"success": False, "error": _make_error("list_segments", None, str(e))}


@mcp.tool()
async def ix_create_domain_targeting_values(domains: list[str], account_id: int | None = None) -> dict:
    """Resolve domain/app-bundle strings to Index Exchange targetingValueIDs.

    Calls PUT /api/supply-configuration/v1/inventory-groups/targets/ext/domain/values.

    When account_id is provided, this tool resolves marketplace.legacyMarketplaceID
    and sends it as publisherAccountID to match IX UI behavior.

    The returned targetingValueIDs are useful for validation/evidence, but should NOT
    be sent directly in deal create payload targeting values for key 120. The create
    endpoint expects literal domain/app-bundle strings for Domain targeting.

    Do NOT use ix_list_targeting_values for domain/app-bundle lookups -- that route
    does not support this key and may return errors.

    Args:
        domains: List of domain strings or app bundle strings to resolve.
            Example: ["generalmills.com", "eonline.com", "outdoorlife.com"]
        account_id: Optional account ID used to resolve publisherAccountID scope.
            For marketplace accounts, this resolves legacyMarketplaceID and sends
            that value as publisherAccountID (the IX query parameter name is
            misleading in this flow).

    Returns:
        On success: {"success": True, "targeting_value_ids": [...], "raw": <API response>}
        On auth error (code 1100): {"success": False, "error": {...},
            "hint": "Marketplace accounts may not have publisher-scoped targeting value
                      routes. Confirm your account has supply configuration access."}
        On other failure: {"success": False, "error": ...}

    Note:
        Marketplace-only accounts may receive a 403 / error code 1100 if they do not
        have publisher-scoped access to this route. Deal creation can still succeed
        without domain targeting -- domains are optional unless you need to restrict
        inventory to specific sites.
    """
    if not domains:
        return {
            "success": False,
            "error": _make_error("create_domain_targeting_values", None, "domains list cannot be empty"),
        }

    client = get_ix_client()
    endpoint = "/api/supply-configuration/v1/inventory-groups/targets/ext/domain/values"
    params: dict[str, Any] | None = None
    if account_id is not None:
        publisher_account_id = await _resolve_targeting_publisher_account_id(client, account_id)
        params = {"publisherAccountID": publisher_account_id}
    batch_size = 200
    targeting_value_ids: list[Any] = []
    raw_batches: list[Any] = []
    failed_domains: list[str] = []

    for start in range(0, len(domains), batch_size):
        batch = domains[start : start + batch_size]
        try:
            data = await client.request(
                "PUT",
                endpoint,
                params=params,
                json_body={"domains": batch},
            )
            raw_batches.append(data)
            ids = data.get("targetingValueIDs") or data.get("targetingValueIds") or data.get("ids") or []
            if isinstance(ids, list):
                for entry in ids:
                    value_id = entry.get("valueID") if isinstance(entry, dict) else entry
                    if value_id is not None:
                        targeting_value_ids.append(str(value_id))
            else:
                logger.warning("Unexpected targeting ID payload type: %s", type(ids).__name__)
        except Exception as e:
            err_str = str(e)
            err_lower = err_str.lower()
            is_403_1100 = "http 403" in err_lower and "1100" in err_lower
            if is_403_1100 or "1100" in err_str or "not authorized" in err_lower:
                return {
                    "success": False,
                    "error": _make_error("create_domain_targeting_values", None, err_str),
                    "hint": (
                        "Marketplace accounts may not have publisher-scoped supply "
                        "configuration access (error code 1100). Domain targeting is "
                        "optional -- deal creation can proceed without it."
                    ),
                }

            logger.error("Domain targeting batch failed for %d domains: %s", len(batch), err_str)
            failed_domains.extend(batch)

    result: dict[str, Any] = {
        "success": True,
        "targeting_value_ids": targeting_value_ids,
        "raw": raw_batches,
    }
    if failed_domains:
        result["partial_success"] = True
        result["failed_domains"] = failed_domains
    return result


@mcp.tool()
async def ix_list_targeting_keys(account_id: int) -> dict:
    """List available targeting keys for dynamic targeting discovery.

    Calls GET /api/supply-configuration/v1/inventory-groups/targets
    with publisherAccountID derived from account lookup.

    For marketplace accounts, this tool first calls
    GET /api/accounts/v2/accounts/?accountIDs={account_id} and uses
    marketplace.legacyMarketplaceID as publisherAccountID. For non-marketplace
    accounts (or if legacyMarketplaceID is unavailable), it falls back to the
    provided account_id.

    Args:
        account_id: Account ID used to resolve publisherAccountID scope.

    Returns:
        {"success": True, "targeting_keys": [...]} or {"success": False, "error": ...}
    """
    try:
        client = get_ix_client()
        publisher_account_id = await _resolve_targeting_publisher_account_id(client, account_id)
        data = await client.request(
            "GET",
            "/api/supply-configuration/v1/inventory-groups/targets",
            params={"publisherAccountID": publisher_account_id},
        )
        keys = data if isinstance(data, list) else data.get("data", data)
        return {"success": True, "targeting_keys": keys}
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("list_targeting_keys", None, str(e)),
        }


@mcp.tool()
async def ix_list_targeting_values(
    account_id: int, key_id: int, search: str | None = None, value: str | None = None
) -> dict:
    """List available targeting values for a targeting key.

    Calls GET /api/supply-configuration/v1/inventory-groups/targets/{key_id}/values
    with publisherAccountID derived from account lookup.

    For marketplace accounts, this tool first calls
    GET /api/accounts/v2/accounts/?accountIDs={account_id} and uses
    marketplace.legacyMarketplaceID as publisherAccountID. For non-marketplace
    accounts (or if legacyMarketplaceID is unavailable), it falls back to the
    provided account_id.

    Note: IX expects the lookup text in the query parameter named "values"
    (not "search") for this endpoint.

    Args:
        account_id: Account ID used to resolve publisherAccountID scope.
        key_id: Targeting key ID from ix_list_targeting_keys.
        search: Optional lookup string mapped to the API query parameter
            "values". Kept for backward compatibility.
        value: Optional alias for search. If both are provided they must match.

    Returns:
        {"success": True, "targeting_values": [...]} or {"success": False, "error": ...}
    """
    try:
        client = get_ix_client()
        publisher_account_id = await _resolve_targeting_publisher_account_id(client, account_id)
        params: dict[str, Any] = {"publisherAccountID": publisher_account_id}
        if search and value and search != value:
            raise ValueError("search and value must match when both are provided")
        lookup_value = value or search
        if lookup_value:
            params["values"] = lookup_value
        data = await client.request(
            "GET",
            f"/api/supply-configuration/v1/inventory-groups/targets/{key_id}/values",
            params=params,
        )
        values = data if isinstance(data, list) else data.get("data", data)
        return {"success": True, "targeting_values": values}
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("list_targeting_values", None, str(e)),
        }


# =============================================================================
# MCP Tools — Accounts
# =============================================================================


@mcp.tool()
async def ix_list_account_information(
    account_ids: list[int] | None = None,
    account_type_ids: list[int] | None = None,
    legacy_marketplace_ids: list[int] | None = None,
) -> dict[str, Any]:
    """List account information from Index Exchange.

    Args:
        account_ids: Optional list of account IDs to filter by.
        account_type_ids: Optional list of account type IDs to filter by.
        legacy_marketplace_ids: Optional list of legacy marketplace IDs.

    Returns:
        Dictionary with success and accounts, or error.
    """
    try:
        client = get_ix_client()
        params: dict[str, Any] = {}
        if account_ids:
            params["accountIDs"] = ",".join(str(x) for x in account_ids)
        if account_type_ids:
            params["accountTypeIDs"] = ",".join(str(x) for x in account_type_ids)
        if legacy_marketplace_ids:
            params["legacyMarketplaceIDs"] = ",".join(str(x) for x in legacy_marketplace_ids)

        data = await client.request("GET", "/api/accounts/v2/accounts/", params=params or None)
        return {"success": True, "accounts": data}
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("list_account_information", None, str(e)),
        }


# =============================================================================
# MCP Tools — Reporting
# =============================================================================


@mcp.tool()
async def ix_list_marketplace_report_presets() -> dict[str, Any]:
    """List curated Index Marketplace reporting presets.

    Returns:
        Dictionary with success and available preset definitions.
    """
    presets: dict[str, Any] = {}
    for preset_name, preset in IX_MARKETPLACE_REPORT_PRESETS.items():
        presets[preset_name] = {
            "description": preset["description"],
            "dimensions": preset["dimensions"],
            "measures": preset["measures"],
        }
    return {
        "success": True,
        "account_type": "marketplace_partner",
        "presets": presets,
        "known_accounts": {
            "Reklaim": 1485234,
            "Elcano": 1491166,
            "The Weather Company, LLC": 1499155,
            "Raptive": 1502939,
            "Stirista": 1503605,
        },
    }


@mcp.tool()
async def ix_reporting_healthcheck(
    account_ids: list[int] | None = None,
    account_group_ids: list[int] | None = None,
    report_status: str | None = "saved",
) -> dict[str, Any]:
    """Validate Index Exchange reporting connectivity end-to-end.

    This is the fastest way to verify that configured credentials can
    authenticate and access the reporting API.

    Args:
        account_ids: Optional list of account IDs to scope the check.
        account_group_ids: Optional list of account group IDs to scope the check.
        report_status: Optional report status filter. Defaults to "saved" to keep
            the check bounded to common active reports.

    Returns:
        Dictionary with success, auth_mode, token state, and a small summary of
        the reporting endpoint response.
    """
    try:
        client = get_ix_client()
        params: dict[str, Any] = {}
        if account_ids:
            params["accountIDs"] = ",".join(str(x) for x in account_ids)
        if account_group_ids:
            params["accountGroupIDs"] = ",".join(str(x) for x in account_group_ids)
        if report_status:
            params["reportStatus"] = report_status

        data = await client.request(
            "GET",
            "/api/reporting/agg/v1/report-specs/info",
            params=params or None,
        )

        report_count = len(data) if isinstance(data, list) else None
        return {
            "success": True,
            "configured": client._is_configured(),
            "auth_mode": client.auth_mode,
            "token_cached": bool(client._access_token),
            "token_valid": client._token_is_valid(),
            "report_count": report_count,
            "reports": data,
        }
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("reporting_healthcheck", None, str(e)),
        }


@mcp.tool()
async def ix_create_marketplace_report_spec(
    account_id: int | str,
    report_title: str,
    date_range: dict[str, Any],
    preset: str = "deal_summary",
    extra_fields: list[str] | None = None,
    delivery: dict[str, Any] | None = None,
    schedule: dict[str, Any] | None = None,
    file_type: str | None = None,
    region_setting: str | None = None,
    report_status: str | None = None,
    query_spec_extras: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Create a Marketplace Partner report spec from a curated preset.

    Args:
        account_id: Marketplace Partner account ID or known account name.
        report_title: Report title to save in Index Exchange.
        date_range: IX reporting dateRange object.
        preset: Curated preset name. One of: deal_summary, deal_labels,
            supply_breakdown, segment_performance.
        extra_fields: Optional additional Marketplace reporting fields to append.
        delivery: Optional delivery configuration.
        schedule: Optional schedule configuration.
        file_type: Optional file type such as csv.
        region_setting: Optional region setting.
        report_status: Optional report status such as saved or draft.
        query_spec_extras: Optional extra fields merged into querySpec.

    Returns:
        Dictionary with success and report spec metadata, or error.
    """
    selected_preset = IX_MARKETPLACE_REPORT_PRESETS.get(preset)
    if selected_preset is None:
        return {
            "success": False,
            "error": _make_error(
                "create_marketplace_report_spec",
                None,
                "Unknown Marketplace report preset. Use ix_list_marketplace_report_presets() to see valid preset names.",
                {"preset": preset, "valid_presets": sorted(IX_MARKETPLACE_REPORT_PRESETS)},
            ),
        }

    try:
        resolved_account_id = _resolve_marketplace_account_id(account_id)
    except ValueError as e:
        return {
            "success": False,
            "error": _make_error("create_marketplace_report_spec", None, str(e)),
        }

    fields = _unique_fields([*selected_preset["dimensions"], *selected_preset["measures"], *(extra_fields or [])])

    result = await ix_create_report_spec(
        report_title=report_title,
        accounts=[resolved_account_id],
        fields=fields,
        date_range=date_range,
        delivery=delivery,
        schedule=schedule,
        file_type=file_type,
        region_setting=region_setting,
        report_status=report_status,
        query_spec_extras=query_spec_extras,
    )
    if not result.get("success"):
        return result

    return {
        "success": True,
        "account_id": resolved_account_id,
        "preset": preset,
        "selected_fields": fields,
        "report_spec": result["report_spec"],
    }


@mcp.tool()
async def ix_run_marketplace_draft_report(
    account_id: int | str,
    report_title: str,
    date_range: dict[str, Any],
    preset: str = "deal_summary",
    extra_fields: list[str] | None = None,
    file_type: str = "csv.zip",
    region_setting: str | None = None,
    query_spec_extras: dict[str, Any] | None = None,
    download: bool = True,
    filename_hint: str | None = None,
    output_dir: str | None = None,
    poll_timeout_seconds: float = DEFAULT_REPORT_FILE_POLL_TIMEOUT_SECONDS,
    poll_interval_seconds: float = DEFAULT_REPORT_FILE_POLL_INTERVAL_SECONDS,
) -> dict[str, Any]:
    """Create, run, and optionally download a Marketplace draft report.

    This wraps the standard ad hoc reporting workflow for Marketplace Partner
    accounts: create a draft spec, run it immediately, then poll for the output
    file and download it.

    output_dir: Optional absolute path the downloaded file is written into.
        Pass the per-conversation workspace dir so other tools can read the
        result. Defaults to ~/Victoria/indexexchange_reports for local use;
        on a hardened production host that path is read-only.
    """
    create_result = await ix_create_marketplace_report_spec(
        account_id=account_id,
        report_title=report_title,
        date_range=date_range,
        preset=preset,
        extra_fields=extra_fields,
        file_type=file_type,
        region_setting=region_setting,
        report_status="draft",
        query_spec_extras=query_spec_extras,
    )
    if not create_result.get("success"):
        return create_result

    report_spec = create_result["report_spec"]
    report_spec_id = report_spec.get("reportSpecID")
    if not isinstance(report_spec_id, int):
        return {
            "success": False,
            "error": _make_error(
                "run_marketplace_draft_report",
                None,
                "Draft report was created but reportSpecID was missing from the response.",
                {"report_spec": report_spec},
            ),
        }

    run_result = await ix_run_report_download(report_id=report_spec_id, report_status="draft")
    if not run_result.get("success"):
        return run_result

    response: dict[str, Any] = {
        "success": True,
        "account_id": create_result["account_id"],
        "preset": preset,
        "selected_fields": create_result["selected_fields"],
        "report_spec": report_spec,
        "report_run": run_result["report_run"],
    }

    if not download:
        response["download_pending"] = True
        return response

    deadline = time.time() + max(poll_timeout_seconds, 0)
    sleep_seconds = max(poll_interval_seconds, 0.1)
    last_files: list[dict[str, Any]] = []

    while True:
        files_result = await ix_list_report_files(report_ids=[str(report_spec_id)])
        if not files_result.get("success"):
            return files_result

        files = _extract_report_files(files_result.get("files"))
        last_files = files
        completed_file = next(
            (file_info for file_info in files if _report_file_ready(file_info) and _extract_report_file_id(file_info)),
            None,
        )
        if completed_file is not None:
            file_id = _extract_report_file_id(completed_file)
            if file_id is None:
                break
            download_result = await ix_download_report_file(
                file_id=file_id,
                filename_hint=filename_hint or file_type,
                output_dir=output_dir,
            )
            if not download_result.get("success"):
                return download_result
            response["report_file"] = completed_file
            response["download"] = download_result
            return response

        if time.time() >= deadline:
            response["download_pending"] = True
            response["polled_files"] = last_files
            response["warning"] = (
                "Draft report run started, but no completed report file was available "
                f"before polling timed out after {poll_timeout_seconds:.0f}s. "
                "To finish: re-call ix_run_marketplace_draft_report with the SAME "
                "arguments and poll_timeout_seconds=600 — the existing report run "
                "will be reused (IX dedupes by spec) and the longer wait should land "
                "the file. Do NOT loop ix_list_report_files manually — that just "
                "burns tokens without speeding up the IX API."
            )
            return response

        await asyncio.sleep(sleep_seconds)


@mcp.tool()
async def ix_update_report_spec(
    report_id: int,
    report_title: str,
    accounts: list[int],
    fields: list[str],
    date_range: dict[str, Any],
    delivery: dict[str, Any] | None = None,
    schedule: dict[str, Any] | None = None,
    file_type: str | None = None,
    region_setting: str | None = None,
    report_status: str | None = None,
    query_spec_extras: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Update an existing report specification on Index Exchange via PATCH."""
    if len(accounts) != 1:
        return {
            "success": False,
            "error": _make_error(
                "update_report_spec",
                None,
                f"Exactly 1 account is supported per Index Exchange docs, got {len(accounts)}",
            ),
        }

    if not any(k in date_range for k in ("from", "current", "previous")):
        return {
            "success": False,
            "error": _make_error(
                "update_report_spec", None, "date_range must contain 'from'+'to', 'current', or 'previous'"
            ),
        }

    query_spec: dict[str, Any] = {"fields": fields}
    if query_spec_extras:
        query_spec.update(query_spec_extras)

    payload: dict[str, Any] = {
        "reportTitle": report_title,
        "accounts": accounts,
        "querySpec": query_spec,
        "dateRange": date_range,
    }
    if delivery is not None:
        payload["delivery"] = delivery
    if schedule is not None:
        payload["schedule"] = schedule
    if file_type is not None:
        payload["fileType"] = file_type
    if region_setting is not None:
        payload["regionSetting"] = region_setting
    if report_status is not None:
        payload["reportStatus"] = report_status

    try:
        client = get_ix_client()
        data = await client.request(
            "PATCH",
            f"/api/reporting/agg/v1/report-specs/{report_id}",
            json_body=payload,
        )
        return {"success": True, "report_spec": data}
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("update_report_spec", None, str(e)),
        }


@mcp.tool()
async def ix_update_marketplace_draft_report_spec(
    report_id: int,
    account_id: int | str,
    report_title: str,
    date_range: dict[str, Any],
    preset: str = "deal_summary",
    extra_fields: list[str] | None = None,
    delivery: dict[str, Any] | None = None,
    schedule: dict[str, Any] | None = None,
    file_type: str | None = None,
    region_setting: str | None = None,
    query_spec_extras: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Update a Marketplace report spec in draft mode from a curated preset."""
    selected_preset = IX_MARKETPLACE_REPORT_PRESETS.get(preset)
    if selected_preset is None:
        return {
            "success": False,
            "error": _make_error(
                "update_marketplace_draft_report_spec",
                None,
                "Unknown Marketplace report preset. Use ix_list_marketplace_report_presets() to see valid preset names.",
                {"preset": preset, "valid_presets": sorted(IX_MARKETPLACE_REPORT_PRESETS)},
            ),
        }

    try:
        resolved_account_id = _resolve_marketplace_account_id(account_id)
    except ValueError as e:
        return {
            "success": False,
            "error": _make_error("update_marketplace_draft_report_spec", None, str(e)),
        }

    fields = _unique_fields([*selected_preset["dimensions"], *selected_preset["measures"], *(extra_fields or [])])

    result = await ix_update_report_spec(
        report_id=report_id,
        report_title=report_title,
        accounts=[resolved_account_id],
        fields=fields,
        date_range=date_range,
        delivery=delivery,
        schedule=schedule,
        file_type=file_type,
        region_setting=region_setting,
        report_status="draft",
        query_spec_extras=query_spec_extras,
    )
    if not result.get("success"):
        return result

    return {
        "success": True,
        "account_id": resolved_account_id,
        "preset": preset,
        "selected_fields": fields,
        "report_spec": result["report_spec"],
    }


@mcp.tool()
async def ix_run_updated_marketplace_draft_report(
    report_id: int,
    account_id: int | str,
    report_title: str,
    date_range: dict[str, Any],
    preset: str = "deal_summary",
    extra_fields: list[str] | None = None,
    file_type: str = "csv.zip",
    region_setting: str | None = None,
    query_spec_extras: dict[str, Any] | None = None,
    download: bool = True,
    filename_hint: str | None = None,
    output_dir: str | None = None,
    poll_timeout_seconds: float = DEFAULT_REPORT_FILE_POLL_TIMEOUT_SECONDS,
    poll_interval_seconds: float = DEFAULT_REPORT_FILE_POLL_INTERVAL_SECONDS,
) -> dict[str, Any]:
    """Update an existing report into draft mode, run it, and optionally download it.

    output_dir: Optional absolute path the downloaded file is written into.
        Pass the per-conversation workspace dir so other tools can read the
        result.
    """
    update_result = await ix_update_marketplace_draft_report_spec(
        report_id=report_id,
        account_id=account_id,
        report_title=report_title,
        date_range=date_range,
        preset=preset,
        extra_fields=extra_fields,
        file_type=file_type,
        region_setting=region_setting,
        query_spec_extras=query_spec_extras,
    )
    if not update_result.get("success"):
        return update_result

    report_spec = update_result["report_spec"]
    run_result = await ix_run_report_download(report_id=report_id, report_status="draft")
    if not run_result.get("success"):
        return run_result

    response: dict[str, Any] = {
        "success": True,
        "account_id": update_result["account_id"],
        "preset": preset,
        "selected_fields": update_result["selected_fields"],
        "report_spec": report_spec,
        "report_run": run_result["report_run"],
    }

    if not download:
        response["download_pending"] = True
        return response

    deadline = time.time() + max(poll_timeout_seconds, 0)
    sleep_seconds = max(poll_interval_seconds, 0.1)
    last_files: list[dict[str, Any]] = []

    while True:
        files_result = await ix_list_report_files(report_ids=[str(report_id)])
        if not files_result.get("success"):
            return files_result

        files = _extract_report_files(files_result.get("files"))
        last_files = files
        completed_file = next(
            (file_info for file_info in files if _report_file_ready(file_info) and _extract_report_file_id(file_info)),
            None,
        )
        if completed_file is not None:
            file_id = _extract_report_file_id(completed_file)
            if file_id is None:
                break
            download_result = await ix_download_report_file(
                file_id=file_id,
                filename_hint=filename_hint or file_type,
                output_dir=output_dir,
            )
            if not download_result.get("success"):
                return download_result
            response["report_file"] = completed_file
            response["download"] = download_result
            return response

        if time.time() >= deadline:
            response["download_pending"] = True
            response["polled_files"] = last_files
            response["warning"] = (
                "Draft report update and run succeeded, but no completed report file "
                f"was available before polling timed out after {poll_timeout_seconds:.0f}s. "
                "To finish: re-call ix_run_updated_marketplace_draft_report with the "
                "SAME arguments and poll_timeout_seconds=600 — the existing report run "
                "will be reused and the longer wait should land the file. Do NOT loop "
                "ix_list_report_files manually."
            )
            return response

        await asyncio.sleep(sleep_seconds)


def _normalize_ix_report_date_range(date_range: dict[str, Any]) -> dict[str, Any]:
    """Make single-day ``{from, to}`` ranges valid against IX's exclusive end.

    Index Exchange treats ``dateRange.to`` as EXCLUSIVE, so ``{"from": D,
    "to": D}`` is an empty window that the API rejects with RSE-4009
    ("Report date range is empty / not valid"). A same-day from/to plainly
    means "just that day", so bump ``to`` forward one day. Multi-day ranges and
    relative ``{current}`` / ``{previous}`` ranges are returned unchanged.
    """
    if not isinstance(date_range, dict):
        return date_range
    start = date_range.get("from")
    end = date_range.get("to")
    if not (isinstance(start, str) and isinstance(end, str)):
        return date_range
    try:
        start_d = date.fromisoformat(start.strip())
        end_d = date.fromisoformat(end.strip())
    except ValueError:
        # Leave timestamped / non-ISO-date values untouched.
        return date_range
    if end_d <= start_d:
        normalized = dict(date_range)
        normalized["to"] = (start_d + timedelta(days=1)).isoformat()
        return normalized
    return date_range


@mcp.tool()
async def ix_create_report_spec(
    report_title: str,
    accounts: list[int],
    fields: list[str],
    date_range: dict[str, Any],
    delivery: dict[str, Any] | None = None,
    schedule: dict[str, Any] | None = None,
    file_type: str | None = None,
    region_setting: str | None = None,
    report_status: str | None = None,
    query_spec_extras: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Create a report specification on Index Exchange.

    Args:
        report_title: Title of the report (required).
        accounts: List of account IDs (exactly 1 supported per IX docs).
        fields: List of field names for querySpec.fields (required).
        date_range: Date range object. Supports:
            {"from": "...", "to": "..."} OR {"current": "..."} OR {"previous": {...}}.
        delivery: Optional delivery configuration.
        schedule: Optional schedule configuration.
        file_type: Optional file type (e.g., "csv").
        region_setting: Optional region setting.
        report_status: Optional report status ("saved", "draft").
        query_spec_extras: Optional extra fields merged into querySpec.

    Returns:
        Dictionary with success and report_spec, or error.
    """
    # Validate accounts
    if len(accounts) != 1:
        return {
            "success": False,
            "error": _make_error(
                "create_report_spec",
                None,
                f"Exactly 1 account is supported per Index Exchange docs, got {len(accounts)}",
            ),
        }

    # Validate date_range has at least one recognized key
    if not any(k in date_range for k in ("from", "current", "previous")):
        return {
            "success": False,
            "error": _make_error(
                "create_report_spec", None, "date_range must contain 'from'+'to', 'current', or 'previous'"
            ),
        }

    # IX's dateRange.to is exclusive, so a same-day {from, to} is an empty
    # window (RSE-4009). Normalize single-day ranges to span one full day.
    date_range = _normalize_ix_report_date_range(date_range)

    # Build payload
    query_spec: dict[str, Any] = {"fields": fields}
    if query_spec_extras:
        query_spec.update(query_spec_extras)

    payload: dict[str, Any] = {
        "reportTitle": report_title,
        "accounts": accounts,
        "querySpec": query_spec,
        "dateRange": date_range,
    }
    if delivery is not None:
        payload["delivery"] = delivery
    if schedule is not None:
        payload["schedule"] = schedule
    if file_type is not None:
        payload["fileType"] = file_type
    if region_setting is not None:
        payload["regionSetting"] = region_setting
    if report_status is not None:
        payload["reportStatus"] = report_status

    try:
        client = get_ix_client()
        data = await client.request(
            "POST",
            "/api/reporting/agg/v1/report-specs",
            json_body=payload,
        )
        return {"success": True, "report_spec": data}
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("create_report_spec", None, str(e)),
        }


@mcp.tool()
async def ix_list_active_reports(
    account_ids: list[int] | None = None,
    account_group_ids: list[int] | None = None,
    report_status: str | None = None,
) -> dict[str, Any]:
    """List active report specifications on Index Exchange.

    Args:
        account_ids: Optional list of account IDs to filter.
        account_group_ids: Optional list of account group IDs.
        report_status: Optional status filter (e.g., "saved", "draft").

    Returns:
        Dictionary with success and reports list, or error.
    """
    try:
        client = get_ix_client()
        params: dict[str, Any] = {}
        if account_ids:
            params["accountIDs"] = ",".join(str(x) for x in account_ids)
        if account_group_ids:
            params["accountGroupIDs"] = ",".join(str(x) for x in account_group_ids)
        if report_status:
            params["reportStatus"] = report_status

        data = await client.request(
            "GET",
            "/api/reporting/agg/v1/report-specs/info",
            params=params or None,
        )
        return {"success": True, "reports": data}
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("list_active_reports", None, str(e)),
        }


@mcp.tool()
async def ix_run_report_download(
    report_id: int,
    report_status: str | None = None,
) -> dict[str, Any]:
    """Run an ad hoc report on Index Exchange (triggers report generation).

    Args:
        report_id: The report spec ID to run.
        report_status: Optional status ("saved" or "draft").

    Returns:
        Dictionary with success and report run info, or error.
    """
    try:
        client = get_ix_client()
        payload: dict[str, Any] = {"reportID": report_id}
        if report_status:
            payload["reportStatus"] = report_status

        data = await client.request(
            "POST",
            "/api/reporting/agg/v1/report-runs",
            json_body=payload,
        )
        return {"success": True, "report_run": data}
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("run_report_download", None, str(e)),
        }


@mcp.tool()
async def ix_list_report_files(
    account_ids: list[int] | None = None,
    account_group_ids: list[int] | None = None,
    status: str | None = None,
    file_ids: list[str] | None = None,
    report_ids: list[str] | None = None,
) -> dict[str, Any]:
    """List available report files on Index Exchange.

    Use this for INSPECTION — what reports have already been generated and
    are downloadable. DO NOT loop on this tool to wait for a report run to
    complete. The all-in-one report tools (ix_run_marketplace_draft_report,
    ix_run_updated_marketplace_draft_report) already poll for completion
    and download internally with download=True (the default), and their
    poll_timeout_seconds covers the typical 30-120s queue window. If one of
    them returns download_pending=True you should re-call the SAME all-in-one
    tool (or pass a longer poll_timeout_seconds) — not poll list_report_files
    yourself, which burns LLM tokens without speeding the IX API up.

    Args:
        account_ids: Optional list of account IDs.
        account_group_ids: Optional list of account group IDs.
        status: Optional status filter.
        file_ids: Optional list of file IDs.
        report_ids: Optional list of report IDs.

    Returns:
        Dictionary with success and files list, or error.
    """
    try:
        client = get_ix_client()
        params: dict[str, Any] = {}
        if account_ids:
            params["accountIDs"] = ",".join(str(x) for x in account_ids)
        if account_group_ids:
            params["accountGroupIDs"] = ",".join(str(x) for x in account_group_ids)
        if status:
            params["status"] = status
        if file_ids:
            params["fileIDs"] = ",".join(file_ids)
        if report_ids:
            params["reportIDs"] = ",".join(report_ids)

        data = await client.request(
            "GET",
            "/api/reporting/agg/v1/report-files/list",
            params=params or None,
        )
        return {"success": True, "files": data}
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("list_report_files", None, str(e)),
        }


@mcp.tool()
async def ix_download_report_file(
    file_id: str,
    filename_hint: str | None = None,
    output_dir: str | None = None,
) -> dict[str, Any]:
    """Download a report file from Index Exchange to disk.

    Args:
        file_id: The file ID to download.
        filename_hint: Optional filename hint (used for extension detection).
        output_dir: Optional absolute path to save into. Pass the
            per-conversation workspace dir so the agent's other tools can
            read the file. Defaults to ~/Victoria/indexexchange_reports
            for legacy/local use; on a hardened production host that path
            is read-only and the call will fail unless output_dir is set.

    Returns:
        Dictionary with path, bytes, sha256, content_type.
    """
    try:
        client = get_ix_client()
        download_dir = Path(output_dir) if output_dir else Path(client._download_dir)
        download_dir = download_dir.expanduser()
        download_dir.mkdir(parents=True, exist_ok=True)

        filepath, size, sha256, content_type = await client.download_file(
            f"/api/reporting/agg/v1/report-files/download/{file_id}",
            download_dir,
            f"ix_report_{file_id}",
            filename_hint,
        )

        logger.info("Downloaded report file to %s (%d bytes)", filepath, size)

        return {
            "success": True,
            "path": str(filepath),
            "bytes": size,
            "sha256": sha256,
            "content_type": content_type,
        }
    except Exception as e:
        return {
            "success": False,
            "error": _make_error("download_report_file", None, str(e)),
        }


# =============================================================================
# Main Entry Point
# =============================================================================

if __name__ == "__main__":
    logger.info("Starting Index Exchange MCP Server")

    has_service = bool(os.environ.get("INDEXEXCHANGE_SERVICE_ID")) and bool(
        os.environ.get("INDEXEXCHANGE_SERVICE_SECRET")
    )
    has_user = bool(os.environ.get("INDEXEXCHANGE_USERNAME")) and bool(os.environ.get("INDEXEXCHANGE_PASSWORD"))

    if not has_service and not has_user:
        logger.warning(
            "Index Exchange not configured. Set INDEXEXCHANGE_SERVICE_ID + "
            "INDEXEXCHANGE_SERVICE_SECRET, or INDEXEXCHANGE_USERNAME + "
            "INDEXEXCHANGE_PASSWORD to enable."
        )

    try:
        mcp.run(transport="stdio")
    except Exception as e:
        logger.error("Failed to start server: %s", e)
        sys.exit(1)
