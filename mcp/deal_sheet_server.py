#!/usr/bin/env python3
"""Deal Sheet MCP Server.

Generates a branded XLSX deal sheet from a list of created deals. Used at
the end of a multi-deal-creation task to produce the trader-and-client
forwardable deliverable.

Outputs match the canonical Elcano deal-sheet layout:
- 6 columns (Partner / Deal Name / Deal ID / SSP / Channel / Recommended Bid)
- Channel banner rows (CTV / OLV / Display) grouping deals by channel
- Themed header + footer fill colors and embedded logo per client

The structural skeleton lives in
``protocols/email_styles/deal_sheet_template.xlsx`` (header row, column
widths, font choices, footer banner). Themes under
``protocols/email_styles/deal_sheet_themes/<theme>/`` provide a logo image
and the header/footer color overrides.
"""

from __future__ import annotations

import logging
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml
from mcp.server.fastmcp import FastMCP
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    stream=sys.stderr,
)
logger = logging.getLogger(__name__)

mcp = FastMCP("deal_sheet")

# Layout constants — derived from the canonical template. Update in lockstep
# with deal_sheet_template.xlsx if the layout ever shifts.
HEADER_ROW = 10
FIRST_DATA_ROW = 11
COL_PARTNER = 2
COL_DEAL_NAME = 3
COL_DEAL_ID = 4
COL_SSP = 5
COL_CHANNEL = 6
COL_BID = 7
LAST_DATA_COL = COL_BID

# Channel display order. Channels not in this list sort alphabetically after.
CHANNEL_ORDER = ["CTV", "OLV", "Display"]

DEFAULT_THEME = "elcano"
PROJECT_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = PROJECT_ROOT / "protocols" / "email_styles" / "deal_sheet_template.xlsx"
THEMES_ROOT = PROJECT_ROOT / "protocols" / "email_styles" / "deal_sheet_themes"

# Where built sheets land. Override with env var for containerized runs.
OUTPUT_DIR = Path(os.environ.get("DEAL_SHEET_OUTPUT_DIR", "/tmp/cutlass-deal-sheets"))

# Failed-row styling
FAIL_FONT_COLOR = "9A4040"  # muted red

# Section banner font (matches existing template — preserved across themes)
BANNER_FONT_NAME = "Aptos Narrow"
BANNER_FONT_SIZE = 11
DATA_FONT_SIZE = 12

# SSPs the multi-deal protocol can actually create on. Magnite joined in
# June 2026 when its ClearLine Curation Demand Management API shipped
# (magnite_execute_deal_from_prompt_inputs in magnite_mcp.py).
# Kept in sync with protocols/deal-brief.schema.yaml's enum and the
# scope.supported_ssps block in protocols/multi-deal-creation.yaml.
SUPPORTED_SSPS = frozenset(
    {
        "openx",
        "indexexchange",
        "index",  # display form sometimes used by traders
        "pubmatic",
        "xandr",
        "triplelift",
        "medianet",
        "mediadotnet",  # alt slug
        "magnite",
    }
)


def _normalize_ssp(ssp: Any) -> str:
    """Normalize an SSP token to the canonical form stored in SUPPORTED_SSPS.

    Traders supply SSPs in many forms — "Index Exchange", "indexexchange",
    "Media.net", "medianet" — and the canonical token is the lowercase
    no-whitespace no-punctuation form. Validating against the raw input
    rejected human-friendly spellings ("Index Exchange" became
    "index exchange" with a space and missed the set); normalizing first
    lets both "indexexchange" and "Index Exchange" resolve to the same
    canonical "indexexchange".
    """
    if not ssp:
        return ""
    return "".join(c for c in str(ssp).strip().lower() if c.isalnum())


# Required per-deal fields. Validated by validate_brief and again by
# build_deal_sheet as defense in depth.
REQUIRED_DEAL_FIELDS = ("ssp", "channel", "deal_name", "recommended_bid")

# Error-field keys we'll probe (in priority order) when extracting a
# failure reason from a raw SSP create response. Most SSPs use one of
# the first three; the fallback walks an `errors` list if present.
_ERROR_KEYS = ("error", "error_message", "message", "detail")


def _extract_failure_reason(ssp_response: Any) -> str:
    """Best-effort extraction of a human-readable failure reason from an
    SSP create response. Each SSP returns errors in a slightly different
    shape; this helper normalizes them so the deal sheet's failed-row
    rendering is consistent regardless of which SSP failed.

    Probe order:
      1. Top-level scalar error key (`error`, `error_message`, ...).
      2. `errors` list — first entry's `message`/`error`/`detail` if dict,
         else the string itself.
      3. Generic `"create returned success=false with no error detail"`
         when the payload clearly indicates failure but offers no string.
      4. `"unknown failure"` as the last-resort fallback.
    """
    if not isinstance(ssp_response, dict):
        return "unknown (no response payload captured)"
    for key in _ERROR_KEYS:
        v = ssp_response.get(key)
        if isinstance(v, str) and v.strip():
            return v.strip()
    errors = ssp_response.get("errors")
    if isinstance(errors, list) and errors:
        first = errors[0]
        if isinstance(first, str) and first.strip():
            return first.strip()
        if isinstance(first, dict):
            for key in _ERROR_KEYS:
                v = first.get(key)
                if isinstance(v, str) and v.strip():
                    return v.strip()
    if ssp_response.get("success") is False:
        return "create returned success=false with no error detail"
    return "unknown failure"


def _slugify(text: str) -> str:
    """Filesystem-safe slug for filenames. Lowercase, alnum + dashes only."""
    out = []
    last_dash = True
    for ch in text.lower():
        if ch.isalnum():
            out.append(ch)
            last_dash = False
        elif not last_dash:
            out.append("-")
            last_dash = True
    return "".join(out).strip("-") or "deal-sheet"


def _load_theme(theme: str) -> tuple[str, dict[str, Any]]:
    """Load a theme by name. Falls back to the default theme if the named
    one is missing. Returns (resolved_theme_name, theme_dict).
    """
    theme_dir = THEMES_ROOT / theme
    if not (theme_dir / "theme.yaml").exists():
        logger.warning("Theme %r not found at %s, falling back to %r", theme, theme_dir, DEFAULT_THEME)
        theme = DEFAULT_THEME
        theme_dir = THEMES_ROOT / theme
    with (theme_dir / "theme.yaml").open() as f:
        cfg = yaml.safe_load(f)
    cfg["_dir"] = theme_dir
    return theme, cfg


def _hex_for_openpyxl(hex_color: str) -> str:
    """Normalize a CSS-style hex color to openpyxl's preferred AARRGGBB form
    with explicit FF (fully opaque) alpha. openpyxl will accept a 6-char
    RRGGBB but serializes it with a leading '00' alpha which mismatches what
    Excel writes natively (FF). The visible color is identical either way,
    but pinning FF here keeps the generated bytes bit-comparable to
    trader-supplied templates."""
    base = hex_color.lstrip("#").upper()
    if len(base) == 6:
        return "FF" + base
    return base


def _channel_sort_key(channel: str) -> tuple[int, str]:
    """Sort order: known channels in CHANNEL_ORDER first, then alphabetical."""
    if channel in CHANNEL_ORDER:
        return (CHANNEL_ORDER.index(channel), "")
    return (len(CHANNEL_ORDER), channel.lower())


def _group_deals_by_channel(deals: list[dict[str, Any]]) -> list[tuple[str, list[dict[str, Any]]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for i, d in enumerate(deals):
        raw = d.get("channel")
        ch = (raw or "Unknown").strip() if isinstance(raw, str) else "Unknown"
        if not ch:
            ch = "Unknown"
        if not raw or (isinstance(raw, str) and not raw.strip()):
            logger.warning(
                "Deal %d (deal_name=%r, ssp=%r) has empty/missing channel — grouped under 'Unknown'.",
                i,
                d.get("deal_name"),
                d.get("ssp"),
            )
        grouped.setdefault(ch, []).append(d)
    return sorted(grouped.items(), key=lambda kv: _channel_sort_key(kv[0]))


def _set_cell(ws, row: int, col: int, value: Any, *, font: Font | None = None) -> None:
    """Set a cell value, skipping merged-cell members."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value
    if font is not None:
        cell.font = font


def _ensure_merge(ws, top: int, left_col: int, right_col: int) -> None:
    """Merge top:left to top:right if not already merged."""
    from openpyxl.utils import get_column_letter

    coord = f"{get_column_letter(left_col)}{top}:{get_column_letter(right_col)}{top}"
    if coord not in {str(r) for r in ws.merged_cells.ranges}:
        ws.merge_cells(coord)


def _unmerge_overlapping(ws, row: int) -> None:
    """If a row is part of an existing merge range, unmerge it. Used so we
    can rewrite section banners cleanly when the template ships with merges
    only at original banner rows.
    """
    to_remove = []
    for r in list(ws.merged_cells.ranges):
        if r.min_row <= row <= r.max_row:
            to_remove.append(str(r))
    for coord in to_remove:
        ws.unmerge_cells(coord)


def _write_header(ws, theme: dict[str, Any]) -> None:
    """Apply theme colors to header row 10 — keeps font name + alignment
    from the template, only swaps fill + text color."""
    bg = _hex_for_openpyxl(theme["header"]["background"])
    fg = _hex_for_openpyxl(theme["header"]["text_color"])
    fill = PatternFill(patternType="solid", fgColor=bg)
    for col in range(COL_PARTNER, LAST_DATA_COL + 1):
        cell = ws.cell(row=HEADER_ROW, column=col)
        if isinstance(cell, MergedCell):
            continue
        cell.fill = fill
        # Preserve existing font (Aptos Narrow 11 bold) and only swap color
        existing = cell.font
        cell.font = Font(
            name=existing.name or BANNER_FONT_NAME,
            size=existing.size or BANNER_FONT_SIZE,
            bold=True,
            color=fg,
        )


def _write_data_rows(ws, deals: list[dict[str, Any]]) -> int:
    """Write deals starting at FIRST_DATA_ROW, grouped by channel with
    banner rows. Returns the row number of the last written row + 1
    (caller can place the footer there).
    """
    grouped = _group_deals_by_channel(deals)
    row = FIRST_DATA_ROW

    banner_fill = PatternFill(patternType="solid", fgColor="D9D9D9")
    banner_font = Font(name=BANNER_FONT_NAME, size=BANNER_FONT_SIZE, bold=True, color="000000")
    data_font = Font(name=BANNER_FONT_NAME, size=DATA_FONT_SIZE, color="000000")
    fail_font = Font(name=BANNER_FONT_NAME, size=DATA_FONT_SIZE, color=FAIL_FONT_COLOR)
    center = Alignment(horizontal="center", vertical="center")

    for channel, channel_deals in grouped:
        # Banner row spans B:G
        _unmerge_overlapping(ws, row)
        _ensure_merge(ws, row, COL_PARTNER, LAST_DATA_COL)
        cell = ws.cell(row=row, column=COL_PARTNER)
        cell.value = channel
        cell.fill = banner_fill
        cell.font = banner_font
        cell.alignment = center
        row += 1

        for d in channel_deals:
            failed = (d.get("status") or "created").lower() == "failed"
            font = fail_font if failed else data_font
            _set_cell(ws, row, COL_PARTNER, d.get("partner") or "—", font=font)
            _set_cell(ws, row, COL_DEAL_NAME, d.get("deal_name") or "—", font=font)
            # When the agent passed an explicit failure_reason, use it
            # verbatim. Otherwise probe the raw ssp_response payload (if
            # supplied) for a normalized error string. Falling back to
            # "unknown error" preserves prior behavior.
            if failed:
                reason = d.get("failure_reason")
                if not (isinstance(reason, str) and reason.strip()):
                    reason = _extract_failure_reason(d.get("ssp_response"))
                deal_id_text = f"FAILED: {reason}"
            else:
                deal_id_text = d.get("deal_id") or "—"
            _set_cell(ws, row, COL_DEAL_ID, deal_id_text, font=font)
            _set_cell(ws, row, COL_SSP, d.get("ssp") or "—", font=font)
            _set_cell(ws, row, COL_CHANNEL, channel, font=font)
            _set_cell(ws, row, COL_BID, d.get("recommended_bid") or "—", font=font)
            for col in range(COL_PARTNER, LAST_DATA_COL + 1):
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    cell.alignment = center
            row += 1

    return row


def _write_footer(ws, theme: dict[str, Any], row: int) -> None:
    """Footer banner row — themed colors, merged across data columns."""
    bg = _hex_for_openpyxl(theme["footer"]["background"])
    fg = _hex_for_openpyxl(theme["footer"]["text_color"])
    fill = PatternFill(patternType="solid", fgColor=bg)
    font = Font(name=BANNER_FONT_NAME, size=BANNER_FONT_SIZE, bold=True, color=fg)

    _unmerge_overlapping(ws, row)
    _ensure_merge(ws, row, COL_PARTNER, LAST_DATA_COL)
    cell = ws.cell(row=row, column=COL_PARTNER)
    cell.value = ""
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _embed_logo(ws, theme: dict[str, Any]) -> None:
    """Place the theme's logo at the configured anchor cell, scaled to
    max_width while preserving aspect ratio. Skips silently if the logo
    file is missing — the rest of the sheet still renders.
    """
    cfg = theme.get("logo") or {}
    logo_path = (theme["_dir"] / cfg.get("file", "logo.png")).resolve()
    if not logo_path.exists():
        logger.warning("Logo file not found: %s", logo_path)
        return
    try:
        img = XLImage(str(logo_path))
    except Exception as exc:
        logger.warning("Failed to embed logo %s: %s", logo_path, exc)
        return

    max_w = int(cfg.get("max_width_px", 360))
    if img.width and img.width > max_w:
        ratio = max_w / img.width
        img.width = max_w
        img.height = int(img.height * ratio)
    ws.add_image(img, cfg.get("anchor_cell", "B1"))


@mcp.tool()
async def build_deal_sheet(
    deals: list[dict[str, Any]],
    client_name: str,
    theme: str = DEFAULT_THEME,
    output_filename: str | None = None,
) -> dict[str, Any]:
    """Build a branded deal-sheet XLSX from a list of created deals.

    Parameters
    ----------
    deals
        List of per-deal records. Each dict should include:
        ``partner`` (str, e.g. "Soundwave"),
        ``deal_name`` (str),
        ``deal_id`` (str, e.g. "OX-bef-V06sGg"),
        ``ssp`` (str, e.g. "OpenX"),
        ``channel`` (str: CTV / OLV / Display / etc.),
        ``recommended_bid`` (str). Canonical per-channel ranges: Display "$2-$5",
            OLV "$8-$12", CTV "$25-$35". Free text — the SSP doesn't see this.
        ``status`` ("created" or "failed", default "created"),
        ``failure_reason`` (str, only when status="failed").
    client_name
        Display name of the end client (used for filename + email subject).
        E.g. "Aruba Tourism Authority" or "Carmichael Lynch".
    theme
        Theme key matching a folder under
        ``protocols/email_styles/deal_sheet_themes/``. Unknown values fall
        back to the ``elcano`` default.
    output_filename
        Optional override. Defaults to
        ``{client_slug}_deal_sheet_{YYYYMMDD}.xlsx``.

    Returns
    -------
    dict
        ``{"success": True, "path": str, "deal_count": int, "theme_used": str}``
        on success, or ``{"success": False, "error": str}`` on failure.
    """
    if not deals:
        return {"success": False, "error": "deals list is empty — nothing to build"}
    if not client_name or not client_name.strip():
        return {"success": False, "error": "client_name is required"}
    if not TEMPLATE_PATH.exists():
        return {"success": False, "error": f"deal_sheet_template.xlsx missing at {TEMPLATE_PATH}"}

    # Defense in depth — the multi-deal protocol calls validate_brief
    # before audit, but if some other caller skips that step, refuse to
    # build a sheet that includes unsupported SSPs. Otherwise we'd quietly
    # render rows for deals cutlass cannot create.
    unsupported = [
        (i, str(d.get("ssp") or "")) for i, d in enumerate(deals) if _normalize_ssp(d.get("ssp")) not in SUPPORTED_SSPS
    ]
    if unsupported:
        return {
            "success": False,
            "error": (
                f"refusing to build sheet: {len(unsupported)} deal(s) reference unsupported SSP(s). "
                f"Supported SSPs: {sorted(SUPPORTED_SSPS)}. "
                f"Unsupported entries: {unsupported}."
            ),
            "unsupported_deals": [{"deal_index": i, "ssp": s} for i, s in unsupported],
        }

    resolved_theme, theme_cfg = _load_theme(theme)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if output_filename is None:
        date = datetime.now().strftime("%Y%m%d")
        output_filename = f"{_slugify(client_name)}_deal_sheet_{date}.xlsx"
    # Path(name).name strips any directory components: an absolute path or
    # ../ in output_filename would otherwise escape OUTPUT_DIR entirely
    # (PosixPath("/out") / "/etc/x.xlsx" == /etc/x.xlsx).
    output_path = OUTPUT_DIR / Path(output_filename).name

    # Copy the template, then mutate the copy.
    shutil.copyfile(TEMPLATE_PATH, output_path)
    wb = load_workbook(output_path)
    ws = wb.active
    if ws is None:
        return {"success": False, "error": "deal_sheet_template.xlsx has no active worksheet"}
    # Drop any embedded images carried over from the template (defensive —
    # the canonical template ships with images stripped already).
    ws._images = []

    # Pre-strip any merges that fall inside the data area so subsequent
    # writes aren't silently dropped by openpyxl's merged-cell guard. The
    # template ships with banners merged at rows 11/16/21; once we unmerge
    # them we redraw the banners we actually want at the right rows.
    for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + 200):
        _unmerge_overlapping(ws, row)

    _embed_logo(ws, theme_cfg)
    _write_header(ws, theme_cfg)
    last_row = _write_data_rows(ws, deals)
    _write_footer(ws, theme_cfg, last_row)

    # Drop any rows the template baked in beyond what this batch needed.
    # The canonical template carries 21 rows of structural scaffolding sized
    # for a typical 8-deal sheet (4 CTV + 4 OLV + footer). A 2-deal batch
    # only uses through row 14, so rows 15..21 would otherwise render as
    # empty scaffolding below the footer with stale row-height + cell-style
    # carried over from the template. delete_rows clears them entirely.
    if ws.max_row > last_row:
        ws.delete_rows(last_row + 1, amount=ws.max_row - last_row)

    wb.save(output_path)

    logger.info(
        "Built deal sheet for %r: %d deals, theme=%s, path=%s",
        client_name,
        len(deals),
        resolved_theme,
        output_path,
    )
    return {
        "success": True,
        "path": str(output_path),
        "deal_count": len(deals),
        "theme_used": resolved_theme,
    }


@mcp.tool()
async def validate_brief(deals: list[dict[str, Any]]) -> dict[str, Any]:
    """Validate a multi-deal brief before audit.

    Call this BEFORE `confirm_audit` in a multi-deal flow. It rejects
    obvious problems — unsupported SSPs, missing required per-deal
    fields — that would otherwise show up only at create time, after
    the audit token has been declared.

    Parameters
    ----------
    deals
        List of per-deal records matching the schema in
        ``protocols/deal-brief.schema.yaml``. Each entry MUST include
        ``ssp``, ``channel``, ``deal_name``, ``recommended_bid`` at
        minimum.

    Returns
    -------
    dict
        ``{"ok": bool, "issues": [...], "deal_count": int}``. When
        ``ok`` is false, ``issues`` lists each problem with
        ``deal_index``, ``field``, and ``message``. A multi-deal flow
        MUST refuse to call ``confirm_audit`` while ``ok`` is false —
        either fix the brief or abort the task.
    """
    if not isinstance(deals, list) or not deals:
        return {
            "ok": False,
            "issues": [{"deal_index": -1, "field": "deals", "message": "brief must contain a non-empty list of deals"}],
            "deal_count": 0,
        }

    issues: list[dict[str, Any]] = []
    for i, d in enumerate(deals):
        if not isinstance(d, dict):
            issues.append(
                {"deal_index": i, "field": "<root>", "message": f"deal entry is not an object (got {type(d).__name__})"}
            )
            continue
        ssp_raw = d.get("ssp")
        ssp = _normalize_ssp(ssp_raw)
        if not ssp:
            issues.append({"deal_index": i, "field": "ssp", "message": "missing required ssp field"})
        elif ssp not in SUPPORTED_SSPS:
            issues.append(
                {
                    "deal_index": i,
                    "field": "ssp",
                    "message": f"unsupported SSP {ssp_raw!r}. Supported: {sorted(SUPPORTED_SSPS)}.",
                }
            )
        for required in REQUIRED_DEAL_FIELDS:
            if required == "ssp":
                continue
            v = d.get(required)
            if v is None or (isinstance(v, str) and not v.strip()):
                issues.append({"deal_index": i, "field": required, "message": f"missing required field {required!r}"})

    return {"ok": len(issues) == 0, "issues": issues, "deal_count": len(deals)}


@mcp.tool()
async def list_deal_sheet_themes() -> dict[str, Any]:
    """List available deal-sheet themes. Useful when an agent needs to
    confirm a theme name exists before building a sheet."""
    if not THEMES_ROOT.exists():
        return {"themes": [], "error": f"themes directory missing at {THEMES_ROOT}"}
    themes = []
    for entry in sorted(THEMES_ROOT.iterdir()):
        if not entry.is_dir():
            continue
        cfg_path = entry / "theme.yaml"
        if not cfg_path.exists():
            continue
        with cfg_path.open() as f:
            cfg = yaml.safe_load(f) or {}
        themes.append(
            {
                "key": entry.name,
                "name": cfg.get("name", entry.name),
                "description": cfg.get("description", ""),
            }
        )
    return {"themes": themes}


if __name__ == "__main__":
    logger.info("Starting Deal Sheet MCP Server")
    try:
        mcp.run(transport="stdio")
    except Exception as exc:
        logger.exception("Failed to start server: %s", exc)
        sys.exit(1)
