#!/usr/bin/env python3
"""
PubMatic MCP Server

Custom MCP integration for PubMatic curated deals (auction packages).
"""

import csv
import hashlib
import json
import logging
import os
import re
import sys
import urllib.parse
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any
from uuid import uuid4

import httpx
from mcp.server.fastmcp import FastMCP

_NON_ALNUM_RE = re.compile(r"[^a-z0-9]+")
_SAFE_KEY_RE = re.compile(r"[^A-Za-z0-9_-]")
_SCHEME_RE = re.compile(r"^[a-z]+://")
_IAB_CODE_RE = re.compile(r"^IAB\d+(-\d+)?$", re.IGNORECASE)


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    stream=sys.stderr,
)
logger = logging.getLogger(__name__)

mcp = FastMCP("pubmatic_mcp")

USER_AGENT = "cutlass/1.0"
DEFAULT_TIMEOUT = 60.0
DEFAULT_BASE_URL = "https://api.pubmatic.com"
DEFAULT_REPORT_DOWNLOAD_DIR = os.path.expanduser("~/Victoria/pubmatic_reports")

# PubMatic owner ID for this MCP variant. Drives loggedInOwnerId on deal creates,
# the security validation guard (briefs must declare an owner that matches this
# subprocess's account, no cross-account leaks), and the recipientId on curator
# fees (curator fee always flows to the owner doing the curation).
#
# Defaults to Elcano's id (60067). Per-variant override via the variant-suffix
# loader: setting PUBMATIC_OWNER_ID_REKLAIM=50751 in the operator's env causes
# the openx_mcp.py / pubmatic_mcp_reklaim subprocess to see PUBMATIC_OWNER_ID=50751.
PUBMATIC_OWNER_ID = int(os.environ.get("PUBMATIC_OWNER_ID", "60067").strip() or "60067")
# Back-compat alias retained because many call sites and docstrings reference the
# old name; both bind to the same per-subprocess value.
ELCANO_OWNER_ID = PUBMATIC_OWNER_ID
ALLOWED_OWNER_TYPE_IDS = {5, 7}
REQUEST_TYPE_CREATE = "CREATE"
REQUEST_TYPE_UPDATE = "UPDATE"

# Deal status codes per the Curated Deals API. Only Active/Inactive are
# agent-settable; the rest are system lifecycle states reported by PubMatic.
PUBMATIC_DEAL_STATUS_NAMES = {
    1: "Active",
    2: "Inactive",
    4: "Scheduled",
    5: "Completed",
    10: "InProgress",
    11: "Failed",
    12: "PartiallyFailed",
}
PUBMATIC_SETTABLE_STATUS_IDS = {1, 2}
ZULU_MILLIS_RE = re.compile(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\.\d{3}Z$")
ZULU_NO_MILLIS_RE = re.compile(r"^(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2})Z$")
DATE_ONLY_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
DOMAIN_PATTERN = re.compile(r"^(?:[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?\.)+[a-z]{2,63}$")
# App-bundle identifiers are not web domains: CTV/OTT inventory lists carry
# reverse-DNS bundle IDs (com.zumobi.msnbc — also matches DOMAIN_PATTERN, plus
# dotted IDs with numeric final labels that DOMAIN_PATTERN rejects) and bare
# numeric store IDs (Roku/Apple/Amazon, e.g. 523428113). PubMatic targets web
# domains and app bundles through one undifferentiated domainList, so accept
# all three shapes here — a web-domain-only validator silently dropped bundles.
APP_BUNDLE_PATTERN = re.compile(r"^[a-z0-9_]+(?:\.[a-z0-9_]+)+$")
APP_STORE_NUMERIC_ID_PATTERN = re.compile(r"^\d{6,}$")


def _is_acceptable_domain_or_bundle(value: str) -> bool:
    """True if value is a web domain OR an app-bundle/store-id shape."""
    return bool(
        DOMAIN_PATTERN.match(value) or APP_BUNDLE_PATTERN.match(value) or APP_STORE_NUMERIC_ID_PATTERN.match(value)
    )


# PubMatic curated-deal detail URL. The PubMatic UI expects three query
# params: dealId (numeric internal id), dealName (full deal name,
# url-encoded), dealCategoryId (3 = Curated). The earlier path-style URL
# (`/deals/curated/{id}`) routes to a list page that doesn't surface the
# specific deal — it appeared in deal-creation responses but the link
# didn't open the deal in the trader's UI.
PUBMATIC_DEAL_URL_TEMPLATE = os.environ.get(
    "PUBMATIC_DEAL_URL_TEMPLATE",
    "https://apps.pubmatic.com/v3/common/pmc/deals?dealId={id}&dealName={name}&dealCategoryId=3",
)

PUBMATIC_REPORT_ACCOUNT_IDS: dict[str, int] = {
    "elcano": 60067,
    "elcano formerly hyphatec": 60067,
    "hyphatec": 60067,
}

PUBMATIC_REPORT_PRESETS: dict[str, dict[str, Any]] = {
    "deal_summary": {
        "description": "Daily buyer-side deal and DSP performance using core spend metrics.",
        "dimensions": ["date", "dealMetaId", "dspId"],
        "metrics": ["paidImpressions", "spend", "ecpm"],
    },
    "dsp_daily": {
        "description": "Daily DSP performance with impression and spend trends.",
        "dimensions": ["date", "dspId"],
        "metrics": ["paidImpressions", "spend", "ecpm"],
    },
    "site_daily": {
        "description": "Daily publisher performance with impression and spend trends.",
        "dimensions": ["date", "pubId"],
        "metrics": ["paidImpressions", "spend", "ecpm"],
    },
}

PUBMATIC_REPORT_DIMENSION_ALIASES: dict[str, str] = {
    "date": "date",
    "day": "date",
    "daily": "date",
    "month": "month",
    "monthly": "month",
    "hour": "hour",
    "hourly": "hour",
    "deal": "dealMetaId",
    "deal id": "dealMetaId",
    "deal name": "dealMetaId",
    "bundled deal": "bundledDealMetaId",
    "dsp": "dspId",
    "buyer": "atdId",
    "atd": "atdId",
    "publisher": "pubId",
    "site": "siteId",
    "platform": "platformId",
    "channel": "channelId",
    "domain": "winningPublisherDomain",
    "app": "appName",
}

PUBMATIC_REPORT_METRIC_ALIASES: dict[str, str] = {
    "impressions": "paidImpressions",
    "paid impressions": "paidImpressions",
    "spend": "spend",
    "ecpm": "ecpm",
    "e c p m": "ecpm",
    "clicks": "clicks",
    "ctr": "ctr",
    "media spend": "mediaSpend",
    "revenue": "revenue",
    "net revenue": "netRevenue",
    "fees": "transactionRevenue",
    "fee": "transactionRevenue",
    "total marketplace fees": "transactionRevenue",
    "marketplace fees": "transactionRevenue",
    "total fees": "transactionRevenue",
    "vendor fees": "transactionRevenue",
    "transaction revenue": "transactionRevenue",
    "transaction ecpm": "transactionEcpm",
    "total requests": "totalRequests",
    "win rate": "winRate",
    "bid rate": "dpBidRate",
    "video through rate": "vtr",
    "vtr": "vtr",
    "video clicks": "videoClicksCount",
    "completed": "videoCompleteCount",
    "25% view": "video25PercentCount",
    "50% view": "video50PercentCount",
    "75% view": "video75PercentCount",
}

PUBMATIC_REPORT_TYPE_PRESET_ALIASES: dict[str, str] = {
    "deal performance": "deal_summary",
    "deal summary": "deal_summary",
    "deal report": "deal_summary",
    "dsp performance": "dsp_daily",
    "dsp report": "dsp_daily",
    "publisher performance": "site_daily",
    "publisher report": "site_daily",
    "site performance": "site_daily",
    "site report": "site_daily",
}

PUBMATIC_DEVICE_TYPE_ALIASES = {
    "desktop": "DESKTOP",
    "mobile": "MOBILE",
    "tablet": "TABLET",
    "ctv": "CTV",
    "connectedtv": "CTV",
    "connected tv": "CTV",
    "ott": "CTV",
    "phone": "MOBILE",
}

PUBMATIC_TARGETING_FIELD_ALIASES = {
    "ownerId": "ownerId",
    "ownerType": "ownerType",
    "publishers": "publishers",
    "publisherIds": "publisherIds",
    "segments": "segments",
    "segmentIds": "segmentIds",
    "domains": "domains",
    "domainList": "domainList",
    "includedDomains": "includedDomains",
    "geo": "geo",
    "geography": "geography",
    "countries": "countries",
    "states": "states",
    "devices": "devices",
    "deviceTypes": "deviceTypes",
    "iabCategories": "iabCategories",
    "categories": "categories",
    "viewability": "viewability",
    "viewabilityThreshold": "viewabilityThreshold",
}

PUBMATIC_PUBLISHER_KEYS = ("id", "publisherId", "accountId")
# PubMatic domainMatchType enum: 1 = include (allowlist / serve only these),
# 2 = exclude (blocklist). Confirmed against the PubMatic targeting API.
PUBMATIC_DOMAIN_MATCH_TYPE_INCLUDE = 1
PUBMATIC_DOMAIN_MATCH_TYPE_EXCLUDE = 2
# Default when a list file is supplied without an explicit operator: include
# (allowlist), preserving the historical hardcoded behavior.
PUBMATIC_DEFAULT_DOMAIN_MATCH_TYPE = PUBMATIC_DOMAIN_MATCH_TYPE_INCLUDE


def _resolve_domain_match_type(operator: str | None) -> int:
    """Map a human-readable match operator to PubMatic's domainMatchType enum.

    ``"allowlist"`` (or None — the default) → 1 (include); ``"blocklist"`` → 2
    (exclude). Raises on any other value so a typo never silently mis-routes.
    """
    if operator is None:
        return PUBMATIC_DEFAULT_DOMAIN_MATCH_TYPE
    normalized = str(operator).strip().lower()
    if normalized == "allowlist":
        return PUBMATIC_DOMAIN_MATCH_TYPE_INCLUDE
    if normalized == "blocklist":
        return PUBMATIC_DOMAIN_MATCH_TYPE_EXCLUDE
    raise ValueError(f"Unsupported domain_match_operator '{operator}'. Pass 'blocklist' or 'allowlist'.")


PUBMATIC_DEFAULT_DEVICE_MAKE_TARGETING = 0
PUBMATIC_DEVICE_TYPE_IDS = {
    "desktop": 1,
    "mobile": 2,
    "tablet": 3,
    "mobileapp": 4,
    "mobileweb": 5,
    "ctv": 7,
    "connectedtv": 7,
    "connected tv": 7,
    "ott": 7,
}
# Platform ids per the Curated Deals API docs (audited 2026-06-12):
# 1 Web, 2 Mobile Web, 3 Not Defined, 4 Mobile App iOS, 5 Mobile App Android,
# 7 CTV. The earlier set {1,2,4,5} was built on a mislabeled enum that called
# 5 "CTV" — 5 is Android in-app; real CTV is 7 and was being REJECTED here.
PUBMATIC_ALLOWED_PLATFORM_IDS = {1, 2, 4, 5, 7}
PUBMATIC_CTV_PLATFORM_ID = 7
PUBMATIC_ALLOWED_AD_FORMAT_IDS = {3, 12, 13}

# Channel-aware device + ad-format defaults. When the caller passes `channel`
# (or a recognized channel hint) and omits `device_types`/`ad_formats`, the
# prepare flow auto-fills both with PubMatic's canonical values for that
# channel. Mirrors the OpenX `DEFAULT_RENDERING_CONTEXTS` and the IX
# `_ensure_deal_type_targeting_defaults` pattern so the agent never has to
# enumerate device IDs or format integers by hand.
#
# Per the Elcano trader spec (canonical across all SSPs):
#
#   Channel | ad_formats   | devices                  | Notes
#   --------|--------------|--------------------------|----------------------
#   display | [3] Banner   | desktop+mobile+tablet    |
#   olv     | [12] Video   | desktop+mobile+tablet    | NOT a display variant
#   ctv     | [12] Video   | ctv                      | Forced app-only inventory
#   ott     | [12] Video   | desktop+mobile+tablet    | Rare; in-app video on mobile
#
# PubMatic ad-format integers (from PUBMATIC_ALLOWED_AD_FORMAT_IDS):
#   3  = Banner / Display
#   12 = Video
#   13 = Native
PM_DEVICE_VALUES_DISPLAY: tuple[str, ...] = ("desktop", "mobile", "tablet")
PM_DEVICE_VALUES_OLV: tuple[str, ...] = ("desktop", "mobile", "tablet")
PM_DEVICE_VALUES_CTV: tuple[str, ...] = ("ctv",)
PM_DEVICE_VALUES_OTT: tuple[str, ...] = ("desktop", "mobile", "tablet")
PM_AD_FORMATS_DISPLAY: tuple[int, ...] = (3,)  # Banner
PM_AD_FORMATS_VIDEO: tuple[int, ...] = (12,)  # Video — used by olv / ctv / ott
PM_CHANNELS_DISPLAY: frozenset[str] = frozenset({"display"})
PM_CHANNELS_OLV: frozenset[str] = frozenset({"olv", "display_olv", "display/olv"})
PM_CHANNELS_CTV: frozenset[str] = frozenset({"ctv"})
PM_CHANNELS_OTT: frozenset[str] = frozenset({"ott"})

# Default curator-fee shape for this MCP variant. PubMatic expresses curator
# fees via the `dealFees` array on the deal payload; the 30% Percent-of-Media
# transaction-fee STRUCTURE is the business-level default for every Cutlass
# curator deal regardless of seat, while the recipient NAME and ID are
# account-specific. Field shape per PubMatic's "Create a curated deal" docs:
# feeValueType=0 (Percentage), feeType=0 (Transaction Fee),
# recipientTypeId=7 (Buyer / Data Provider).
#
# Recipient name is derived from MCP_VARIANT_CLIENT (injected by the mcp loader
# from the `client=` arg passed to mcp_load_servers). Empty/unset → "Elcano"
# (the historical default). The capitalize() handles "reklaim" → "Reklaim" so
# the name shown in PubMatic's UI matches the seat's account name.
PM_FEE_VALUE_TYPE_PERCENTAGE: int = 0
PM_FEE_VALUE_TYPE_CPM: int = 1
PM_FEE_TYPE_TRANSACTION: int = 0
PM_FEE_RECIPIENT_TYPE_BUYER: int = 7
_MCP_VARIANT_CLIENT = os.environ.get("MCP_VARIANT_CLIENT", "").strip()
DEFAULT_FEE_RECIPIENT_NAME: str = _MCP_VARIANT_CLIENT.capitalize() if _MCP_VARIANT_CLIENT else "Elcano"
# Back-compat alias retained for code that still references the old name.
ELCANO_FEE_RECIPIENT_NAME: str = DEFAULT_FEE_RECIPIENT_NAME
ELCANO_DEFAULT_FEE_VALUE_PERCENT: float = 30.0


def _normalize_pm_channel(channel: str | None) -> str | None:
    """Return canonical "display", "olv", "ctv", "ott", or None for an input channel hint."""
    if channel is None:
        return None
    normalized = str(channel).strip().lower()
    if normalized in PM_CHANNELS_DISPLAY:
        return "display"
    if normalized in PM_CHANNELS_OLV:
        return "olv"
    if normalized in PM_CHANNELS_CTV:
        return "ctv"
    if normalized in PM_CHANNELS_OTT:
        return "ott"
    return None


_PM_CHANNEL_DEVICE_DEFAULTS: dict[str, tuple[str, ...]] = {
    "display": PM_DEVICE_VALUES_DISPLAY,
    "olv": PM_DEVICE_VALUES_OLV,
    "ctv": PM_DEVICE_VALUES_CTV,
    "ott": PM_DEVICE_VALUES_OTT,
}

_PM_CHANNEL_AD_FORMAT_DEFAULTS: dict[str, tuple[int, ...]] = {
    "display": PM_AD_FORMATS_DISPLAY,
    "olv": PM_AD_FORMATS_VIDEO,
    "ctv": PM_AD_FORMATS_VIDEO,
    "ott": PM_AD_FORMATS_VIDEO,
}

# Channel-aware PLATFORM defaults (Curated Deals enum: 1 Web, 2 Mobile Web,
# 4 Mobile App iOS, 5 Mobile App Android, 7 CTV). display/olv keep the
# long-standing [1] (Web) default; ctv maps to the real CTV platform 7 and
# ott to the mobile in-app pair.
_PM_CHANNEL_PLATFORM_DEFAULTS: dict[str, tuple[int, ...]] = {
    "display": (1,),
    "olv": (1,),
    "ctv": (PUBMATIC_CTV_PLATFORM_ID,),
    "ott": (4, 5),
}


def _apply_pm_channel_platform_defaults(
    platforms: list[int] | None,
    channel: str | None,
) -> tuple[list[int], bool]:
    """Auto-fill platforms from a channel hint when caller didn't supply any.

    Returns (resolved_platforms, applied_default). Falls back to the legacy
    [1] (Web) default when no channel hint resolves, preserving historical
    behavior for channel-less prompts.
    """
    if platforms:
        return platforms, False
    canonical = _normalize_pm_channel(channel)
    if canonical and canonical in _PM_CHANNEL_PLATFORM_DEFAULTS:
        return list(_PM_CHANNEL_PLATFORM_DEFAULTS[canonical]), True
    return [1], False


def _apply_pm_channel_device_defaults(
    device_types: list[str] | None,
    channel: str | None,
) -> tuple[list[str] | None, bool]:
    """Auto-fill device_types from a channel hint when caller didn't supply any.

    Returns (resolved_device_types, applied_default). `applied_default` is True
    only when channel defaults filled in an empty list — used to emit the
    `pm_default_channel_devices_applied` quality flag.
    """
    if device_types:
        return device_types, False
    canonical = _normalize_pm_channel(channel)
    if canonical and canonical in _PM_CHANNEL_DEVICE_DEFAULTS:
        return list(_PM_CHANNEL_DEVICE_DEFAULTS[canonical]), True
    return device_types, False


def _apply_pm_channel_ad_format_defaults(
    ad_formats: list[int] | None,
    channel: str | None,
) -> tuple[list[int] | None, bool]:
    """Auto-fill ad_formats from a channel hint when caller didn't supply any.

    Returns (resolved_ad_formats, applied_default). When `channel` resolves to
    one of the canonical channels and `ad_formats` is None/empty, returns the
    PubMatic canonical ad-format integer list for that channel. Display →
    [3] (Banner); olv/ctv/ott → [12] (Video). Used to emit the
    `pm_default_channel_ad_formats_applied` quality flag.
    """
    if ad_formats:
        return ad_formats, False
    canonical = _normalize_pm_channel(channel)
    if canonical and canonical in _PM_CHANNEL_AD_FORMAT_DEFAULTS:
        return list(_PM_CHANNEL_AD_FORMAT_DEFAULTS[canonical]), True
    return ad_formats, False


def _build_elcano_curator_fee_entry(fee_percent: float) -> dict[str, Any]:
    """Build a single Percentage-of-Media curator-fee entry for Elcano."""
    return {
        "recipientName": ELCANO_FEE_RECIPIENT_NAME,
        "recipientId": ELCANO_OWNER_ID,
        "recipientTypeId": PM_FEE_RECIPIENT_TYPE_BUYER,
        "feeValue": float(fee_percent),
        "feeValueType": PM_FEE_VALUE_TYPE_PERCENTAGE,
        "feeType": PM_FEE_TYPE_TRANSACTION,
    }


def _resolve_pm_deal_fees(
    fee: list[dict[str, Any]] | dict[str, Any] | None,
) -> tuple[list[dict[str, Any]], bool]:
    """Resolve the dealFees array from an explicit caller value or default.

    Returns (deal_fees, applied_default). When `fee` is None, returns the
    standard 30% Elcano curator-fee entry and applied_default=True so the
    caller can emit the `pm_default_curator_fee_applied` quality flag. When
    `fee` is a single dict it is wrapped into a list. When `fee` is already
    a list it is used verbatim.
    """
    if fee is None:
        return [_build_elcano_curator_fee_entry(ELCANO_DEFAULT_FEE_VALUE_PERCENT)], True
    if isinstance(fee, dict):
        return [fee], False
    if isinstance(fee, list):
        return [entry for entry in fee if isinstance(entry, dict)], False
    raise ValueError(f"fee must be a dict, list of dicts, or None — got {type(fee).__name__}")


def _make_pm_quality_flag(flag: str, impact: str, **context: Any) -> dict[str, Any]:
    """Build a structured quality_flags entry. Mirrors the IX pattern."""
    entry: dict[str, Any] = {"flag": flag, "impact": impact}
    for key, value in context.items():
        if value is not None:
            entry[key] = value
    return entry


def _blockers_to_quality_flags(blockers: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Surface existing blockers as structured quality_flags.

    Each blocker already carries a `code` and `message`; map them to the
    `flag`/`impact` shape so the protocol's `quality_flags_contract` has data
    to render.
    """
    quality_flags: list[dict[str, Any]] = []
    for blocker in blockers:
        if not isinstance(blocker, dict):
            continue
        code = blocker.get("code") or "pm_blocker"
        message = blocker.get("message") or ""
        context = {key: value for key, value in blocker.items() if key not in {"code", "message"}}
        quality_flags.append(
            _make_pm_quality_flag(f"pm_{code}" if not code.startswith("pm_") else code, message, **context)
        )
    return quality_flags


def _normalize_error(err: Any) -> dict[str, Any]:
    if isinstance(err, list):
        return {
            "message": "; ".join(item.get("errorMessage", str(item)) for item in err if isinstance(item, dict)),
            "errors": err,
        }
    if isinstance(err, dict):
        if "errorMessage" in err:
            return {"message": err.get("errorMessage"), "errors": [err]}
        return {
            "message": err.get("message") or err.get("error") or str(err),
            "errors": [err],
        }
    return {"message": str(err), "errors": []}


def _ok(payload: dict[str, Any]) -> dict[str, Any]:
    return {"success": True, **payload}


def _err(prefix: str, error: Any) -> dict[str, Any]:
    normalized = error if isinstance(error, dict) else _normalize_error(error)
    return {
        "success": False,
        "error": f"{prefix}: {normalized.get('message', 'unknown error')}",
        "details": normalized.get("errors", []),
    }


def _validate_owner_context(owner_id: int, owner_type_id: int) -> None:
    if owner_id != ELCANO_OWNER_ID:
        raise ValueError(f"SECURITY: loggedInOwnerId must be {ELCANO_OWNER_ID}")
    if owner_type_id not in ALLOWED_OWNER_TYPE_IDS:
        raise ValueError(f"SECURITY: loggedInOwnerTypeId must be one of {sorted(ALLOWED_OWNER_TYPE_IDS)}")


def _normalize_zulu_datetime(field_name: str, value: str) -> str:
    """Validate and normalize a Zulu datetime string, adding .000 milliseconds if missing."""
    if DATE_ONLY_RE.match(value):
        return f"{value}T00:00:00.000Z"
    if ZULU_MILLIS_RE.match(value):
        return value
    m = ZULU_NO_MILLIS_RE.match(value)
    if m:
        return m.group(1) + ".000Z"
    raise ValueError(f"{field_name} must use format YYYY-MM-DDTHH:MM:SS[.000]Z (e.g. 2026-03-01T00:00:00Z)")


def _extract_ownership_value(payload: Any) -> int | None:
    keys = {
        "ownedById",
        "ownerId",
        "loggedInOwnerId",
        "publisherId",
    }
    if isinstance(payload, dict):
        for key in keys:
            val = payload.get(key)
            if isinstance(val, int):
                return val
            if isinstance(val, str) and val.isdigit():
                return int(val)
        for v in payload.values():
            nested = _extract_ownership_value(v)
            if nested is not None:
                return nested
    if isinstance(payload, list):
        for item in payload:
            nested = _extract_ownership_value(item)
            if nested is not None:
                return nested
    return None


def _normalize_lookup_text(value: Any) -> str:
    return _NON_ALNUM_RE.sub("", str(value or "").strip().lower())


def _resolve_pubmatic_report_account_id(account_id: int | str) -> int:
    if isinstance(account_id, int):
        return account_id
    if isinstance(account_id, str):
        normalized = account_id.strip()
        if normalized.isdigit():
            return int(normalized)
        resolved = PUBMATIC_REPORT_ACCOUNT_IDS.get(normalized.lower())
        if resolved is not None:
            return resolved
    raise ValueError("Unknown PubMatic data provider account. Use Elcano or provide a numeric account ID.")


def _normalize_pubmatic_datetime(value: str, *, end_of_day: bool = False) -> str:
    normalized = str(value).strip()
    if DATE_ONLY_RE.match(normalized):
        return f"{normalized}T23:59" if end_of_day else f"{normalized}T00:00"
    try:
        parsed = datetime.fromisoformat(normalized.replace("Z", "+00:00"))
    except ValueError as exc:
        raise ValueError("PubMatic report datetimes must use YYYY-MM-DD or YYYY-MM-DDTHH:MM format") from exc
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(UTC).replace(tzinfo=None)
    return parsed.strftime("%Y-%m-%dT%H:%M")


def _resolve_pubmatic_date_range(date_range: dict[str, Any]) -> tuple[str, str]:
    if "from" in date_range and "to" in date_range:
        return (
            _normalize_pubmatic_datetime(str(date_range["from"])),
            _normalize_pubmatic_datetime(str(date_range["to"]), end_of_day=True),
        )

    now = datetime.now(UTC)
    if "current" in date_range:
        current = str(date_range["current"]).strip().lower()
        if current == "month":
            start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            return start.strftime("%Y-%m-%dT%H:%M"), now.strftime("%Y-%m-%dT%H:%M")
        if current == "day":
            start = now.replace(hour=0, minute=0, second=0, microsecond=0)
            return start.strftime("%Y-%m-%dT%H:%M"), now.strftime("%Y-%m-%dT%H:%M")
        raise ValueError("Unsupported current date range for PubMatic. Use 'day' or 'month'.")

    previous = date_range.get("previous")
    if isinstance(previous, dict):
        if previous.get("days"):
            days = int(previous["days"])
            end = now.replace(hour=23, minute=59, second=0, microsecond=0)
            start = (end - timedelta(days=max(days - 1, 0))).replace(hour=0, minute=0)
            return start.strftime("%Y-%m-%dT%H:%M"), end.strftime("%Y-%m-%dT%H:%M")
        if previous.get("weeks"):
            weeks = int(previous["weeks"])
            days = weeks * 7
            end = now.replace(hour=23, minute=59, second=0, microsecond=0)
            start = (end - timedelta(days=max(days - 1, 0))).replace(hour=0, minute=0)
            return start.strftime("%Y-%m-%dT%H:%M"), end.strftime("%Y-%m-%dT%H:%M")
        if previous.get("months"):
            months = int(previous["months"])
            end = now.replace(hour=23, minute=59, second=0, microsecond=0)
            start = (end - timedelta(days=max((months * 30) - 1, 0))).replace(hour=0, minute=0)
            return start.strftime("%Y-%m-%dT%H:%M"), end.strftime("%Y-%m-%dT%H:%M")

    raise ValueError("PubMatic date_range must include from/to, current, or previous.")


def _apply_pubmatic_date_dimension(dimensions: list[str], date_unit: str) -> list[str]:
    requested_unit = str(date_unit).strip().lower()
    if requested_unit == "month":
        return ["month" if dimension == "date" else dimension for dimension in dimensions]
    if requested_unit == "hour":
        return ["hour" if dimension == "date" else dimension for dimension in dimensions]
    return dimensions


def _flatten_pubmatic_analytics_rows(payload: dict[str, Any]) -> tuple[list[str], list[dict[str, Any]]]:
    columns = [str(column) for column in payload.get("columns", []) if isinstance(column, str)]
    rows = payload.get("rows", []) if isinstance(payload.get("rows"), list) else []
    display_value = payload.get("displayValue", {}) if isinstance(payload.get("displayValue"), dict) else {}

    flattened_rows: list[dict[str, Any]] = []
    expanded_columns = list(columns)

    for row in rows:
        if not isinstance(row, list):
            continue
        flattened: dict[str, Any] = {}
        for index, column in enumerate(columns):
            value = row[index] if index < len(row) else None
            flattened[column] = value
            display_map = display_value.get(column)
            if isinstance(display_map, dict) and value is not None:
                display_key = str(value)
                display_text = display_map.get(display_key)
                if display_text is not None:
                    display_column = f"{column}_display"
                    flattened[display_column] = display_text
                    if display_column not in expanded_columns:
                        expanded_columns.append(display_column)
        flattened_rows.append(flattened)

    return expanded_columns, flattened_rows


def _resolve_pubmatic_report_fields(
    requested_values: list[str] | None,
    *,
    alias_map: dict[str, str],
    kind: str,
) -> tuple[list[str], list[str]]:
    resolved_values: list[str] = []
    warnings: list[str] = []
    for requested_value in requested_values or []:
        normalized = str(requested_value).strip()
        if not normalized:
            continue
        alias = alias_map.get(normalized.lower())
        if alias is not None:
            resolved_values.append(alias)
            continue
        # Allow direct API field names through unchanged.
        resolved_values.append(normalized)
        warnings.append(f"Unverified PubMatic {kind} '{requested_value}' was passed through as-is.")
    return _dedupe_preserving_order(resolved_values), warnings


def _resolve_pubmatic_report_preset(report_type: str | None) -> str | None:
    if report_type is None:
        return None
    normalized = report_type.strip().lower()
    if not normalized:
        return None
    if normalized in PUBMATIC_REPORT_PRESETS:
        return normalized
    return PUBMATIC_REPORT_TYPE_PRESET_ALIASES.get(normalized)


def _write_pubmatic_report_file(
    *,
    columns: list[str],
    rows: list[dict[str, Any]],
    filename_hint: str | None,
    output_format: str,
    output_dir: str | None = None,
) -> dict[str, Any]:
    # Operators can pass output_dir per-call (typically the per-conversation
    # workspace dir from the system prompt). The default
    # ~/Victoria/pubmatic_reports is fine for local dev but read-only on
    # production hosts running under systemd's ProtectSystem=strict — see
    # magnite_mcp.py for the rationale on dropping the *_DOWNLOAD_DIR
    # env-var override.
    download_dir = Path(output_dir) if output_dir else Path(DEFAULT_REPORT_DOWNLOAD_DIR)
    download_dir = download_dir.expanduser()
    download_dir.mkdir(parents=True, exist_ok=True)

    normalized_format = output_format.strip().lower()
    base_name = (
        filename_hint or f"pubmatic_report.{normalized_format}"
    ).strip() or f"pubmatic_report.{normalized_format}"
    if "." not in Path(base_name).name:
        base_name = f"{base_name}.{normalized_format}"
    filepath = download_dir / Path(base_name).name

    if normalized_format == "csv":
        with filepath.open("w", newline="", encoding="utf-8") as output_file:
            writer = csv.DictWriter(output_file, fieldnames=columns)
            writer.writeheader()
            for row in rows:
                writer.writerow({column: row.get(column) for column in columns})
        content_type = "text/csv"
    elif normalized_format == "json":
        with filepath.open("w", encoding="utf-8") as output_file:
            json.dump(rows, output_file, indent=2)
        content_type = "application/json"
    else:
        raise ValueError("PubMatic output_format must be 'csv' or 'json'.")

    file_bytes = filepath.read_bytes()
    return {
        "success": True,
        "path": str(filepath),
        "bytes": len(file_bytes),
        "sha256": hashlib.sha256(file_bytes).hexdigest(),
        "content_type": content_type,
    }


def _summarize_pubmatic_rows(rows: list[dict[str, Any]], *, sample_limit: int = 25) -> dict[str, Any]:
    return {
        "row_count": len(rows),
        "sample_rows": rows[:sample_limit],
        "rows_truncated": len(rows) > sample_limit,
    }


def _dedupe_preserving_order(values: list[Any]) -> list[Any]:
    deduped: list[Any] = []
    seen: set[Any] = set()
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        deduped.append(value)
    return deduped


def _normalize_domain_candidate(value: Any) -> str | None:
    if value is None:
        return None

    candidate = str(value).strip().lower()
    if not candidate:
        return None
    candidate = _SCHEME_RE.sub("", candidate)
    candidate = candidate.split("/", 1)[0].strip(".")
    if candidate.startswith("www."):
        candidate = candidate[4:]
    return candidate or None


def _extract_domains_from_csv(file_path: str, column_name: str | None = None) -> dict[str, Any]:
    with open(file_path, newline="", encoding="utf-8-sig") as csv_file:
        rows = list(csv.reader(csv_file))

    if not rows:
        raise ValueError("CSV file is empty.")

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized_headers = {_normalize_lookup_text(header): index for index, header in enumerate(header_row) if header}

    if column_name:
        requested_column_key = _normalize_lookup_text(column_name)
        if requested_column_key not in normalized_headers:
            raise ValueError(f"Column not found in CSV: {column_name}")
        domain_column_index = normalized_headers[requested_column_key]
    else:
        preferred_columns = ["sites", "domains", "domain", "url", "urls"]
        domain_column_index = next(
            (normalized_headers[column_key] for column_key in preferred_columns if column_key in normalized_headers),
            0,
        )

    extracted_domains: list[str] = []
    invalid_values: list[str] = []
    for row in rows[1:]:
        if domain_column_index >= len(row):
            continue
        cell = row[domain_column_index]
        if isinstance(cell, float) and cell.is_integer():
            cell = int(cell)
        normalized_domain = _normalize_domain_candidate(cell)
        if normalized_domain is None:
            continue
        if _is_acceptable_domain_or_bundle(normalized_domain):
            extracted_domains.append(normalized_domain)
        else:
            invalid_values.append(str(cell).strip())

    return {
        "sheet_name": None,
        "column_name": header_row[domain_column_index] if header_row else None,
        "row_count": max(len(rows) - 1, 0),
        "domains": sorted(dict.fromkeys(extracted_domains)),
        "invalid_values": invalid_values,
    }


def _extract_domains_from_xlsx(
    file_path: str,
    sheet_name: str | None = None,
    column_name: str | None = None,
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read Excel domain files.") from exc

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found in workbook: {sheet_name}")
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Workbook sheet is empty.")

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized_headers = {_normalize_lookup_text(header): index for index, header in enumerate(header_row) if header}

    if column_name:
        requested_column_key = _normalize_lookup_text(column_name)
        if requested_column_key not in normalized_headers:
            raise ValueError(f"Column not found in workbook: {column_name}")
        domain_column_index = normalized_headers[requested_column_key]
    else:
        preferred_columns = ["sites", "domains", "domain", "url", "urls"]
        domain_column_index = next(
            (normalized_headers[column_key] for column_key in preferred_columns if column_key in normalized_headers),
            0,
        )

    extracted_domains: list[str] = []
    invalid_values: list[str] = []
    for row in rows[1:]:
        if domain_column_index >= len(row):
            continue
        cell = row[domain_column_index]
        if isinstance(cell, float) and cell.is_integer():
            cell = int(cell)
        normalized_domain = _normalize_domain_candidate(cell)
        if normalized_domain is None:
            continue
        if _is_acceptable_domain_or_bundle(normalized_domain):
            extracted_domains.append(normalized_domain)
        else:
            invalid_values.append(str(cell).strip())

    return {
        "sheet_name": worksheet.title,
        "column_name": header_row[domain_column_index] if header_row else None,
        "row_count": max(len(rows) - 1, 0),
        "domains": sorted(dict.fromkeys(extracted_domains)),
        "invalid_values": invalid_values,
    }


def _extract_domains_from_file(
    file_path: str,
    sheet_name: str | None = None,
    column_name: str | None = None,
) -> dict[str, Any]:
    path = Path(file_path).expanduser()
    path = (Path.cwd() / path).resolve() if not path.is_absolute() else path.resolve()

    if not path.exists() or not path.is_file():
        raise ValueError(f"Domain file does not exist or is not a file: {path}")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _extract_domains_from_csv(str(path), column_name=column_name)
    if suffix in {".xlsx", ".xlsm"}:
        return _extract_domains_from_xlsx(str(path), sheet_name=sheet_name, column_name=column_name)

    raise ValueError(f"Unsupported domain file format: {path.suffix or '<none>'}")


def _candidate_lookup_values(item: dict[str, Any], fields: tuple[str, ...]) -> list[str]:
    values: list[str] = []
    for field in fields:
        raw_value = item.get(field)
        if raw_value is None:
            continue
        text = str(raw_value).strip()
        if text:
            values.append(text)
    return values


def _resolve_unique_match(
    items: list[dict[str, Any]],
    requested_value: str,
    field_label: str,
    *,
    lookup_fields: tuple[str, ...],
    allow_contains_match: bool = True,
) -> dict[str, Any]:
    requested_text = str(requested_value).strip()
    normalized_requested = _normalize_lookup_text(requested_text)

    exact_matches = [
        item
        for item in items
        if any(
            _normalize_lookup_text(candidate) == normalized_requested
            for candidate in _candidate_lookup_values(item, lookup_fields)
        )
    ]
    if len(exact_matches) == 1:
        return exact_matches[0]
    if len(exact_matches) > 1:
        raise ValueError(f"Ambiguous {field_label}: {requested_value}")

    if allow_contains_match:
        contains_matches = [
            item
            for item in items
            if any(
                normalized_requested in _normalize_lookup_text(candidate)
                for candidate in _candidate_lookup_values(item, lookup_fields)
            )
        ]
        if len(contains_matches) == 1:
            return contains_matches[0]
        if len(contains_matches) > 1:
            raise ValueError(f"Ambiguous {field_label}: {requested_value}")

    raise LookupError(f"{field_label.capitalize()} not found: {requested_value}")


def _extract_tool_error_message(result: dict[str, Any], fallback: str) -> str:
    error = result.get("error")
    if isinstance(error, str) and error.strip():
        return error
    details = result.get("details")
    if isinstance(details, list) and details:
        normalized = _normalize_error(details)
        if normalized.get("message"):
            return str(normalized["message"])
    return fallback


def _coerce_int(value: Any, field_name: str) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{field_name} must be an integer")
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    raise ValueError(f"{field_name} must be an integer")


def _build_pubmatic_deal_url(curated_id: Any, deal_name: str | None = None) -> str | None:
    """Build a PubMatic deal-detail URL.

    The current PubMatic UI requires `dealId`, `dealName`, and
    `dealCategoryId` as query params to navigate directly to the deal.
    The deal_name is URL-encoded (PubMatic deal names contain underscores
    and sometimes other punctuation that must be safe in a query string).
    """
    if curated_id is None:
        return None
    encoded_name = urllib.parse.quote(str(deal_name or ""), safe="")
    return PUBMATIC_DEAL_URL_TEMPLATE.format(id=curated_id, name=encoded_name)


def _normalize_integer_list(values: list[Any] | None, field_name: str) -> list[int]:
    normalized: list[int] = []
    for value in values or []:
        normalized.append(_coerce_int(value, field_name))
    return _dedupe_preserving_order(normalized)


def _validate_allowed_ids(values: list[int], allowed_values: set[int], field_name: str) -> list[int]:
    invalid_values = [value for value in values if value not in allowed_values]
    if invalid_values:
        raise ValueError(
            f"{field_name} contains unsupported values: {invalid_values}. Allowed values: {sorted(allowed_values)}"
        )
    return values


def _select_id_from_item(item: dict[str, Any], keys: tuple[str, ...]) -> int | None:
    for key in keys:
        value = item.get(key)
        if isinstance(value, int):
            return value
        if isinstance(value, str) and value.isdigit():
            return int(value)
    return None


def _collect_items(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if not isinstance(payload, dict):
        return []

    items: list[dict[str, Any]] = []
    for key in ("items", "results", "data", "publishers", "segments"):
        candidate = payload.get(key)
        if isinstance(candidate, list):
            items.extend(item for item in candidate if isinstance(item, dict))
    if not items:
        items.extend(value for value in payload.values() if isinstance(value, dict))
    return items


def _find_targeting_field_name(available_fields: set[str], preferred_names: tuple[str, ...]) -> str | None:
    alias_map = {_normalize_lookup_text(name): name for name in available_fields}
    for preferred_name in preferred_names:
        normalized = _normalize_lookup_text(preferred_name)
        alias = PUBMATIC_TARGETING_FIELD_ALIASES.get(preferred_name, preferred_name)
        if normalized in alias_map:
            return alias_map[normalized]
        alias_normalized = _normalize_lookup_text(alias)
        if alias_normalized in alias_map:
            return alias_map[alias_normalized]
    return None


def _set_first_matching_field(
    payload: dict[str, Any], available_fields: set[str], names: tuple[str, ...], value: Any
) -> bool:
    field_name = _find_targeting_field_name(available_fields, names)
    if field_name is None:
        return False
    payload[field_name] = value
    return True


def _set_nested_targeting_value(targeting_payload: dict[str, Any], path: tuple[str, ...], value: Any) -> None:
    current = targeting_payload
    for key in path[:-1]:
        nested = current.get(key)
        if not isinstance(nested, dict):
            nested = {}
            current[key] = nested
        current = nested
    current[path[-1]] = value


class PubMaticClient:
    def __init__(self):
        self.base_url = os.environ.get("PUBMATIC_BASE_URL", DEFAULT_BASE_URL).rstrip("/")
        self.api_key = os.environ.get("PUBMATIC_API_KEY", "")
        self.username = os.environ.get("PUBMATIC_USERNAME", "")
        self.password = os.environ.get("PUBMATIC_PASSWORD", "")
        self.api_product = os.environ.get("PUBMATIC_API_PRODUCT", "PUBLISHER")
        # Optional account identifiers. Per PubMatic's docs, passing BOTH
        # accountId and accountType suppresses the token-notification email on
        # token generation AND refresh (for accounts mapped to that feature).
        self._account_id = os.environ.get("PUBMATIC_ACCOUNT_ID", "")
        self._account_type = os.environ.get("PUBMATIC_ACCOUNT_TYPE", "")
        self._access_token = os.environ.get("PUBMATIC_ACCESS_TOKEN", "")
        self._refresh_token = os.environ.get("PUBMATIC_REFRESH_TOKEN", "")
        self._user_email = ""
        self._http_client: httpx.AsyncClient | None = None
        # The token cache lineage is keyed to the env access token, so dropping a
        # new PUBMATIC_ACCESS_TOKEN into the environment invalidates a stale cache.
        self._token_seed_hash = _seed_hash(self._access_token)
        cached = _load_token_cache(self._token_seed_hash)
        if cached:
            # Reuse a previously renewed token instead of logging in (and emailing
            # a fresh token) on every run.
            self._access_token = cached["access_token"]
            self._refresh_token = cached.get("refresh_token") or self._refresh_token

    def _is_configured(self) -> bool:
        return bool(self._access_token) or (bool(self.username) and bool(self.password))

    def _account_email_suppression_fields(self) -> dict[str, str]:
        """accountId/accountType, which suppress PubMatic's token email when both set."""
        if self._account_id and self._account_type:
            return {"accountId": self._account_id, "accountType": self._account_type}
        return {}

    async def _get_http_client(self) -> httpx.AsyncClient:
        if self._http_client is None:
            self._http_client = httpx.AsyncClient(timeout=DEFAULT_TIMEOUT)
        return self._http_client

    def _base_headers(self) -> dict[str, str]:
        headers = {
            "accept": "application/json",
            "content-type": "application/json",
            "user-agent": USER_AGENT,
        }
        if self.api_key:
            headers["x-api-key"] = self.api_key
        return headers

    def _auth_headers(self) -> dict[str, str]:
        if not self._access_token:
            raise ValueError("PubMatic access token not available")
        headers = self._base_headers()
        headers["authorization"] = f"Bearer {self._access_token}"
        return headers

    async def authenticate(self, *, force: bool = False) -> dict[str, Any]:
        """Acquire a PubMatic access token.

        By default returns the env-provided token if one is set. Pass
        ``force=True`` to discard the current token and exchange the
        configured username/password for a fresh one — used by the
        request layer when the server reports the token as expired.
        Requires PUBMATIC_USERNAME and PUBMATIC_PASSWORD when forcing
        (or when no env token is set in the first place).
        """
        if self._access_token and not force:
            return {
                "userEmail": self._user_email,
                "tokenType": "Bearer",
                "hasAccessToken": True,
                "hasRefreshToken": bool(self._refresh_token),
            }
        if not self.username or not self.password:
            raise ValueError("PUBMATIC_ACCESS_TOKEN or PUBMATIC_USERNAME/PUBMATIC_PASSWORD are required")

        # Force-refresh path: discard the stale env token so a failed
        # exchange surfaces a clean error rather than silently retrying
        # against the bad token on the next request.
        if force:
            self._access_token = ""

        payload: dict[str, Any] = {"apiProduct": self.api_product}
        if self.username:
            payload["userName"] = self.username
        if self.password:
            payload["password"] = self.password
        payload.update(self._account_email_suppression_fields())

        client = await self._get_http_client()
        response = await client.post(
            f"{self.base_url}/v1/developer-integrations/developer/token",
            headers=self._base_headers(),
            json=payload,
        )
        response.raise_for_status()
        data = response.json()
        self._access_token = data.get("accessToken", "")
        self._refresh_token = data.get("refreshToken", "") or self._refresh_token
        self._user_email = data.get("userEmail", "")
        _save_token_cache(self._token_seed_hash, self._access_token, self._refresh_token)

        return {
            "userEmail": self._user_email,
            "tokenType": data.get("tokenType"),
            "hasAccessToken": bool(self._access_token),
            "hasRefreshToken": bool(self._refresh_token),
        }

    async def refresh_access_token(self) -> dict[str, Any]:
        """Renew the access token using the refresh token — no re-login.

        PubMatic's documented renewal path (PUT
        /v1/developer-integrations/developer/refreshToken). Unlike
        ``authenticate(force=True)`` (username/password), it does not resend
        credentials, and — for accounts mapped to "Disable plain text access
        token", or when accountId/accountType are supplied — does not send the
        token-notification email. Returns a new access AND refresh token, which
        we persist (the old refresh token is rotated out). Raises if no refresh
        token / user email is configured or the request fails, so the caller can
        fall back to the username/password login.

        Docs: https://help.pubmatic.com/publisher/reference/get-started-with-pubmatic-apis
        """
        if not self._refresh_token:
            raise ValueError("No PubMatic refresh token available to renew the access token.")
        email = self.username or self._user_email
        if not email:
            raise ValueError("PubMatic token refresh requires the API user email (set PUBMATIC_USERNAME).")

        client = await self._get_http_client()
        headers = self._base_headers()
        # The refresh call authenticates with the *previous* access token.
        if self._access_token:
            headers["authorization"] = f"Bearer {self._access_token}"

        payload: dict[str, Any] = {
            "email": email,
            "apiProduct": self.api_product,
            "refreshToken": self._refresh_token,
        }
        payload.update(self._account_email_suppression_fields())

        response = await client.put(
            f"{self.base_url}{PUBMATIC_REFRESH_PATH}",
            headers=headers,
            json=payload,
        )
        response.raise_for_status()
        data = response.json()
        access = data.get("accessToken", "")
        if not access:
            raise ValueError("PubMatic token refresh returned no accessToken.")
        self._access_token = access
        self._refresh_token = data.get("refreshToken", "") or self._refresh_token
        self._user_email = data.get("userEmail", "") or self._user_email
        _save_token_cache(self._token_seed_hash, self._access_token, self._refresh_token)
        return {
            "tokenType": data.get("tokenType", "Bearer"),
            "hasAccessToken": True,
            "hasRefreshToken": bool(self._refresh_token),
        }

    async def _ensure_token(self) -> None:
        if not self._access_token:
            await self.authenticate()

    def _is_token_expired_response(self, response: httpx.Response) -> bool:
        """Detect PubMatic's token-expired signal across response shapes.

        PubMatic surfaces an expired token in two ways depending on the
        endpoint family:
          - HTTP 401 with no body (REST gateways)
          - HTTP 200/4xx where the JSON body contains an `errorCode`
            of `access_token_expired` (curated-deal + reporting APIs)
        Both forms are observed in production; both must trigger the
        username/password re-auth path.
        """
        if response.status_code == 401:
            return True
        try:
            body = response.json()
        except (ValueError, AttributeError):
            return False

        # The body shape varies: sometimes a top-level dict, sometimes
        # under `errors` or `details`. We accept any reasonable nesting.
        def _walk(node: Any) -> bool:
            if isinstance(node, dict):
                code = node.get("errorCode") or node.get("code")
                if isinstance(code, str) and code.lower() == "access_token_expired":
                    return True
                msg = node.get("errorMessage") or node.get("message") or ""
                if isinstance(msg, str) and "access token expired" in msg.lower():
                    return True
                return any(_walk(v) for v in node.values())
            if isinstance(node, list):
                return any(_walk(item) for item in node)
            return False

        return _walk(body)

    async def _request(
        self,
        method: str,
        path: str,
        json_data: dict[str, Any] | None = None,
        params: Any | None = None,
    ) -> dict[str, Any]:
        await self._ensure_token()
        client = await self._get_http_client()

        async def _send() -> httpx.Response:
            return await client.request(
                method=method,
                url=f"{self.base_url}{path}",
                headers=self._auth_headers(),
                json=json_data,
                params=params,
            )

        response = await _send()
        # Auto-recovery on an expired token, retried ONCE. Preferred path is the
        # refresh token (quiet — no notification email, no credential resend);
        # username/password login is the fallback when no refresh token is set or
        # the refresh token has itself expired. If neither path is available we
        # surface the original 401 so the operator knows to regenerate manually.
        if self._is_token_expired_response(response):
            recovered = False
            if self._refresh_token:
                try:
                    logger.info("PubMatic access token expired; renewing via refresh token")
                    await self.refresh_access_token()
                    recovered = True
                except Exception as exc:  # noqa: BLE001 - fall back to credential login
                    logger.info("PubMatic refresh-token renewal failed (%s); falling back to username/password", exc)
            if not recovered and self.username and self.password:
                logger.info("PubMatic re-authenticating via username/password")
                await self.authenticate(force=True)
                recovered = True
            if recovered:
                response = await _send()

        response.raise_for_status()
        return response.json() if response.content else {}

    async def create_targeting(self, payload: dict[str, Any]) -> dict[str, Any]:
        return await self._request("POST", "/v1/inventory/targeting", json_data=payload)

    async def get_targeting(self, targeting_id: int) -> dict[str, Any]:
        return await self._request("GET", f"/v1/inventory/targeting/{targeting_id}")

    async def create_curated_deal(self, payload: dict[str, Any]) -> dict[str, Any]:
        return await self._request("POST", "/curateddeals/create", json_data=payload)

    async def update_curated_deal(self, curated_id: int, payload: dict[str, Any]) -> dict[str, Any]:
        """PUT /curateddeals/{id} — FULL-REPLACEMENT update (create-shaped body)."""
        return await self._request("PUT", f"/curateddeals/{curated_id}", json_data=payload)

    async def update_curated_deal_status(self, curated_deal_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        """PUT /curateddeals/updateStatus/{curatedDealId} — pause/resume the parent deal."""
        return await self._request("PUT", f"/curateddeals/updateStatus/{curated_deal_id}", json_data=payload)

    async def get_curated_deal(
        self,
        curated_id: int,
        logged_in_owner_id: int,
        logged_in_owner_type_id: int,
        view: str = "SUMMARY",
    ) -> dict[str, Any]:
        params = {
            "loggedInOwnerId": logged_in_owner_id,
            "loggedInOwnerTypeId": logged_in_owner_type_id,
            "view": view,
        }
        return await self._request("GET", f"/curateddeals/{curated_id}", params=params)

    async def list_pmp_dsps(
        self,
        atd_id: int = -1,
        page_number: int = 1,
        page_size: int = 100,
        sort: str = "name",
    ) -> dict[str, Any]:
        """List PMP-enabled DSPs.

        Defaults to atd_id=-1 (Data Provider mode), which is what Elcano
        is registered as in PubMatic. Passing atd_id=60067 (the publisher
        owner ID) returns 0 DSPs for our token's scope. Owner-id filters
        are not applied because they reduce results without adding
        relevant scoping for a curator workflow.
        """
        params: list[tuple[str, Any]] = [
            ("atdId", atd_id),
            ("pmpEnabled", 1),
            ("pageNumber", page_number),
            ("pageSize", page_size),
            ("sort", sort),
        ]
        return await self._request("GET", "/v1/common/advertisingEntity", params=params)

    async def list_dsp_buyer_map(
        self,
        dsp_id: int,
        query: str | None = None,
        buyer_page_number: int = 1,
        buyer_page_size: int = 1000,
    ) -> dict[str, Any]:
        params: list[tuple[str, Any]] = [
            ("dspId", dsp_id),
            ("buyerPageNumber", buyer_page_number),
            ("buyerPageSize", buyer_page_size),
        ]
        if query:
            params.append(("query", query))
        return await self._request("GET", "/v1/common/advertisingEntity/dspBuyerMap", params=params)

    async def list_publishers(
        self,
        page_number: int = 1,
        page_size: int = 200,
    ) -> dict[str, Any]:
        params: list[tuple[str, Any]] = [
            ("pageNumber", page_number),
            ("pageSize", page_size),
        ]
        return await self._request("GET", "/v1/common/publisher", params=params)

    async def list_buyer_audiences(
        self,
        *,
        search_key: str,
        account_id: int = ELCANO_OWNER_ID,
        account_type: str = "BUYER",
        audience_type: str = "ALL_AUDIENCE",
        segment_status: str = "ENABLED",
        page_number: int = 1,
        page_size: int = 100,
    ) -> dict[str, Any]:
        """Search buyer-targetable audience segments by name substring.

        Hits POST /v1/audience/buyerInsights/audiences with `searchKey` in
        the body — the documented Audience API for Buyers endpoint that
        replaces the permission-gated /v1/audience/segments. Returns
        items with `audienceId`, `audienceName`, `providerAudienceId`.
        """
        body = {
            "accountId": account_id,
            "accountType": account_type,
            "audienceType": audience_type,
            "segmentStatus": segment_status,
            "searchKey": search_key,
        }
        params = {"pageNumber": page_number, "pageSize": page_size}
        return await self._request("POST", "/v1/audience/buyerInsights/audiences", json_data=body, params=params)

    async def list_geos(
        self,
        *,
        name_like: str | None = None,
        country_code: str | None = None,
        geo_level: int | None = None,
        page_number: int = 1,
        page_size: int = 50,
        sort: str = "geoLevel,name",
        hide_invalid: bool = True,
    ) -> dict[str, Any]:
        """Search PubMatic geo entities using the documented filter syntax.

        Pass narrow filters (`name like *<value>*`, `countryCode eq US`,
        `geoLevel eq 1|2|3`) — the unfiltered endpoint times out / returns
        502 because the result set is too large.

        geo_level: 1=Country, 2=Region/State, 3=City.
        """
        params: list[tuple[str, Any]] = [
            ("hideInvalid", str(hide_invalid).lower()),
            ("pageNumber", page_number),
            ("pageSize", page_size),
            ("sort", sort),
        ]
        if name_like:
            params.append(("filters", f"name like *{name_like}*"))
        if country_code:
            params.append(("filters", f"countryCode eq {country_code}"))
        if geo_level is not None:
            params.append(("filters", f"geoLevel eq {geo_level}"))
        return await self._request("GET", "/v1/common/geo", params=params)

    async def list_iab_categories(self, page_number: int = 1, page_size: int = 100) -> dict[str, Any]:
        """Fetch PubMatic's IAB content taxonomy with PubMatic-internal numeric IDs.

        Returns 26 top-level IAB categories, each with a nested
        subCategoryList. The `id` field on each entry is what the
        targeting payload's iabCategories[] expects.
        """
        return await self._request(
            "GET",
            "/v1/common/iabCategories",
            params=[("pageNumber", page_number), ("pageSize", page_size)],
        )

    async def query_standard_analytics(
        self,
        account_id: int,
        *,
        from_date: str,
        to_date: str,
        metrics: list[str],
        dimensions: list[str] | None = None,
        filters: str | None = None,
        sort: str | None = None,
        date_unit: str | None = None,
        data_type: str | None = None,
        page_size: int | None = None,
        fluid_timezone: str | None = None,
    ) -> dict[str, Any]:
        params: dict[str, Any] = {
            "fromDate": from_date,
            "toDate": to_date,
            "metrics": ",".join(metrics),
        }
        if dimensions:
            params["dimensions"] = ",".join(dimensions)
        if filters:
            params["filters"] = filters
        if sort:
            params["sort"] = sort
        if date_unit:
            params["dateUnit"] = date_unit
        if data_type:
            params["dataType"] = data_type
        if page_size is not None:
            params["pageSize"] = page_size
        if fluid_timezone:
            params["isFluidTimezoneSupported"] = "true"
            params["fluidTimezone"] = fluid_timezone
        return await self._request("GET", f"/v1/analytics/data/dataprovider/{account_id}", params=params)


_pubmatic_client: PubMaticClient | None = None
_prepared_pubmatic_deals: dict[str, dict[str, Any]] = {}
_iab_taxonomy_cache: list[dict[str, Any]] | None = None
_dsp_list_cache: list[dict[str, Any]] | None = None


# ──────────────────────────────────────────────────────────────────────────────
# Disk cache for stable PubMatic lookup results (DSPs, IAB taxonomy, geo).
#
# Most catalog lookups are stable for hours, but every cutlass run re-fetches
# them from scratch — costing ~1.5s of wall time and a few hundred tokens
# (the agent reads the response). Caching to disk with a 4h TTL eliminates
# the second-and-onwards run cost.
#
# Disable with PUBMATIC_CACHE_TTL_SECONDS=0.
# ──────────────────────────────────────────────────────────────────────────────


def _cache_dir() -> Path:
    base = os.environ.get("XDG_CACHE_HOME") or os.path.expanduser("~/.cache")
    return Path(base) / "cutlass" / "pubmatic"


def _cache_ttl_seconds() -> int:
    raw = os.environ.get("PUBMATIC_CACHE_TTL_SECONDS", "14400")
    try:
        return max(0, int(raw))
    except ValueError:
        return 14400


def _cache_get(key: str) -> Any | None:
    """Return the cached value for `key` if present and fresh; else None."""
    ttl = _cache_ttl_seconds()
    if ttl <= 0:
        return None
    path = _cache_dir() / f"{key}.json"
    if not path.is_file():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    if not isinstance(payload, dict) or "stored_at" not in payload or "value" not in payload:
        return None
    stored_at = payload.get("stored_at")
    if not isinstance(stored_at, (int, float)):
        return None
    if (datetime.now(UTC).timestamp() - stored_at) > ttl:
        return None
    return payload.get("value")


def _cache_put(key: str, value: Any) -> None:
    """Best-effort write of `value` to the cache; ignore IO errors."""
    if _cache_ttl_seconds() <= 0:
        return
    try:
        cache_dir = _cache_dir()
        cache_dir.mkdir(parents=True, exist_ok=True)
        path = cache_dir / f"{key}.json"
        tmp = path.with_suffix(".tmp")
        tmp.write_text(
            json.dumps({"stored_at": datetime.now(UTC).timestamp(), "value": value}),
            encoding="utf-8",
        )
        tmp.replace(path)
    except OSError:
        pass


# --- Developer access-token persistence & renewal --------------------------
# The PubMatic developer access token is valid ~60 days. Persisting it (and its
# refresh token) across runs lets every cutlass invocation reuse the same valid
# token instead of re-minting one each run — a username/password login both
# generates a new token AND triggers a notification email, so re-minting per run
# is noisy and wasteful. Renewal at expiry uses the refresh token (see
# PubMaticClient.refresh_access_token), which is quiet.
#
# PUBMATIC_REFRESH_PATH centralizes the refresh endpoint (PUT). Confirmed
# against PubMatic's API docs; overridable via env if the path ever changes.
PUBMATIC_REFRESH_PATH = os.environ.get("PUBMATIC_REFRESH_PATH", "/v1/developer-integrations/developer/refreshToken")
_TOKEN_CACHE_KEY = "auth_token"
_DEFAULT_TOKEN_CACHE_TTL = 55 * 24 * 60 * 60  # 55d — renew before the ~60d hard expiry


def _token_cache_ttl_seconds() -> int:
    raw = os.environ.get("PUBMATIC_TOKEN_CACHE_TTL_SECONDS", str(_DEFAULT_TOKEN_CACHE_TTL))
    try:
        return max(0, int(raw))
    except ValueError:
        return _DEFAULT_TOKEN_CACHE_TTL


def _seed_hash(token: str) -> str:
    """Stable fingerprint of the env access token used to key the cache lineage."""
    return hashlib.sha256(token.encode("utf-8")).hexdigest() if token else ""


def _load_token_cache(seed_hash: str) -> dict[str, str] | None:
    """Return persisted {access_token, refresh_token} for this env-token lineage.

    Returns None when caching is disabled, the file is missing/expired, or the
    cached lineage was seeded from a different PUBMATIC_ACCESS_TOKEN (i.e. the
    operator rotated the env token, so the cache is stale).
    """
    ttl = _token_cache_ttl_seconds()
    if ttl <= 0 or not seed_hash:
        return None
    path = _cache_dir() / f"{_TOKEN_CACHE_KEY}.json"
    if not path.is_file():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    if not isinstance(payload, dict) or payload.get("seed_hash") != seed_hash:
        return None
    stored_at = payload.get("stored_at")
    if not isinstance(stored_at, (int, float)) or (datetime.now(UTC).timestamp() - stored_at) > ttl:
        return None
    access = payload.get("access_token")
    if not isinstance(access, str) or not access:
        return None
    refresh = payload.get("refresh_token")
    return {"access_token": access, "refresh_token": refresh if isinstance(refresh, str) else ""}


def _save_token_cache(seed_hash: str, access_token: str, refresh_token: str) -> None:
    """Persist the current token lineage with owner-only (0600) permissions."""
    if _token_cache_ttl_seconds() <= 0 or not access_token or not seed_hash:
        return
    try:
        cache_dir = _cache_dir()
        cache_dir.mkdir(parents=True, exist_ok=True)
        path = cache_dir / f"{_TOKEN_CACHE_KEY}.json"
        tmp = path.with_suffix(".tmp")
        tmp.write_text(
            json.dumps(
                {
                    "stored_at": datetime.now(UTC).timestamp(),
                    "seed_hash": seed_hash,
                    "access_token": access_token,
                    "refresh_token": refresh_token,
                }
            ),
            encoding="utf-8",
        )
        os.chmod(tmp, 0o600)
        tmp.replace(path)
    except OSError:
        pass


def get_pubmatic_client() -> PubMaticClient:
    global _pubmatic_client
    if _pubmatic_client is None:
        _pubmatic_client = PubMaticClient()
    return _pubmatic_client


def _make_blocker(code: str, message: str, **details: Any) -> dict[str, Any]:
    blocker: dict[str, Any] = {"code": code, "message": message}
    if details:
        blocker["details"] = details
    return blocker


class PubMaticResolutionError(Exception):
    """Raised when an MCP-side name -> ID resolution fails.

    Carries structured `details` so callers can surface candidate matches
    in the prepared-deal artifact's blocker, giving the agent an escape
    hatch when its name guess doesn't match what PubMatic has registered.
    """

    def __init__(self, message: str, **details: Any) -> None:
        super().__init__(message)
        self.message = message
        self.details: dict[str, Any] = details


@mcp.tool()
async def pm_auth_status() -> dict[str, Any]:
    client = get_pubmatic_client()
    if not client._is_configured():
        return {
            "configured": False,
            "authenticated": False,
            "error": "PUBMATIC_ACCESS_TOKEN or PUBMATIC_USERNAME/PUBMATIC_PASSWORD are required",
        }
    try:
        auth = await client.authenticate()
        return {
            "configured": True,
            "authenticated": True,
            "user_email": auth.get("userEmail"),
            "token_type": auth.get("tokenType"),
            "has_refresh_token": auth.get("hasRefreshToken", False),
        }
    except httpx.HTTPStatusError as e:
        try:
            body = e.response.json()
        except Exception:
            body = e.response.text
        normalized = _normalize_error(body)
        return {
            "configured": True,
            "authenticated": False,
            "error": normalized.get("message"),
            "details": normalized.get("errors", []),
        }
    except Exception as e:
        return {"configured": True, "authenticated": False, "error": str(e)}


@mcp.tool()
async def pm_list_reporting_presets() -> dict[str, Any]:
    presets: dict[str, Any] = {}
    for preset_name, preset in PUBMATIC_REPORT_PRESETS.items():
        presets[preset_name] = {
            "description": preset["description"],
            "dimensions": preset["dimensions"],
            "metrics": preset["metrics"],
        }
    return {
        "success": True,
        "presets": presets,
        "known_accounts": {"Elcano": 60067},
    }


@mcp.tool()
async def pm_reporting_healthcheck(
    account_id: int | str = "Elcano",
    date_range: dict[str, Any] | None = None,
    metrics: list[str] | None = None,
) -> dict[str, Any]:
    try:
        resolved_account_id = _resolve_pubmatic_report_account_id(account_id)
        resolved_range = date_range or {"previous": {"days": 7}}
        from_date, to_date = _resolve_pubmatic_date_range(resolved_range)
        selected_metrics = metrics or ["spend", "paidImpressions"]
        client = get_pubmatic_client()
        data = await client.query_standard_analytics(
            resolved_account_id,
            from_date=from_date,
            to_date=to_date,
            metrics=selected_metrics,
            dimensions=["date"],
            date_unit="date",
            page_size=1,
        )
        columns, rows = _flatten_pubmatic_analytics_rows(data)
        return {
            "success": True,
            "account_id": resolved_account_id,
            "from_date": from_date,
            "to_date": to_date,
            "columns": columns,
            "row_count": len(rows),
            "sample_rows": rows[:1],
            "alert": data.get("alert") if isinstance(data, dict) else None,
            "data_freshness": data.get("dataFreshness") if isinstance(data, dict) else None,
        }
    except httpx.HTTPStatusError as e:
        try:
            body = e.response.json()
        except Exception:
            body = e.response.text
        return _err("Failed PubMatic dataprovider reporting healthcheck", body)
    except Exception as e:
        return _err("Failed PubMatic dataprovider reporting healthcheck", e)


@mcp.tool()
async def pm_run_standard_report(
    account_id: int | str,
    date_range: dict[str, Any],
    metrics: list[str],
    dimensions: list[str] | None = None,
    filters: str | None = None,
    sort: str | None = None,
    date_unit: str = "date",
    data_type: str | None = None,
    page_size: int | None = None,
    fluid_timezone: str | None = None,
    download: bool = True,
    filename_hint: str | None = None,
    output_format: str = "csv",
    include_rows: bool = False,
    output_dir: str | None = None,
) -> dict[str, Any]:
    """Run a PubMatic dataprovider standard analytics report.

    output_dir: Optional absolute path the report file is written into.
        Pass the per-conversation workspace dir so the agent's other
        tools can read it. Defaults to ~/Victoria/pubmatic_reports for
        local use; on a hardened production host that path is read-only.
    """
    try:
        resolved_account_id = _resolve_pubmatic_report_account_id(account_id)
        from_date, to_date = _resolve_pubmatic_date_range(date_range)
        selected_dimensions = _apply_pubmatic_date_dimension(list(dimensions or []), date_unit)
        client = get_pubmatic_client()
        data = await client.query_standard_analytics(
            resolved_account_id,
            from_date=from_date,
            to_date=to_date,
            metrics=metrics,
            dimensions=selected_dimensions,
            filters=filters,
            sort=sort,
            date_unit=date_unit,
            data_type=data_type,
            page_size=page_size,
            fluid_timezone=fluid_timezone,
        )
        columns, rows = _flatten_pubmatic_analytics_rows(data)
        row_summary = _summarize_pubmatic_rows(rows)
        response = {
            "success": True,
            "account_id": resolved_account_id,
            "from_date": from_date,
            "to_date": to_date,
            "dimensions": selected_dimensions,
            "metrics": metrics,
            "columns": columns,
            "row_count": row_summary["row_count"],
            "sample_rows": row_summary["sample_rows"],
            "rows_truncated": row_summary["rows_truncated"],
            "alert": data.get("alert") if isinstance(data, dict) else None,
            "data_freshness": data.get("dataFreshness") if isinstance(data, dict) else None,
        }
        if include_rows:
            response["rows"] = rows
        if download:
            response["download"] = _write_pubmatic_report_file(
                columns=columns,
                rows=rows,
                filename_hint=filename_hint,
                output_format=output_format,
                output_dir=output_dir,
            )
        return response
    except httpx.HTTPStatusError as e:
        try:
            body = e.response.json()
        except Exception:
            body = e.response.text
        return _err("Failed to run PubMatic dataprovider standard report", body)
    except Exception as e:
        return _err("Failed to run PubMatic dataprovider standard report", e)


@mcp.tool()
async def pm_run_preset_report(
    account_id: int | str,
    date_range: dict[str, Any],
    preset: str = "deal_summary",
    extra_dimensions: list[str] | None = None,
    extra_metrics: list[str] | None = None,
    filters: str | None = None,
    sort: str | None = None,
    date_unit: str = "date",
    data_type: str | None = None,
    page_size: int | None = None,
    fluid_timezone: str | None = None,
    download: bool = True,
    filename_hint: str | None = None,
    output_format: str = "csv",
    output_dir: str | None = None,
) -> dict[str, Any]:
    """Run a preset PubMatic report. Pass output_dir = workspace path so
    the agent's other tools can read the result on hardened hosts."""
    preset_config = PUBMATIC_REPORT_PRESETS.get(preset)
    if preset_config is None:
        return _err(
            "Failed to run PubMatic preset report",
            {"message": f"Unknown preset: {preset}. Valid presets: {sorted(PUBMATIC_REPORT_PRESETS)}"},
        )

    dimensions = _dedupe_preserving_order(preset_config["dimensions"] + (extra_dimensions or []))
    metrics = _dedupe_preserving_order(preset_config["metrics"] + (extra_metrics or []))

    result = await pm_run_standard_report(
        account_id=account_id,
        date_range=date_range,
        metrics=metrics,
        dimensions=dimensions,
        filters=filters,
        sort=sort,
        date_unit=date_unit,
        data_type=data_type,
        page_size=page_size,
        fluid_timezone=fluid_timezone,
        download=download,
        filename_hint=filename_hint,
        output_format=output_format,
        include_rows=False,
        output_dir=output_dir,
    )
    if not result.get("success"):
        return result
    result["preset"] = preset
    return result


@mcp.tool()
async def pm_run_report_from_prompt_inputs(
    account_id: int | str,
    date_range: dict[str, Any],
    breakdowns: list[str] | None = None,
    metrics: list[str] | None = None,
    report_type: str | None = None,
    filters: str | None = None,
    sort: str | None = None,
    date_unit: str | None = None,
    fluid_timezone: str | None = None,
    page_size: int | None = None,
    download: bool = True,
    filename_hint: str | None = None,
    output_format: str = "csv",
    output_dir: str | None = None,
) -> dict[str, Any]:
    """Run a PubMatic dataprovider analytics report from human-readable prompt inputs.

    Use human terms like:
    - breakdowns: ["day", "deal", "DSP"]
    - metrics: ["impressions", "spend", "eCPM"]
    - report_type: "deal performance"

    output_dir: Optional absolute path the report file is written into.
        Pass the per-conversation workspace dir so other tools can read it.
    """

    warnings: list[str] = []

    selected_preset = _resolve_pubmatic_report_preset(report_type)
    if selected_preset and not breakdowns and not metrics:
        result = await pm_run_preset_report(
            account_id=account_id,
            date_range=date_range,
            preset=selected_preset,
            filters=filters,
            sort=sort,
            date_unit=date_unit or "date",
            fluid_timezone=fluid_timezone,
            page_size=page_size,
            download=download,
            filename_hint=filename_hint,
            output_format=output_format,
            output_dir=output_dir,
        )
        if result.get("success"):
            result["warnings"] = warnings
        return result

    resolved_breakdowns, breakdown_warnings = _resolve_pubmatic_report_fields(
        breakdowns,
        alias_map=PUBMATIC_REPORT_DIMENSION_ALIASES,
        kind="dimension",
    )
    resolved_metrics, metric_warnings = _resolve_pubmatic_report_fields(
        metrics,
        alias_map=PUBMATIC_REPORT_METRIC_ALIASES,
        kind="metric",
    )
    warnings.extend(breakdown_warnings)
    warnings.extend(metric_warnings)

    if not resolved_breakdowns and selected_preset:
        preset = PUBMATIC_REPORT_PRESETS[selected_preset]
        resolved_breakdowns = list(preset["dimensions"])
        resolved_metrics = resolved_metrics or list(preset["metrics"])

    if not resolved_metrics:
        if selected_preset:
            resolved_metrics = list(PUBMATIC_REPORT_PRESETS[selected_preset]["metrics"])
        else:
            resolved_metrics = ["paidImpressions", "spend", "ecpm"]

    if not resolved_breakdowns:
        resolved_breakdowns = ["date", "dealMetaId", "dspId"]

    inferred_date_unit = date_unit or (
        "month" if "month" in resolved_breakdowns else "hour" if "hour" in resolved_breakdowns else "date"
    )

    result = await pm_run_standard_report(
        account_id=account_id,
        date_range=date_range,
        metrics=resolved_metrics,
        dimensions=resolved_breakdowns,
        filters=filters,
        sort=sort,
        date_unit=inferred_date_unit,
        fluid_timezone=fluid_timezone,
        page_size=page_size,
        download=download,
        filename_hint=filename_hint,
        output_format=output_format,
        output_dir=output_dir,
    )
    if result.get("success"):
        result["resolved_breakdowns"] = resolved_breakdowns
        result["resolved_metrics"] = resolved_metrics
        result["warnings"] = _dedupe_preserving_order(warnings)
        if selected_preset:
            result["selected_preset"] = selected_preset
    return result


@mcp.tool()
async def pm_create_targeting(payload: dict[str, Any]) -> dict[str, Any]:
    try:
        owner_id = payload.get("ownerId")
        owner_type = payload.get("ownerType")
        if owner_id is not None and owner_type is not None:
            _validate_owner_context(int(owner_id), int(owner_type))

        client = get_pubmatic_client()
        data = await client.create_targeting(payload)
        return _ok({"result": data, "targeting_id": data.get("id")})
    except httpx.HTTPStatusError as e:
        try:
            body = e.response.json()
        except Exception:
            body = e.response.text
        return _err("Failed to create targeting", body)
    except Exception as e:
        return _err("Failed to create targeting", e)


@mcp.tool()
async def pm_get_targeting(targeting_id: int) -> dict[str, Any]:
    try:
        client = get_pubmatic_client()
        data = await client.get_targeting(targeting_id)
        return _ok({"result": data})
    except httpx.HTTPStatusError as e:
        try:
            body = e.response.json()
        except Exception:
            body = e.response.text
        return _err("Failed to get targeting", body)
    except Exception as e:
        return _err("Failed to get targeting", e)


def _validate_curated_payload(payload: dict[str, Any]) -> dict[str, Any]:
    owner_id_raw = payload.get("loggedInOwnerId")
    owner_type_raw = payload.get("loggedInOwnerTypeId")
    if owner_id_raw is None or owner_type_raw is None:
        raise ValueError("loggedInOwnerId and loggedInOwnerTypeId are required")
    owner_id = int(owner_id_raw)
    owner_type = int(owner_type_raw)
    _validate_owner_context(owner_id, owner_type)

    start_date = payload.get("startDate")
    end_date = payload.get("endDate")
    if not isinstance(start_date, str) or not isinstance(end_date, str):
        raise ValueError("startDate and endDate are required string fields")
    start_date = _normalize_zulu_datetime("startDate", start_date)
    end_date = _normalize_zulu_datetime("endDate", end_date)

    payload = dict(payload)
    payload["startDate"] = start_date
    payload["endDate"] = end_date
    payload["requestTypeEnum"] = REQUEST_TYPE_CREATE
    return payload


async def pm_create_curated_deal(payload: dict[str, Any]) -> dict[str, Any]:
    try:
        safe_payload = _validate_curated_payload(payload)
        client = get_pubmatic_client()
        data = await client.create_curated_deal(safe_payload)
        return _ok({"result": data, "id": data.get("id"), "dealId": data.get("dealId")})
    except httpx.HTTPStatusError as e:
        try:
            body = e.response.json()
        except Exception:
            body = e.response.text
        return _err("Failed to create curated deal", body)
    except Exception as e:
        return _err("Failed to create curated deal", e)


@mcp.tool()
async def pm_get_curated_deal(
    curated_id: int,
    logged_in_owner_id: int,
    logged_in_owner_type_id: int,
    view: str = "SUMMARY",
) -> dict[str, Any]:
    try:
        _validate_owner_context(logged_in_owner_id, logged_in_owner_type_id)
        client = get_pubmatic_client()
        data = await client.get_curated_deal(
            curated_id=curated_id,
            logged_in_owner_id=logged_in_owner_id,
            logged_in_owner_type_id=logged_in_owner_type_id,
            view=view,
        )
        owner_candidate = _extract_ownership_value(data)
        if owner_candidate is not None and owner_candidate != ELCANO_OWNER_ID:
            return {
                "success": False,
                "error": f"SECURITY FAILURE: ownership mismatch {owner_candidate} != {ELCANO_OWNER_ID}",
                "result": data,
            }
        return _ok({"result": data})
    except httpx.HTTPStatusError as e:
        try:
            body = e.response.json()
        except Exception:
            body = e.response.text
        return _err("Failed to get curated deal", body)
    except Exception as e:
        return _err("Failed to get curated deal", e)


def _curated_response_to_update_payload(deal: dict[str, Any]) -> dict[str, Any]:
    """Map a GET /curateddeals/{id} response bean to a PUT request payload.

    PubMatic's update endpoint is FULL-REPLACEMENT with a create-shaped body,
    so an update must round-trip the deal's current definition. The response
    bean uses object shapes the request bean does not (adFormats/platforms as
    {id,name} objects, labels as objects, status as {id,name}, ttdMainBuyer /
    transactionFeeRecipient as single objects) — this maps them back. Fields
    are copied via an explicit allowlist so response-only fields
    (creationTime, targetedDealPubs, samplingStatus, dealType, ...) can never
    leak into the PUT body. Pure Python, no network — unit-testable.
    """

    def _ids_from_objects(values: Any) -> list[int] | None:
        if not isinstance(values, list):
            return None
        ids: list[int] = []
        for item in values:
            if isinstance(item, dict) and isinstance(item.get("id"), int):
                ids.append(item["id"])
            elif isinstance(item, int):
                ids.append(item)
        return ids

    payload: dict[str, Any] = {}

    # Scalars and already-request-shaped fields copied verbatim when present.
    for field in (
        "name",
        "dealId",
        "description",
        "startDate",
        "endDate",
        "auctionType",
        "flooreCPM",
        "priority",
        "dealSource",
        "hasMaxReach",
        "maxAllowedPublishers",
        "autoUpdatePublishers",
        "customPubDealEnabled",
        "pubIds",
        "publisherBlockList",
        "marketplaceLowPriorityPubIds",
        "targeting",
        "tacticId",
        "podName",
        "videoOrientation",
        "alwaysOnFlag",
        "dealFlights",
        "dealFees",
        "transactionFeeCpm",
        "transactionFeeCpmPercent",
        "firstPartyAudienceCPM",
        "fpAudCPMPercentage",
        "takeRatePercentage",
        "activateAdvertiserId",
        "liveEventEnabled",
        "dv360DealFlag",
        "isTTDMainBuyerSelected",
        "bundled",
        "dealDspBuyerMappings",
        "marketplace",
        "timeZone",
    ):
        if deal.get(field) is not None:
            payload[field] = deal[field]

    ad_format_ids = _ids_from_objects(deal.get("adFormats"))
    if ad_format_ids:
        payload["adFormats"] = ad_format_ids
    platform_ids = _ids_from_objects(deal.get("platforms"))
    if platform_ids:
        payload["platforms"] = platform_ids
    label_ids = _ids_from_objects(deal.get("labels"))
    if label_ids:
        payload["labelIds"] = label_ids

    status = deal.get("status")
    if isinstance(status, dict) and isinstance(status.get("id"), int):
        payload["status"] = status["id"]
    elif isinstance(status, int):
        payload["status"] = status

    # Single objects in the response that the request bean wants as arrays.
    for field in ("ttdMainBuyer", "transactionFeeRecipient"):
        value = deal.get(field)
        if isinstance(value, dict):
            payload[field] = [value]
        elif isinstance(value, list) and value:
            payload[field] = value

    # bundledPMP: response carries objects; the request wants the PMP ids.
    bundled_pmp = deal.get("bundledPMP")
    if isinstance(bundled_pmp, list) and bundled_pmp:
        pmp_ids = [item.get("pmpId", item.get("id")) for item in bundled_pmp if isinstance(item, dict)]
        pmp_ids = [pmp_id for pmp_id in pmp_ids if isinstance(pmp_id, int)]
        if pmp_ids:
            payload["bundledPMP"] = pmp_ids

    return payload


@mcp.tool()
async def pm_update_curated_deal(
    curated_id: int,
    logged_in_owner_id: int,
    logged_in_owner_type_id: int,
    name: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    floor_ecpm: float | None = None,
    auction_type: int | None = None,
    status: int | None = None,
    priority: int | None = None,
    pub_ids: list[int] | None = None,
    publisher_block_list: list[int] | None = None,
    ad_formats: list[int] | None = None,
    platforms: list[int] | None = None,
    targeting: int | None = None,
    max_allowed_publishers: int | None = None,
    auto_update_publishers: bool | None = None,
    description: str | None = None,
    payload_overrides: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Update an existing PubMatic curated deal (PUT /curateddeals/{id}).

    PubMatic's update endpoint is FULL-REPLACEMENT (the body is the same
    schema as create, with the same required fields) — there is no partial
    PATCH. This tool therefore does read-modify-write for you: it fetches the
    deal, maps the response back into a request payload, overlays ONLY the
    arguments you pass, sets requestTypeEnum=UPDATE, submits, and verifies
    with a re-fetch. You still only pass what you want changed.

    Caveat inherited from the API: any field PubMatic's GET response does not
    return cannot be round-tripped — if the fetched deal carries no flooreCPM
    and you don't pass floor_ecpm, the PUT goes out without a floor and a
    warning is emitted. Pass floor_ecpm explicitly when the floor matters.

    For pause/resume prefer pm_update_curated_deal_status — it uses the
    dedicated lightweight updateStatus endpoint instead of replaying the
    whole deal definition.

    Args:
        curated_id: Numeric curated-deal id (the `id` from create/get — NOT
            the PM deal-id string).
        logged_in_owner_id: Must be the Elcano owner id (security-enforced).
        logged_in_owner_type_id: 5 (Demand Partner) or 7 (ATD/Buyer).
        name: New deal name.
        start_date / end_date: Zulu datetimes (YYYY-MM-DDTHH:MM:SS[.000]Z;
            bare dates are normalized).
        floor_ecpm: New floor (flooreCPM). Only honored by PubMatic when the
            deal's auction type is 3 (Fixed Price) — a floor on a First Price
            deal is dropped by PubMatic, and this tool warns when that applies.
        auction_type: 1 (First Price) or 3 (Fixed Price).
        status: 1 (Active) or 2 (Inactive) — but prefer the status tool.
        priority: Deal priority (lower = higher priority).
        pub_ids: Replacement publisher id list (the WHOLE list).
        publisher_block_list: Replacement blocked-publisher id list.
        ad_formats: Replacement ad-format ids — 3 (Banner), 12 (Video),
            13 (Native).
        platforms: Replacement platform ids — 1 Web, 2 Mobile Web,
            4 Mobile App iOS, 5 Mobile App Android, 7 CTV.
        targeting: New targeting id (from pm_create_targeting).
        max_allowed_publishers: For Max Reach deals (hasMaxReach=1).
        auto_update_publishers: For Max Reach deals.
        description: New description.
        payload_overrides: Raw request-bean fields merged LAST (escape hatch
            for fields without a dedicated argument). Cannot override the
            security-pinned loggedInOwnerId/requestTypeEnum.

    Returns:
        {"success": True, "result": <PUT response>, "updated_fields": [...],
         "warnings": [...], "verification": <re-fetched deal>} or
        {"success": False, "error": ...}.
    """
    try:
        _validate_owner_context(logged_in_owner_id, logged_in_owner_type_id)
        client = get_pubmatic_client()
        warnings: list[str] = []

        current = await client.get_curated_deal(
            curated_id=curated_id,
            logged_in_owner_id=logged_in_owner_id,
            logged_in_owner_type_id=logged_in_owner_type_id,
        )
        owner_candidate = _extract_ownership_value(current)
        if owner_candidate is not None and owner_candidate != ELCANO_OWNER_ID:
            return {
                "success": False,
                "error": f"SECURITY FAILURE: ownership mismatch {owner_candidate} != {ELCANO_OWNER_ID}",
            }

        payload = _curated_response_to_update_payload(current)

        changes: dict[str, Any] = {}
        if name is not None:
            changes["name"] = name
        if start_date is not None:
            changes["startDate"] = _normalize_zulu_datetime("start_date", start_date)
        if end_date is not None:
            changes["endDate"] = _normalize_zulu_datetime("end_date", end_date)
        if auction_type is not None:
            if auction_type not in (1, 3):
                raise ValueError("auction_type must be 1 (First Price) or 3 (Fixed Price)")
            changes["auctionType"] = auction_type
        if floor_ecpm is not None:
            changes["flooreCPM"] = float(floor_ecpm)
        if status is not None:
            if status not in PUBMATIC_SETTABLE_STATUS_IDS:
                raise ValueError(
                    "status must be 1 (Active) or 2 (Inactive); other codes are PubMatic system states. "
                    "Prefer pm_update_curated_deal_status for pause/resume."
                )
            changes["status"] = status
        if priority is not None:
            changes["priority"] = int(priority)
        if pub_ids is not None:
            changes["pubIds"] = _normalize_integer_list(pub_ids, "pub_ids")
        if publisher_block_list is not None:
            changes["publisherBlockList"] = _normalize_integer_list(publisher_block_list, "publisher_block_list")
        if ad_formats is not None:
            changes["adFormats"] = _validate_allowed_ids(
                _normalize_integer_list(ad_formats, "ad_formats"), PUBMATIC_ALLOWED_AD_FORMAT_IDS, "ad_formats"
            )
        if platforms is not None:
            changes["platforms"] = _validate_allowed_ids(
                _normalize_integer_list(platforms, "platforms"), PUBMATIC_ALLOWED_PLATFORM_IDS, "platforms"
            )
        if targeting is not None:
            changes["targeting"] = int(targeting)
        if max_allowed_publishers is not None:
            changes["maxAllowedPublishers"] = int(max_allowed_publishers)
        if auto_update_publishers is not None:
            changes["autoUpdatePublishers"] = bool(auto_update_publishers)
        if description is not None:
            changes["description"] = description
        if payload_overrides:
            changes.update(dict(payload_overrides))

        if not changes:
            return {"success": False, "error": "No update fields provided — pass at least one field to change."}

        payload.update(changes)

        # Security pins win over anything in payload_overrides.
        payload["id"] = curated_id
        payload["loggedInOwnerId"] = ELCANO_OWNER_ID
        payload["loggedInOwnerTypeId"] = logged_in_owner_type_id
        payload["requestTypeEnum"] = REQUEST_TYPE_UPDATE

        effective_auction_type = payload.get("auctionType")
        if payload.get("flooreCPM") is not None and effective_auction_type == 1:
            warnings.append(
                "flooreCPM is set but the deal's auction type is 1 (First Price) — PubMatic drops "
                "floors on First Price deals. Set auction_type=3 (Fixed Price) to honor the floor."
            )
        if payload.get("flooreCPM") is None and effective_auction_type == 3:
            warnings.append(
                "The fetched deal carried no flooreCPM, so this full-replacement update goes out "
                "without one on a Fixed Price deal — pass floor_ecpm explicitly if the floor matters."
            )

        result = await client.update_curated_deal(curated_id, payload)

        verification: dict[str, Any] | None = None
        try:
            verification = await client.get_curated_deal(
                curated_id=curated_id,
                logged_in_owner_id=logged_in_owner_id,
                logged_in_owner_type_id=logged_in_owner_type_id,
            )
        except Exception as verify_exc:  # noqa: BLE001 - verification is best-effort
            warnings.append(f"Post-update verification fetch failed: {verify_exc}")

        return _ok(
            {
                "result": result,
                "updated_fields": sorted(changes.keys()),
                "warnings": warnings,
                "verification": verification,
                "deal_url": _build_pubmatic_deal_url(curated_id, payload.get("name")),
            }
        )
    except httpx.HTTPStatusError as e:
        try:
            body = e.response.json()
        except Exception:
            body = e.response.text
        return _err("Failed to update curated deal", body)
    except Exception as e:
        return _err("Failed to update curated deal", e)


@mcp.tool()
async def pm_update_curated_deal_status(
    curated_deal_id: str,
    status: str | int,
    logged_in_owner_id: int,
    logged_in_owner_type_id: int,
) -> dict[str, Any]:
    """Pause or resume a PubMatic curated deal (PUT /curateddeals/updateStatus/{id}).

    This is the lightweight status path — it changes only the deal's status
    without replaying the full deal definition (use pm_update_curated_deal
    for everything else). Only Active/Inactive are settable; the other status
    codes (Scheduled, Completed, InProgress, Failed, PartiallyFailed) are
    PubMatic system states.

    Args:
        curated_deal_id: The curated deal identifier the updateStatus endpoint
            expects (the numeric curated-deal id as a string).
        status: "active"/"resume"/1 to activate, "paused"/"inactive"/"pause"/2
            to pause.
        logged_in_owner_id: Must be the Elcano owner id (security-enforced).
        logged_in_owner_type_id: 5 (Demand Partner) or 7 (ATD/Buyer).
    """
    try:
        _validate_owner_context(logged_in_owner_id, logged_in_owner_type_id)

        status_aliases: dict[str, int] = {
            "active": 1,
            "resume": 1,
            "1": 1,
            "paused": 2,
            "pause": 2,
            "inactive": 2,
            "2": 2,
        }
        resolved_status = status_aliases.get(str(status).strip().lower())
        if resolved_status is None:
            raise ValueError(
                "status must resolve to 1 (Active) or 2 (Inactive); other codes are PubMatic system states."
            )

        client = get_pubmatic_client()
        payload = {
            "status": resolved_status,
            "loggedInOwnerId": ELCANO_OWNER_ID,
            "loggedInOwnerTypeId": logged_in_owner_type_id,
        }
        result = await client.update_curated_deal_status(str(curated_deal_id), payload)
        return _ok(
            {
                "result": result,
                "status": resolved_status,
                "status_name": PUBMATIC_DEAL_STATUS_NAMES.get(resolved_status),
            }
        )
    except httpx.HTTPStatusError as e:
        try:
            body = e.response.json()
        except Exception:
            body = e.response.text
        return _err("Failed to update curated deal status", body)
    except Exception as e:
        return _err("Failed to update curated deal status", e)


@mcp.tool()
async def pm_discover_dsps(
    logged_in_owner_type_id: int,
    atd_id: int = -1,
    page_number: int = 1,
    page_size: int = 100,
    sort: str = "name",
) -> dict[str, Any]:
    """Internal helper: discover PMP-enabled DSP entities.

    Uses atd_id=-1 (Data Provider mode) by default since Elcano is
    registered as a Data Provider in PubMatic; using atd_id=60067
    returns 0 DSPs for our token's scope.
    """
    try:
        _validate_owner_context(ELCANO_OWNER_ID, logged_in_owner_type_id)

        client = get_pubmatic_client()
        data = await client.list_pmp_dsps(
            atd_id=atd_id,
            page_number=page_number,
            page_size=page_size,
            sort=sort,
        )

        items_raw = data.get("items", []) if isinstance(data, dict) else []
        items = [{"id": item.get("id"), "name": item.get("name")} for item in items_raw if isinstance(item, dict)]
        return _ok({"result": data, "items": items})
    except httpx.HTTPStatusError as e:
        try:
            body = e.response.json()
        except Exception:
            body = e.response.text
        return _err("Failed to discover DSPs", body)
    except Exception as e:
        return _err("Failed to discover DSPs", e)


@mcp.tool()
async def pm_discover_dsp_buyer_map(
    dsp_id: int,
    query: str | None = None,
    buyer_page_number: int = 1,
    buyer_page_size: int = 1000,
) -> dict[str, Any]:
    """Internal helper: discover buyer-seat mappings for a DSP."""
    try:
        client = get_pubmatic_client()
        data = await client.list_dsp_buyer_map(
            dsp_id=dsp_id,
            query=query,
            buyer_page_number=buyer_page_number,
            buyer_page_size=buyer_page_size,
        )

        rows_raw: list[Any] = []
        if isinstance(data, dict):
            if isinstance(data.get("items"), list):
                rows_raw = data.get("items", [])
            elif isinstance(data.get("results"), list):
                rows_raw = data.get("results", [])

        rows: list[dict[str, Any]] = []
        for item in rows_raw:
            if not isinstance(item, dict):
                continue
            # Live PubMatic shape: {demandPartner: {id, name}, buyers: [{id, seatId, dspBuyerId, name}, ...]}
            # Each buyer is a separate row in our flattened representation.
            demand_partner = item.get("demandPartner") if isinstance(item.get("demandPartner"), dict) else {}
            buyers_list = item.get("buyers") if isinstance(item.get("buyers"), list) else None
            if buyers_list:
                top_dsp_id = demand_partner.get("id") or item.get("dspId") or dsp_id
                top_dsp_name = demand_partner.get("name") or item.get("dspName")
                for buyer in buyers_list:
                    if not isinstance(buyer, dict):
                        continue
                    rows.append(
                        {
                            "dspId": top_dsp_id,
                            "dspName": top_dsp_name,
                            "buyerId": buyer.get("id"),
                            "buyerName": buyer.get("name"),
                            "seatId": str(buyer.get("seatId") or "").strip() or None,
                            "dspBuyerId": buyer.get("dspBuyerId"),
                        }
                    )
                continue
            # Legacy/fallback flat shape (used by older mocked endpoints).
            rows.append(
                {
                    "dspId": item.get("dspId") or item.get("advertisingEntityId") or dsp_id,
                    "dspName": item.get("dspName") or item.get("advertisingEntityName"),
                    "buyerId": item.get("buyerId") or item.get("atdId") or item.get("id"),
                    "buyerName": item.get("buyerName") or item.get("name"),
                    "seatId": item.get("seatId") or item.get("seat") or item.get("seatName"),
                    "dspBuyerId": item.get("dspBuyerId"),
                }
            )

        return _ok({"result": data, "items": rows})
    except httpx.HTTPStatusError as e:
        try:
            body = e.response.json()
        except Exception:
            body = e.response.text
        return _err("Failed to discover DSP buyer mappings", body)
    except Exception as e:
        return _err("Failed to discover DSP buyer mappings", e)


_MAX_DSP_CANDIDATES = 50
_MAX_BUYER_SEAT_CANDIDATES = 100


async def _list_all_pmp_dsps(logged_in_owner_type_id: int) -> list[dict[str, Any]]:
    """Paginate through every PMP-enabled DSP using atd_id=-1 (Data Provider scope).

    Cached per-process and on disk (~/.cache/cutlass/pubmatic/dsps_atdneg1.json,
    4-hour TTL by default). Disable with PUBMATIC_CACHE_TTL_SECONDS=0.
    """
    global _dsp_list_cache
    if _dsp_list_cache is not None:
        return _dsp_list_cache

    cached = _cache_get("dsps_atdneg1")
    if isinstance(cached, list):
        _dsp_list_cache = cached
        return cached

    all_items: list[dict[str, Any]] = []
    page = 1
    while True:
        result = await pm_discover_dsps(
            logged_in_owner_type_id=logged_in_owner_type_id,
            page_number=page,
            page_size=100,
            sort="name",
        )
        if not result.get("success"):
            raise PubMaticResolutionError(_extract_tool_error_message(result, "Failed to load DSPs"))
        items = [item for item in result.get("items", []) if isinstance(item, dict)]
        all_items.extend(items)
        if len(items) < 100 or page >= 20:
            break
        page += 1
    _dsp_list_cache = all_items
    _cache_put("dsps_atdneg1", all_items)
    return all_items


async def _resolve_dsp_buyer_mapping(
    *,
    dsp_name: str,
    buyer_name: str,
    seat_id: str,
    logged_in_owner_type_id: int,
    dsp_id: int | None = None,
    buyer_id: int | None = None,
) -> tuple[dict[str, Any], list[str]]:
    """Resolve DSP + buyer + seat into PubMatic IDs, with optional bypass.

    If dsp_id and/or buyer_id are passed, the corresponding lookup is
    skipped — useful when the agent already knows the integer ID.
    """
    warnings: list[str] = []

    resolved_dsp_id = dsp_id
    resolved_dsp_name = dsp_name
    if resolved_dsp_id is None:
        dsps = await _list_all_pmp_dsps(logged_in_owner_type_id)
        available_dsp_names = [
            str(dsp.get("name")).strip() for dsp in dsps if dsp.get("name") and str(dsp.get("name")).strip()
        ]
        try:
            dsp = _resolve_unique_match(
                dsps,
                dsp_name,
                "DSP",
                lookup_fields=("name", "id"),
                allow_contains_match=True,
            )
        except (LookupError, ValueError) as exc:
            raise PubMaticResolutionError(
                str(exc),
                dsp_name=dsp_name,
                available_dsps=available_dsp_names[:_MAX_DSP_CANDIDATES],
                available_dsp_count=len(available_dsp_names),
            ) from exc
        resolved_dsp_id = _select_id_from_item(dsp, ("id",))
        if resolved_dsp_id is None:
            raise PubMaticResolutionError(
                f"Resolved DSP is missing an id: {dsp_name}",
                dsp_name=dsp_name,
            )
        resolved_dsp_name = dsp.get("name") or dsp_name

    resolved_buyer_id = buyer_id
    resolved_buyer_name = buyer_name

    # When the agent provides only seat_id (no buyer_name and no buyer_id), or when
    # the supplied buyer_id might actually be a dspBuyerId (the DSP-side reference)
    # rather than PubMatic's internal buyer.id, hit the dspBuyerMap to resolve the
    # canonical buyer.id. PubMatic's deal payload requires the internal id.
    need_buyer_lookup = resolved_buyer_id is None or (resolved_buyer_id is not None and seat_id)
    if need_buyer_lookup:
        mappings_result = await pm_discover_dsp_buyer_map(
            dsp_id=resolved_dsp_id,
            query=seat_id or None,
            buyer_page_number=1,
            buyer_page_size=1000,
        )
        if not mappings_result.get("success"):
            raise PubMaticResolutionError(
                _extract_tool_error_message(mappings_result, "Failed to load DSP buyer mappings"),
                dsp_id=resolved_dsp_id,
                dsp_name=resolved_dsp_name,
            )
        mappings = [row for row in mappings_result.get("items", []) if isinstance(row, dict)]
        available_buyer_seats = [
            {
                "buyerId": row.get("buyerId"),
                "buyerName": row.get("buyerName"),
                "seatId": row.get("seatId"),
                "dspBuyerId": row.get("dspBuyerId"),
            }
            for row in mappings
            if row.get("buyerId") is not None
        ]

        normalized_seat = seat_id.strip() if seat_id else ""
        normalized_buyer_name = _normalize_lookup_text(buyer_name) if buyer_name else ""

        def _matches(row: dict[str, Any]) -> bool:
            row_seat = str(row.get("seatId") or "").strip()
            row_buyer_name = _normalize_lookup_text(row.get("buyerName") or "")
            row_buyer_id = row.get("buyerId")
            row_dsp_buyer_id = row.get("dspBuyerId")
            seat_ok = (not normalized_seat) or row_seat == normalized_seat
            if resolved_buyer_id is not None:
                # Match if input matches either the canonical buyer.id or the dspBuyerId.
                id_ok = row_buyer_id == resolved_buyer_id or row_dsp_buyer_id == resolved_buyer_id
                return seat_ok and id_ok
            if normalized_buyer_name:
                return seat_ok and row_buyer_name == normalized_buyer_name
            # Seat-only path: any row matching seat is acceptable.
            return seat_ok

        candidates = [row for row in mappings if _matches(row)]

        if not candidates:
            raise PubMaticResolutionError(
                f"Buyer-seat mapping not found for buyer={buyer_name or buyer_id!r} seat={seat_id!r}",
                dsp_id=resolved_dsp_id,
                dsp_name=resolved_dsp_name,
                buyer_name=buyer_name,
                buyer_id=buyer_id,
                seat_id=seat_id,
                available_buyer_seats=available_buyer_seats[:_MAX_BUYER_SEAT_CANDIDATES],
                available_buyer_seat_count=len(available_buyer_seats),
            )

        # Prefer real buyers over the synthetic "ANY BUYER" / id=-1 wildcard PubMatic
        # returns when no specific seat filter narrows the result.
        real_candidates = [row for row in candidates if (row.get("buyerId") or 0) > 0]
        mapping_row = real_candidates[0] if real_candidates else candidates[0]

        canonical_buyer_id = mapping_row.get("buyerId")
        if canonical_buyer_id is None:
            raise PubMaticResolutionError(
                f"Resolved buyer mapping is missing a buyer id for buyer={buyer_name or buyer_id!r}",
                dsp_id=resolved_dsp_id,
                buyer_name=buyer_name,
                seat_id=seat_id,
            )

        if resolved_buyer_id is not None and resolved_buyer_id != canonical_buyer_id:
            warnings.append(
                f"Input buyer_id={resolved_buyer_id} matched as the dspBuyerId; using PubMatic's "
                f"canonical buyer id {canonical_buyer_id} ({mapping_row.get('buyerName')!r}) instead."
            )
        resolved_buyer_id = canonical_buyer_id
        resolved_buyer_name = mapping_row.get("buyerName") or buyer_name or str(resolved_buyer_id)
        if not seat_id and mapping_row.get("seatId"):
            seat_id = str(mapping_row.get("seatId"))

    return {
        "dsp_id": int(resolved_dsp_id),
        "dsp_name": resolved_dsp_name or f"dsp_id={resolved_dsp_id}",
        "buyer_id": int(resolved_buyer_id),
        "buyer_name": resolved_buyer_name or f"buyer_id={resolved_buyer_id}",
        "seat_id": (seat_id or "").strip(),
    }, warnings


async def _resolve_named_entities(
    *,
    kind: str,
    names: list[str],
    logged_in_owner_type_id: int,
) -> tuple[list[int], list[str]]:
    warnings: list[str] = []
    if not names:
        return [], warnings

    if kind == "segment":
        return await _resolve_audience_segments(names, logged_in_owner_type_id=logged_in_owner_type_id)

    client = get_pubmatic_client()
    if kind == "publisher":
        payload = await client.list_publishers()
        id_keys = PUBMATIC_PUBLISHER_KEYS
        lookup_fields = ("companyName", "name", "publisherName", "accountName", "id", "publisherId", "accountId")
    else:
        raise ValueError(f"Unsupported entity kind: {kind}")

    items = _collect_items(payload)
    resolved_ids: list[int] = []
    for name in names:
        try:
            match = _resolve_unique_match(
                items,
                name,
                kind,
                lookup_fields=lookup_fields,
                allow_contains_match=True,
            )
        except (LookupError, ValueError):
            warnings.append(f"Could not find {kind} '{name}'.")
            continue
        resolved_id = _select_id_from_item(match, id_keys)
        if resolved_id is None:
            warnings.append(f"Resolved {kind} '{name}' did not include an id.")
            continue
        resolved_ids.append(resolved_id)

    return _dedupe_preserving_order(resolved_ids), warnings


_AUDIENCE_SEARCH_KEY_STOP_WORDS: frozenset[str] = frozenset(
    {"in", "the", "and", "or", "for", "with", "all", "of", "to", "a", "an", "intent", "market"}
)
# Generic nouns that, while long, frequently saturate PubMatic's 100-result
# page with unrelated audiences (e.g. "Enthusiasts" returns "Auto Enthusiasts",
# "Sports Enthusiasts", etc., crowding out the actual target). When they
# appear, prefer them only as a fallback after more distinctive tokens.
_AUDIENCE_SEARCH_KEY_GENERIC_NOUNS: frozenset[str] = frozenset(
    {"enthusiasts", "shoppers", "consumers", "users", "audience", "audiences", "buyers", "owners"}
)


def _audience_search_key_candidates(name: str, *, max_candidates: int = 3) -> list[str]:
    """Pick ordered substring candidates for PubMatic's searchKey.

    PubMatic's searchKey behaves like a substring/full-text match on
    audienceName, capped at 100 results per page. The longest token is
    usually the most distinctive, but generic nouns ("Enthusiasts",
    "Shoppers") saturate the page and crowd out the target. Order:
    distinctive tokens (longest first), then generic-noun tokens, then
    the first 60 chars of the input as a last-ditch fallback.
    """
    words = re.findall(r"[A-Za-z]{3,}", name)
    distinctive: list[str] = []
    generic: list[str] = []
    seen: set[str] = set()
    for word in words:
        lower = word.lower()
        if lower in _AUDIENCE_SEARCH_KEY_STOP_WORDS or lower in seen:
            continue
        seen.add(lower)
        if lower in _AUDIENCE_SEARCH_KEY_GENERIC_NOUNS:
            generic.append(word)
        else:
            distinctive.append(word)
    distinctive.sort(key=len, reverse=True)
    generic.sort(key=len, reverse=True)
    ordered = distinctive + generic
    if not ordered:
        fallback = name.strip()[:60] or name
        return [fallback] if fallback else []
    return ordered[:max_candidates]


def _audience_search_key(name: str) -> str:
    """Return the single best searchKey candidate (backward-compat shim)."""
    candidates = _audience_search_key_candidates(name, max_candidates=1)
    return candidates[0] if candidates else name


async def _cached_buyer_audience_search(search_key: str) -> dict[str, Any]:
    """Disk-cached buyerInsights audience search keyed by searchKey."""
    safe = _SAFE_KEY_RE.sub("_", search_key.lower())[:64]
    cache_key = f"audience_search_{safe}"
    cached = _cache_get(cache_key)
    if isinstance(cached, dict):
        return cached
    client = get_pubmatic_client()
    payload = await client.list_buyer_audiences(search_key=search_key, page_size=100)
    if isinstance(payload, dict):
        _cache_put(cache_key, payload)
    return payload


async def _resolve_audience_segments(
    names: list[str],
    *,
    logged_in_owner_type_id: int,
) -> tuple[list[int], list[str]]:
    """Resolve segment names to PubMatic audienceIds via the Audience API for Buyers.

    Uses POST /v1/audience/buyerInsights/audiences with a per-name
    searchKey to narrow ~43k registered audiences to a tractable
    candidate set. Replaces the GET /v1/audience/segments path which
    returns AUD01_0161 'no permission' for our token.
    """
    warnings: list[str] = []
    resolved_ids: list[int] = []

    for name in names:
        normalized_target = _normalize_lookup_text(name)
        search_keys = _audience_search_key_candidates(name)

        match: dict[str, Any] | None = None
        last_search_key: str | None = None
        last_items: list[Any] = []
        http_error: httpx.HTTPStatusError | None = None
        for search_key in search_keys:
            last_search_key = search_key
            try:
                payload = await _cached_buyer_audience_search(search_key)
            except httpx.HTTPStatusError as exc:
                http_error = exc
                continue

            items = payload.get("items") if isinstance(payload, dict) else None
            if not isinstance(items, list):
                continue
            last_items = items

            # Prefer exact normalized match; fall back to substring.
            for item in items:
                if not isinstance(item, dict):
                    continue
                item_name = item.get("audienceName") or item.get("name") or ""
                if _normalize_lookup_text(item_name) == normalized_target:
                    match = item
                    break
            if match is None:
                for item in items:
                    if not isinstance(item, dict):
                        continue
                    item_name = item.get("audienceName") or item.get("name") or ""
                    if normalized_target and normalized_target in _normalize_lookup_text(item_name):
                        match = item
                        break
            if match is not None:
                break

        if match is None:
            if http_error is not None and not last_items:
                warnings.append(
                    f"PubMatic audience search for {name!r} returned HTTP "
                    f"{http_error.response.status_code}; segment skipped."
                )
                continue
            if not last_items:
                warnings.append(f"Audience search for {name!r} returned no items.")
                continue
            sample = [
                item.get("audienceName") or item.get("name")
                for item in last_items[:5]
                if isinstance(item, dict) and (item.get("audienceName") or item.get("name"))
            ]
            warnings.append(
                f"PubMatic has no audience matching segment {name!r} "
                f"(searchKeys={search_keys!r}, last={last_search_key!r} returned "
                f"{len(last_items)} candidates). Closest: {sample}."
            )
            continue

        audience_id = match.get("audienceId") or match.get("id")
        if audience_id is None:
            warnings.append(f"Audience match for {name!r} is missing an id.")
            continue
        try:
            resolved_ids.append(int(audience_id))
        except (TypeError, ValueError):
            warnings.append(f"Audience match for {name!r} has non-integer id {audience_id!r}.")

    return _dedupe_preserving_order(resolved_ids), warnings


_PUBMATIC_GEO_LEVEL_COUNTRY = 1
_PUBMATIC_GEO_LEVEL_REGION = 2


async def _cached_list_geos(
    *,
    name_like: str | None = None,
    country_code: str | None = None,
    geo_level: int | None = None,
    page_size: int = 20,
) -> dict[str, Any]:
    """Disk-cached wrapper around list_geos. Cache key is the lookup signature."""
    parts = [
        f"nl={(name_like or '').lower().replace('/', '_')}",
        f"cc={(country_code or '').upper()}",
        f"gl={geo_level if geo_level is not None else ''}",
        f"ps={page_size}",
    ]
    key = "geo_" + "_".join(parts)
    cached = _cache_get(key)
    if isinstance(cached, dict):
        return cached
    client = get_pubmatic_client()
    payload = await client.list_geos(
        name_like=name_like,
        country_code=country_code,
        geo_level=geo_level,
        page_size=page_size,
    )
    if isinstance(payload, dict):
        _cache_put(key, payload)
    return payload


async def _resolve_geo_ids(
    *,
    geo_countries: list[str] | None,
    geo_states: list[str] | None,
) -> tuple[list[int], list[str]]:
    """Resolve country/state names to PubMatic geo IDs via filtered lookup.

    Per PubMatic docs, the geo endpoint requires `filters=name like *X*`
    plus optional `countryCode` / `geoLevel` filters. The unfiltered
    list-all variant returns 502 because the dataset is too large.
    """
    warnings: list[str] = []
    countries_in = [str(v).strip() for v in (geo_countries or []) if str(v).strip()]
    states_in = [str(v).strip() for v in (geo_states or []) if str(v).strip()]
    if not countries_in and not states_in:
        return [], warnings

    numeric_geo_ids: list[int] = []
    unresolved: list[str] = []
    state_country_code: str | None = None

    def _pick_match(items: list[dict[str, Any]], raw: str) -> dict[str, Any] | None:
        normalized_raw = _normalize_lookup_text(raw)
        exact = next(
            (
                it
                for it in items
                if _normalize_lookup_text(it.get("name")) == normalized_raw
                or _normalize_lookup_text(it.get("countryCode")) == normalized_raw
                or _normalize_lookup_text(it.get("regionCode")) == normalized_raw
                or _normalize_lookup_text(it.get("stateCode")) == normalized_raw
            ),
            None,
        )
        if exact is not None:
            return exact
        return items[0] if items else None

    for raw in countries_in:
        if raw.isdigit():
            numeric_geo_ids.append(int(raw))
            continue
        try:
            if len(raw) == 2 and raw.isalpha():
                payload = await _cached_list_geos(
                    country_code=raw.upper(), geo_level=_PUBMATIC_GEO_LEVEL_COUNTRY, page_size=10
                )
            else:
                payload = await _cached_list_geos(name_like=raw, geo_level=_PUBMATIC_GEO_LEVEL_COUNTRY, page_size=10)
        except httpx.HTTPStatusError:
            raise
        items = _collect_items(payload)
        match = _pick_match(items, raw)
        if match is None:
            unresolved.append(raw)
            continue
        resolved_id = _select_id_from_item(match, ("id",))
        if resolved_id is None:
            unresolved.append(raw)
            continue
        numeric_geo_ids.append(resolved_id)
        if state_country_code is None and match.get("countryCode"):
            state_country_code = str(match.get("countryCode"))

    if not state_country_code:
        state_country_code = "US"

    for raw in states_in:
        if raw.isdigit():
            numeric_geo_ids.append(int(raw))
            continue
        try:
            payload = await _cached_list_geos(
                name_like=raw,
                country_code=state_country_code,
                geo_level=_PUBMATIC_GEO_LEVEL_REGION,
                page_size=20,
            )
        except httpx.HTTPStatusError:
            raise
        items = _collect_items(payload)
        match = _pick_match(items, raw)
        if match is None:
            unresolved.append(raw)
            continue
        resolved_id = _select_id_from_item(match, ("id",))
        if resolved_id is None:
            unresolved.append(raw)
            continue
        numeric_geo_ids.append(resolved_id)

    if unresolved:
        warnings.append("Could not resolve PubMatic geo IDs for: " + ", ".join(sorted(dict.fromkeys(unresolved))) + ".")

    return _dedupe_preserving_order(numeric_geo_ids), warnings


async def _load_iab_taxonomy() -> list[dict[str, Any]]:
    """Fetch PubMatic's IAB content taxonomy and flatten top-level + subcategories.

    Cached per-process and on disk (~/.cache/cutlass/pubmatic/iab_taxonomy.json,
    4-hour TTL by default). Disable with PUBMATIC_CACHE_TTL_SECONDS=0.
    """
    global _iab_taxonomy_cache
    if _iab_taxonomy_cache is not None:
        return _iab_taxonomy_cache

    cached = _cache_get("iab_taxonomy")
    if isinstance(cached, list):
        _iab_taxonomy_cache = cached
        return cached

    client = get_pubmatic_client()
    data = await client.list_iab_categories(page_size=200)
    items = data.get("items", []) if isinstance(data, dict) else []
    flat: list[dict[str, Any]] = []
    for top in items:
        if not isinstance(top, dict):
            continue
        if top.get("id") is not None:
            flat.append(
                {
                    "id": top.get("id"),
                    "iabId": top.get("iabId"),
                    "name": top.get("iabName") or top.get("name"),
                }
            )
        for sub in top.get("subCategoryList") or []:
            if isinstance(sub, dict) and sub.get("id") is not None:
                flat.append(
                    {
                        "id": sub.get("id"),
                        "iabId": sub.get("iabId"),
                        "name": sub.get("iabName") or sub.get("name"),
                    }
                )
    _iab_taxonomy_cache = flat
    _cache_put("iab_taxonomy", flat)
    return flat


async def _resolve_iab_categories(values: list[Any]) -> tuple[list[int], list[str]]:
    """Resolve mixed IAB inputs (numeric IDs, IAB codes, or names) to PubMatic numeric IDs.

    Inputs accepted per element:
      - int: assumed to be a PubMatic IAB id; validated against the taxonomy
      - "IAB2-1": IAB code lookup
      - "Auto Parts": name lookup (case-insensitive normalized)
      - "IAB2-1 Auto Parts": name takes precedence; warning emitted on
        code/name disagreement so bad codes (e.g. user labeled IAB2-5
        as "Automotive" when IAB2-5 is "Certified Pre-Owned") still
        resolve to the correct ID by name.

    Raises PubMaticResolutionError with `unresolved` details if any
    input can't be matched.
    """
    if not values:
        return [], []

    taxonomy = await _load_iab_taxonomy()
    by_id = {item["id"]: item for item in taxonomy if isinstance(item.get("id"), int)}
    by_iab_code = {str(item["iabId"]).upper(): item for item in taxonomy if item.get("iabId")}
    by_name = {_normalize_lookup_text(item["name"]): item for item in taxonomy if item.get("name")}

    resolved_ids: list[int] = []
    warnings: list[str] = []
    unresolved: list[dict[str, Any]] = []

    for raw in values:
        if isinstance(raw, bool):
            unresolved.append({"input": raw, "reason": "invalid_type"})
            continue
        if isinstance(raw, int):
            if raw in by_id:
                resolved_ids.append(raw)
            else:
                unresolved.append({"input": raw, "reason": "id_not_in_taxonomy"})
            continue

        text = str(raw).strip()
        if not text:
            continue

        if text.isdigit():
            n = int(text)
            if n in by_id:
                resolved_ids.append(n)
            else:
                unresolved.append({"input": text, "reason": "id_not_in_taxonomy"})
            continue

        tokens = text.split(None, 1)
        leading_code = tokens[0] if tokens and _IAB_CODE_RE.match(tokens[0]) else None
        trailing_name = tokens[1] if leading_code and len(tokens) > 1 else (None if leading_code else text)

        code_match = by_iab_code.get(leading_code.upper()) if leading_code else None
        name_match = by_name.get(_normalize_lookup_text(trailing_name)) if trailing_name else None

        chosen = name_match or code_match
        if chosen is None:
            unresolved.append({"input": text, "reason": "no_match"})
            continue

        if name_match and code_match and name_match["id"] != code_match["id"]:
            warnings.append(
                f"IAB input {text!r}: code {code_match['iabId']} maps to id {code_match['id']} "
                f"({code_match['name']!r}), but the trailing name maps to id {name_match['id']} "
                f"({name_match['name']!r}). Using the name match."
            )
        elif code_match and trailing_name and name_match is None:
            warnings.append(
                f"IAB input {text!r}: trailing name {trailing_name!r} did not match a category, but "
                f"code {code_match['iabId']} resolved to id {code_match['id']} ({code_match['name']!r}). "
                f"Using the code match — verify this is what you intended."
            )

        resolved_ids.append(chosen["id"])

    if unresolved:
        raise PubMaticResolutionError(
            "Could not resolve PubMatic IAB categories for: "
            + ", ".join(str(item.get("input")) for item in unresolved),
            unresolved=unresolved,
            taxonomy_size=len(taxonomy),
        )

    return _dedupe_preserving_order(resolved_ids), warnings


async def _build_targeting_payload_from_prompt_inputs(
    *,
    logged_in_owner_type_id: int,
    publisher_ids: list[int] | None,
    segment_ids: list[int] | None,
    domains: list[str] | None,
    geo_countries: list[str] | None,
    geo_states: list[str] | None,
    device_types: list[str] | None,
    iab_category_ids: list[int] | None,
    viewability_threshold: int | None,
    domain_match_type: int = PUBMATIC_DEFAULT_DOMAIN_MATCH_TYPE,
) -> tuple[dict[str, Any], list[str]]:
    warnings: list[str] = []
    has_targeting_inputs = any(
        [
            segment_ids,
            domains,
            geo_countries,
            geo_states,
            device_types,
            iab_category_ids,
            viewability_threshold is not None,
        ]
    )
    if not has_targeting_inputs:
        return {}, warnings

    targeting_payload: dict[str, Any] = {
        "ownerId": ELCANO_OWNER_ID,
        "ownerType": logged_in_owner_type_id,
        "deviceMakeTargeting": PUBMATIC_DEFAULT_DEVICE_MAKE_TARGETING,
    }

    resolved_segments = [segment_id for segment_id in (segment_ids or []) if isinstance(segment_id, int)]
    normalized_domains = _dedupe_preserving_order(
        [domain for domain in (domains or []) if isinstance(domain, str) and domain]
    )

    if resolved_segments:
        targeting_payload["audienceSegments"] = [{"id": segment_id} for segment_id in resolved_segments]

    if normalized_domains:
        targeting_payload["domainList"] = normalized_domains
        targeting_payload["domainMatchType"] = domain_match_type

    geo_ids, geo_warnings = await _resolve_geo_ids(geo_countries=geo_countries, geo_states=geo_states)
    warnings.extend(geo_warnings)
    if geo_ids:
        targeting_payload["geos"] = geo_ids

    if device_types:
        normalized_device_types = _dedupe_preserving_order(
            [
                PUBMATIC_DEVICE_TYPE_IDS.get(_normalize_lookup_text(device_type))
                for device_type in device_types
                if str(device_type).strip()
            ]
        )
        targeting_payload["deviceType"] = [
            device_type for device_type in normalized_device_types if device_type is not None
        ]

    if iab_category_ids:
        targeting_payload["iabCategories"] = list(iab_category_ids)

    if viewability_threshold is not None:
        threshold = _coerce_int(viewability_threshold, "viewability_threshold")
        if threshold < 0 or threshold > 100:
            raise ValueError("viewability_threshold must be between 0 and 100")
        if threshold % 10 != 0 or threshold < 10 or threshold > 90:
            raise ValueError("viewability_threshold must be one of 10,20,30,40,50,60,70,80,90")
        targeting_payload["minViewabilityValue"] = threshold

    return targeting_payload, warnings


async def _build_prepared_pubmatic_deal(
    *,
    name: str,
    start_date: str,
    end_date: str,
    dsp_name: str | None,
    buyer_name: str | None,
    logged_in_owner_type_id: int,
    seat_id: str | None,
    deal_id: str | None,
    deal_source: int,
    floor_ecpm: float | None,
    publisher_names: list[str] | None,
    publisher_ids: list[int] | None,
    segment_names: list[str] | None,
    domain_file_path: str | None,
    domain_sheet: str | None,
    domain_column: str | None,
    domain_match_operator: str | None = None,
    app_bundle_file_path: str | None = None,
    app_bundle_sheet: str | None = None,
    app_bundle_column: str | None = None,
    app_bundle_match_operator: str | None = None,
    geo_countries: list[str] | None,
    geo_states: list[str] | None,
    device_types: list[str] | None,
    iab_categories: list[Any] | None,
    viewability_threshold: int | None,
    ad_formats: list[int] | None,
    platforms: list[int] | None,
    priority: int,
    has_max_reach: int,
    publisher_block_list: list[int] | None,
    max_allowed_publishers: int | None,
    auto_update_publishers: bool | None,
    dsp_id: int | None,
    buyer_id: int | None,
    auction_type: int | None,
    channel: str | None = None,
    fee: list[dict[str, Any]] | dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Resolve all PubMatic deal inputs and build a prepared artifact.

    Resolution failures become structured blockers rather than raised
    exceptions so the prepared artifact always carries a complete picture
    of what the agent needs to fix before submitting.

    DSP/buyer resolution accepts either:
      - dsp_name + buyer_name (server-side fuzzy match, requires the
        /v1/common/advertisingEntity endpoint to be reachable for the
        token's API scope), or
      - dsp_id + buyer_id (raw integers from the PubMatic UI; bypasses
        lookup entirely). seat_id is required either way.
    """
    warnings: list[str] = []
    blockers: list[dict[str, Any]] = []
    resolved_entities: dict[str, Any] = {}
    quality_flags: list[dict[str, Any]] = []

    device_types, applied_channel_default = _apply_pm_channel_device_defaults(device_types, channel)
    if applied_channel_default:
        canonical_channel = _normalize_pm_channel(channel)
        warnings.append(
            f"Applied PubMatic default device targeting for {canonical_channel} channel: {device_types}. "
            "Pass device_types= to override."
        )
        quality_flags.append(
            _make_pm_quality_flag(
                "pm_default_channel_devices_applied",
                f"Auto-filled device_types from channel={canonical_channel!r}.",
                channel=canonical_channel,
                device_types=list(device_types or []),
            )
        )

    try:
        resolved_deal_fees, applied_fee_default = _resolve_pm_deal_fees(fee)
    except ValueError as exc:
        resolved_deal_fees = []
        applied_fee_default = False
        blockers.append(_make_blocker("invalid_fee", str(exc)))
    if applied_fee_default:
        warnings.append(
            f"Applied default {ELCANO_FEE_RECIPIENT_NAME} curator fee: "
            f"{ELCANO_DEFAULT_FEE_VALUE_PERCENT}% of media. Pass fee= to override."
        )
        quality_flags.append(
            _make_pm_quality_flag(
                "pm_default_curator_fee_applied",
                f"Auto-applied flat {ELCANO_DEFAULT_FEE_VALUE_PERCENT}% "
                f"Percent-of-Media curator fee for {ELCANO_FEE_RECIPIENT_NAME}.",
                fee_value_percent=ELCANO_DEFAULT_FEE_VALUE_PERCENT,
                recipient=ELCANO_FEE_RECIPIENT_NAME,
                recipient_id=ELCANO_OWNER_ID,
            )
        )

    try:
        _validate_owner_context(ELCANO_OWNER_ID, logged_in_owner_type_id)
    except ValueError as exc:
        blockers.append(_make_blocker("invalid_owner_context", str(exc)))

    normalized_start: str | None = None
    normalized_end: str | None = None
    try:
        normalized_start = _normalize_zulu_datetime("start_date", start_date)
        normalized_end = _normalize_zulu_datetime("end_date", end_date)
    except ValueError as exc:
        blockers.append(_make_blocker("invalid_dates", str(exc)))

    resolved_platforms: list[int] = []
    try:
        platform_input, applied_platform_default = _apply_pm_channel_platform_defaults(platforms, channel)
        resolved_platforms = _validate_allowed_ids(
            _normalize_integer_list(platform_input, "platforms"),
            PUBMATIC_ALLOWED_PLATFORM_IDS,
            "platforms",
        )
        if applied_platform_default:
            warnings.append(
                f"Applied PubMatic default platforms for {_normalize_pm_channel(channel)} channel: "
                f"{resolved_platforms}. Pass platforms= to override."
            )
            quality_flags.append(
                _make_pm_quality_flag(
                    "pm_default_channel_platforms_applied",
                    f"Auto-applied canonical platforms {resolved_platforms} for channel "
                    f"{_normalize_pm_channel(channel)}.",
                    platforms=resolved_platforms,
                )
            )
        # Mislabeled-enum guard: CTV is platform 7. Legacy prompts (and the
        # pre-fix Manifest UI) sent 5, which is Mobile App Android — flag it
        # loudly but do NOT silently rewrite an explicit caller value.
        if (
            _normalize_pm_channel(channel) == "ctv"
            and resolved_platforms
            and PUBMATIC_CTV_PLATFORM_ID not in resolved_platforms
        ):
            warnings.append(
                f"channel=ctv but platforms={resolved_platforms} does not include {PUBMATIC_CTV_PLATFORM_ID} "
                "(CTV). Platform 5 is Mobile App Android, not CTV — this deal will NOT target CTV "
                f"inventory. Pass platforms=[{PUBMATIC_CTV_PLATFORM_ID}] unless the mismatch is intentional."
            )
            quality_flags.append(
                _make_pm_quality_flag(
                    "pm_ctv_platform_mismatch",
                    f"CTV channel deal without CTV platform {PUBMATIC_CTV_PLATFORM_ID}; "
                    f"submitted platforms={resolved_platforms}.",
                    platforms=resolved_platforms,
                )
            )
    except ValueError as exc:
        blockers.append(_make_blocker("invalid_platforms", str(exc)))

    ad_formats, applied_ad_format_default = _apply_pm_channel_ad_format_defaults(ad_formats, channel)
    if applied_ad_format_default:
        canonical_channel = _normalize_pm_channel(channel)
        warnings.append(
            f"Applied PubMatic default ad_formats for {canonical_channel} channel: {ad_formats}. "
            "Pass ad_formats= to override."
        )
        quality_flags.append(
            _make_pm_quality_flag(
                "pm_default_channel_ad_formats_applied",
                f"Auto-filled ad_formats from channel={canonical_channel!r}.",
                channel=canonical_channel,
                ad_formats=list(ad_formats or []),
            )
        )
    resolved_ad_formats: list[int] = []
    try:
        resolved_ad_formats = _validate_allowed_ids(
            _normalize_integer_list(ad_formats or list(PM_AD_FORMATS_DISPLAY), "ad_formats"),
            PUBMATIC_ALLOWED_AD_FORMAT_IDS,
            "ad_formats",
        )
    except ValueError as exc:
        blockers.append(_make_blocker("invalid_ad_formats", str(exc)))

    resolved_publisher_block_list: list[int] = []
    try:
        resolved_publisher_block_list = _normalize_integer_list(publisher_block_list, "publisher_block_list")
    except ValueError as exc:
        blockers.append(_make_blocker("invalid_publisher_block_list", str(exc)))

    if has_max_reach not in {0, 1}:
        blockers.append(_make_blocker("invalid_has_max_reach", "has_max_reach must be 0 or 1"))
    if has_max_reach == 1 and max_allowed_publishers is not None and not 1 <= int(max_allowed_publishers) <= 200:
        blockers.append(
            _make_blocker("invalid_max_allowed_publishers", "max_allowed_publishers must be between 1 and 200")
        )

    mapping: dict[str, Any] | None = None
    have_dsp = bool(dsp_name) or dsp_id is not None
    have_buyer = bool(buyer_name) or buyer_id is not None or bool(seat_id)
    if not have_dsp or not have_buyer:
        blockers.append(
            _make_blocker(
                "missing_dsp_buyer",
                "Provide DSP and a buyer reference. Buyer may be specified as buyer_name, "
                "buyer_id (raw integer), or seat_id alone — when only seat_id is provided, "
                "the MCP looks up the buyer mapped to that seat under the resolved DSP.",
            )
        )
    else:
        try:
            mapping, mapping_warnings = await _resolve_dsp_buyer_mapping(
                dsp_name=dsp_name or "",
                buyer_name=buyer_name or "",
                seat_id=seat_id or "",
                logged_in_owner_type_id=logged_in_owner_type_id,
                dsp_id=dsp_id,
                buyer_id=buyer_id,
            )
            warnings.extend(mapping_warnings)
            resolved_entities["dsp_buyer_mapping"] = mapping
            if dsp_id is not None and buyer_id is not None:
                resolved_entities["dsp_buyer_source"] = "raw_ids"
            elif dsp_id is not None or buyer_id is not None:
                resolved_entities["dsp_buyer_source"] = "hybrid"
            else:
                resolved_entities["dsp_buyer_source"] = "name_lookup"
        except PubMaticResolutionError as exc:
            blocker_details = {
                "dsp_name": dsp_name,
                "buyer_name": buyer_name,
                "seat_id": seat_id,
                **exc.details,
            }
            blockers.append(_make_blocker("dsp_buyer_unresolved", exc.message, **blocker_details))
        except (LookupError, ValueError) as exc:
            blockers.append(
                _make_blocker(
                    "dsp_buyer_unresolved",
                    str(exc),
                    dsp_name=dsp_name,
                    buyer_name=buyer_name,
                    seat_id=seat_id,
                )
            )

    resolved_publisher_ids = [pid for pid in (publisher_ids or []) if isinstance(pid, int)]
    if publisher_names:
        try:
            named_publisher_ids, publisher_warnings = await _resolve_named_entities(
                kind="publisher",
                names=publisher_names,
                logged_in_owner_type_id=logged_in_owner_type_id,
            )
            resolved_publisher_ids.extend(named_publisher_ids)
            warnings.extend(publisher_warnings)
        except ValueError as exc:
            blockers.append(_make_blocker("publisher_lookup_failed", str(exc)))
    resolved_publisher_ids = _dedupe_preserving_order(resolved_publisher_ids)
    resolved_entities["publisher_ids"] = resolved_publisher_ids

    if has_max_reach == 0 and not resolved_publisher_ids:
        blockers.append(
            _make_blocker(
                "publishers_required_for_manual_mode",
                "publisher_ids or resolvable publisher_names are required when has_max_reach=0",
            )
        )

    resolved_segment_ids: list[int] = []
    if segment_names:
        try:
            resolved_segment_ids, segment_warnings = await _resolve_named_entities(
                kind="segment",
                names=segment_names,
                logged_in_owner_type_id=logged_in_owner_type_id,
            )
            warnings.extend(segment_warnings)
        except ValueError as exc:
            blockers.append(_make_blocker("segment_lookup_failed", str(exc)))
    resolved_entities["segment_ids"] = resolved_segment_ids

    # App-bundle lists ride the same undifferentiated domainList as web domains
    # (the extractor accepts reverse-DNS bundles and numeric store IDs), so
    # app_bundle_* are aliases onto the domain pipeline. A deal supplies one or
    # the other — Manifest emits app_bundle_* for CTV/OTT and domain_* for
    # Display/OLV.
    list_match_operator = domain_match_operator
    if app_bundle_file_path is not None:
        if domain_file_path is not None:
            blockers.append(
                _make_blocker(
                    "conflicting_list_sources",
                    "Provide either domain_file_path or app_bundle_file_path, not both — "
                    "PubMatic targets both via one domainList.",
                )
            )
        else:
            domain_file_path = app_bundle_file_path
            domain_sheet = app_bundle_sheet if app_bundle_sheet is not None else domain_sheet
            domain_column = app_bundle_column if app_bundle_column is not None else domain_column
            list_match_operator = app_bundle_match_operator

    domain_match_type = PUBMATIC_DEFAULT_DOMAIN_MATCH_TYPE
    try:
        domain_match_type = _resolve_domain_match_type(list_match_operator)
    except ValueError as exc:
        blockers.append(_make_blocker("invalid_domain_match_operator", str(exc)))
    resolved_domains: list[str] = []
    if domain_file_path:
        try:
            domain_file_result = _extract_domains_from_file(
                domain_file_path,
                sheet_name=domain_sheet,
                column_name=domain_column,
            )
            resolved_domains = domain_file_result["domains"]
            if not resolved_domains:
                blockers.append(
                    _make_blocker(
                        "no_valid_domains",
                        f"No valid domains found in file: {domain_file_path}",
                    )
                )
            if domain_file_result["invalid_values"]:
                warnings.append(
                    f"Dropped {len(domain_file_result['invalid_values'])} invalid domains from {domain_file_path}."
                )
            if resolved_domains:
                match_label = (
                    "include/allowlist"
                    if domain_match_type == PUBMATIC_DOMAIN_MATCH_TYPE_INCLUDE
                    else "exclude/blocklist"
                )
                warnings.append(
                    f"Applied domainMatchType={domain_match_type} ({match_label}) to "
                    f"{len(resolved_domains)} values from {domain_file_path}."
                )
        except (ValueError, RuntimeError) as exc:
            blockers.append(_make_blocker("domain_file_error", str(exc)))
    resolved_entities["domains"] = resolved_domains

    resolved_iab_category_ids: list[int] = []
    if iab_categories:
        try:
            resolved_iab_category_ids, iab_warnings = await _resolve_iab_categories(iab_categories)
            warnings.extend(iab_warnings)
        except PubMaticResolutionError as exc:
            blockers.append(
                _make_blocker(
                    "iab_categories_unresolved",
                    exc.message,
                    iab_categories=iab_categories,
                    **exc.details,
                )
            )
        except httpx.HTTPStatusError as exc:
            blockers.append(
                _make_blocker(
                    "iab_lookup_failed",
                    f"PubMatic IAB taxonomy lookup returned HTTP {exc.response.status_code}.",
                    status=exc.response.status_code,
                    iab_categories=iab_categories,
                )
            )
    resolved_entities["iab_category_ids"] = resolved_iab_category_ids

    targeting_payload: dict[str, Any] = {}
    if not blockers:
        try:
            targeting_payload, targeting_warnings = await _build_targeting_payload_from_prompt_inputs(
                logged_in_owner_type_id=logged_in_owner_type_id,
                publisher_ids=resolved_publisher_ids,
                segment_ids=resolved_segment_ids,
                domains=resolved_domains,
                geo_countries=geo_countries,
                geo_states=geo_states,
                device_types=device_types,
                iab_category_ids=resolved_iab_category_ids,
                viewability_threshold=viewability_threshold,
                domain_match_type=domain_match_type,
            )
            warnings.extend(targeting_warnings)
        except ValueError as exc:
            blockers.append(_make_blocker("targeting_build_failed", str(exc)))
        except httpx.HTTPStatusError as exc:
            url = str(exc.request.url) if exc.request else ""
            kind = "geo" if "/geo" in url else "segment" if "/segments" in url else "lookup"
            blockers.append(
                _make_blocker(
                    f"{kind}_lookup_failed",
                    f"PubMatic {kind} lookup returned HTTP {exc.response.status_code}. "
                    f"This is a transient PubMatic API failure. Re-run prepare with the affected "
                    f"input (e.g. drop geo_countries/geo_states) or wait and retry.",
                    status=exc.response.status_code,
                    url=url,
                    geo_countries=geo_countries,
                    geo_states=geo_states,
                )
            )

    deal_intent: dict[str, Any] = {}
    if not blockers and mapping is not None and normalized_start and normalized_end:
        # PubMatic's deal-id field is capped (server returns "Auction Package ID exceeds
        # max length" for ours when long). Truncate to 64 chars while keeping the trailing
        # tag (e.g. attribution suffix) so dedup is stable.
        candidate_deal_id = (deal_id or name).strip()
        if len(candidate_deal_id) > 64:
            candidate_deal_id = candidate_deal_id[:64]

        # PubMatic deals default to First Price (auctionType=1) with NO custom
        # media floor — the publisher's existing minimums apply via the
        # auction's own defaults. Brief writers who actually want a deal-level
        # floor override pass `floor_ecpm` explicitly; in that case the deal
        # switches to Fixed Price (auctionType=3) so PubMatic accepts the
        # flooreCPM field (it rejects flooreCPM under First Price). UI ground
        # truth (Reklaim Marriott Travel deal PM-ZWKR-2887): auctionType=1,
        # flooreCPM=0.0, no deal-level floor override.
        explicit_floor_requested = floor_ecpm is not None
        if auction_type is not None:
            resolved_auction_type = int(auction_type)
        elif explicit_floor_requested:
            resolved_auction_type = 3  # Fixed Price so flooreCPM survives
        else:
            resolved_auction_type = 1  # First Price, no custom floor
        deal_intent = {
            "name": name,
            "dealId": candidate_deal_id,
            "auctionType": resolved_auction_type,
            "startDate": normalized_start,
            "endDate": normalized_end,
            "timeZone": 1,
            "targeting": None,
            "platforms": resolved_platforms,
            "adFormats": resolved_ad_formats,
            "hasMaxReach": has_max_reach,
            "priority": int(priority),
            "dealSource": int(deal_source),
            "loggedInOwnerId": ELCANO_OWNER_ID,
            "loggedInOwnerTypeId": logged_in_owner_type_id,
            "dealDspBuyerMappings": [
                {
                    "dspId": mapping["dsp_id"],
                    "buyerId": mapping["buyer_id"],
                    "seatId": mapping["seat_id"],
                }
            ],
        }
        if explicit_floor_requested and resolved_auction_type != 1:
            deal_intent["flooreCPM"] = float(floor_ecpm)
        elif explicit_floor_requested and resolved_auction_type == 1:
            # Caller explicitly forced First Price AND passed a floor — PubMatic
            # would silently drop the floor. Surface it as a warning so the
            # trader doesn't think their floor is in effect.
            warnings.append(
                f"floor_ecpm={floor_ecpm} ignored: caller forced auction_type=1 (First Price), "
                "which doesn't support custom floors. Drop auction_type=1 to honor the floor, "
                "or drop floor_ecpm to use PubMatic's default First Price behavior."
            )
        if resolved_publisher_ids:
            deal_intent["pubIds"] = resolved_publisher_ids
        if resolved_publisher_block_list:
            deal_intent["publisherBlockList"] = resolved_publisher_block_list
        if has_max_reach == 1:
            # PubMatic returns "Invalid value for maxAllowedPublishers" if the field is
            # absent under hasMaxReach=1. Default to 200 (the documented upper bound).
            deal_intent["maxAllowedPublishers"] = (
                int(max_allowed_publishers) if max_allowed_publishers is not None else 200
            )
        if has_max_reach == 1 and auto_update_publishers is not None:
            deal_intent["autoUpdatePublishers"] = auto_update_publishers
        if resolved_deal_fees:
            deal_intent["dealFees"] = resolved_deal_fees

    quality_flags.extend(_blockers_to_quality_flags(blockers))

    prepared_deal_id = f"pubmatic-prepared-{uuid4()}"
    prepared = {
        "prepared_deal_id": prepared_deal_id,
        "ready_to_create": not blockers,
        "blocking_issues": [blocker["message"] for blocker in blockers],
        "blockers": blockers,
        "warnings": _dedupe_preserving_order(warnings),
        "quality_flags": quality_flags,
        "resolved_entities": resolved_entities,
        "targeting_intent": targeting_payload or None,
        "deal_intent": deal_intent,
        "logged_in_owner_type_id": logged_in_owner_type_id,
    }
    _prepared_pubmatic_deals[prepared_deal_id] = prepared
    return prepared


@mcp.tool()
async def pm_prepare_deal_from_prompt_inputs(
    name: str,
    start_date: str,
    end_date: str,
    logged_in_owner_type_id: int,
    dsp_name: str | None = None,
    buyer_name: str | None = None,
    dsp_id: int | None = None,
    buyer_id: int | None = None,
    seat_id: str | None = None,
    deal_id: str | None = None,
    deal_source: int = 8,
    floor_ecpm: float | None = None,
    auction_type: int | None = None,
    publisher_names: list[str] | None = None,
    publisher_ids: list[int] | None = None,
    segment_names: list[str] | None = None,
    domain_file_path: str | None = None,
    domain_sheet: str | None = None,
    domain_column: str | None = None,
    domain_match_operator: str | None = None,
    app_bundle_file_path: str | None = None,
    app_bundle_sheet: str | None = None,
    app_bundle_column: str | None = None,
    app_bundle_match_operator: str | None = None,
    geo_countries: list[str] | None = None,
    geo_states: list[str] | None = None,
    device_types: list[str] | None = None,
    iab_categories: list[Any] | None = None,
    viewability_threshold: int | None = None,
    ad_formats: list[int] | None = None,
    platforms: list[int] | None = None,
    priority: int = 10,
    has_max_reach: int = 0,
    publisher_block_list: list[int] | None = None,
    max_allowed_publishers: int | None = None,
    auto_update_publishers: bool | None = None,
    channel: str | None = None,
    fee: list[dict[str, Any]] | dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Resolve human-readable PubMatic deal inputs into a server-side draft.

    All name -> ID resolution (DSP, buyer, seat, publishers, segments, geo,
    device types) and file extraction happens server-side. The returned
    prepared_deal_id is submitted via pm_create_prepared_deal to actually
    create the deal. Inspect ready_to_create and blocking_issues before
    submitting.

    DSP/buyer inputs accept any of:
      - dsp_name + buyer_name (server-side fuzzy resolution)
      - dsp_id + buyer_id (raw integers)
      - dsp_name + seat_id (looks up the buyer mapped to the seat under the DSP)
      - any combination; if buyer_id matches a dspBuyerId rather than a
        canonical buyer.id, the resolver corrects it and emits a warning.

    auction_type defaults to 1 (First Price) when no floor is requested so
    PubMatic's own publisher minimums apply via the auction. Pass `floor_ecpm`
    to override that — when a floor is supplied without an explicit
    auction_type the resolver upgrades to 3 (Fixed Price) so PubMatic accepts
    the flooreCPM field. Forcing auction_type=1 alongside a non-None
    floor_ecpm surfaces a warning because PubMatic silently drops floors
    under First Price.

    deal_source defaults to 8 (DataProvider); PubMatic external users must
    supply one of 5 (DSP), 7 (Buyer), or 8 (DataProvider).

    channel accepts "display", "olv", "ctv", or "ott" (case-insensitive). When
    device_types is omitted, the prepare flow auto-fills the canonical
    PubMatic device set for that channel (display/olv/ott -> desktop+mobile+
    tablet; ctv -> ctv) and emits a `pm_default_channel_devices_applied`
    quality flag.

    Domain / app-bundle targeting: web-domain lists (Display/OLV) use
    domain_file_path / domain_sheet / domain_column / domain_match_operator;
    CTV/OTT app-bundle lists use app_bundle_file_path / app_bundle_sheet /
    app_bundle_column / app_bundle_match_operator. Both land on PubMatic's one
    domainList (the extractor accepts reverse-DNS bundles and numeric store
    IDs), so supply EITHER a domain file OR an app-bundle file, not both. The
    match operator maps to domainMatchType: "allowlist" -> 1 (include, the
    default when omitted), "blocklist" -> 2 (exclude).

    fee accepts a single dict or list of dealFees entries (see PubMatic
    "Create a curated deal" docs). When omitted, a flat 30% Percentage-of-
    Media transaction fee for Elcano (recipientId=ELCANO_OWNER_ID,
    recipientTypeId=7, feeType=0, feeValueType=0) is auto-applied and a
    `pm_default_curator_fee_applied` quality flag is emitted.
    """
    logger.info("pm_prepare_deal_from_prompt_inputs called with name: %s file: %s", name, domain_file_path)
    try:
        prepared = await _build_prepared_pubmatic_deal(
            name=name,
            start_date=start_date,
            end_date=end_date,
            dsp_name=dsp_name,
            buyer_name=buyer_name,
            logged_in_owner_type_id=logged_in_owner_type_id,
            seat_id=seat_id,
            deal_id=deal_id,
            deal_source=deal_source,
            floor_ecpm=floor_ecpm,
            publisher_names=publisher_names,
            publisher_ids=publisher_ids,
            segment_names=segment_names,
            domain_file_path=domain_file_path,
            domain_sheet=domain_sheet,
            domain_column=domain_column,
            domain_match_operator=domain_match_operator,
            app_bundle_file_path=app_bundle_file_path,
            app_bundle_sheet=app_bundle_sheet,
            app_bundle_column=app_bundle_column,
            app_bundle_match_operator=app_bundle_match_operator,
            geo_countries=geo_countries,
            geo_states=geo_states,
            device_types=device_types,
            iab_categories=iab_categories,
            viewability_threshold=viewability_threshold,
            ad_formats=ad_formats,
            platforms=platforms,
            priority=priority,
            has_max_reach=has_max_reach,
            publisher_block_list=publisher_block_list,
            max_allowed_publishers=max_allowed_publishers,
            auto_update_publishers=auto_update_publishers,
            dsp_id=dsp_id,
            buyer_id=buyer_id,
            auction_type=auction_type,
            channel=channel,
            fee=fee,
        )
        return {
            "success": True,
            "prepared_deal_id": prepared["prepared_deal_id"],
            "ready_to_create": prepared["ready_to_create"],
            "blocking_issues": prepared["blocking_issues"],
            "blockers": prepared["blockers"],
            "warnings": prepared["warnings"],
            "quality_flags": prepared.get("quality_flags", []),
            "resolved_entities": prepared["resolved_entities"],
            "targeting_intent_preview": prepared["targeting_intent"],
            "deal_intent_preview": prepared["deal_intent"],
        }
    except Exception as exc:
        logger.error("pm_prepare_deal_from_prompt_inputs failed: %s", exc)
        return {
            "success": False,
            "ready_to_create": False,
            "blocking_issues": [str(exc)],
            "blockers": [_make_blocker("preparation_error", str(exc))],
            "warnings": [],
            "quality_flags": [
                _make_pm_quality_flag(
                    "pm_preparation_error",
                    str(exc),
                )
            ],
            "error": str(exc),
        }


@mcp.tool()
async def pm_create_prepared_deal(prepared_deal_id: str) -> dict[str, Any]:
    """Submit a previously prepared PubMatic deal artifact and verify it.

    Calls the targeting API (when targeting was prepared), then the curated
    deal API, then verifies the created deal. Will refuse to submit if the
    prepared artifact has unresolved blocking_issues.
    """
    logger.info("pm_create_prepared_deal called with prepared_deal_id: %s", prepared_deal_id)

    prepared = _prepared_pubmatic_deals.get(prepared_deal_id)
    if prepared is None:
        return {
            "success": False,
            "deal_url": None,
            "deal": None,
            "targeting_id": None,
            "warnings": [],
            "quality_flags": [
                _make_pm_quality_flag(
                    "pm_prepared_deal_not_found",
                    f"Prepared PubMatic deal not found: {prepared_deal_id}",
                    prepared_deal_id=prepared_deal_id,
                )
            ],
            "error": f"Prepared PubMatic deal not found: {prepared_deal_id}",
            "verification": None,
        }

    if not prepared["ready_to_create"]:
        return {
            "success": False,
            "prepared_deal_id": prepared_deal_id,
            "deal_url": None,
            "deal": None,
            "targeting_id": None,
            "warnings": prepared["warnings"],
            "quality_flags": list(prepared.get("quality_flags", [])),
            "error": "Prepared PubMatic deal is blocked and cannot be created.",
            "blocking_issues": prepared["blocking_issues"],
            "blockers": prepared["blockers"],
            "verification": None,
        }

    if prepared.get("created"):
        # The create POST for this artifact already succeeded. Re-submitting
        # would create a duplicate live deal on PubMatic, so return the
        # recorded outcome instead. (A prior attempt may have reported
        # success=False if post-create verification failed — the deal still
        # exists in that case.)
        recorded = prepared.get("created_result")
        if isinstance(recorded, dict):
            replay = dict(recorded)
            replay["replayed"] = True
            return replay
        return {
            "success": False,
            "prepared_deal_id": prepared_deal_id,
            "deal_url": None,
            "deal": None,
            "targeting_id": None,
            "warnings": list(prepared.get("warnings", [])),
            "quality_flags": [
                _make_pm_quality_flag(
                    "pm_deal_already_created",
                    "This prepared deal was already submitted and the deal exists on PubMatic. "
                    "Do NOT prepare/submit it again; use pm_list_curated_deals to locate it.",
                    prepared_deal_id=prepared_deal_id,
                )
            ],
            "error": "Prepared deal already submitted; refusing to create a duplicate.",
            "verification": None,
        }

    warnings = list(prepared["warnings"])
    quality_flags: list[dict[str, Any]] = list(prepared.get("quality_flags", []))
    targeting_id: int | None = None
    deal_response: dict[str, Any] | None = None
    deal_url: str | None = None
    verification: dict[str, Any] | None = None

    try:
        targeting_payload = prepared.get("targeting_intent") or {}
        if targeting_payload:
            targeting_result = await pm_create_targeting(targeting_payload)
            if not targeting_result.get("success"):
                raise ValueError(_extract_tool_error_message(targeting_result, "Failed to create targeting"))
            targeting_id = _coerce_int(targeting_result.get("targeting_id"), "targeting_id")

        deal_payload = dict(prepared["deal_intent"])
        deal_payload["targeting"] = targeting_id

        create_result = await pm_create_curated_deal(deal_payload)
        if not create_result.get("success"):
            raise ValueError(_extract_tool_error_message(create_result, "Failed to create curated deal"))

        # The deal now exists on PubMatic. Mark the artifact consumed BEFORE
        # verification: any exception past this point must not let a retry
        # re-run the create POST and double-create the deal.
        prepared["created"] = True

        deal_response = create_result.get("result") if isinstance(create_result.get("result"), dict) else create_result
        curated_id = deal_response.get("id") if isinstance(deal_response, dict) else None
        # Prefer the name from the create response (echoes what PubMatic
        # actually persisted, including any server-side normalization);
        # fall back to the submitted payload name.
        resolved_deal_name = (
            deal_response.get("name") if isinstance(deal_response, dict) else None
        ) or deal_payload.get("name")
        deal_url = _build_pubmatic_deal_url(curated_id, deal_name=resolved_deal_name)

        if curated_id is not None:
            verification = await pm_get_curated_deal(
                curated_id=_coerce_int(curated_id, "curated id"),
                logged_in_owner_id=ELCANO_OWNER_ID,
                logged_in_owner_type_id=prepared["logged_in_owner_type_id"],
            )
            if isinstance(verification, dict) and not verification.get("success"):
                quality_flags.append(
                    _make_pm_quality_flag(
                        "pm_verification_failed",
                        _extract_tool_error_message(verification, "PubMatic verification re-fetch failed."),
                        curated_id=curated_id,
                    )
                )
        else:
            verification = {"success": False, "error": "Create response did not include a curated deal id."}
            quality_flags.append(
                _make_pm_quality_flag(
                    "pm_verification_failed",
                    "Create response did not include a curated deal id.",
                )
            )

        result = {
            "success": True,
            "prepared_deal_id": prepared_deal_id,
            "deal_url": deal_url,
            "deal": deal_response,
            "targeting_id": targeting_id,
            "warnings": _dedupe_preserving_order(warnings),
            "quality_flags": quality_flags,
            "error": None,
            "verification": verification,
        }
        prepared["created_result"] = result
        return result
    except Exception as exc:
        logger.error("pm_create_prepared_deal failed: %s", exc)
        quality_flags.append(
            _make_pm_quality_flag(
                "pm_create_call_failed",
                str(exc),
            )
        )
        if prepared.get("created"):
            # The create POST succeeded before this exception (e.g. the
            # verification re-fetch raised). Make that unmissable so the
            # agent does not retry the creation.
            quality_flags.append(
                _make_pm_quality_flag(
                    "pm_deal_already_created",
                    "The curated deal WAS created before this error; do NOT submit again. "
                    "Use pm_list_curated_deals to confirm it.",
                    prepared_deal_id=prepared_deal_id,
                )
            )
        result = {
            "success": False,
            "prepared_deal_id": prepared_deal_id,
            "deal_url": deal_url,
            "deal": deal_response,
            "targeting_id": targeting_id,
            "warnings": _dedupe_preserving_order(warnings),
            "quality_flags": quality_flags,
            "error": str(exc),
            "verification": verification,
        }
        if prepared.get("created"):
            prepared["created_result"] = result
        return result


@mcp.tool()
async def pm_execute_deal_from_prompt_inputs(
    name: str,
    start_date: str,
    end_date: str,
    logged_in_owner_type_id: int,
    dsp_name: str | None = None,
    buyer_name: str | None = None,
    dsp_id: int | None = None,
    buyer_id: int | None = None,
    seat_id: str | None = None,
    deal_id: str | None = None,
    deal_source: int = 8,
    floor_ecpm: float | None = None,
    auction_type: int | None = None,
    publisher_names: list[str] | None = None,
    publisher_ids: list[int] | None = None,
    segment_names: list[str] | None = None,
    domain_file_path: str | None = None,
    domain_sheet: str | None = None,
    domain_column: str | None = None,
    domain_match_operator: str | None = None,
    app_bundle_file_path: str | None = None,
    app_bundle_sheet: str | None = None,
    app_bundle_column: str | None = None,
    app_bundle_match_operator: str | None = None,
    geo_countries: list[str] | None = None,
    geo_states: list[str] | None = None,
    device_types: list[str] | None = None,
    iab_categories: list[Any] | None = None,
    viewability_threshold: int | None = None,
    ad_formats: list[int] | None = None,
    platforms: list[int] | None = None,
    priority: int = 10,
    has_max_reach: int = 0,
    publisher_block_list: list[int] | None = None,
    max_allowed_publishers: int | None = None,
    auto_update_publishers: bool | None = None,
    channel: str | None = None,
    fee: list[dict[str, Any]] | dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Prepare, submit, and verify a PubMatic auction package deal in one call.

    Thin wrapper over pm_prepare_deal_from_prompt_inputs and
    pm_create_prepared_deal. Use the two-step prepare/submit form when you
    want to inspect the resolved artifact before committing.

    DSP/buyer inputs accept dsp_name+buyer_name, dsp_id+buyer_id, or
    dsp_name+seat_id (the resolver looks up the buyer mapped to the
    seat under the DSP).

    auction_type defaults to 1 (First Price) when no floor is requested.
    Pass `floor_ecpm` to set a custom deal-level floor; the resolver will
    automatically upgrade auction_type to 3 (Fixed Price) so PubMatic
    accepts the flooreCPM field.
    deal_source defaults to 8 (DataProvider); allowed values are 5 (DSP),
    7 (Buyer), 8 (DataProvider).

    channel accepts "display", "olv", or "ctv". When device_types is
    omitted, the canonical PubMatic device set for that channel is
    auto-applied.

    fee accepts a single dealFees entry dict or a list of them. When
    omitted, a flat 30% Percent-of-Media curator fee for Elcano is
    auto-applied.
    """
    logger.info("pm_execute_deal_from_prompt_inputs called with name: %s file: %s", name, domain_file_path)

    preparation = await pm_prepare_deal_from_prompt_inputs(
        name=name,
        start_date=start_date,
        end_date=end_date,
        dsp_name=dsp_name,
        buyer_name=buyer_name,
        dsp_id=dsp_id,
        buyer_id=buyer_id,
        logged_in_owner_type_id=logged_in_owner_type_id,
        seat_id=seat_id,
        deal_id=deal_id,
        deal_source=deal_source,
        floor_ecpm=floor_ecpm,
        auction_type=auction_type,
        publisher_names=publisher_names,
        publisher_ids=publisher_ids,
        segment_names=segment_names,
        domain_file_path=domain_file_path,
        domain_sheet=domain_sheet,
        domain_column=domain_column,
        domain_match_operator=domain_match_operator,
        app_bundle_file_path=app_bundle_file_path,
        app_bundle_sheet=app_bundle_sheet,
        app_bundle_column=app_bundle_column,
        app_bundle_match_operator=app_bundle_match_operator,
        geo_countries=geo_countries,
        geo_states=geo_states,
        device_types=device_types,
        iab_categories=iab_categories,
        viewability_threshold=viewability_threshold,
        ad_formats=ad_formats,
        platforms=platforms,
        priority=priority,
        has_max_reach=has_max_reach,
        publisher_block_list=publisher_block_list,
        max_allowed_publishers=max_allowed_publishers,
        auto_update_publishers=auto_update_publishers,
        channel=channel,
        fee=fee,
    )

    if not preparation.get("ready_to_create"):
        blocking_issues = preparation.get("blocking_issues") or []
        error_message = (
            blocking_issues[0]
            if blocking_issues
            else (preparation.get("error") or "Preparation did not produce a creatable artifact.")
        )
        return {
            "success": False,
            "phase": "prepare",
            "deal_url": None,
            "deal": None,
            "targeting_id": None,
            "warnings": preparation.get("warnings", []),
            "quality_flags": list(preparation.get("quality_flags", [])),
            "error": error_message,
            "verification": None,
            "preparation": preparation,
        }

    creation = await pm_create_prepared_deal(preparation["prepared_deal_id"])
    combined_warnings = _dedupe_preserving_order(
        list(preparation.get("warnings", [])) + list(creation.get("warnings", []))
    )
    # pm_create_prepared_deal already seeds its quality_flags from the
    # prepared artifact, so concatenating preparation's flags here would
    # double them. Take creation's list as the canonical merged result.
    combined_quality_flags = list(creation.get("quality_flags", []))
    verification_result = creation.get("verification") or {}
    return {
        "success": creation.get("success", False),
        "phase": "verify" if verification_result.get("success") else "create",
        "deal_url": creation.get("deal_url"),
        "deal": creation.get("deal"),
        "targeting_id": creation.get("targeting_id"),
        "warnings": combined_warnings,
        "quality_flags": combined_quality_flags,
        "error": creation.get("error"),
        "verification": creation.get("verification"),
        "preparation": preparation,
        "creation": creation,
    }


if __name__ == "__main__":
    logger.info("Starting PubMatic MCP Server")
    if not os.environ.get("PUBMATIC_ACCESS_TOKEN") and (
        not os.environ.get("PUBMATIC_USERNAME") or not os.environ.get("PUBMATIC_PASSWORD")
    ):
        logger.warning("PUBMATIC_ACCESS_TOKEN or PUBMATIC_USERNAME/PUBMATIC_PASSWORD are not configured")
    try:
        mcp.run(transport="stdio")
    except Exception as e:
        logger.error("Failed to start server: %s", e)
        sys.exit(1)
