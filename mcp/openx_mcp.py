#!/usr/bin/env python3
"""
OpenX MCP Server

A Model Context Protocol (MCP) server for programmatic deal creation on OpenX.
This is a dedicated MCP for the OpenX Select GraphQL API.

Runs within the Victoria Terminal container environment.
"""

import json
import logging
import os
import re
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Literal
from uuid import uuid4

import httpx
from mcp.server.fastmcp import FastMCP

_HTTP_RE = re.compile(r"^https?://")
_NON_ALNUM_RE = re.compile(r"[^a-z0-9]+")


# Configure logging to stderr (not stdout for STDIO transport)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    stream=sys.stderr,
)
logger = logging.getLogger(__name__)

# Initialize FastMCP server
mcp = FastMCP("openx_mcp")

# Constants
USER_AGENT = "victoria-terminal/1.0"
DEFAULT_TIMEOUT = 60.0

# Regex pattern for valid domain
# - No protocol prefix
# - Each label: starts/ends with alphanumeric, can contain hyphens
# - At least one dot separating labels
DOMAIN_PATTERN = re.compile(
    r"^(?!-)"  # Cannot start with hyphen
    r"(?:[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?\.)"  # Labels with dots
    r"+[a-zA-Z]{2,}$"  # TLD (at least 2 chars)
)

# App bundle identifiers accepted by OpenX url_targeting alongside web FQDNs.
# Verified against UI-created deals: OpenX's url_targeting.urls accepts a mix of
# web domains (iwastesomuchtime.com), Android-style reverse-DNS package names
# (com.app.real.flashlight, net.uploss.water_app, gridplus.nut.sort.puz3d), and
# bare numeric Apple App Store IDs (1589762792). Earlier MCP versions dropped
# the latter two shapes as "invalid domains"; the UI did not.
APP_BUNDLE_PATTERN = re.compile(
    r"^[A-Za-z0-9_]+(?:\.[A-Za-z0-9_]+)+$"  # Reverse-DNS package names: letters/digits/underscores
)
APP_STORE_NUMERIC_ID_PATTERN = re.compile(r"^\d{6,}$")  # Apple App Store IDs (6+ digits)


def _is_acceptable_url_targeting_value(value: str) -> bool:
    """Return True if value is a shape OpenX url_targeting.urls accepts."""
    return bool(
        DOMAIN_PATTERN.match(value) or APP_BUNDLE_PATTERN.match(value) or APP_STORE_NUMERIC_ID_PATTERN.match(value)
    )


# OpenX API Configuration
# Production endpoint per OpenX documentation: https://api.openx.com/oa/graphql
OPENX_GRAPHQL_ENDPOINT = "https://api.openx.com/oa/graphql"


def _build_openx_deal_url(deal: dict[str, Any]) -> str | None:
    """Build an operator-facing OpenX deal URL from the public Deal.id.

    OpenXSelect deal details pages accept the public Deal.id returned by dealCreate and
    dealById, so the MCP can build a direct clickable URL without relying on private
    exchange identifiers.
    """
    template = os.environ.get("OPENX_DEAL_URL_TEMPLATE", "").strip()
    deal_id = str(deal.get("deal_id") or "").strip()
    internal_id = str(deal.get("id") or "").strip()

    if template:
        return template.format(id=internal_id, deal_id=deal_id)

    if internal_id:
        # OpenX dropped the /v2 prefix from OpenXSelect console routes
        # (mid-2026): /v2/deals/{id}/details now dead-ends while
        # /deals/{id}/details resolves. Confirmed live by Elyse on deal
        # 564835301, 2026-06-11.
        return f"https://select.openx.com/deals/{internal_id}/details"

    return None


# PMP Deal Type mapping: human-readable names to numeric codes per OpenRTB 2.5 Section 7.3
# The OpenX API accepts pmp_deal_type as a String!, using numeric codes.
PMP_DEAL_TYPE_MAP: dict[str, str] = {
    "PREFERRED_DEAL": "3",
    "PRIVATE_AUCTION": "2",
    "PROGRAMMATIC_GUARANTEED": "1",
}

OPENX_OPTIONS_PATHS: dict[str, str] = {
    "demand_partner": "deal.deal_participants.demand_partner",
    "buyer_ids": "deal.deal_participants.buyer_ids",
    "fee_partner": "deal.third_party_fees_config.partner_id",
    "language": "deal.package.targeting.technographic.language",
    "iab_categories": "deal.package.targeting.domain.categories_iab_v2",
    # Legacy IAB v1 taxonomy: OpenX's UI emits both v1 (`domain.categories`) and v2
    # (`domain.categories_iab_v2`) in parallel joined with inter_dimension_operator=OR
    # so a deal matches inventory tagged under either taxonomy. Mirrored from
    # UI-created Reklaim Marriott Travel deals where v2 "Travel" expanded to v1
    # IDs {28, 2030, 2804, 2816, 2819, 2824, 2825, 2841}.
    "iab_categories_legacy": "deal.package.targeting.domain.categories",
    "inventory_categories": "deal.package.targeting.metacategory.includes",
    "audience": "deal.package.targeting.audience.openaudience_custom",
    "state": "deal.package.targeting.geographic.state",
    "country": "deal.package.targeting.geographic.country",
    "continent": "deal.package.targeting.geographic.continent",
    "region": "deal.package.targeting.geographic.region",
    "city": "deal.package.targeting.geographic.city",
    "dma": "deal.package.targeting.geographic.dma",
    "msa": "deal.package.targeting.geographic.msa",
    "postal_code": "deal.package.targeting.geographic.postal_code",
}

TTD_MAIN_BUYER_DEMAND_PARTNER_IDS = {"537073292"}

VERIFIED_GEOGRAPHIC_IDS: dict[str, dict[str, str]] = {
    "state": {
        "AZ": "3586",
        "CA": "3588",
        "FL": "3593",
        "HI": "3595",
        "NV": "3610",
        "NM": "3613",
        "TX": "3625",
        "ARIZONA": "3586",
        "CALIFORNIA": "3588",
        "FLORIDA": "3593",
        "HAWAII": "3595",
        "NEVADA": "3610",
        "NEWMEXICO": "3613",
        "TEXAS": "3625",
    }
}

COUNTRY_NAME_ALIASES: dict[str, str] = {
    "US": "united states",
    "USA": "united states",
    "UNITEDSTATES": "united states",
    "CA": "canada",
    "CAN": "canada",
    "CANADA": "canada",
}

GEOGRAPHIC_FILTER_KEYS: dict[str, str] = {
    "country": "country",
    "continent": "continent",
    "region": "region",
    "state": "state",
    "city": "city",
    "dma": "dma",
    "msa": "msa",
    "postal_code": "postal_code",
}

US_STATE_NAMES: dict[str, str] = {
    "AL": "Alabama",
    "AK": "Alaska",
    "AZ": "Arizona",
    "AR": "Arkansas",
    "CA": "California",
    "CO": "Colorado",
    "CT": "Connecticut",
    "DE": "Delaware",
    "FL": "Florida",
    "GA": "Georgia",
    "HI": "Hawaii",
    "ID": "Idaho",
    "IL": "Illinois",
    "IN": "Indiana",
    "IA": "Iowa",
    "KS": "Kansas",
    "KY": "Kentucky",
    "LA": "Louisiana",
    "ME": "Maine",
    "MD": "Maryland",
    "MA": "Massachusetts",
    "MI": "Michigan",
    "MN": "Minnesota",
    "MS": "Mississippi",
    "MO": "Missouri",
    "MT": "Montana",
    "NE": "Nebraska",
    "NV": "Nevada",
    "NH": "New Hampshire",
    "NJ": "New Jersey",
    "NM": "New Mexico",
    "NY": "New York",
    "NC": "North Carolina",
    "ND": "North Dakota",
    "OH": "Ohio",
    "OK": "Oklahoma",
    "OR": "Oregon",
    "PA": "Pennsylvania",
    "RI": "Rhode Island",
    "SC": "South Carolina",
    "SD": "South Dakota",
    "TN": "Tennessee",
    "TX": "Texas",
    "UT": "Utah",
    "VT": "Vermont",
    "VA": "Virginia",
    "WA": "Washington",
    "WV": "West Virginia",
    "WI": "Wisconsin",
    "WY": "Wyoming",
    "DC": "District of Columbia",
}

CANADA_PROVINCE_NAMES: dict[str, str] = {
    "AB": "Alberta",
    "BC": "British Columbia",
    "MB": "Manitoba",
    "NB": "New Brunswick",
    "NL": "Newfoundland and Labrador",
    "NS": "Nova Scotia",
    "NT": "Northwest Territories",
    "NU": "Nunavut",
    "ON": "Ontario",
    "PE": "Prince Edward Island",
    "QC": "Quebec",
    "SK": "Saskatchewan",
    "YT": "Yukon",
}

# Verified OpenX US state ids from live geoSearch responses.
OPENX_VERIFIED_STATE_IDS: dict[str, str] = {
    "AZ": "3586",
    "CA": "3588",
    "FL": "3593",
    "NV": "3610",
    "NM": "3613",
    "TX": "3625",
}

# Canonical rendering-context defaults per channel. Mirrors the Elcano trader
# spec used across every SSP MCP:
#
#   Channel  | Format  | Distribution | Devices                           | Notes
#   ---------|---------|--------------|-----------------------------------|----------------------
#   DISPLAY  | BANNER  | WEB+APP      | DESKTOP+MOBILE+TABLET             | Banner everywhere
#   OLV      | VIDEO   | WEB+APP      | DESKTOP+MOBILE+TABLET             | Video on web/app
#   CTV      | CTV     | APP          | CTV + SET_TOP_BOX                 | App-only, TV devices
#   OTT      | VIDEO   | APP          | DESKTOP+MOBILE+TABLET             | App-only mobile video
#
# CTV is always in-app even if a brief asks for "all inventory" — `distribution_channel="APP"`
# is non-negotiable for that channel. OTT mirrors OLV's device set but forces app-only
# inventory; rare in practice but the trader-side definition for the in-app video case.
DEFAULT_RENDERING_CONTEXTS: dict[str, dict[str, Any]] = {
    "DISPLAY": {
        "op": "AND",
        "ad_placement": {"op": "==", "val": "BANNER"},
        "distribution_channel": {"op": "INTERSECTS", "val": "WEB,APP"},
        "device_types": ["DESKTOP", "MOBILE", "TABLET"],
    },
    "OLV": {
        "op": "AND",
        "ad_placement": {"op": "==", "val": "VIDEO"},
        "distribution_channel": {"op": "INTERSECTS", "val": "WEB,APP"},
        "device_types": ["DESKTOP", "MOBILE", "TABLET"],
    },
    "CTV": {
        "op": "AND",
        "ad_placement": {"op": "==", "val": "CTV"},
        "distribution_channel": {"op": "INTERSECTS", "val": "APP"},
        "device_types": ["CTV", "SET_TOP_BOX"],
    },
    "OTT": {
        "op": "AND",
        "ad_placement": {"op": "==", "val": "VIDEO"},
        "distribution_channel": {"op": "INTERSECTS", "val": "APP"},
        "device_types": ["DESKTOP", "MOBILE", "TABLET"],
    },
}

# Device type V2 groups: maps business-facing device names to the V2 sub-field and value.
# Values must match what optionsByPath returns for each sub-path:
#   deal.package.targeting.rendering_context.device_type.desktop_devices
#   deal.package.targeting.rendering_context.device_type.mobile_devices
#   deal.package.targeting.rendering_context.device_type.tv_devices
DEVICE_TYPE_V2_GROUPS: dict[str, tuple[str, str]] = {
    "DESKTOP": ("desktop_devices", "desktop"),
    "MOBILE": ("mobile_devices", "phone"),
    "PHONE": ("mobile_devices", "phone"),
    "TABLET": ("mobile_devices", "tablet"),
    "FEATURE_PHONE": ("mobile_devices", "feature_phone"),
    "CTV": ("tv_devices", "tv"),
    "CONNECTED_DEVICE": ("tv_devices", "tv"),
    "SET_TOP_BOX": ("tv_devices", "set-top-box"),
}


def _format_decimal_string(value: Any) -> str:
    """Serialize Decimal-compatible values as strings for GraphQL Decimal inputs."""
    return str(value)


def _make_blocker(code: str, message: str, **details: Any) -> dict[str, Any]:
    """Build a structured preparation blocker."""
    blocker = {"code": code, "message": message}
    if details:
        blocker["details"] = details
    return blocker


def _normalize_lookup_text(value: Any) -> str:
    """Normalize ids and names for exact matching without punctuation sensitivity."""
    return _NON_ALNUM_RE.sub("", str(value).strip().lower())


def _normalize_domain_candidate(value: Any) -> str | None:
    """Normalize a raw value into a bare domain if possible."""
    raw_value = str(value or "").strip().lower()
    if not raw_value:
        return None

    normalized_value = _HTTP_RE.sub("", raw_value)
    normalized_value = normalized_value.split("/")[0].split("?")[0].split(":")[0].strip()
    normalized_value = normalized_value.lstrip(".")
    if normalized_value.startswith("www."):
        normalized_value = normalized_value[4:]
    return normalized_value or None


def _serialize_percentage_fraction(value: Any) -> str:
    """Convert user-friendly percentages like 40 into OpenX decimal fractions like 0.4."""
    numeric_value = float(value)
    if numeric_value < 0:
        raise ValueError("Percentage values must be non-negative")
    if numeric_value > 1:
        numeric_value = numeric_value / 100.0
    return format(numeric_value, ".6f").rstrip("0").rstrip(".") or "0"


# Default OpenX curator fee. Applied automatically when the caller supplies no
# `fee` argument so every curated deal carries the curator share without operator
# follow-up in the OpenX UI. Override per-deal by passing an explicit `fee=` argument.
#
# The 30% PoM fee STRUCTURE is the business-level default for every Cutlass curator
# deal regardless of seat. The fee PARTNER name, however, is account-specific:
# each OpenX seat's curator fee flows to a partner record named after the seat
# itself (Reklaim seat → "Reklaim" partner, Elcano seat → "Elcano" partner). The
# mcp loader passes the variant client name through as MCP_VARIANT_CLIENT, which
# we use as the default partner name. The OpenX resolver's case-insensitive
# normalization handles the lowercase-from-loader → "Reklaim" mapping automatically.
# Empty MCP_VARIANT_CLIENT (the default Elcano variant) falls back to "Elcano".
#
# (The A/B/C/D00/B00 attribution code in the deal name is internal Elcano rev-share
# accounting between Elcano and partners — it does NOT determine the SSP curator fee.)
_MCP_VARIANT_CLIENT = os.environ.get("MCP_VARIANT_CLIENT", "").strip()
DEFAULT_FEE_PARTNER_NAME = _MCP_VARIANT_CLIENT or "Elcano"
DEFAULT_FEE_REVENUE_METHOD = "PoM"
DEFAULT_FEE_GROSS_SHARE_PERCENT = 30.0

# OpenX's own platform-share partner id. UI-created partner-sourced deals attach this
# alongside the third-party partner fee (visible in dealCreate responses as the second
# entry in third_party_fees with fee_type="0"). Override via OPENX_PLATFORM_PARTNER_ID
# if OpenX migrates the platform partner record to a new id.
OPENX_PLATFORM_PARTNER_ID = os.environ.get("OPENX_PLATFORM_PARTNER_ID", "").strip() or "540278980"


def _default_curator_fee() -> tuple[dict[str, Any], str]:
    """Return the standard curator fee for this MCP variant plus a warning describing what was applied."""
    derived_fee = {
        "partner_name_or_id": DEFAULT_FEE_PARTNER_NAME,
        "revenue_method": DEFAULT_FEE_REVENUE_METHOD,
        "gross_share_percent": DEFAULT_FEE_GROSS_SHARE_PERCENT,
    }
    warning = (
        f"Applied default OpenX curator fee ({DEFAULT_FEE_PARTNER_NAME}): "
        f"{DEFAULT_FEE_REVENUE_METHOD} {DEFAULT_FEE_GROSS_SHARE_PERCENT:g}%. Pass fee= to override."
    )
    return derived_fee, warning


def _default_package_name(deal_name: str) -> str:
    """Generate a unique package name in OpenX UI's default format.

    OpenX's UI auto-creates packages with the pattern
    `Package for {deal: <name>} [<ISO timestamp .millis Z>]`. Mirror that
    pattern so MCP-created deals look like UI-created deals and — more
    importantly — so a failed dealCreate (which leaves an orphan package)
    never blocks a subsequent retry on package-name uniqueness.
    """
    now = datetime.now(UTC)
    timestamp = now.strftime("%Y-%m-%dT%H:%M:%S")
    millis = now.microsecond // 1000
    return f"Package for {{deal: {deal_name}}} [{timestamp}.{millis:03d}Z]"


def _make_ox_quality_flag(flag: str, impact: str, **context: Any) -> dict[str, Any]:
    """Build a structured quality_flags entry. Mirrors the IX/PM/MN/Xandr pattern."""
    entry: dict[str, Any] = {"flag": flag, "impact": impact}
    for key, value in context.items():
        if value is not None:
            entry[key] = value
    return entry


def _blockers_to_ox_quality_flags(blockers: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Surface existing blockers as structured quality_flags.

    Each blocker already carries a `code` and `message`; map them to the
    `flag`/`impact` shape so the protocol's `quality_flags_contract` has data
    to render. The `details` blob is splatted into the flag's context fields.
    """
    quality_flags: list[dict[str, Any]] = []
    for blocker in blockers:
        if not isinstance(blocker, dict):
            continue
        code = blocker.get("code") or "ox_blocker"
        message = blocker.get("message") or ""
        details = blocker.get("details") or {}
        flag_name = code if code.startswith("ox_") else f"ox_{code}"
        quality_flags.append(_make_ox_quality_flag(flag_name, message, **details))
    return quality_flags


def _extract_allowed_buyer_ids(options: list[dict[str, Any]]) -> list[str]:
    """Extract valid third-party DSP seat ids from optionsByPath results.

    OpenX's deal_participants.buyer_ids field expects the **third-party DSP seat id**
    (the value displayed as "Spark Foundry > 849135" in the UI), NOT the OpenX-internal
    buyer record id from the options catalog. Earlier versions of this helper also
    accepted the buyer record id, which OpenX silently rejected at deal-create time
    with a misleading "Seat <record_id> does not exist" error.

    Only `extra.third_party_buyer_id[*]` is valid for buyer_ids.
    """
    buyer_ids: list[str] = []
    for option in options:
        extra = option.get("extra")
        if isinstance(extra, dict) and isinstance(extra.get("third_party_buyer_id"), list):
            for buyer_id in extra["third_party_buyer_id"]:
                buyer_id_str = str(buyer_id)
                if buyer_id_str not in buyer_ids:
                    buyer_ids.append(buyer_id_str)
    return buyer_ids


def _resolve_buyer_seat_to_third_party_ids(
    options: list[dict[str, Any]], requested_ids: list[str]
) -> tuple[list[str], list[str]]:
    """Map caller-supplied buyer ids to valid third-party DSP seat ids.

    Accepts either:
      - The third-party DSP seat id directly (e.g. "849135") — pass through unchanged.
      - The OpenX-internal buyer record id (e.g. "562224857") — substitute with the
        record's third_party_buyer_id values (one record may expose multiple seats).

    Returns ``(resolved_seat_ids, warnings)``. Unrecognized ids are passed through
    unchanged so the GraphQL error path retains its detail.
    """
    resolved: list[str] = []
    warnings: list[str] = []
    by_record_id: dict[str, list[str]] = {}
    valid_seat_ids: set[str] = set()
    for option in options:
        extra = option.get("extra")
        if not isinstance(extra, dict):
            continue
        record_id = str(option.get("id") or "")
        seat_ids = [str(seat) for seat in (extra.get("third_party_buyer_id") or [])]
        if record_id:
            by_record_id[record_id] = seat_ids
        valid_seat_ids.update(seat_ids)

    for raw in requested_ids:
        raw_str = str(raw)
        if raw_str in valid_seat_ids:
            if raw_str not in resolved:
                resolved.append(raw_str)
            continue
        substitute_seats = by_record_id.get(raw_str)
        if substitute_seats:
            warnings.append(
                f"buyer_ids value '{raw_str}' looks like an OpenX-internal buyer record id; "
                f"substituted the record's third-party DSP seat id(s) {substitute_seats}. "
                "OpenX deal_participants.buyer_ids expects the DSP seat id (third_party_buyer_id), "
                "not the buyer record id."
            )
            for seat in substitute_seats:
                if seat not in resolved:
                    resolved.append(seat)
            continue
        # Unrecognized — let it through so OpenX returns its own error.
        if raw_str not in resolved:
            resolved.append(raw_str)
    return resolved, warnings


def _should_set_main_buyer_id(demand_partner_id: str | None, buyer_ids: list[str] | None) -> bool:
    """Return whether the participant should include OpenX main_buyer_id.

    OpenX UI-created The Trade Desk - RTB deals persist the seat both as the
    participant main_buyer_id and in buyer_ids. Mirror that shape when a single
    TTD buyer seat is supplied so automated creates match UI behavior.
    """
    return bool(demand_partner_id in TTD_MAIN_BUYER_DEMAND_PARTNER_IDS and buyer_ids and len(buyer_ids) == 1)


def _choose_unique_option(
    options: list[dict[str, Any]],
    requested_value: str,
    field_label: str,
    *,
    aliases: dict[str, str] | None = None,
    allow_contains_match: bool = False,
) -> dict[str, Any]:
    """Resolve a single OpenX option by exact id/name, optional aliases, then optional unique contains match."""
    normalized_requested = _normalize_lookup_text(requested_value)
    if aliases and normalized_requested in aliases:
        requested_value = aliases[normalized_requested]
        normalized_requested = _normalize_lookup_text(requested_value)

    id_matches = [option for option in options if _normalize_lookup_text(option.get("id", "")) == normalized_requested]
    if len(id_matches) == 1:
        return id_matches[0]
    if len(id_matches) > 1:
        raise ValueError(f"Ambiguous {field_label}: {requested_value}")

    name_matches = [
        option for option in options if _normalize_lookup_text(option.get("name", "")) == normalized_requested
    ]
    if len(name_matches) == 1:
        return name_matches[0]
    if len(name_matches) > 1:
        raise ValueError(f"Ambiguous {field_label}: {requested_value}")

    if allow_contains_match:
        contains_matches = [
            option
            for option in options
            if normalized_requested in _normalize_lookup_text(option.get("name", ""))
            or _normalize_lookup_text(option.get("name", "")) in normalized_requested
        ]
        if len(contains_matches) == 1:
            return contains_matches[0]
        if len(contains_matches) > 1:
            raise ValueError(f"Ambiguous {field_label}: {requested_value}")

    raise LookupError(f"No OpenX match found for {field_label}: {requested_value}")


def _parse_demand_partner_reference(
    demand_partner: str, buyer_ids: list[str] | None
) -> tuple[str, list[str] | None, list[str]]:
    """Support operator-friendly partner aliases like 'bidswitch / 393'."""
    warnings: list[str] = []
    normalized_buyer_ids = [str(buyer_id).strip() for buyer_id in (buyer_ids or []) if str(buyer_id).strip()] or None
    demand_partner_reference = str(demand_partner).strip()

    if "/" in demand_partner_reference:
        left, right = [part.strip() for part in demand_partner_reference.split("/", 1)]
        if left and right and right.isdigit():
            demand_partner_reference = left
            if not normalized_buyer_ids:
                normalized_buyer_ids = [right]
                warnings.append(f"Extracted buyer_id '{right}' from demand partner reference '{demand_partner}'.")

    alias_map = {
        "tradr": "Bidswitch - RTB",
        "bidswitch": "Bidswitch - RTB",
        "bidswitchrtb": "Bidswitch - RTB",
    }
    alias_key = _normalize_lookup_text(demand_partner_reference)
    if alias_key in alias_map:
        resolved_alias = alias_map[alias_key]
        if resolved_alias != demand_partner_reference:
            warnings.append(f"Mapped demand partner alias '{demand_partner_reference}' to '{resolved_alias}'.")
        demand_partner_reference = resolved_alias

    return demand_partner_reference, normalized_buyer_ids, warnings


def _extract_domains_from_xlsx(
    file_path: str,
    sheet_name: str | None = None,
    column_name: str | None = None,
) -> dict[str, Any]:
    """Extract, normalize, and deduplicate domains from an Excel workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx domain files.") from exc

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found in workbook: {sheet_name}")
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Workbook sheet is empty.")

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized_headers = {_normalize_lookup_text(header): index for index, header in enumerate(header_row) if header}

    if column_name:
        requested_column_key = _normalize_lookup_text(column_name)
        if requested_column_key not in normalized_headers:
            raise ValueError(f"Column not found in workbook: {column_name}")
        domain_column_index = normalized_headers[requested_column_key]
    else:
        preferred_columns = ["sites", "domains", "domain", "url", "urls"]
        domain_column_index = next(
            (normalized_headers[column_key] for column_key in preferred_columns if column_key in normalized_headers),
            0,
        )

    extracted_domains: list[str] = []
    invalid_values: list[str] = []
    for row in rows[1:]:
        if domain_column_index >= len(row):
            continue
        normalized_domain = _normalize_domain_candidate(row[domain_column_index])
        if normalized_domain is None:
            continue
        if _is_acceptable_url_targeting_value(normalized_domain):
            extracted_domains.append(normalized_domain)
        else:
            invalid_values.append(str(row[domain_column_index]).strip())

    return {
        "sheet_name": worksheet.title,
        "column_name": header_row[domain_column_index] if header_row else None,
        "row_count": max(len(rows) - 1, 0),
        "domains": sorted(dict.fromkeys(extracted_domains)),
        "invalid_values": invalid_values,
    }


def _extract_domains_from_csv(file_path: str, column_name: str | None = None) -> dict[str, Any]:
    """Extract, normalize, and deduplicate domains from a CSV file."""
    import csv

    with open(file_path, newline="", encoding="utf-8") as csv_file:
        reader = csv.reader(csv_file)
        rows = list(reader)

    if not rows:
        raise ValueError("CSV file is empty.")

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized_headers = {_normalize_lookup_text(header): index for index, header in enumerate(header_row) if header}

    if column_name:
        requested_column_key = _normalize_lookup_text(column_name)
        if requested_column_key not in normalized_headers:
            raise ValueError(f"Column not found in CSV: {column_name}")
        domain_column_index = normalized_headers[requested_column_key]
    else:
        preferred_columns = ["sites", "domains", "domain", "url", "urls"]
        domain_column_index = next(
            (normalized_headers[column_key] for column_key in preferred_columns if column_key in normalized_headers),
            0,
        )

    extracted_domains: list[str] = []
    invalid_values: list[str] = []
    for row in rows[1:]:
        if domain_column_index >= len(row):
            continue
        normalized_domain = _normalize_domain_candidate(row[domain_column_index])
        if normalized_domain is None:
            continue
        if _is_acceptable_url_targeting_value(normalized_domain):
            extracted_domains.append(normalized_domain)
        else:
            invalid_values.append(str(row[domain_column_index]).strip())

    return {
        "sheet_name": None,
        "column_name": header_row[domain_column_index] if header_row else None,
        "row_count": max(len(rows) - 1, 0),
        "domains": sorted(dict.fromkeys(extracted_domains)),
        "invalid_values": invalid_values,
    }


def _extract_domains_from_file(
    file_path: str,
    sheet_name: str | None = None,
    column_name: str | None = None,
) -> dict[str, Any]:
    """Extract domains from a supported attachment file format."""
    suffix = Path(file_path).suffix.lower()
    if suffix == ".csv":
        return _extract_domains_from_csv(file_path, column_name)
    return _extract_domains_from_xlsx(file_path, sheet_name, column_name)


# App-bundle file column headers recognized when no explicit column is given.
# Keys are normalized via _normalize_lookup_text (lowercased, punctuation stripped),
# so "App Bundle IDs", "app_bundle_id", and "Bundle" all match.
APP_BUNDLE_PREFERRED_COLUMNS = (
    "appbundleids",
    "appbundleid",
    "appbundles",
    "appbundle",
    "bundleids",
    "bundleid",
    "bundles",
    "bundle",
    "appstoreids",
    "appids",
    "appid",
)


def _coerce_bundle_cell(value: Any) -> str:
    """Coerce a spreadsheet cell into a clean app-bundle identifier string.

    openpyxl returns numeric-typed cells as int/float. A 13-digit Apple App Store ID
    entered in Excel is stored as a float, so a naive str() yields "3.2024e+12" or a
    trailing ".0" — exactly the corruption seen in hand-pasted UI bundle lists. Render
    integral numerics as plain integers; fall back to a stripped string otherwise.
    """
    if isinstance(value, bool):
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else repr(value)
    return str(value).strip()


def _extract_app_bundles_from_file(
    file_path: str,
    sheet_name: str | None = None,
    column_name: str | None = None,
) -> dict[str, Any]:
    """Extract app-bundle identifiers from a CSV/XLSX file.

    Unlike domain extraction, bundle values are NOT normalized to bare domains or
    filtered against url_targeting shapes. OpenX ``targeting.app_inventory.app_bundle_id``
    accepts reverse-DNS package names (``com.fubotv.vix``), bare numeric Apple App Store
    IDs (``162057``), Amazon ASINs (``B072QYQ43R``), Roku channel IDs (``G20196015164``),
    and more. Case is preserved (bundle ids are case-sensitive); values are de-duplicated
    preserving first-seen order.
    """
    suffix = Path(file_path).suffix.lower()
    if suffix == ".csv":
        import csv

        with open(file_path, newline="", encoding="utf-8") as csv_file:
            rows: list[list[Any]] = [list(row) for row in csv.reader(csv_file)]
        worksheet_title = None
    else:
        from openpyxl import load_workbook

        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet not found in workbook: {sheet_name}")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[workbook.sheetnames[0]]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        worksheet_title = worksheet.title

    if not rows:
        raise ValueError("App-bundle file is empty.")

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized_headers = {_normalize_lookup_text(header): index for index, header in enumerate(header_row) if header}

    if column_name:
        requested_column_key = _normalize_lookup_text(column_name)
        if requested_column_key not in normalized_headers:
            raise ValueError(f"Column not found in app-bundle file: {column_name}")
        bundle_column_index = normalized_headers[requested_column_key]
    else:
        bundle_column_index = next(
            (normalized_headers[key] for key in APP_BUNDLE_PREFERRED_COLUMNS if key in normalized_headers),
            0,
        )

    bundles: list[str] = []
    seen: set[str] = set()
    for row in rows[1:]:
        if bundle_column_index >= len(row):
            continue
        cell = row[bundle_column_index]
        if cell is None:
            continue
        value = _coerce_bundle_cell(cell)
        if not value or value.lower() == "nan":
            continue
        if value not in seen:
            seen.add(value)
            bundles.append(value)

    return {
        "sheet_name": worksheet_title,
        "column_name": header_row[bundle_column_index] if header_row else None,
        "row_count": max(len(rows) - 1, 0),
        "app_bundles": bundles,
    }


def _normalize_publisher_id_candidate(value: Any) -> str | None:
    """Normalize a raw workbook/CSV cell into a publisher id string when possible."""
    if value is None:
        return None

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    publisher_id = str(value).strip()
    if not publisher_id:
        return None

    if re.fullmatch(r"\d+\.0+", publisher_id):
        return publisher_id.split(".", 1)[0]

    return publisher_id


def _extract_publisher_ids_from_xlsx(
    file_path: str,
    sheet_name: str | None = None,
    column_name: str | None = None,
) -> dict[str, Any]:
    """Extract, normalize, and deduplicate publisher ids from an Excel workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx publisher files.") from exc

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found in workbook: {sheet_name}")
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Workbook sheet is empty.")

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized_headers = {_normalize_lookup_text(header): index for index, header in enumerate(header_row) if header}

    if column_name:
        requested_column_key = _normalize_lookup_text(column_name)
        if requested_column_key not in normalized_headers:
            raise ValueError(f"Column not found in workbook: {column_name}")
        publisher_column_index = normalized_headers[requested_column_key]
    else:
        preferred_columns = [
            "publisherids",
            "publisherid",
            "publisher",
            "publisheraccount",
            "accountids",
            "accountid",
            "accounts",
            "account",
            "ids",
            "id",
        ]
        publisher_column_index = next(
            (normalized_headers[column_key] for column_key in preferred_columns if column_key in normalized_headers),
            0,
        )

    publisher_ids: list[str] = []
    for row in rows[1:]:
        if publisher_column_index >= len(row):
            continue
        publisher_id = _normalize_publisher_id_candidate(row[publisher_column_index])
        if publisher_id and publisher_id not in publisher_ids:
            publisher_ids.append(publisher_id)

    return {
        "sheet_name": worksheet.title,
        "column_name": header_row[publisher_column_index] if header_row else None,
        "row_count": max(len(rows) - 1, 0),
        "publisher_ids": publisher_ids,
    }


def _extract_publisher_ids_from_csv(file_path: str, column_name: str | None = None) -> dict[str, Any]:
    """Extract, normalize, and deduplicate publisher ids from a CSV file."""
    import csv

    with open(file_path, newline="", encoding="utf-8") as csv_file:
        reader = csv.reader(csv_file)
        rows = list(reader)

    if not rows:
        raise ValueError("CSV file is empty.")

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    normalized_headers = {_normalize_lookup_text(header): index for index, header in enumerate(header_row) if header}

    if column_name:
        requested_column_key = _normalize_lookup_text(column_name)
        if requested_column_key not in normalized_headers:
            raise ValueError(f"Column not found in CSV: {column_name}")
        publisher_column_index = normalized_headers[requested_column_key]
    else:
        preferred_columns = [
            "publisherids",
            "publisherid",
            "publisher",
            "publisheraccount",
            "accountids",
            "accountid",
            "accounts",
            "account",
            "ids",
            "id",
        ]
        publisher_column_index = next(
            (normalized_headers[column_key] for column_key in preferred_columns if column_key in normalized_headers),
            0,
        )

    publisher_ids: list[str] = []
    for row in rows[1:]:
        if publisher_column_index >= len(row):
            continue
        publisher_id = _normalize_publisher_id_candidate(row[publisher_column_index])
        if publisher_id and publisher_id not in publisher_ids:
            publisher_ids.append(publisher_id)

    return {
        "sheet_name": None,
        "column_name": header_row[publisher_column_index] if header_row else None,
        "row_count": max(len(rows) - 1, 0),
        "publisher_ids": publisher_ids,
    }


def _extract_publisher_ids_from_file(
    file_path: str,
    sheet_name: str | None = None,
    column_name: str | None = None,
) -> dict[str, Any]:
    """Extract publisher ids from a supported attachment file format."""
    suffix = Path(file_path).suffix.lower()
    if suffix == ".csv":
        return _extract_publisher_ids_from_csv(file_path, column_name)
    return _extract_publisher_ids_from_xlsx(file_path, sheet_name, column_name)


def _autodetect_single_xlsx_attachment() -> tuple[str | None, dict[str, Any] | None]:
    """Auto-detect a single XLSX attachment in the canonical input directory.

    Only ``/input/`` is scanned — the dedicated MOC attachment drop. Previously this also
    walked ``Path.cwd()``, which under MOC's ``/app/`` working directory picked up
    ``/app/protocols/email_styles/deal_sheet_template.xlsx`` (a system template, not a
    trader attachment) and produced a misleading ``Auto-detected domain file: …`` warning
    on every prepare call that didn't ship its own file. Restricting the scope to
    ``/input/`` eliminates the source-tree footgun; local-dev callers should pass
    ``domain_file_path`` explicitly.
    """
    input_root = Path("/input")
    candidates: list[str] = []
    seen: set[str] = set()
    if input_root.exists():
        for match in input_root.rglob("*.xlsx"):
            path_str = str(match.resolve())
            if path_str not in seen:
                seen.add(path_str)
                candidates.append(path_str)

    if len(candidates) == 1:
        return candidates[0], None
    if not candidates:
        return None, _make_blocker(
            "missing_domain_file",
            "No .xlsx domain file was provided and no single attachment could be auto-detected.",
        )
    return None, _make_blocker(
        "ambiguous_domain_file",
        "Multiple .xlsx files were found. Provide domain_file_path explicitly.",
        candidates=candidates,
    )


def _normalize_rendering_context_device_type(device_types: list[str] | dict[str, Any] | None) -> dict[str, Any]:
    """Normalize device types into TargetingDeviceTypeV2CreateParams."""
    if isinstance(device_types, dict):
        normalized_device_type = dict(device_types)
        normalized_device_type.setdefault("op", "INTERSECTS")
        return normalized_device_type

    normalized_device_type: dict[str, Any] = {"op": "INTERSECTS"}
    grouped_values: dict[str, list[str]] = {
        "desktop_devices": [],
        "mobile_devices": [],
        "tv_devices": [],
    }

    for device_type in device_types or []:
        normalized_name = str(device_type).strip().upper()
        if not normalized_name:
            continue

        device_group = DEVICE_TYPE_V2_GROUPS.get(normalized_name)
        if device_group is None:
            raise ValueError(f"Unsupported OpenX device_type for rendering_context: {device_type}")

        field_name, field_value = device_group
        if field_value not in grouped_values[field_name]:
            grouped_values[field_name].append(field_value)

    for field_name, values in grouped_values.items():
        if values:
            normalized_device_type[field_name] = ",".join(values)

    if len(normalized_device_type) == 1:
        normalized_device_type["desktop_devices"] = DEVICE_TYPE_V2_GROUPS["DESKTOP"][1]

    return normalized_device_type


_CTV_DEVICE_NAMES = {"CTV", "SET_TOP_BOX", "SET-TOP BOX", "SETTOPBOX", "CONNECTED_DEVICE", "CONNECTED TV"}


def _infer_channel(targeting: dict[str, Any], rendering_context: dict[str, Any] | None) -> str:
    """Infer the deal channel used to auto-construct rendering_context defaults.

    Resolution order:
      1. explicit `targeting.channel`
      2. explicit `rendering_context.ad_placement == CTV` (strong CTV signal)
      3. explicit `rendering_context.distribution_channel == CTV`
      4. CTV-only `targeting.device_type` list (set-top-box, connected TV)
      5. explicit `rendering_context.ad_placement == VIDEO` → OLV
      6. fallback DISPLAY

    The device-list fallback catches the case where an agent emits
    ``device_type=["CTV"]`` without setting ``channel`` — pre-fix that
    silently fell through to DISPLAY (BANNER format) or, if the agent
    also passed ``ad_placement=VIDEO``, to OLV. Both produced the
    Format=Video-on-TV-devices warning OpenX surfaces on CTV deals.
    """
    explicit_channel = targeting.get("channel")
    if isinstance(explicit_channel, str) and explicit_channel.strip():
        return explicit_channel.strip().upper()

    if rendering_context:
        ad_placement_value = str(rendering_context.get("ad_placement", {}).get("val", "")).upper()
        if ad_placement_value == "CTV":
            return "CTV"

        distribution_value = str(rendering_context.get("distribution_channel", {}).get("val", "")).upper()
        if distribution_value == "CTV":
            return "CTV"

    device_type = targeting.get("device_type")
    if isinstance(device_type, list) and device_type:
        normalized_devices = {str(item).strip().upper() for item in device_type if str(item).strip()}
        if normalized_devices and normalized_devices.issubset(_CTV_DEVICE_NAMES):
            return "CTV"

    if rendering_context:
        ad_placement_value = str(rendering_context.get("ad_placement", {}).get("val", "")).upper()
        if ad_placement_value == "VIDEO":
            return "OLV"

    return "DISPLAY"


def _build_rendering_context(
    targeting: dict[str, Any], device_types: list[str] | dict[str, Any] | None
) -> dict[str, Any]:
    """Build a complete rendering_context object required by OpenX."""
    raw_rendering_context = targeting.get("rendering_context")
    normalized_rendering_context = dict(raw_rendering_context) if isinstance(raw_rendering_context, dict) else {}

    channel = _infer_channel(targeting, normalized_rendering_context)
    channel_defaults = DEFAULT_RENDERING_CONTEXTS.get(channel, DEFAULT_RENDERING_CONTEXTS["DISPLAY"])

    if channel == "CTV":
        # CTV has hard wire-shape requirements: Format MUST be CTV (not VIDEO,
        # not BANNER), Distribution MUST be APP-only. OpenX surfaces a "Video
        # on TV devices and App is a CTV configuration" warning and the deal
        # bids as the wrong format when ad_placement is anything other than
        # CTV on TV-device inventory. Override explicit values rather than
        # let an agent's manual rendering_context produce an invalid deal.
        rendering_context = {
            "op": channel_defaults["op"],
            "ad_placement": channel_defaults["ad_placement"],
            "distribution_channel": channel_defaults["distribution_channel"],
        }
    elif channel == "OTT":
        # OTT shares CTV's APP-only constraint — in-app mobile/connected-device
        # video. Force distribution_channel=APP so a caller that emits
        # WEB,APP (the OLV default) doesn't accidentally include web
        # inventory on a channel that's defined as app-only. ad_placement
        # is configurable so the rare non-VIDEO OTT case still works.
        rendering_context = {
            "op": normalized_rendering_context.get("op", channel_defaults["op"]),
            "ad_placement": normalized_rendering_context.get("ad_placement", channel_defaults["ad_placement"]),
            "distribution_channel": channel_defaults["distribution_channel"],
        }
    else:
        rendering_context = {
            "op": normalized_rendering_context.get("op", channel_defaults["op"]),
            "ad_placement": normalized_rendering_context.get("ad_placement", channel_defaults["ad_placement"]),
            "distribution_channel": normalized_rendering_context.get(
                "distribution_channel", channel_defaults["distribution_channel"]
            ),
        }

    # Device defaults must be channel-aware: an OLV or DISPLAY deal with no explicit
    # device_type should still reach mobile inventory, and a CTV deal should reach TV
    # devices, not desktop. Without this, every channel silently fell back to desktop-only.
    explicit_device_input = normalized_rendering_context.get("device_type", device_types)
    if explicit_device_input is None or (isinstance(explicit_device_input, list) and not explicit_device_input):
        explicit_device_input = list(channel_defaults["device_types"])

    rendering_context["device_type"] = _normalize_rendering_context_device_type(explicit_device_input)
    return rendering_context


def _copy_structured_targeting_branches(targeting_brief: dict[str, Any]) -> dict[str, Any]:
    """Preserve already-structured targeting branches passed by the caller.

    Note: rendering_context is NOT preserved here because the MCP always builds it
    from business-facing fields (channel, device_types). This prevents stale or
    incorrectly-formatted rendering_context from being passed through.
    """
    preserved_targeting: dict[str, Any] = {}
    for key in (
        "attention",
        "content",
        "custom",
        "domain",
        "audience",
        "technographic",
    ):
        value = targeting_brief.get(key)
        if isinstance(value, dict):
            preserved_targeting[key] = dict(value)
        elif value is not None and key == "attention":
            preserved_targeting[key] = value
    return preserved_targeting


def _normalize_geographic_item(
    item: dict[str, Any], *, warn_pre_resolved_state_ids: bool = True
) -> tuple[dict[str, Any], list[str]]:
    """Normalize a geographic includes/excludes item for OpenX create payloads."""
    normalized_item = dict(item)
    warnings: list[str] = []

    country_value = normalized_item.get("country")
    if isinstance(country_value, str):
        normalized_country = ",".join(part.strip().lower() for part in country_value.split(",") if part.strip())
        if normalized_country != country_value:
            warnings.append(f"Normalized geographic country '{country_value}' to '{normalized_country}'.")
        normalized_item["country"] = normalized_country or None

    state_value = normalized_item.get("state")
    has_state = isinstance(state_value, str) and state_value.strip()
    if has_state and warn_pre_resolved_state_ids:
        state_parts = [part.strip() for part in str(state_value).split(",") if part.strip()]
        if state_parts and all(part.isdigit() for part in state_parts):
            warnings.append(
                "Received pre-resolved geographic state ids. The MCP did not resolve abbreviations/names for these values; verify them with ox_list_states to avoid incorrect geography."
            )

    return normalized_item, warnings


def _normalize_geographic_targeting(
    geographic: Any, *, warn_pre_resolved_state_ids: bool = True
) -> tuple[Any, list[str]]:
    """Normalize geographic targeting payloads before OpenX create."""
    if isinstance(geographic, list):
        normalized_countries = [str(value).strip().lower() for value in geographic if str(value).strip()]
        warnings = []
        if normalized_countries != [str(value).strip() for value in geographic if str(value).strip()]:
            warnings.append("Normalized geographic country list to lowercase OpenX values.")
        return normalized_countries, warnings

    if not isinstance(geographic, dict):
        return geographic, []

    normalized_geographic = dict(geographic)
    warnings: list[str] = []
    for key in ("includes", "excludes"):
        value = normalized_geographic.get(key)
        if isinstance(value, dict):
            normalized_value, value_warnings = _normalize_geographic_item(
                value, warn_pre_resolved_state_ids=warn_pre_resolved_state_ids
            )
            normalized_geographic[key] = normalized_value
            warnings.extend(value_warnings)

    return normalized_geographic, warnings


async def _wrap_country_list_as_structured_geo(
    client: "OpenXClient", country_values: list[str]
) -> tuple[dict[str, Any] | list[str], list[str]]:
    """Wrap a flat list of country codes/names into the structured form OpenX expects.

    OpenX's create endpoint silently drops geographic targeting when supplied as a flat
    list — it requires {"includes": {<field_name>: "us,ca,..."}}. Mirrors the wrapping
    logic ox_create_deal already performs so the prep path produces a creatable artifact.
    Returns the original empty list when no values are supplied.
    """
    if not country_values:
        return country_values, []
    geo_field_name = await client.get_geographic_field_name()
    structured = {"includes": {geo_field_name: ",".join(country_values)}}
    return _normalize_geographic_targeting(structured)


def _cleanup_structured_geographic_targeting(geographic: dict[str, Any]) -> dict[str, Any]:
    """Remove lookup-only geographic hints that should not survive into final payloads."""
    cleaned_geographic = dict(geographic)
    includes_branch = cleaned_geographic.get("includes")
    excludes_branch = cleaned_geographic.get("excludes")

    if not isinstance(includes_branch, dict) or not isinstance(excludes_branch, dict):
        return cleaned_geographic

    include_country_hint = includes_branch.get("country")
    exclude_country_hint = excludes_branch.get("country")
    exclude_state = excludes_branch.get("state")
    include_countries = [part.strip() for part in str(include_country_hint or "").split(",") if part.strip()]

    if (
        isinstance(exclude_state, str)
        and exclude_state.strip()
        and len(include_countries) == 1
        and isinstance(exclude_country_hint, str)
        and _normalize_country_hint(exclude_country_hint) == _normalize_country_hint(include_countries[0])
    ):
        cleaned_excludes = dict(excludes_branch)
        cleaned_excludes.pop("country", None)
        cleaned_geographic["excludes"] = cleaned_excludes

    return cleaned_geographic


def _extract_pre_resolved_state_ids(geographic: Any) -> list[str]:
    """Return numeric state ids supplied directly in geographic includes/excludes, if any."""
    if not isinstance(geographic, dict):
        return []

    state_ids: list[str] = []
    for key in ("includes", "excludes"):
        value = geographic.get(key)
        if not isinstance(value, dict):
            continue
        state_value = value.get("state")
        if not isinstance(state_value, str) or not state_value.strip():
            continue
        state_parts = [part.strip() for part in state_value.split(",") if part.strip()]
        if state_parts and all(part.isdigit() for part in state_parts):
            state_ids.extend(part for part in state_parts if part not in state_ids)
    return state_ids


def _normalize_third_party_fees_config(
    third_party_fees_config: dict[str, Any] | list[dict[str, Any]] | None,
) -> list[dict[str, Any]] | None:
    """Normalize fee config shape and serialize Decimal fields as strings."""
    if not third_party_fees_config:
        return None

    raw_fee_configs = (
        [third_party_fees_config] if isinstance(third_party_fees_config, dict) else third_party_fees_config
    )
    normalized_fee_configs = []

    for fee_config in raw_fee_configs:
        normalized_fee_config = dict(fee_config)
        # Strip response-only fields that callers sometimes copy from a
        # dealCreate response back into a new request. `ThirdPartyFeesConfigCreateParams`
        # only accepts `revenue_method`, `gross_share`, `gross_cpm_cap`,
        # `partner_id` — extra fields are rejected by GraphQL coercion
        # (e.g. "Field 'platform_share' is not defined by type ...").
        for response_only_field in ("platform_share", "platform_partner_id"):
            normalized_fee_config.pop(response_only_field, None)
        for decimal_field in ("gross_share", "gross_cpm_cap"):
            if decimal_field in normalized_fee_config and normalized_fee_config[decimal_field] is not None:
                normalized_fee_config[decimal_field] = _format_decimal_string(normalized_fee_config[decimal_field])
        normalized_fee_configs.append(normalized_fee_config)

    return normalized_fee_configs


class OpenXGraphQLError(Exception):
    """Raised when OpenX returns GraphQL errors."""

    def __init__(self, *, operation_name: str | None, errors: list[dict[str, Any]]):
        self.operation_name = operation_name
        self.errors = errors
        super().__init__(self._build_message())

    def _build_message(self) -> str:
        formatted_errors: list[str] = []
        for error in self.errors:
            message = error.get("message", str(error))
            code = error.get("extensions", {}).get("code")
            path = error.get("path")
            locations = error.get("locations")

            details: list[str] = []
            if code:
                details.append(f"code={code}")
            if path:
                details.append(f"path={'.'.join(str(part) for part in path)}")
            if locations:
                rendered_locations = ",".join(
                    f"{location.get('line')}:{location.get('column')}" for location in locations
                )
                details.append(f"locations={rendered_locations}")

            if details:
                formatted_errors.append(f"{message} ({', '.join(details)})")
            else:
                formatted_errors.append(message)

        operation_label = f" during {self.operation_name}" if self.operation_name else ""
        return f"GraphQL errors{operation_label}: {'; '.join(formatted_errors)}"


def _summarize_create_payload(deal_params: dict[str, Any]) -> dict[str, Any]:
    """Return a compact, log-friendly summary of the final dealCreate input."""
    package = deal_params.get("package", {}) if isinstance(deal_params.get("package"), dict) else {}
    targeting = package.get("targeting", {}) if isinstance(package.get("targeting"), dict) else {}
    url_targeting = package.get("url_targeting", {}) if isinstance(package.get("url_targeting"), dict) else {}
    fee_configs = deal_params.get("third_party_fees_config") or []

    summary: dict[str, Any] = {
        "name": deal_params.get("name"),
        "currency": deal_params.get("currency"),
        "deal_price": deal_params.get("deal_price"),
        "pmp_deal_type": deal_params.get("pmp_deal_type"),
        "start_date": deal_params.get("start_date"),
        "end_date": deal_params.get("end_date"),
        "deal_participants": deal_params.get("deal_participants"),
        "package_name": package.get("name"),
        "targeting_keys": sorted(targeting.keys()),
        "has_rendering_context": "rendering_context" in targeting,
        "url_targeting": None,
        "third_party_fees_config": None,
    }

    if url_targeting:
        summary["url_targeting"] = {
            "type": url_targeting.get("type"),
            "urls_count": len(url_targeting.get("urls", [])) if isinstance(url_targeting.get("urls"), list) else None,
            "domain_targeting_option": url_targeting.get("domain_targeting_option"),
        }

    if fee_configs:
        summary["third_party_fees_config"] = [
            {
                "partner_id": fee_config.get("partner_id"),
                "revenue_method": fee_config.get("revenue_method"),
                "gross_share": fee_config.get("gross_share"),
                "gross_cpm_cap": fee_config.get("gross_cpm_cap"),
            }
            for fee_config in fee_configs
            if isinstance(fee_config, dict)
        ]

    return summary


class OpenXClient:
    """
    Client for interacting with the OpenX Select GraphQL API.

    Uses Bearer token authentication with an API key.
    """

    def __init__(self):
        self.api_key = os.environ.get("OPENX_API_KEY", "")
        self.client = httpx.AsyncClient(timeout=DEFAULT_TIMEOUT)
        # Cache for introspected type field names
        self._type_field_cache: dict[str, str | None] = {}

    async def close(self):
        """Close the HTTP client."""
        await self.client.aclose()

    def _is_configured(self) -> bool:
        """Check if OpenX API key is configured."""
        return bool(self.api_key)

    def _get_headers(self) -> dict[str, str]:
        """Get the request headers including authentication."""
        if not self._is_configured():
            raise ValueError("OpenX API key not configured. Set OPENX_API_KEY environment variable.")
        return {
            "x-apikey": self.api_key,
            "Content-Type": "application/json",
            "User-Agent": USER_AGENT,
        }

    async def _graphql_request(
        self,
        query: str,
        variables: dict[str, Any] | None = None,
        operation_name: str | None = None,
    ) -> dict[str, Any]:
        """
        Execute a GraphQL request against the OpenX Select API.

        Args:
            query: The GraphQL query or mutation string
            variables: Optional variables for the query

        Returns:
            The 'data' portion of the GraphQL response

        Raises:
            Exception: If the request fails or returns errors
        """
        headers = self._get_headers()

        payload: dict[str, Any] = {"query": query}
        if variables:
            payload["variables"] = variables

        logger.info("Executing GraphQL request to OpenX")
        logger.debug(f"Query: {query[:200]}...")

        response = await self.client.post(
            OPENX_GRAPHQL_ENDPOINT,
            json=payload,
            headers=headers,
        )

        # Status check FIRST: a 502/503 from a gateway carries an HTML body,
        # and response.json() on it would raise a bare JSONDecodeError that
        # hides the status code from every caller.
        if response.status_code >= 400:
            raise Exception(f"GraphQL HTTP {response.status_code}: {response.text[:2000]}")
        result = response.json()

        # Check for GraphQL errors
        if "errors" in result:
            raise OpenXGraphQLError(operation_name=operation_name, errors=result["errors"])

        return result.get("data", {})

    async def create_deal(self, deal_params: dict[str, Any]) -> dict[str, Any]:
        """
        Create a new deal using the dealCreate mutation.

        Args:
            deal_params: DealCreateParams input object containing:
                - name: Deal name following convention (e.g., Elcano_OpenX_Partner_...)
                - currency: Currency code (e.g., "USD")
                - deal_price: Price in CPM
                - pmp_deal_type: PMP deal type (e.g., "PREFERRED_DEAL")
                - start_date: ISO 8601 formatted start date
                - deal_participants: List of demand partner configurations
                  (using "demand_partner" as the key)
                - package: PackageCreateParams with targeting
                  (targeting includes "inter_dimension_operator": "AND")
                - third_party_fees_config: Optional fee configuration

        Returns:
            The created deal object with deal_id, name, status, etc.
        """
        mutation = """
        mutation CreateDeal($input: DealCreateParams!) {
            dealCreate(input: $input) {
                id
                deal_id
                name
                status
                currency
                deal_price
                pmp_deal_type
                start_date
                end_date
                created_date
                modified_date
                package {
                    uid
                }
            }
        }
        """

        logger.info(f"Creating deal: {deal_params.get('name', 'unnamed')}")
        result = await self._graphql_request(mutation, {"input": deal_params}, operation_name="dealCreate")
        return result.get("dealCreate", {})

    async def geo_search(self, query_text: str) -> list[dict[str, Any]]:
        """Run OpenX geoSearch for a freeform location query."""
        query = """
        query GeoSearch($query: String!) {
            geoSearch(query: $query) {
                id
                name
                state
                country
                type
                type_id
            }
        }
        """

        logger.info("Running OpenX geoSearch for query=%s", query_text)
        result = await self._graphql_request(query, {"query": query_text}, operation_name="GeoSearch")
        return result.get("geoSearch", [])

    async def list_geographic_options(self, kind: str, filter: dict[str, Any]) -> list[dict[str, Any]]:
        """Query optionsByPath for a geographic dimension."""
        path = OPENX_OPTIONS_PATHS[kind]
        logger.info("Listing geographic options for kind=%s filter=%s", kind, filter)
        return await self.list_options_by_path(path, filter)

    # dealById field selection used since the original integration. Every
    # field here is empirically validated against our account's schema —
    # notably device_type's V2 fields (desktop_devices/mobile_devices/
    # tv_devices), which is how this schema models devices.
    _DEAL_LEGACY_QUERY = """
        query GetDeal($id: String!) {
            dealById(id: $id) {
                id
                deal_id
                name
                status
                currency
                deal_price
                pmp_deal_type
                start_date
                end_date
                created_date
                modified_date
                deal_participants {
                    demand_partner
                    buyer_ids
                    brand_ids
                }
                package {
                    uid
                    name
                    targeting {
                        inter_dimension_operator
                        rendering_context {
                            op
                            ad_placement {
                                op
                                val
                            }
                            distribution_channel {
                                op
                                val
                            }
                            device_type {
                                op
                                desktop_devices
                                mobile_devices
                                tv_devices
                            }
                        }
                        domain {
                            categories_iab_v2 {
                                op
                                val
                            }
                        }
                        audience {
                            openaudience_custom {
                                op
                                val
                            }
                        }
                        content {
                            account {
                                op
                                val
                            }
                        }
                        geographic {
                            includes {
                                country
                                state
                                region
                            }
                            excludes {
                                country
                                state
                                region
                            }
                        }
                    }
                    url_targeting {
                        type
                        urls
                        domain_targeting_option
                    }
                }
                third_party_fees_config {
                    partner_id
                    revenue_method
                    gross_share
                    gross_cpm_cap
                }
            }
        }
    """

    # Full targeting selection for read-modify-write updates. Field-for-field
    # validated against the LIVE schema via output-type introspection on
    # 2026-06-11 (the published "Working with Deals" guide drifts from the
    # live schema: app_bundle_id lives under targeting.app_inventory — not
    # content — and the live Targeting type carries extra branches like
    # metacategory, instream_content, inventory_quality, attention, and
    # app_inventory that the guide omits). Reading a deal with this query
    # lets package.targeting be edited and resubmitted on dealUpdate without
    # nulling branches. geographic.circles/circles_v2 are deliberately not
    # selected (geo-radius targeting is unused in our curation flows; their
    # val shape is undocumented) — a deal using circles would lose them on a
    # blind resubmit, which the update tool's warning calls out.
    _DEAL_FULL_QUERY = """
        query GetDealFull($id: String!) {
            dealById(id: $id) {
                id
                deal_id
                name
                status
                currency
                deal_price
                pmp_deal_type
                start_date
                end_date
                created_date
                modified_date
                enable_dv360_inventory_package
                is_archived
                deal_participants {
                    demand_partner
                    buyer_ids
                    brand_ids
                }
                package {
                    uid
                    name
                    created_date
                    modified_date
                    targeting {
                        inter_dimension_operator
                        exclude_non_direct
                        include_open_bidding_inventory
                        attention
                        audience {
                            openaudience_custom {
                                op
                                val
                            }
                        }
                        content {
                            content_inter_dimension_operator
                            _idsByInstances
                            account {
                                op
                                val
                            }
                            adunit_size {
                                op
                                val
                            }
                            site {
                                op
                                val
                            }
                            page_url {
                                op
                                val {
                                    op
                                    val
                                }
                            }
                            instance {
                                op
                                val
                            }
                            includes {
                                instance
                                site
                                page
                                account
                                section
                                adunit
                                network
                                sitesection
                            }
                            excludes {
                                instance
                                site
                                page
                                account
                                section
                                adunit
                                network
                                sitesection
                            }
                            keywords {
                                op
                                val
                            }
                        }
                        custom {
                            op
                            val {
                                op
                                attr
                                val
                            }
                        }
                        domain {
                            inter_dimension_operator
                            categories {
                                op
                                val
                            }
                            categories_iab_v2 {
                                op
                                val
                            }
                        }
                        geographic {
                            includes {
                                city
                                state
                                country
                                continent
                                region
                                msa
                                dma
                                postal_code
                            }
                            excludes {
                                city
                                state
                                country
                                continent
                                region
                                msa
                                dma
                                postal_code
                            }
                        }
                        rendering_context {
                            op
                            ad_placement {
                                op
                                val
                            }
                            distribution_channel {
                                op
                                val
                            }
                            device_type {
                                op
                                desktop_devices
                                mobile_devices
                                tv_devices
                            }
                        }
                        technographic {
                            os {
                                op
                                val
                            }
                            language {
                                op
                                val
                            }
                            browser {
                                op
                                val
                            }
                            api_framework {
                                op
                                val
                            }
                        }
                        video {
                            rewarded_video {
                                op
                                val
                            }
                            video_format {
                                op
                                val
                            }
                            video_plcmt {
                                op
                                val
                            }
                            skip {
                                op
                                val
                            }
                            screen_location {
                                op
                                val
                            }
                            adunit_max_duration_range {
                                start
                                end
                            }
                        }
                        viewability {
                            viewability_score {
                                op
                                val
                            }
                        }
                        vtr {
                            vtr_score {
                                op
                                val
                            }
                        }
                        inventory_quality {
                            iq_score {
                                op
                                val
                            }
                        }
                        metacategory {
                            exclude_mfa
                            excludes
                            inter_dimension_operator
                            includes {
                                op
                                val
                            }
                            keywords {
                                includes
                                excludes
                            }
                        }
                        instream_content {
                            inter_dimension_operator
                            livestream
                            genre {
                                allow
                                block
                            }
                            contentrating {
                                allow
                                block
                            }
                            language {
                                op
                                val
                            }
                            channel {
                                name {
                                    op
                                    val
                                }
                            }
                            network {
                                name {
                                    op
                                    val
                                }
                            }
                            series {
                                op
                                val
                            }
                            episode {
                                op
                                val
                            }
                        }
                        app_inventory {
                            app_inventory_inter_dimension_operator
                            app_bundle_id {
                                op
                                val {
                                    op
                                    val
                                }
                                inventorylist {
                                    type
                                    id
                                }
                            }
                            appgroup {
                                op
                                val
                            }
                        }
                    }
                    url_targeting {
                        type
                        urls
                        domain_targeting_option
                    }
                }
                third_party_fees_config {
                    partner_id
                    revenue_method
                    gross_share
                    gross_cpm_cap
                }
            }
        }
    """

    async def get_deal(self, deal_id: str, full_targeting: bool = False) -> dict[str, Any]:
        """
        Fetch full details for a specific deal.

        Args:
            deal_id: The ID of the deal to retrieve
            full_targeting: When True, request the complete documented targeting
                selection (required before a read-modify-write targeting update).
                Falls back to the legacy selection if the schema rejects any
                documented field; the result then carries
                _targeting_selection="legacy" so callers know branches may be
                missing and a blind targeting resubmit is NOT safe.

        Returns:
            Complete deal object with all fields
        """
        logger.info(f"Fetching deal: {deal_id} (full_targeting={full_targeting})")

        if full_targeting:
            try:
                result = await self._graphql_request(self._DEAL_FULL_QUERY, {"id": deal_id})
                deal = result.get("dealById", {})
                if deal:
                    deal["_targeting_selection"] = "full"
                return deal
            except Exception as exc:
                # A schema drift on any one field fails the whole query —
                # degrade to the proven legacy selection rather than erroring
                # the read, but mark the result so update flows know a full
                # targeting resubmit cannot be built from it.
                # Broad except is deliberate: Apollo returns GraphQL
                # validation failures as HTTP 400, which _graphql_request
                # raises as a plain Exception (observed live 2026-06-11) —
                # catching only OpenXGraphQLError misses exactly the schema
                # rejection this fallback exists for. A transport outage just
                # fails the legacy retry too and surfaces from there.
                logger.warning("Full-targeting dealById failed (%s); falling back to legacy selection", exc)
                result = await self._graphql_request(self._DEAL_LEGACY_QUERY, {"id": deal_id})
                deal = result.get("dealById", {})
                if deal:
                    deal["_targeting_selection"] = "legacy"
                return deal

        result = await self._graphql_request(self._DEAL_LEGACY_QUERY, {"id": deal_id})
        return result.get("dealById", {})

    async def update_deal(self, deal_id: str, update_params: dict[str, Any]) -> dict[str, Any]:
        """
        Update a deal using the dealUpdate mutation.

        Partial updates are supported only for fields at the top level of
        `deal` and `deal.package` — nested objects (e.g. package.targeting)
        are replaced wholesale, with unspecified fields assumed null.

        Args:
            deal_id: The internal deal id (dealById id, not the OX-... deal_id).
            update_params: DealUpdateParams input object.

        Returns:
            The updated deal object.
        """
        mutation = """
        mutation DealUpdate($id: String!, $input: DealUpdateParams!) {
            dealUpdate(id: $id, input: $input) {
                id
                deal_id
                name
                status
                currency
                deal_price
                pmp_deal_type
                start_date
                end_date
                modified_date
                package {
                    uid
                    modified_date
                }
            }
        }
        """

        logger.info(f"Updating deal {deal_id}: fields={sorted(update_params.keys())}")
        result = await self._graphql_request(
            mutation, {"id": deal_id, "input": update_params}, operation_name="dealUpdate"
        )
        return result.get("dealUpdate", {})

    # NOTE: the OpenXSelect API also exposes a dealArchive mutation. It is
    # DELIBERATELY not implemented here and MUST NOT be exposed as an MCP
    # tool: archiving is effectively irreversible (the guide documents no
    # un-archive) and giving the agent an irreversible destructive action is
    # a security decision Elcano made against (2026-06-11, Elyse). To stop a
    # deal's delivery, pause it via update_deal({"status": "Paused"}).

    async def list_deals(self, limit: int = 10, offset: int = 0) -> list[dict[str, Any]]:
        """
        List existing deals with pagination.

        Args:
            limit: Maximum number of deals to return
            offset: Number of deals to skip (for pagination)

        Returns:
            List of deal objects
        """
        query = """
        query ListDeals($limit: Int!, $offset: Int!) {
            deals(limit: $limit, offset: $offset) {
                id
                deal_id
                name
                status
                currency
                deal_price
                pmp_deal_type
                start_date
                end_date
                created_date
                modified_date
            }
        }
        """

        logger.info(f"Listing deals (limit={limit}, offset={offset})")
        result = await self._graphql_request(query, {"limit": limit, "offset": offset})
        return result.get("deals", [])

    async def list_demand_partners(self) -> list[dict[str, Any]]:
        """
        Query and return available demand partners.

        Returns:
            List of demand partner objects with id and name
        """
        # Use optionsByPath to get demand partners as per OpenX API schema
        query = """
        query ListDemandPartners {
            optionsByPath(path: "deal.deal_participants.demand_partner") {
                id
                name
                path
                extra
            }
        }
        """

        logger.info("Listing demand partners via optionsByPath")
        result = await self._graphql_request(query)
        return result.get("optionsByPath", [])

    async def list_options_by_path(self, path: str, filter: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        """
        Query optionsByPath for a specific OpenX schema path.

        This is primarily a debug/discovery helper. Prefer the high-level deal preparation
        tools for normal agent workflows.
        """
        query = """
        query ListOptionsByPath($path: String!, $filter: JSON) {
            optionsByPath(path: $path, filter: $filter) {
                id
                name
                path
                extra
            }
        }
        """

        logger.info("Listing OpenX options via optionsByPath for path=%s", path)
        result = await self._graphql_request(query, {"path": path, "filter": filter})
        return result.get("optionsByPath", [])

    async def introspect_type(self, type_name: str) -> dict[str, Any]:
        """
        Run a GraphQL introspection query to discover fields of a named type.

        Args:
            type_name: The name of the GraphQL type to introspect
                       (e.g., "TargetingCreateParams", "TargetingGeographicItemCreateParams")

        Returns:
            The introspection result containing type name, kind, and inputFields
        """
        query = """
        query IntrospectType($typeName: String!) {
            __type(name: $typeName) {
                name
                kind
                inputFields {
                    name
                    type {
                        name
                        kind
                        ofType { name kind ofType { name kind } }
                    }
                }
            }
        }
        """

        logger.info(f"Introspecting type: {type_name}")
        result = await self._graphql_request(query, {"typeName": type_name})
        return result.get("__type", {})

    async def get_geographic_field_name(self) -> str:
        """
        Discover the field name for country-level geographic targeting via introspection.

        Used by the country-list-to-structured-geo wrapper. Prefers `country` when
        the introspected type exposes it (the only sensible mapping for a list of
        ISO country codes). Falls back to the first inputField only when `country`
        is absent — earlier versions used `inputFields[0]` unconditionally, which
        on this account happened to be `city`, silently mistargeting every deal.
        """
        cache_key = "TargetingGeographicItemCreateParams"

        if cache_key in self._type_field_cache:
            return self._type_field_cache[cache_key] or "country"

        try:
            type_info = await self.introspect_type(cache_key)
            input_fields = type_info.get("inputFields", [])
            field_names = [field.get("name") for field in input_fields if field.get("name")]

            if "country" in field_names:
                self._type_field_cache[cache_key] = "country"
                logger.info("Discovered geographic field name: country")
                return "country"

            if field_names:
                field_name = field_names[0]
                self._type_field_cache[cache_key] = field_name
                logger.info(f"Discovered geographic field name (fallback, no 'country' in type): {field_name}")
                return field_name

            logger.warning(f"No input fields found for {cache_key}, using 'country'")
            self._type_field_cache[cache_key] = None
            return "country"
        except Exception as e:
            logger.warning(f"Introspection failed for {cache_key}: {e}, using 'country'")
            self._type_field_cache[cache_key] = None
            return "country"


# Global client instance
_openx_client: OpenXClient | None = None
_prepared_openx_deals: dict[str, dict[str, Any]] = {}


def get_openx_client() -> OpenXClient:
    """Get or create the OpenX client singleton."""
    global _openx_client
    if _openx_client is None:
        _openx_client = OpenXClient()
    return _openx_client


def _mapping_hints_for_type(type_name: str, type_info: dict[str, Any]) -> list[str]:
    """Return targeted guidance for commonly-confused OpenX types."""
    hints: list[str] = []
    field_names = {field.get("name") for field in type_info.get("inputFields", [])}

    if type_name == "TargetingCreateParams":
        if "technographic" in field_names:
            hints.append("Language targeting belongs under targeting.technographic.language, not targeting.content.")
        if "domain" in field_names:
            hints.append(
                "IAB category targeting belongs under targeting.domain.categories_iab_v2, not targeting.content."
            )
        hints.append(
            "This MCP accepts targeting.device_type as a convenience alias and normalizes it into targeting.rendering_context.device_type."
        )

    if type_name == "TargetingContentCreateParams":
        hints.append(
            "This type is for content filters such as page/app/site/keywords. It does not expose OpenX language targeting."
        )
        hints.append(
            "If you need Language, inspect TargetingTechnographicCreateParams. If you need IAB categories, inspect TargetingDomainCreateParams and TargetingDomainCategoriesIabV2CreateParams."
        )

    if type_name == "TargetingTechnographicCreateParams" and "language" in field_names:
        hints.append(
            "This is the OpenX input type for request-time language targeting via targeting.technographic.language."
        )

    if type_name == "TargetingDomainCreateParams" and "categories_iab_v2" in field_names:
        hints.append(
            "Use targeting.domain.categories_iab_v2 for IAB v2 category targeting. Resolve allowed values with ox_list_iab_categories or ox_list_options_by_path before create."
        )

    if type_name == "TargetingAudienceCreateParams" and field_names == {"openaudience_custom"}:
        hints.append(
            "Audience targeting is expressed through targeting.audience.openaudience_custom, not a freeform segment-name field."
        )
        hints.append(
            "Resolve valid audience option values with ox_list_options_by_path before create instead of guessing from the segment display name."
        )

    if type_name == "DealCreateParams" and "deal_participants" in field_names and "package" in field_names:
        hints.append(
            "deal_participants and package.targeting are required to create a meaningful deal. Prefer ox_prepare_deal_from_prompt_inputs or ox_prepare_deal_from_brief for normal workflows."
        )

    return hints


async def _resolve_option_ids(
    client: OpenXClient,
    *,
    path: str,
    values: list[str],
    field_label: str,
    filter: dict[str, Any] | None = None,
    aliases: dict[str, str] | None = None,
    allow_contains_match: bool = False,
) -> tuple[list[str], list[str]]:
    """Resolve business-facing values to OpenX option ids, returning ids and warnings."""
    if not values:
        return [], []

    options = await client.list_options_by_path(path, filter)
    resolved_ids: list[str] = []
    warnings: list[str] = []
    for raw_value in values:
        selected_option = _choose_unique_option(
            options,
            raw_value,
            field_label,
            aliases=aliases,
            allow_contains_match=allow_contains_match,
        )
        selected_id = str(selected_option["id"])
        if selected_id not in resolved_ids:
            resolved_ids.append(selected_id)
        if str(raw_value) != selected_id:
            warnings.append(f"Resolved {field_label} '{raw_value}' to OpenX id '{selected_id}'.")
    return resolved_ids, warnings


async def _resolve_option(
    client: OpenXClient,
    *,
    path: str,
    value: str,
    field_label: str,
    filter: dict[str, Any] | None = None,
    aliases: dict[str, str] | None = None,
    allow_contains_match: bool = False,
) -> tuple[dict[str, Any], list[str]]:
    """Resolve a single business-facing value to an OpenX option plus warnings."""
    options = await client.list_options_by_path(path, filter)
    selected_option = _choose_unique_option(
        options,
        value,
        field_label,
        aliases=aliases,
        allow_contains_match=allow_contains_match,
    )
    warnings: list[str] = []
    selected_id = str(selected_option["id"])
    if str(value) != selected_id:
        warnings.append(f"Resolved {field_label} '{value}' to OpenX id '{selected_id}'.")
    return selected_option, warnings


def _filter_options_by_query(options: list[dict[str, Any]], query: str | None) -> list[dict[str, Any]]:
    """Filter OpenX option results by normalized id/name substring."""
    if not query:
        return options

    normalized_query = _normalize_lookup_text(query)
    if not normalized_query:
        return options

    return [
        option
        for option in options
        if normalized_query in _normalize_lookup_text(option.get("name", ""))
        or normalized_query in _normalize_lookup_text(option.get("id", ""))
        or normalized_query in _normalize_lookup_text((option.get("extra") or {}).get("full_name", ""))
        or normalized_query in _normalize_lookup_text((option.get("extra") or {}).get("state", ""))
        or normalized_query in _normalize_lookup_text((option.get("extra") or {}).get("abbreviation", ""))
    ]


def _state_lookup_aliases() -> dict[str, str]:
    """Build normalized aliases for US states and Canadian provinces."""
    aliases: dict[str, str] = {}
    for geographic_names in (US_STATE_NAMES, CANADA_PROVINCE_NAMES):
        for abbreviation, full_name in geographic_names.items():
            aliases[_normalize_lookup_text(abbreviation)] = abbreviation
            aliases[_normalize_lookup_text(full_name)] = abbreviation
    return aliases


# Standard IAB 1.0 taxonomy code-to-OpenX category name mapping.
# OpenX uses its own category names on the legacy `domain.categories` path.
# This table maps the industry-standard IAB codes that traders use to the
# exact OpenX category names returned by optionsByPath.
_LEGACY_IAB_CODE_PATTERN = re.compile(r"^IAB\d+(-\d+)?$", re.IGNORECASE)


def _normalize_country_hint(value: str | None) -> str | None:
    """Normalize a country code/name into the OpenX country label used in option metadata."""
    normalized_value = _normalize_lookup_text(value)
    if not normalized_value:
        return None
    return COUNTRY_NAME_ALIASES.get(normalized_value.upper()) or COUNTRY_NAME_ALIASES.get(normalized_value)


def _country_code_from_hint(country_hint: str | None) -> str | None:
    """Map a normalized country hint back to a compact country code when possible."""
    if country_hint == "united states":
        return "us"
    if country_hint == "canada":
        return "ca"
    return None


def _normalize_state_query(raw_value: str, country_hint: str | None = None) -> tuple[str, str]:
    """Normalize a state/province input into abbreviation and full-name query text."""
    aliases = _state_lookup_aliases()
    normalized_state = aliases.get(_normalize_lookup_text(raw_value), str(raw_value).strip().upper())
    if country_hint == "canada":
        full_name = CANADA_PROVINCE_NAMES.get(normalized_state, str(raw_value).strip())
    elif country_hint == "united states":
        full_name = US_STATE_NAMES.get(normalized_state, str(raw_value).strip())
    else:
        full_name = (
            US_STATE_NAMES.get(normalized_state)
            or CANADA_PROVINCE_NAMES.get(normalized_state)
            or str(raw_value).strip()
        )
    return normalized_state, full_name


def _normalize_geographic_lookup_input(kind: str, raw_value: str, country_hint: str | None = None) -> tuple[str, str]:
    """Normalize a geographic input into a lookup key and preferred query text."""
    if kind == "state":
        normalized_state, full_name = _normalize_state_query(raw_value, country_hint)
        return normalized_state, full_name

    normalized_value = str(raw_value).strip()
    return normalized_value, normalized_value


def _build_geographic_options_filter(
    kind: str, query_text: str, *, country_hint: str | None = None, include_country_filter: bool = False
) -> dict[str, Any]:
    """Build the documented optionsByPath filter for a geographic dimension."""
    filter_key = GEOGRAPHIC_FILTER_KEYS[kind]
    query_value = str(query_text).strip().lower()
    options_filter = {filter_key: f"{query_value}*"}
    if include_country_filter and country_hint:
        options_filter["country"] = f"{country_hint}*"
    return options_filter


def _get_verified_geographic_id(kind: str, normalized_value: str) -> str | None:
    """Return a verified geographic id when available."""
    verified_values = VERIFIED_GEOGRAPHIC_IDS.get(kind, {})
    direct_match = verified_values.get(str(normalized_value).strip().upper())
    if direct_match:
        return direct_match

    return verified_values.get(_normalize_lookup_text(normalized_value).upper())


async def _resolve_geographic_ids(
    client: OpenXClient,
    *,
    kind: str,
    values: list[str],
    country_hint: str | None = "united states",
) -> tuple[list[str], list[str]]:
    """Resolve geographic ids using optionsByPath and verified fallbacks where needed."""
    warnings: list[str] = []
    resolved_ids: list[str] = []
    include_country_filter = kind == "state" and country_hint not in {None, "united states"}

    for raw_value in values:
        normalized_value, query_text = _normalize_geographic_lookup_input(kind, str(raw_value), country_hint)
        verified_id = _get_verified_geographic_id(kind, normalized_value) if country_hint == "united states" else None
        if verified_id:
            if verified_id not in resolved_ids:
                resolved_ids.append(verified_id)
            warnings.append(f"Resolved {kind} '{raw_value}' to verified OpenX id '{verified_id}'.")
            continue

        try:
            requested_option_value = query_text if kind == "state" else normalized_value
            options_filter = _build_geographic_options_filter(
                kind,
                query_text,
                country_hint=country_hint,
                include_country_filter=include_country_filter,
            )
            options = await client.list_geographic_options(kind, options_filter)
            filtered_options = [
                option
                for option in options
                if country_hint is None
                or str((option.get("extra") or {}).get("country", "")).lower() == country_hint.lower()
            ]
            selected_option = _choose_unique_option(
                filtered_options,
                requested_option_value,
                kind,
                allow_contains_match=True,
            )
            selected_id = str(selected_option["id"])
            if selected_id not in resolved_ids:
                resolved_ids.append(selected_id)
            warnings.append(f"Resolved {kind} '{raw_value}' to OpenX id '{selected_id}' via optionsByPath.")
        except (OpenXGraphQLError, LookupError, ValueError) as exc:
            raise LookupError(f"No OpenX match found for {kind}: {raw_value}") from exc

    return resolved_ids, warnings


def _detect_legacy_iab_codes(values: list[str]) -> list[str]:
    """Detect legacy IAB codes like IAB2-5 that won't resolve against live OpenX options."""
    return [value for value in values if _LEGACY_IAB_CODE_PATTERN.match(value.strip())]


def _legacy_iab_query_hint(legacy_codes: list[str]) -> str | None:
    """Return a suggested query hint for ox_list_iab_categories when legacy codes are supplied."""
    if legacy_codes and all(code.upper().startswith("IAB2") for code in legacy_codes):
        return "automotive"
    return None


def _targeting_has_audience(targeting: dict[str, Any]) -> bool:
    """Return whether the current targeting includes an audience clause."""
    audience = targeting.get("audience")
    return isinstance(audience, dict) and bool(audience.get("openaudience_custom"))


def _geographic_includes_explicit_country(geographic: Any) -> bool:
    """Return whether geographic targeting includes at least one explicit include country."""
    if not isinstance(geographic, dict):
        return False
    includes = geographic.get("includes")
    if not isinstance(includes, dict):
        return False
    country_value = includes.get("country")
    return isinstance(country_value, str) and bool([part for part in country_value.split(",") if part.strip()])


def _geographic_includes_only_us(geographic: Any) -> bool:
    """Return whether geographic targeting includes only United States country values."""
    if not isinstance(geographic, dict):
        return False
    includes = geographic.get("includes")
    if not isinstance(includes, dict):
        return False
    country_value = includes.get("country")
    if not isinstance(country_value, str):
        return False
    countries = [part.strip() for part in country_value.split(",") if part.strip()]
    return bool(countries) and all(_normalize_country_hint(country) == "united states" for country in countries)


def _audience_export_type_requires_us(audience_option: dict[str, Any] | None) -> bool:
    """Return whether the resolved audience export type is restricted to US geo in OpenX."""
    if not isinstance(audience_option, dict):
        return False
    extra = audience_option.get("extra") or {}
    export_type = _normalize_lookup_text(extra.get("export_type", ""))
    return export_type in {"oamatch", "bidstreamplus"}


_OPENAUDIENCE_ID_PREFIX = "openaudience-"


def _normalize_openaudience_id(raw_id: Any) -> str:
    """Return an audience id with the ``openaudience-`` prefix required by dealCreate.

    Why: ``optionsByPath`` for ``deal.package.targeting.audience.openaudience_custom`` returns
    the bare audience UUID, but the ``dealCreate`` mutation rejects it as ``invalid audience id``
    unless the value is prefixed with ``openaudience-``. UI deal JSONs confirm the wire format.
    """
    value = str(raw_id).strip()
    if not value:
        return value
    if value.startswith(_OPENAUDIENCE_ID_PREFIX):
        return value
    return f"{_OPENAUDIENCE_ID_PREFIX}{value}"


def _subnational_country_hint_from_item(item: dict[str, Any]) -> str | None:
    """Extract a normalized country hint for state/province resolution from a geo item."""
    country_value = item.get("country")
    if not isinstance(country_value, str):
        return None
    countries = [part.strip() for part in country_value.split(",") if part.strip()]
    if len(countries) != 1:
        return None
    return _normalize_country_hint(countries[0])


async def _resolve_state_ids(
    client: OpenXClient, values: list[str], *, country_hint: str = "united states"
) -> tuple[list[str], list[str]]:
    return await _resolve_geographic_ids(client, kind="state", values=values, country_hint=country_hint)


async def _resolve_structured_geographic_targeting(
    client: OpenXClient, geographic: dict[str, Any]
) -> tuple[dict[str, Any], list[str], list[dict[str, Any]]]:
    """Normalize and resolve structured geographic targeting, including state/province ids."""
    normalized_geographic, warnings = _normalize_geographic_targeting(geographic)
    blockers: list[dict[str, Any]] = []

    for branch_name in ("includes", "excludes"):
        branch = normalized_geographic.get(branch_name)
        if not isinstance(branch, dict):
            continue

        state_value = branch.get("state")
        if not isinstance(state_value, str) or not state_value.strip():
            continue

        state_parts = [part.strip() for part in state_value.split(",") if part.strip()]
        if not state_parts or all(part.isdigit() for part in state_parts):
            continue

        country_hint = _subnational_country_hint_from_item(branch)
        if country_hint is None:
            blockers.append(
                _make_blocker(
                    "subnational_geo_requires_country",
                    f"Geographic {branch_name}.state targeting requires an explicit {branch_name}.country.",
                    branch=branch_name,
                    values=state_parts,
                )
            )
            continue

        try:
            state_ids, state_warnings = await _resolve_state_ids(client, state_parts, country_hint=country_hint)
            warnings.extend(state_warnings)
            branch["state"] = ",".join(state_ids)
            if not branch.get("country"):
                country_code = _country_code_from_hint(country_hint)
                if country_code:
                    branch["country"] = country_code
        except (LookupError, ValueError) as exc:
            blockers.append(
                _make_blocker(
                    "unresolved_state",
                    str(exc),
                    branch=branch_name,
                    values=state_parts,
                    country=country_hint,
                )
            )

    return _cleanup_structured_geographic_targeting(normalized_geographic), warnings, blockers


async def _build_prepared_openx_deal(
    *,
    client: OpenXClient,
    name: str,
    currency: str,
    deal_price: float,
    start_date: str,
    package_name: str | None,
    demand_partner: str,
    targeting_brief: dict[str, Any] | None,
    publisher_ids: list[str] | None = None,
    excluded_publisher_ids: list[str] | None = None,
    buyer_ids: list[str] | None = None,
    end_date: str | None = None,
    pmp_deal_type: str = "3",
    fee: dict[str, Any] | None = None,
    domains_allowlist: list[str] | None = None,
    domains_blocklist: list[str] | None = None,
    domain_targeting_option: str | None = None,
    viewability_threshold: float | None = None,
    status: str | None = None,
    expected_ad_category: str | None = None,
    inventory_categories: list[str] | None = None,
    app_bundles: list[str] | None = None,
) -> dict[str, Any]:
    """Resolve a business-level deal brief into a safe ox_create_deal request."""
    blockers: list[dict[str, Any]] = []
    warnings: list[str] = []
    targeting_brief = dict(targeting_brief or {})

    if end_date and end_date < start_date:
        blockers.append(_make_blocker("invalid_date_order", "end_date must be greater than or equal to start_date."))
    if domains_allowlist and domains_blocklist:
        blockers.append(
            _make_blocker(
                "conflicting_domain_lists",
                "Specify only one of domains_allowlist or domains_blocklist per prepared deal.",
            )
        )
    if publisher_ids and excluded_publisher_ids:
        # OpenX content.account is a single object with one op + one val — it supports either
        # INTERSECTS (include) or NOT INTERSECTS (exclude), not both at once. If a brief needs
        # both inclusion and exclusion logic, the trader has to split it into two deals.
        blockers.append(
            _make_blocker(
                "conflicting_publisher_lists",
                "Specify only one of publisher_ids (INTERSECTS) or excluded_publisher_ids "
                "(NOT INTERSECTS) per prepared deal. OpenX content.account supports a single "
                "op per deal; combining inclusion and exclusion requires separate deals.",
            )
        )

    demand_partner_reference, normalized_buyer_ids, alias_warnings = _parse_demand_partner_reference(
        demand_partner, buyer_ids
    )
    warnings.extend(alias_warnings)

    demand_partner_options = await client.list_options_by_path(OPENX_OPTIONS_PATHS["demand_partner"])
    resolved_partner = None
    resolved_demand_partner_id = str(demand_partner_reference)
    try:
        resolved_partner = _choose_unique_option(demand_partner_options, demand_partner_reference, "demand partner")
        resolved_demand_partner_id = str(resolved_partner["id"])
        if str(demand_partner_reference) != resolved_demand_partner_id:
            warnings.append(
                f"Resolved demand partner '{demand_partner_reference}' to OpenX id '{resolved_demand_partner_id}'."
            )
    except (LookupError, ValueError) as exc:
        blockers.append(_make_blocker("demand_partner_unresolved", str(exc), input=demand_partner_reference))

    if normalized_buyer_ids and resolved_partner is not None:
        buyer_options = await client.list_options_by_path(
            OPENX_OPTIONS_PATHS["buyer_ids"],
            {"demand_partner_id": resolved_demand_partner_id},
        )
        resolved_buyer_ids, buyer_warnings = _resolve_buyer_seat_to_third_party_ids(buyer_options, normalized_buyer_ids)
        for buyer_warning in buyer_warnings:
            warnings.append(buyer_warning)
        normalized_buyer_ids = resolved_buyer_ids or normalized_buyer_ids
        allowed_buyer_ids = _extract_allowed_buyer_ids(buyer_options)
        invalid_buyer_ids = [buyer_id for buyer_id in normalized_buyer_ids if buyer_id not in allowed_buyer_ids]
        if invalid_buyer_ids:
            blockers.append(
                _make_blocker(
                    "unresolved_buyer_id",
                    f"Unknown buyer_ids for demand partner {resolved_demand_partner_id}: {', '.join(invalid_buyer_ids)}. "
                    "Expected the third-party DSP seat id (the value shown after '>' in the OpenX UI, "
                    "e.g. 'Spark Foundry > 849135' → '849135').",
                    demand_partner=resolved_demand_partner_id,
                    buyer_ids=invalid_buyer_ids,
                )
            )

    quality_flags: list[dict[str, Any]] = []
    if fee is None:
        fee, default_fee_warning = _default_curator_fee()
        warnings.append(default_fee_warning)
        quality_flags.append(
            _make_ox_quality_flag(
                "ox_default_curator_fee_applied",
                f"Auto-applied default OpenX curator fee ({DEFAULT_FEE_PARTNER_NAME}): "
                f"{DEFAULT_FEE_REVENUE_METHOD} {DEFAULT_FEE_GROSS_SHARE_PERCENT:g}%.",
                revenue_method=DEFAULT_FEE_REVENUE_METHOD,
                gross_share_percent=DEFAULT_FEE_GROSS_SHARE_PERCENT,
                partner=DEFAULT_FEE_PARTNER_NAME,
            )
        )

    resolved_fee = None
    if fee:
        resolved_fee = dict(fee)
        if "partner" in resolved_fee and "partner_name_or_id" not in resolved_fee and "partner_id" not in resolved_fee:
            resolved_fee["partner_name_or_id"] = resolved_fee.pop("partner")
        partner_name_or_id = resolved_fee.pop("partner_name_or_id", None) or resolved_fee.get("partner_id")
        if partner_name_or_id is None:
            blockers.append(
                _make_blocker(
                    "missing_fee_partner",
                    "fee.partner_name_or_id or fee.partner_id is required when fee is provided.",
                )
            )
        else:
            try:
                fee_partner = _choose_unique_option(
                    await client.list_options_by_path(OPENX_OPTIONS_PATHS["fee_partner"]),
                    str(partner_name_or_id),
                    "fee partner",
                    allow_contains_match=True,
                )
                resolved_fee["partner_id"] = str(fee_partner["id"])
                if str(partner_name_or_id) != resolved_fee["partner_id"]:
                    warnings.append(
                        f"Resolved fee partner '{partner_name_or_id}' to OpenX id '{resolved_fee['partner_id']}'."
                    )
                # NOTE: do NOT inject `platform_share` or `platform_partner_id`
                # here. Those fields appear in OpenX's dealCreate *response*
                # but `ThirdPartyFeesConfigCreateParams` (the input type) only
                # accepts `revenue_method` / `gross_share` / `gross_cpm_cap` /
                # `partner_id`. OpenX computes the platform share server-side
                # from the partner record. Passing `platform_share` on the
                # input is rejected with
                #   "Field "platform_share" is not defined by type
                #    "ThirdPartyFeesConfigCreateParams"."
                # which broke every `ox_create_prepared_deal` call until the
                # operator manually bypassed the prepare flow and called
                # `ox_create_deal` directly with a stripped-down fee block.
            except (LookupError, ValueError) as exc:
                blockers.append(_make_blocker("unresolved_fee_partner", str(exc), input=str(partner_name_or_id)))

        if "gross_share_percent" in resolved_fee:
            try:
                original_gross_share_percent = resolved_fee.pop("gross_share_percent")
                resolved_fee["gross_share"] = _serialize_percentage_fraction(original_gross_share_percent)
                warnings.append(
                    f"Normalized fee gross_share_percent '{original_gross_share_percent}' to OpenX gross_share '{resolved_fee['gross_share']}'."
                )
            except (TypeError, ValueError) as exc:
                blockers.append(_make_blocker("invalid_fee_percentage", str(exc)))
        elif "gross_share" in resolved_fee and resolved_fee["gross_share"] is not None:
            try:
                original_gross_share = resolved_fee["gross_share"]
                resolved_fee["gross_share"] = _serialize_percentage_fraction(original_gross_share)
                if str(original_gross_share) != resolved_fee["gross_share"]:
                    warnings.append(
                        f"Normalized fee gross_share '{original_gross_share}' to OpenX gross_share '{resolved_fee['gross_share']}'."
                    )
            except (TypeError, ValueError) as exc:
                blockers.append(_make_blocker("invalid_fee_percentage", str(exc)))

    normalized_allowlist: list[str] = []
    invalid_allowlist: list[dict[str, str]] = []
    if domains_allowlist:
        validation = await ox_validate_domains([str(domain) for domain in domains_allowlist])
        normalized_allowlist = sorted(dict.fromkeys(validation["valid"]))
        invalid_allowlist = validation["invalid"]
        if invalid_allowlist:
            warnings.append(f"Dropped {len(invalid_allowlist)} invalid allowlist domains during preparation.")

    normalized_blocklist: list[str] = []
    invalid_blocklist: list[dict[str, str]] = []
    if domains_blocklist:
        validation = await ox_validate_domains([str(domain) for domain in domains_blocklist])
        normalized_blocklist = sorted(dict.fromkeys(validation["valid"]))
        invalid_blocklist = validation["invalid"]
        if invalid_blocklist:
            warnings.append(f"Dropped {len(invalid_blocklist)} invalid blocklist domains during preparation.")

    targeting: dict[str, Any] = _copy_structured_targeting_branches(targeting_brief)
    channel = targeting_brief.get("channel")
    if channel:
        targeting["channel"] = str(channel).upper()

    device_types = targeting_brief.get("device_types") or targeting_brief.get("device_type")
    if device_types:
        if not isinstance(device_types, list):
            blockers.append(
                _make_blocker("invalid_device_types", "targeting.device_types must be a list of device type strings.")
            )
        else:
            targeting["device_type"] = [
                str(device_type).upper() for device_type in device_types if str(device_type).strip()
            ]

    geographic = targeting_brief.get("geographic") or targeting_brief.get("geo")
    if isinstance(geographic, dict):
        supplied_state_ids = _extract_pre_resolved_state_ids(geographic)
        if supplied_state_ids:
            blockers.append(
                _make_blocker(
                    "unvalidated_state_ids",
                    "State/province targeting must be supplied as abbreviations or names, not pre-resolved numeric OpenX ids.",
                    values=supplied_state_ids,
                )
            )
            warnings.append(
                "Rejected pre-resolved geographic state ids. Use state/province abbreviations or names and let the MCP resolve them via ox_list_states."
            )
        else:
            (
                normalized_geographic,
                geographic_warnings,
                geographic_blockers,
            ) = await _resolve_structured_geographic_targeting(client, geographic)
            targeting["geographic"] = normalized_geographic
            warnings.extend(geographic_warnings)
            blockers.extend(geographic_blockers)
    elif isinstance(geographic, list):
        normalized_geographic_values = [str(value).strip() for value in geographic if str(value).strip()]
        all_us_state_names_lower = {name.lower() for name in US_STATE_NAMES.values()}
        us_state_abbreviations_upper = set(US_STATE_NAMES.keys())
        is_national_us = normalized_geographic_values == ["US"]
        # Treat as US states only when every token is recognizable as a US state — either a
        # known abbreviation (case-insensitive match against US_STATE_NAMES keys) or full
        # name. Previously this also matched ANY 2-letter token, which silently routed
        # Canadian provinces ("AB", "ON"), UK country code ("UK"), etc. through the
        # state-resolution path and produced unresolved_state blockers (or worse, wrong
        # resolutions). The tightened check still accepts mixed-case input ("ca", "Tx").
        is_states = (
            bool(normalized_geographic_values)
            and not is_national_us
            and all(
                value.upper() in us_state_abbreviations_upper or value.lower() in all_us_state_names_lower
                for value in normalized_geographic_values
            )
        )
        if is_states:
            try:
                state_ids, state_warnings = await _resolve_state_ids(client, normalized_geographic_values)
                warnings.extend(state_warnings)
                normalized_geographic, geographic_warnings = _normalize_geographic_targeting(
                    {
                        "includes": {
                            "state": ",".join(state_ids),
                            "country": "us",
                        }
                    },
                    warn_pre_resolved_state_ids=False,
                )
                warnings.extend(geographic_warnings)
                targeting["geographic"] = normalized_geographic
            except (LookupError, ValueError) as exc:
                blockers.append(_make_blocker("unresolved_state", str(exc), values=normalized_geographic_values))
        else:
            # Wrap flat country list into the structured includes.country form OpenX expects.
            # Without this wrap, OpenX silently drops the value at create.
            normalized_geographic, geographic_warnings = await _wrap_country_list_as_structured_geo(
                client, normalized_geographic_values
            )
            targeting["geographic"] = normalized_geographic
            warnings.extend(geographic_warnings)
    elif isinstance(geographic, str) and geographic.strip():
        normalized_geographic, geographic_warnings = await _wrap_country_list_as_structured_geo(
            client, [geographic.strip()]
        )
        targeting["geographic"] = normalized_geographic
        warnings.extend(geographic_warnings)

    languages = targeting_brief.get("languages") or targeting_brief.get("language")
    if isinstance(languages, str):
        languages = [languages]
    if languages:
        try:
            language_ids, language_warnings = await _resolve_option_ids(
                client,
                path=OPENX_OPTIONS_PATHS["language"],
                values=[str(language) for language in languages],
                field_label="language",
                allow_contains_match=True,
            )
            warnings.extend(language_warnings)
            targeting.setdefault("technographic", {})["language"] = {
                "op": "INTERSECTS",
                "val": ",".join(language_ids),
            }
        except (LookupError, ValueError) as exc:
            blockers.append(_make_blocker("unresolved_language", str(exc), values=languages))

    iab_categories = (
        targeting_brief.get("iab_categories")
        or targeting_brief.get("domain_categories_iab_v2")
        or targeting_brief.get("domain_categories")
    )
    if iab_categories:
        raw_category_values = [str(category) for category in iab_categories]
        legacy_codes = _detect_legacy_iab_codes(raw_category_values)
        if legacy_codes:
            blockers.append(
                _make_blocker(
                    "legacy_iab_codes_not_supported",
                    f"Legacy IAB codes cannot be resolved against live OpenX categories: {', '.join(legacy_codes)}. "
                    "Use ox_list_iab_categories to discover the correct OpenX IAB v2 category names or IDs, "
                    "then pass those names or IDs directly.",
                    legacy_codes=legacy_codes,
                    suggested_action="Use ox_list_iab_categories to discover correct OpenX names.",
                    query_hint=_legacy_iab_query_hint(legacy_codes),
                )
            )
        else:
            try:
                category_ids, category_warnings = await _resolve_option_ids(
                    client,
                    path=OPENX_OPTIONS_PATHS["iab_categories"],
                    values=raw_category_values,
                    field_label="IAB category",
                    allow_contains_match=True,
                )
                warnings.extend(category_warnings)
                targeting.setdefault("domain", {})["categories_iab_v2"] = {
                    "op": "INTERSECTS",
                    "val": ",".join(category_ids),
                }
            except (LookupError, ValueError) as exc:
                blockers.append(_make_blocker("unresolved_iab_category", str(exc), values=raw_category_values))

            # Mirror OpenX UI behavior by ALSO resolving against the legacy IAB v1
            # taxonomy on domain.categories. UI-created Reklaim deals emit both
            # taxonomies joined with inter_dimension_operator=OR so the deal serves
            # against inventory tagged under either generation. Failures here are
            # informational — the v2-only payload above still creates a valid deal.
            try:
                legacy_ids, legacy_warnings = await _resolve_option_ids(
                    client,
                    path=OPENX_OPTIONS_PATHS["iab_categories_legacy"],
                    values=raw_category_values,
                    field_label="IAB category (legacy v1)",
                    allow_contains_match=True,
                )
                if legacy_ids:
                    warnings.extend(legacy_warnings)
                    domain_targeting = targeting.setdefault("domain", {})
                    domain_targeting["categories"] = {
                        "op": "INTERSECTS",
                        "val": ",".join(legacy_ids),
                    }
                    # OR is REQUIRED here so a deal matches inventory tagged under
                    # either taxonomy; AND would require BOTH and starve the deal.
                    domain_targeting["inter_dimension_operator"] = "OR"
                    warnings.append(
                        f"Resolved {len(legacy_ids)} legacy IAB v1 category id(s) "
                        "alongside IAB v2 (OR'd) to mirror OpenX UI reach."
                    )
            except (LookupError, ValueError) as exc:
                # Legacy v1 names often won't match v2-flavored input ("travel and travel
                # related" vs. "Travel"); that's expected, not a blocker.
                logger.info("Legacy IAB v1 resolution skipped (no matches): %s", exc)
            except Exception as exc:  # noqa: BLE001 — network/API hiccups must not block deal
                logger.warning("Legacy IAB v1 resolution errored, falling back to v2-only: %s", exc)

    raw_inventory_categories = inventory_categories or targeting_brief.get("inventory_categories")
    if raw_inventory_categories:
        raw_inventory_values = [str(category) for category in raw_inventory_categories]
        try:
            inventory_ids, inventory_warnings = await _resolve_option_ids(
                client,
                path=OPENX_OPTIONS_PATHS["inventory_categories"],
                values=raw_inventory_values,
                field_label="inventory category",
                allow_contains_match=True,
            )
            warnings.extend(inventory_warnings)
            # Set the full metacategory block so the create-side defaulting at
            # ox_create_deal (which only fires when "metacategory" not in targeting)
            # doesn't strip exclude_mfa or the other policy defaults. Inventory category
            # includes use a "OR" op with a list value — distinct from the comma-string
            # shape used by IAB categories under domain.
            targeting["metacategory"] = {
                "includes": {"op": "OR", "val": inventory_ids},
                "excludes": [],
                "keywords": None,
                "exclude_mfa": True,
                "inter_dimension_operator": "AND",
            }
        except (LookupError, ValueError) as exc:
            blockers.append(_make_blocker("unresolved_inventory_category", str(exc), values=raw_inventory_values))

    raw_audience_input = targeting_brief.get("audience")
    audience_segments = None
    audience_alias_used: str | None = None
    for alias in (
        "audience_segments",
        "audience_segment",
        "audience_includes",
        "audience_segments_include",
        "audience_segments_includes",
    ):
        value = targeting_brief.get(alias)
        if value:
            audience_segments = value
            audience_alias_used = alias
            break
    if audience_segments is None and raw_audience_input is not None and not isinstance(raw_audience_input, dict):
        audience_segments = raw_audience_input
    # Brief-shape audience dict — e.g. the multi-deal protocol emits
    #   {"openaudience_custom": {"includes": [<segment_name>, ...]}}
    # which is the convenient input shape for the brief, but it's NOT the
    # GraphQL wire shape (which is `{op: INTERSECTS, val: <audience_id>}`).
    # Extract the segment names so the existing resolver below converts
    # them to the wire shape; otherwise the unresolved-names dict ends up
    # in the create payload and `dealCreate` rejects with
    #   "Field "includes" is not defined by type
    #    "TargetingOpenaudienceCustomCreateParams"."
    if audience_segments is None and isinstance(raw_audience_input, dict):
        openaudience_custom = raw_audience_input.get("openaudience_custom")
        if isinstance(openaudience_custom, dict):
            includes_value = openaudience_custom.get("includes")
            if includes_value:
                audience_segments = includes_value
                audience_alias_used = "audience.openaudience_custom.includes"
    if audience_alias_used and audience_alias_used not in {"audience_segments", "audience_segment"}:
        warnings.append(
            f"Accepted '{audience_alias_used}' as alias for 'audience_segments'. "
            "Use 'audience_segments' going forward for forward compatibility."
        )
    if isinstance(audience_segments, str):
        audience_segments = [audience_segments]
    if audience_segments:
        if len(audience_segments) != 1:
            blockers.append(
                _make_blocker(
                    "unsupported_audience_count",
                    "Only one audience segment is currently supported per prepared deal.",
                    values=audience_segments,
                )
            )
        else:
            try:
                audience_option, audience_warnings = await _resolve_option(
                    client,
                    path=OPENX_OPTIONS_PATHS["audience"],
                    value=str(audience_segments[0]),
                    field_label="audience segment",
                    allow_contains_match=True,
                )
                warnings.extend(audience_warnings)
                targeting.setdefault("audience", {})["openaudience_custom"] = {
                    "op": "INTERSECTS",
                    "val": _normalize_openaudience_id(audience_option["id"]),
                }
                if _audience_export_type_requires_us(audience_option) and not _geographic_includes_only_us(
                    targeting.get("geographic")
                ):
                    blockers.append(
                        _make_blocker(
                            "audience_export_type_requires_us_geo",
                            "This OpenX audience export type can target audiences only in US geography.",
                            value=audience_segments[0],
                            export_type=(audience_option.get("extra") or {}).get("export_type"),
                        )
                    )
            except (LookupError, ValueError) as exc:
                blockers.append(_make_blocker("unresolved_audience", str(exc), value=audience_segments[0]))

    if _targeting_has_audience(targeting) and not _geographic_includes_explicit_country(targeting.get("geographic")):
        blockers.append(
            _make_blocker(
                "audience_requires_explicit_country",
                "Audience targeting requires an explicit geographic.includes.country.",
            )
        )

    if publisher_ids:
        normalized_publisher_ids = [
            str(publisher_id).strip() for publisher_id in publisher_ids if str(publisher_id).strip()
        ]
        if normalized_publisher_ids:
            targeting.setdefault("content", {})["account"] = {
                "op": "INTERSECTS",
                "val": ",".join(normalized_publisher_ids),
            }
    elif excluded_publisher_ids:
        # Wire format mirrors the UI deal JSON observed on Reklaim CTV deals:
        # targeting.content.account = {op: "NOT INTERSECTS", val: "<comma-list>"}.
        normalized_excluded_publisher_ids = [
            str(publisher_id).strip() for publisher_id in excluded_publisher_ids if str(publisher_id).strip()
        ]
        if normalized_excluded_publisher_ids:
            targeting.setdefault("content", {})["account"] = {
                "op": "NOT INTERSECTS",
                "val": ",".join(normalized_excluded_publisher_ids),
            }

    requested_targeting_fields = {
        key
        for key in (
            "rendering_context",
            "technographic",
            "domain",
            "audience",
            "language",
            "languages",
            "iab_categories",
            "domain_categories_iab_v2",
            "audience_segments",
            "audience_segment",
            "audience_includes",
            "audience_segments_include",
            "audience_segments_includes",
            "inventory_categories",
        )
        if targeting_brief.get(key)
    }
    applied_targeting_fields = set(targeting.keys())
    omitted_targeting_fields = sorted(
        requested_targeting_fields
        - applied_targeting_fields
        - {
            "language",
            "languages",
            "iab_categories",
            "domain_categories_iab_v2",
            "audience_segments",
            "audience_segment",
            "audience_includes",
            "audience_segments_include",
            "audience_segments_includes",
            "inventory_categories",
        }
    )
    if omitted_targeting_fields:
        warnings.append("Preserved targeting omitted from prepared payload: " + ", ".join(omitted_targeting_fields))

    if viewability_threshold is not None:
        if not (0 < viewability_threshold <= 1):
            if 1 < viewability_threshold <= 100:
                viewability_threshold = viewability_threshold / 100.0
                warnings.append(
                    f"Normalized viewability threshold from percentage to decimal: {viewability_threshold:.2f}"
                )
            else:
                blockers.append(
                    _make_blocker(
                        "invalid_viewability_threshold",
                        f"Viewability threshold must be between 0 and 100 (or 0.0 and 1.0): {viewability_threshold}",
                    )
                )
        if not blockers or all(b["code"] != "invalid_viewability_threshold" for b in blockers):
            targeting["viewability"] = {
                "viewability_score": {
                    "op": ">=",
                    "val": f"{viewability_threshold:.2f}",
                }
            }

    if app_bundles:
        normalized_app_bundles: list[str] = []
        seen_app_bundles: set[str] = set()
        for bundle in app_bundles:
            bundle_value = str(bundle).strip()
            if bundle_value and bundle_value not in seen_app_bundles:
                seen_app_bundles.add(bundle_value)
                normalized_app_bundles.append(bundle_value)
        if normalized_app_bundles:
            # App bundles are their OWN OpenX targeting dimension, distinct from
            # web-domain url_targeting. Verified against UI-created deals:
            #   targeting.app_inventory.app_bundle_id = {op:"OR", val:[{op:"==", val:b}, ...]}
            # OpenX auto-generates the OXTL bridge so web inventory (which carries no
            # bundle) still serves: "(... app.bundle intersects "…" OR not
            # (distribution_channel intersects "APP"))". Bundle values are NOT url shapes
            # (reverse-DNS package names, numeric Apple App Store IDs, Amazon ASINs, Roku
            # ids) and must never be routed into url_targeting, which silently drops them.
            targeting["app_inventory"] = {
                "app_bundle_id": {
                    "op": "OR",
                    "val": [{"op": "==", "val": bundle} for bundle in normalized_app_bundles],
                },
                "app_inventory_inter_dimension_operator": "AND",
            }

    participant: dict[str, Any] = {"demand_partner": resolved_demand_partner_id}
    if normalized_buyer_ids:
        participant["buyer_ids"] = normalized_buyer_ids
        if _should_set_main_buyer_id(resolved_demand_partner_id, normalized_buyer_ids):
            participant["main_buyer_id"] = normalized_buyer_ids[0]

    url_targeting = None
    if normalized_allowlist:
        url_targeting = {"allowlist": normalized_allowlist}
    elif normalized_blocklist:
        url_targeting = {"blocklist": normalized_blocklist}
    if url_targeting and domain_targeting_option:
        url_targeting["domain_targeting_option"] = domain_targeting_option

    create_args = {
        "name": name,
        "currency": currency,
        "deal_price": deal_price,
        "start_date": start_date,
        "deal_participants": [participant],
        "package_name": package_name,
        "targeting": targeting,
        "pmp_deal_type": pmp_deal_type,
        "url_targeting": url_targeting,
        "third_party_fees_config": resolved_fee,
        "status": status,
        "end_date": end_date,
        "expected_ad_category": expected_ad_category,
    }

    quality_flags.extend(_blockers_to_ox_quality_flags(blockers))

    prepared_deal_id = f"openx-prepared-{uuid4()}"
    prepared_deal = {
        "prepared_deal_id": prepared_deal_id,
        "ready_to_create": not blockers,
        "blocking_issues": [blocker["message"] for blocker in blockers],
        "blockers": blockers,
        "warnings": warnings,
        "quality_flags": quality_flags,
        "resolved_entities": {
            "demand_partner": {
                "input": demand_partner,
                "resolved_id": resolved_demand_partner_id,
                "resolved_name": resolved_partner.get("name") if resolved_partner else None,
            },
            "buyer_ids": normalized_buyer_ids or [],
            "fee_partner_id": resolved_fee.get("partner_id") if resolved_fee else None,
            "language": targeting.get("technographic", {}).get("language")
            if isinstance(targeting.get("technographic"), dict)
            else None,
            "iab_categories": targeting.get("domain", {}).get("categories_iab_v2")
            if isinstance(targeting.get("domain"), dict)
            else None,
            "audience": targeting.get("audience", {}).get("openaudience_custom")
            if isinstance(targeting.get("audience"), dict)
            else None,
        },
        "invalid_domains": {
            "allowlist": invalid_allowlist,
            "blocklist": invalid_blocklist,
        },
        "create_args": create_args,
    }
    _prepared_openx_deals[prepared_deal_id] = prepared_deal
    return prepared_deal


# =============================================================================
# MCP Tools
# =============================================================================


@mcp.tool()
async def ox_create_deal(
    name: str,
    currency: str,
    deal_price: float,
    start_date: str,
    deal_participants: list[dict[str, Any]],
    targeting: dict[str, Any],
    package_name: str | None = None,
    pmp_deal_type: str = "3",
    url_targeting: dict[str, Any] | None = None,
    third_party_fees_config: dict[str, Any] | list[dict[str, Any]] | None = None,
    end_date: str | None = None,
    status: str | None = None,
    expected_ad_category: str | None = None,
) -> dict[str, Any]:
    """
    Create a new programmatic deal on OpenX.

    This is a CRITICAL action that requires self-audit confirmation before execution.
    This is the low-level write tool. For normal prompt-driven workflows, prefer:
    - ox_prepare_deal_from_prompt_inputs
    - ox_prepare_deal_from_brief
    - ox_create_prepared_deal
    - ox_execute_deal_from_prompt_inputs

    Use this direct tool mainly for debugging or controlled low-level calls.

    Required MCP inputs for this tool:
    - name
    - currency
    - deal_price
    - start_date
    - deal_participants
    - targeting

    `package_name` is optional — omit to let the MCP auto-generate a unique
    OpenX-UI-style name (preferred for automated flows; avoids orphan-package
    not-unique conflicts on retries).

    **Deal Naming Convention:**
    Deals should follow the structured naming pattern:
    `Elcano_OpenX_<DemandPartner>_<TargetingType>_<DealID>_<Version>`

    Example: `Elcano_OpenX_Crimtan_US_CuratedDomains_ELC00001_A0`

    **Curated Deal Fees (third_party_fees_config):**
    For curated deals, configure third-party fees using:
    ```
    {
        "partner_id": "your-partner-id",       # Required: third-party partner identifier
        "revenue_method": "PoM",               # Required: "CPM" or "PoM" (Percent of Media)
        "gross_share": "30.0",                # Decimal fields must be strings
        "gross_cpm_cap": "5.0"                # Optional Decimal string CPM cap
    }
    ```

    Args:
        name: Deal name following the naming convention
        currency: Currency code (e.g., "USD", "EUR")
        deal_price: CPM price for the deal (sent as Decimal string with 2 decimal places)
        start_date: ISO 8601 formatted start date (e.g., "2024-01-15T00:00:00Z")
        deal_participants: List of demand partner configurations. Each should contain:
            - demand_partner: The partner's ID (also accepts demand_partner_id, which
              will be automatically renamed to demand_partner)
            - Optional additional configuration (e.g., buyer_ids, brand_ids)
        package_name: Optional. Name for the underlying OpenX package containing the
            targeting rules. When omitted (or empty), the MCP auto-generates a unique
            name in OpenX UI format `Package for {deal: <name>} [<ISO timestamp Z>]`,
            mirroring how UI-created deals name their packages and guaranteeing
            uniqueness so an orphan package from a prior failed retry cannot block a
            new attempt. Pass an explicit name only when there is a reason to control
            it (e.g. matching an existing custom-named package).
        targeting: TargetingCreateParams object. Common fields:
            - geographic: List of country codes (e.g., ["US", "CA"]) — automatically
              converted to {"includes": {"country": "US,CA"}} (comma-delimited string)
            - device_type: List of device types (e.g., ["DESKTOP", "MOBILE", "CTV"]) —
              automatically converted into rendering_context.device_type V2 fields
            - rendering_context: Required by OpenX. If omitted, the tool auto-builds a
              default from targeting.channel (DISPLAY, OLV, CTV) or defaults to DISPLAY.
            - Additional targeting parameters
            Note: OpenX MCP always sends "inter_dimension_operator": "AND" so deals use
            "Use Unique Targeting Settings".
        pmp_deal_type: PMP deal type as a numeric string code per OpenRTB 2.5:
            "3" = Preferred Deal (default), "2" = Private Auction,
            "1" = Programmatic Guaranteed.
            Also accepts human-readable names ("PREFERRED_DEAL", "PRIVATE_AUCTION",
            "PROGRAMMATIC_GUARANTEED") which are automatically mapped to numeric codes.
        url_targeting: URLTargetingCreateParams for domain targeting. Optional.
            Convenience format:
            - allowlist: List of allowed domains (converted to type="whitelist", urls=[...])
            - blocklist: List of blocked domains (converted to type="blacklist", urls=[...])
            Raw API format:
            - type: "whitelist" or "blacklist"
            - urls: List of domain names
            - domain_targeting_option: "SUBDOMAIN" or "ROOT" (optional)
        third_party_fees_config: ExchangeThirdPartyFeesConfigCreateParams. Optional.
            Wrapped in a list automatically if a single dict is provided.
            - partner_id: Third-party partner identifier (required)
            - revenue_method: "CPM" or "PoM" (Percent of Media) (required)
            - gross_share: Percentage share (for PoM), sent as a string
            - gross_cpm_cap: CPM cap amount, sent as a string
            - platform_share: Platform share, sent as a string
        end_date: Optional ISO 8601 formatted end date
        expected_ad_category: Optional OpenX Expected Sensitive Category (top-level free-text
            enum, e.g. "Politics"). Required by policy for political deals and others the
            trader brief flags as sensitive. OpenX validates the value server-side.

    Returns:
        Dictionary containing:
            - success: Boolean indicating if the deal was created
            - deal: The created deal object (if successful)
            - error: Error message (if failed)
    """
    logger.info(f"create_deal called with name: {name}")

    try:
        client = get_openx_client()

        # Resolve pmp_deal_type: map human-readable names to numeric codes
        resolved_pmp_deal_type = PMP_DEAL_TYPE_MAP.get(pmp_deal_type.upper(), pmp_deal_type)

        # Normalize deal_participants: use "demand_partner" as the key name
        normalized_participants = []
        for participant in deal_participants:
            normalized = dict(participant)
            if "demand_partner_id" in normalized:
                normalized["demand_partner"] = normalized.pop("demand_partner_id")
            normalized_participants.append(normalized)

        # Discover the geographic field name via introspection (cached after first call)
        geo_field_name = await client.get_geographic_field_name()

        device_types = None
        if isinstance(targeting.get("device_type"), (list, dict)):
            device_types = targeting.get("device_type")

        # Restructure targeting: add inter_dimension_operator and transform geographic
        normalized_targeting: dict[str, Any] = {
            "inter_dimension_operator": "AND",
            "rendering_context": _build_rendering_context(targeting, device_types),
        }

        if "metacategory" not in targeting:
            normalized_targeting["metacategory"] = {
                "excludes": [],
                "includes": None,
                "keywords": None,
                "exclude_mfa": True,
                "inter_dimension_operator": "AND",
            }

        for key, value in targeting.items():
            if key == "geographic" and isinstance(value, list):
                # Per API docs: includes is a single TargetingGeographicItemCreateParams object.
                # Multiple countries use comma-delimited string: "US,CA,UK"
                normalized_targeting["geographic"] = {
                    "includes": {geo_field_name: ",".join(value)},
                }
            elif key == "geographic" and isinstance(value, dict):
                normalized_geographic, _ = _normalize_geographic_targeting(value)
                normalized_targeting["geographic"] = normalized_geographic
            elif key in {"device_type", "rendering_context", "channel", "inter_dimension_operator"}:
                continue
            elif key == "technographic":
                # Already structured — pass through
                technographic = dict(value)
                technographic.pop("device_type", None)
                if technographic:
                    normalized_targeting["technographic"] = technographic
            else:
                normalized_targeting[key] = value

        # Generate a unique OpenX-UI-style package name when caller omits it (or passes
        # empty). Keeps the canonical deal `name` clean while ensuring package uniqueness
        # so a prior failed dealCreate (which leaves orphan packages) cannot block retries.
        effective_package_name = (package_name or "").strip() or _default_package_name(name)

        # Build the DealCreateParams structure
        deal_params: dict[str, Any] = {
            "name": name,
            "currency": currency,
            "deal_price": f"{deal_price:.2f}",
            "pmp_deal_type": resolved_pmp_deal_type,
            "start_date": start_date,
            "deal_participants": normalized_participants,
            "package": {
                "name": effective_package_name,
                "targeting": normalized_targeting,
            },
        }

        # Add optional URL targeting with allowlist/blocklist normalization
        if url_targeting:
            normalized_url_targeting = dict(url_targeting)
            # Normalize allowlist/blocklist to the API's type+urls format
            if "allowlist" in normalized_url_targeting:
                normalized_url_targeting["type"] = "whitelist"
                normalized_url_targeting["urls"] = normalized_url_targeting.pop("allowlist")
            elif "blocklist" in normalized_url_targeting:
                normalized_url_targeting["type"] = "blacklist"
                normalized_url_targeting["urls"] = normalized_url_targeting.pop("blocklist")
            deal_params["package"]["url_targeting"] = normalized_url_targeting

        # Add optional end date
        if end_date:
            deal_params["end_date"] = end_date

        # Add optional third-party fees (must be a list per API schema)
        normalized_fees_config = _normalize_third_party_fees_config(third_party_fees_config)
        if normalized_fees_config:
            deal_params["third_party_fees_config"] = normalized_fees_config

        # Add optional sensitive-category declaration (top-level free-text enum;
        # OpenX validates the value server-side). Required by policy for political
        # deals and others that the trader brief flags as sensitive.
        if expected_ad_category:
            deal_params["expected_ad_category"] = expected_ad_category

        # Status is handled via dealUpdate after creation, not in DealCreateParams
        requested_status = status

        logger.debug(f"Create deal payload: {json.dumps(deal_params, indent=2, default=str)}")

        payload_preview = _summarize_create_payload(deal_params)

        # Create the deal
        deal = await client.create_deal(deal_params)
        deal_url = _build_openx_deal_url(deal)

        # If a specific status was requested (e.g., "Paused"), update the deal after creation
        status_update_result = None
        if requested_status and deal.get("id"):
            try:
                status_update_result = await client.update_deal(deal["id"], {"status": requested_status})
                logger.info(f"Deal status updated to {requested_status}")
            except Exception as status_exc:
                logger.warning(f"Deal created but status update to '{requested_status}' failed: {status_exc}")
                status_update_result = {"error": str(status_exc)}

        logger.info(f"Deal created successfully: {deal.get('deal_id')}")
        return {
            "success": True,
            "deal": deal,
            "deal_url": deal_url,
            "status_update": status_update_result,
        }

    except OpenXGraphQLError as e:
        error_msg = f"Failed to create deal: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "error": error_msg,
            "error_details": {
                "operation_name": e.operation_name,
                "errors": e.errors,
            },
            "create_payload_preview": payload_preview,
            "create_payload": deal_params,
        }

    except Exception as e:
        error_msg = f"Failed to create deal: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "error": error_msg,
            "create_payload_preview": payload_preview if "payload_preview" in locals() else None,
            "create_payload": deal_params if "deal_params" in locals() else None,
        }


@mcp.tool()
async def ox_prepare_deal_from_brief(
    name: str,
    currency: str,
    deal_price: float,
    start_date: str,
    demand_partner: str,
    targeting: dict[str, Any] | None = None,
    publisher_ids: list[str] | None = None,
    excluded_publisher_ids: list[str] | None = None,
    buyer_ids: list[str] | None = None,
    end_date: str | None = None,
    pmp_deal_type: str = "3",
    fee: dict[str, Any] | None = None,
    domains_allowlist: list[str] | None = None,
    domains_blocklist: list[str] | None = None,
    domain_targeting_option: str | None = None,
    viewability_threshold: float | None = None,
    status: str | None = None,
    expected_ad_category: str | None = None,
    inventory_categories: list[str] | None = None,
    app_bundles: list[str] | None = None,
    package_name: str | None = None,
) -> dict[str, Any]:
    """Prepare an OpenX deal from business-facing inputs.

    Fee normalization note:
    - `fee.gross_share_percent=40` becomes OpenX `gross_share="0.4"`
    - `fee.gross_share="40.0"` is also normalized to `"0.4"`

    Geographic note:
    - Supply states/provinces as abbreviations or full names (for example `CA`, `Texas`, `AB`, `Alberta`), not raw numeric ids.
    - Subnational targeting requires an explicit matching country on the same includes/excludes branch.

    IAB note:
    - Pass OpenX IAB v2 category names or IDs directly (e.g. "Certified Pre-Owned Cars").
    - Legacy IAB codes like "IAB2-5" are not supported; use ox_list_iab_categories to
      discover the correct OpenX category names first.

    Viewability note:
    - viewability_threshold accepts percentage (70) or decimal (0.70).
    - Percentages > 1 are automatically normalized to decimals.

    Status note:
    - status can be "Paused" to create the deal in paused state instead of Active.

    Sensitive category note:
    - ``expected_ad_category`` declares the OpenX Expected Sensitive Category (top-level
      free-text enum, e.g. ``"Politics"``). Required by policy for political deals and
      others the trader brief flags as sensitive. OpenX validates the value server-side.

    Publisher targeting note:
    - ``publisher_ids`` adds an INTERSECTS clause: only inventory from these publisher ids
      will be eligible.
    - ``excluded_publisher_ids`` adds a NOT INTERSECTS clause: inventory from these
      publisher ids will be excluded.
    - These are mutually exclusive on a single deal — OpenX ``content.account`` is one
      object with one ``op``. Briefs needing both inclusion AND exclusion must split into
      two deals.

    Inventory category note:
    - ``inventory_categories`` declares OpenX inventory metacategories to include
      (e.g. ``["premiumctv"]``). Names are resolved via ``ox_list_options_by_path`` on
      the metacategory path; pass human-readable names or codes — the resolver handles
      both. Wire format: ``targeting.metacategory.includes = {op: "OR", val: [...]}``.
      Setting this also preserves OpenX policy defaults (``exclude_mfa: True``, etc.).

    App-bundle note:
    - ``app_bundles`` targets in-app inventory by store identifier and is a SEPARATE
      dimension from web-domain ``domains_allowlist`` — a deal may carry both. Pass a
      clean list of bundle ids (reverse-DNS like ``com.fubotv.vix``, numeric Apple App
      Store ids like ``162057``, Amazon ASINs, Roku ids). Wire format:
      ``targeting.app_inventory.app_bundle_id = {op:"OR", val:[{op:"==", val:b}, ...]}``.
      These are NOT url_targeting values and must not be passed via domains_allowlist.
    """
    logger.info("ox_prepare_deal_from_brief called with name: %s", name)

    try:
        client = get_openx_client()
        prepared_deal = await _build_prepared_openx_deal(
            client=client,
            name=name,
            currency=currency,
            deal_price=deal_price,
            start_date=start_date,
            package_name=package_name,
            demand_partner=demand_partner,
            targeting_brief=targeting,
            publisher_ids=publisher_ids,
            excluded_publisher_ids=excluded_publisher_ids,
            buyer_ids=buyer_ids,
            end_date=end_date,
            pmp_deal_type=pmp_deal_type,
            fee=fee,
            domains_allowlist=domains_allowlist,
            domains_blocklist=domains_blocklist,
            domain_targeting_option=domain_targeting_option,
            viewability_threshold=viewability_threshold,
            status=status,
            expected_ad_category=expected_ad_category,
            inventory_categories=inventory_categories,
            app_bundles=app_bundles,
        )
        return {
            "success": True,
            "prepared_deal_id": prepared_deal["prepared_deal_id"],
            "ready_to_create": prepared_deal["ready_to_create"],
            "blocking_issues": prepared_deal["blocking_issues"],
            "blockers": prepared_deal["blockers"],
            "warnings": prepared_deal["warnings"],
            "resolved_entities": prepared_deal["resolved_entities"],
            "invalid_domains": prepared_deal["invalid_domains"],
            "create_args_preview": prepared_deal["create_args"],
        }
    except Exception as e:
        error_msg = f"Failed to prepare OpenX deal: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "error": error_msg,
        }


@mcp.tool()
async def ox_prepare_deal_from_prompt_inputs(
    name: str,
    currency: str,
    deal_price: float,
    start_date: str,
    demand_partner: str,
    domain_file_path: str | None = None,
    targeting: dict[str, Any] | None = None,
    publisher_file_path: str | None = None,
    publisher_sheet: str | None = None,
    publisher_column: str | None = None,
    publisher_ids: list[str] | None = None,
    excluded_publisher_ids: list[str] | None = None,
    buyer_ids: list[str] | None = None,
    end_date: str | None = None,
    pmp_deal_type: str = "3",
    fee: dict[str, Any] | None = None,
    domain_sheet: str | None = None,
    domain_column: str | None = None,
    domain_targeting_option: str | None = None,
    domain_match_operator: Literal["blocklist", "allowlist"] | None = None,
    app_bundle_file_path: str | None = None,
    app_bundle_sheet: str | None = None,
    app_bundle_column: str | None = None,
    app_bundle_match_operator: Literal["blocklist", "allowlist"] | None = None,
    viewability_threshold: float | None = None,
    status: str | None = None,
    expected_ad_category: str | None = None,
    inventory_categories: list[str] | None = None,
    package_name: str | None = None,
) -> dict[str, Any]:
    """Prepare an OpenX deal from prompt fields plus an attached domain file.

    Supported domain file formats:
    - `.xlsx` / `.xlsm` / other openpyxl-supported Excel workbooks
    - `.csv` files containing a domain column

    Supported publisher file formats:
    - `.xlsx` / `.xlsm` / other openpyxl-supported Excel workbooks
    - `.csv` files containing a publisher/account id column

    Fee normalization note:
    - `fee.gross_share_percent=40` becomes OpenX `gross_share="0.4"`
    - `fee.gross_share="40.0"` is also normalized to `"0.4"`

    Geographic note:
    - Supply states/provinces as abbreviations or full names (for example `CA`, `Texas`, `AB`, `Alberta`), not raw numeric ids.
    - Subnational targeting requires an explicit matching country on the same includes/excludes branch.

    IAB note:
    - Pass OpenX IAB v2 category names or IDs directly (e.g. "Certified Pre-Owned Cars").
    - Legacy IAB codes like "IAB2-5" are not supported; use ox_list_iab_categories to
      discover the correct OpenX category names first.

    Viewability note:
    - viewability_threshold accepts percentage (70) or decimal (0.70).

    Status note:
    - status can be "Paused" to create the deal in paused state.

    Domain match note:
    - When ``domain_file_path`` is supplied, ``domain_match_operator`` is REQUIRED. There is no
      default — the caller MUST explicitly state whether the file is a ``"blocklist"`` (domains
      to exclude) or an ``"allowlist"`` (domains to include). Briefs ship both list types and
      neither is privileged; omitting the operator returns a structured blocker so a domain
      file is never silently routed to the wrong wire type.

    App-bundle note:
    - For in-app inventory, use ``app_bundle_file_path`` / ``app_bundle_sheet`` /
      ``app_bundle_column`` / ``app_bundle_match_operator``. App bundles are a SEPARATE
      OpenX targeting dimension (``targeting.app_inventory.app_bundle_id``) from web-domain
      ``url_targeting`` — a single deal MAY supply BOTH (domains via ``domain_file_path``,
      bundles via ``app_bundle_file_path``); they no longer conflict. Bundle ids are taken
      verbatim (reverse-DNS like ``com.foo.bar``, numeric Apple App Store ids, Amazon ASINs,
      Roku ids) — they are NOT url shapes and must never go through ``domain_file_path``,
      which would silently drop the non-FQDN values. Only an allowlist (include set) is
      supported today; ``app_bundle_match_operator='blocklist'`` returns a structured blocker.

    Sensitive category note:
    - ``expected_ad_category`` declares the OpenX Expected Sensitive Category (top-level
      free-text enum, e.g. ``"Politics"``). Required by policy for political deals and
      others the trader brief flags as sensitive. OpenX validates the value server-side.
    """
    logger.info("ox_prepare_deal_from_prompt_inputs called with name: %s file: %s", name, domain_file_path)

    try:
        targeting = dict(targeting or {})
        channel = str(targeting.get("channel") or "").strip().upper()

        # App bundles are a DISTINCT OpenX targeting dimension
        # (targeting.app_inventory.app_bundle_id) — NOT web-domain url_targeting.
        # Extract them up front so a deal may carry BOTH a domain allowlist
        # (url_targeting) and an app-bundle list (app_inventory) at once, which OpenX
        # supports. Bundle ids (reverse-DNS, numeric Apple ids, Amazon ASINs, Roku ids)
        # are NOT valid url_targeting shapes and would be silently dropped there.
        extracted_app_bundles: list[str] = []
        app_bundle_source: dict[str, Any] | None = None
        if app_bundle_file_path is not None:
            if app_bundle_match_operator is not None and str(app_bundle_match_operator).lower() == "blocklist":
                bundle_blocklist_blocker = _make_blocker(
                    "app_bundle_blocklist_unsupported",
                    "App-bundle blocklists are not yet supported — OpenX app_inventory.app_bundle_id "
                    "targets an include (allowlist) set. Pass app_bundle_match_operator='allowlist' or omit it.",
                )
                return {
                    "success": True,
                    "ready_to_create": False,
                    "blocking_issues": [bundle_blocklist_blocker["message"]],
                    "blockers": [bundle_blocklist_blocker],
                    "warnings": [],
                    "quality_flags": _blockers_to_ox_quality_flags([bundle_blocklist_blocker]),
                }
            if not os.path.exists(app_bundle_file_path):
                bundle_file_blocker = _make_blocker(
                    "missing_app_bundle_file", f"App-bundle file not found: {app_bundle_file_path}"
                )
                return {
                    "success": True,
                    "ready_to_create": False,
                    "blocking_issues": [bundle_file_blocker["message"]],
                    "blockers": [bundle_file_blocker],
                    "warnings": [],
                    "quality_flags": _blockers_to_ox_quality_flags([bundle_file_blocker]),
                }
            bundle_result = _extract_app_bundles_from_file(
                file_path=app_bundle_file_path,
                sheet_name=app_bundle_sheet,
                column_name=app_bundle_column,
            )
            extracted_app_bundles = bundle_result["app_bundles"]
            app_bundle_source = {
                "file_path": app_bundle_file_path,
                "sheet_name": bundle_result["sheet_name"],
                "column_name": bundle_result["column_name"],
                "row_count": bundle_result["row_count"],
            }
            if not extracted_app_bundles:
                empty_bundle_blocker = _make_blocker(
                    "missing_app_bundle_ids",
                    f"No app bundle ids found in file: {app_bundle_file_path}",
                )
                return {
                    "success": True,
                    "ready_to_create": False,
                    "blocking_issues": [empty_bundle_blocker["message"]],
                    "blockers": [empty_bundle_blocker],
                    "warnings": [],
                    "quality_flags": _blockers_to_ox_quality_flags([empty_bundle_blocker]),
                }

        resolved_publisher_ids = publisher_ids
        resolved_publisher_file_path = publisher_file_path
        publisher_source = None
        publisher_auto_detection_warning = None
        if (
            not resolved_publisher_ids
            and channel == "CTV"
            and not resolved_publisher_file_path
            and not domain_file_path
            and not extracted_app_bundles
        ):
            resolved_publisher_file_path, publisher_auto_detection_blocker = _autodetect_single_xlsx_attachment()
            if resolved_publisher_file_path:
                publisher_auto_detection_warning = f"Auto-detected publisher file: {resolved_publisher_file_path}"
        else:
            publisher_auto_detection_blocker = None

        if publisher_auto_detection_blocker:
            return {
                "success": True,
                "ready_to_create": False,
                "blocking_issues": [publisher_auto_detection_blocker["message"]],
                "blockers": [publisher_auto_detection_blocker],
                "warnings": [],
                "quality_flags": _blockers_to_ox_quality_flags([publisher_auto_detection_blocker]),
            }

        if resolved_publisher_file_path:
            if not os.path.exists(resolved_publisher_file_path):
                pub_file_blocker = _make_blocker(
                    "missing_publisher_file", f"Publisher file not found: {resolved_publisher_file_path}"
                )
                return {
                    "success": True,
                    "ready_to_create": False,
                    "blocking_issues": [pub_file_blocker["message"]],
                    "blockers": [pub_file_blocker],
                    "warnings": [],
                    "quality_flags": _blockers_to_ox_quality_flags([pub_file_blocker]),
                }

            publisher_file_result = _extract_publisher_ids_from_file(
                file_path=resolved_publisher_file_path,
                sheet_name=publisher_sheet,
                column_name=publisher_column,
            )
            resolved_publisher_ids = publisher_file_result["publisher_ids"]
            publisher_source = {
                "file_path": resolved_publisher_file_path,
                "sheet_name": publisher_file_result["sheet_name"],
                "column_name": publisher_file_result["column_name"],
                "row_count": publisher_file_result["row_count"],
            }
            if not resolved_publisher_ids:
                missing_pub_blocker = _make_blocker(
                    "missing_publisher_ids",
                    f"No publisher ids found in file: {resolved_publisher_file_path}",
                )
                return {
                    "success": True,
                    "ready_to_create": False,
                    "blocking_issues": [missing_pub_blocker["message"]],
                    "blockers": [missing_pub_blocker],
                    "warnings": [] if not publisher_auto_detection_warning else [publisher_auto_detection_warning],
                    "quality_flags": _blockers_to_ox_quality_flags([missing_pub_blocker]),
                }

        resolved_domain_file_path = domain_file_path
        auto_detection_warning = None
        auto_detection_blocker = None
        if not resolved_domain_file_path and not resolved_publisher_ids and not extracted_app_bundles:
            resolved_domain_file_path, auto_detection_blocker = _autodetect_single_xlsx_attachment()
            if resolved_domain_file_path:
                auto_detection_warning = f"Auto-detected domain file: {resolved_domain_file_path}"

        if auto_detection_blocker:
            return {
                "success": True,
                "ready_to_create": False,
                "blocking_issues": [auto_detection_blocker["message"]],
                "blockers": [auto_detection_blocker],
                "warnings": [],
                "quality_flags": _blockers_to_ox_quality_flags([auto_detection_blocker]),
            }

        if resolved_domain_file_path and not os.path.exists(resolved_domain_file_path):
            domain_file_blocker = _make_blocker(
                "missing_domain_file", f"Domain file not found: {resolved_domain_file_path}"
            )
            return {
                "success": True,
                "ready_to_create": False,
                "blocking_issues": [domain_file_blocker["message"]],
                "blockers": [domain_file_blocker],
                "warnings": [],
                "quality_flags": _blockers_to_ox_quality_flags([domain_file_blocker]),
            }

        if not resolved_domain_file_path and not resolved_publisher_ids and not extracted_app_bundles:
            missing_attach_blocker = _make_blocker(
                "missing_prompt_input_attachment",
                "A domain file, app-bundle file, or publisher ids are required for prompt-input preparation.",
            )
            return {
                "success": True,
                "ready_to_create": False,
                "blocking_issues": [missing_attach_blocker["message"]],
                "blockers": [missing_attach_blocker],
                "warnings": [],
                "quality_flags": _blockers_to_ox_quality_flags([missing_attach_blocker]),
            }

        file_result = None
        if resolved_domain_file_path:
            file_result = _extract_domains_from_file(
                file_path=resolved_domain_file_path,
                sheet_name=domain_sheet,
                column_name=domain_column,
            )

        extracted_domains = file_result["domains"] if file_result else None
        domains_allowlist_arg: list[str] | None = None
        domains_blocklist_arg: list[str] | None = None
        if extracted_domains:
            if domain_match_operator is None:
                missing_operator_blocker = _make_blocker(
                    "missing_domain_match_operator",
                    "A domain file was supplied without domain_match_operator. Pass 'blocklist' "
                    "(domains to exclude) or 'allowlist' (domains to include) explicitly — there "
                    "is no default because both list types are equally legitimate per brief.",
                    file_path=resolved_domain_file_path,
                )
                return {
                    "success": True,
                    "ready_to_create": False,
                    "blocking_issues": [missing_operator_blocker["message"]],
                    "blockers": [missing_operator_blocker],
                    "warnings": [],
                    "quality_flags": _blockers_to_ox_quality_flags([missing_operator_blocker]),
                }
            operator = str(domain_match_operator).lower()
            if operator not in {"blocklist", "allowlist"}:
                raise ValueError(
                    f"Unsupported domain_match_operator '{domain_match_operator}'. Pass 'blocklist' or 'allowlist'."
                )
            if operator == "allowlist":
                domains_allowlist_arg = extracted_domains
            else:
                domains_blocklist_arg = extracted_domains

        client = get_openx_client()
        prepared_deal = await _build_prepared_openx_deal(
            client=client,
            name=name,
            currency=currency,
            deal_price=deal_price,
            start_date=start_date,
            package_name=package_name,
            demand_partner=demand_partner,
            targeting_brief=targeting,
            publisher_ids=resolved_publisher_ids,
            excluded_publisher_ids=excluded_publisher_ids,
            buyer_ids=buyer_ids,
            end_date=end_date,
            pmp_deal_type=pmp_deal_type,
            fee=fee,
            domains_allowlist=domains_allowlist_arg,
            domains_blocklist=domains_blocklist_arg,
            domain_targeting_option=domain_targeting_option,
            viewability_threshold=viewability_threshold,
            status=status,
            expected_ad_category=expected_ad_category,
            inventory_categories=inventory_categories,
            app_bundles=extracted_app_bundles or None,
        )

        warnings = list(prepared_deal["warnings"])
        if publisher_auto_detection_warning:
            warnings.insert(0, publisher_auto_detection_warning)
        if auto_detection_warning:
            warnings.insert(0, auto_detection_warning)
        quality_flags = list(prepared_deal.get("quality_flags", []))
        if file_result and file_result["invalid_values"]:
            dropped_count = len(file_result["invalid_values"])
            warnings.append(f"Dropped {dropped_count} non-domain values from the attached file during extraction.")
            # Emit a structured quality_flag when a significant share of the file was dropped
            # so the trader doesn't miss a mostly-failed extraction buried in a warning string.
            # The Optimum political-compliance file (Reklaim mixed-inventory list) dropped 425
            # of 1,185 rows (~36%) as app bundle IDs — silent in warnings, loud as a flag.
            valid_count = len(file_result["domains"])
            total_count = dropped_count + valid_count
            if total_count > 0:
                dropped_percent = round((dropped_count / total_count) * 100, 1)
                if dropped_percent > 10:
                    quality_flags.append(
                        _make_ox_quality_flag(
                            "ox_domain_extraction_partial",
                            (
                                f"Domain file extraction dropped {dropped_count} of {total_count} rows "
                                f"({dropped_percent}%) as non-domain values. The dropped rows are likely "
                                "app bundle IDs or other non-URL identifiers that OpenX url_targeting "
                                "does not accept. If app-bundle targeting was intended, pass those entries "
                                "via app_bundle_file_path (targeting.app_inventory.app_bundle_id) instead — "
                                "it accepts reverse-DNS bundles, numeric store IDs, ASINs, and Roku ids."
                            ),
                            dropped_count=dropped_count,
                            valid_count=valid_count,
                            dropped_percent=dropped_percent,
                            file_path=resolved_domain_file_path,
                        )
                    )
        _prepared_openx_deals[prepared_deal["prepared_deal_id"]]["warnings"] = warnings
        _prepared_openx_deals[prepared_deal["prepared_deal_id"]]["quality_flags"] = quality_flags

        return {
            "success": True,
            "prepared_deal_id": prepared_deal["prepared_deal_id"],
            "ready_to_create": prepared_deal["ready_to_create"],
            "blocking_issues": prepared_deal["blocking_issues"],
            "blockers": prepared_deal["blockers"],
            "warnings": warnings,
            "quality_flags": quality_flags,
            "extracted_domain_count": len(file_result["domains"]) if file_result else 0,
            "invalid_domain_rows": file_result["invalid_values"] if file_result else [],
            "domain_source": {
                "file_path": resolved_domain_file_path,
                "sheet_name": file_result["sheet_name"],
                "column_name": file_result["column_name"],
                "row_count": file_result["row_count"],
            }
            if file_result
            else None,
            "extracted_publisher_count": len(resolved_publisher_ids or []),
            "publisher_source": publisher_source,
            "extracted_app_bundle_count": len(extracted_app_bundles),
            "app_bundle_source": app_bundle_source,
            "resolved_entities": prepared_deal["resolved_entities"],
            "invalid_domains": prepared_deal["invalid_domains"],
            "create_args_preview": prepared_deal["create_args"],
        }
    except Exception as e:
        error_msg = f"Failed to prepare OpenX deal from prompt inputs: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "error": error_msg,
        }


@mcp.tool()
async def ox_create_prepared_deal(prepared_deal_id: str) -> dict[str, Any]:
    """Create an OpenX deal from a previously prepared server-side artifact."""
    logger.info("ox_create_prepared_deal called with prepared_deal_id: %s", prepared_deal_id)

    prepared_deal = _prepared_openx_deals.get(prepared_deal_id)
    if prepared_deal is None:
        return {
            "success": False,
            "error": f"Prepared OpenX deal not found: {prepared_deal_id}",
            "quality_flags": [
                _make_ox_quality_flag(
                    "ox_prepared_deal_not_found",
                    f"Prepared OpenX deal not found: {prepared_deal_id}",
                    prepared_deal_id=prepared_deal_id,
                )
            ],
        }

    if not prepared_deal["ready_to_create"]:
        return {
            "success": False,
            "prepared_deal_id": prepared_deal_id,
            "error": "Prepared OpenX deal is blocked and cannot be created.",
            "blocking_issues": prepared_deal["blocking_issues"],
            "blockers": prepared_deal["blockers"],
            "quality_flags": list(prepared_deal.get("quality_flags", [])),
        }

    if prepared_deal.get("created"):
        # This artifact was already submitted and the deal exists on OpenX.
        # Re-running ox_create_deal would create a duplicate live deal, so
        # return the recorded outcome instead.
        recorded = prepared_deal.get("created_result")
        if isinstance(recorded, dict):
            replay = dict(recorded)
            replay["replayed"] = True
            return replay
        return {
            "success": False,
            "prepared_deal_id": prepared_deal_id,
            "error": "Prepared deal already submitted; refusing to create a duplicate.",
            "quality_flags": [
                _make_ox_quality_flag(
                    "ox_deal_already_created",
                    "This prepared deal was already submitted and the deal exists on OpenX. "
                    "Do NOT prepare/submit it again; use ox_list_deals to locate it.",
                    prepared_deal_id=prepared_deal_id,
                )
            ],
        }

    create_args = {key: value for key, value in prepared_deal["create_args"].items() if value is not None}
    result = await ox_create_deal(**create_args)
    quality_flags = list(prepared_deal.get("quality_flags", []))
    if result.get("success"):
        # Mark the artifact consumed so a retry cannot double-create.
        prepared_deal["created"] = True
        result["prepared_deal_id"] = prepared_deal_id
        result["warnings"] = prepared_deal["warnings"]
    else:
        quality_flags.append(
            _make_ox_quality_flag(
                "ox_create_call_failed",
                result.get("error") or "OpenX create_deal returned a non-success response.",
            )
        )
    result["quality_flags"] = quality_flags
    if prepared_deal.get("created"):
        prepared_deal["created_result"] = result
    return result


@mcp.tool()
async def ox_execute_deal_from_prompt_inputs(
    name: str,
    currency: str,
    deal_price: float,
    start_date: str,
    demand_partner: str,
    domain_file_path: str | None = None,
    targeting: dict[str, Any] | None = None,
    publisher_file_path: str | None = None,
    publisher_sheet: str | None = None,
    publisher_column: str | None = None,
    publisher_ids: list[str] | None = None,
    excluded_publisher_ids: list[str] | None = None,
    buyer_ids: list[str] | None = None,
    end_date: str | None = None,
    pmp_deal_type: str = "3",
    fee: dict[str, Any] | None = None,
    domain_sheet: str | None = None,
    domain_column: str | None = None,
    domain_targeting_option: str | None = None,
    domain_match_operator: Literal["blocklist", "allowlist"] | None = None,
    app_bundle_file_path: str | None = None,
    app_bundle_sheet: str | None = None,
    app_bundle_column: str | None = None,
    app_bundle_match_operator: Literal["blocklist", "allowlist"] | None = None,
    viewability_threshold: float | None = None,
    status: str | None = None,
    expected_ad_category: str | None = None,
    inventory_categories: list[str] | None = None,
    package_name: str | None = None,
) -> dict[str, Any]:
    """Prepare, create, and verify an OpenX deal in one call."""
    preparation = await ox_prepare_deal_from_prompt_inputs(
        name=name,
        currency=currency,
        deal_price=deal_price,
        start_date=start_date,
        package_name=package_name,
        demand_partner=demand_partner,
        domain_file_path=domain_file_path,
        targeting=targeting,
        publisher_file_path=publisher_file_path,
        publisher_sheet=publisher_sheet,
        publisher_column=publisher_column,
        publisher_ids=publisher_ids,
        excluded_publisher_ids=excluded_publisher_ids,
        buyer_ids=buyer_ids,
        end_date=end_date,
        pmp_deal_type=pmp_deal_type,
        fee=fee,
        domain_sheet=domain_sheet,
        domain_column=domain_column,
        domain_targeting_option=domain_targeting_option,
        domain_match_operator=domain_match_operator,
        app_bundle_file_path=app_bundle_file_path,
        app_bundle_sheet=app_bundle_sheet,
        app_bundle_column=app_bundle_column,
        app_bundle_match_operator=app_bundle_match_operator,
        viewability_threshold=viewability_threshold,
        status=status,
        expected_ad_category=expected_ad_category,
        inventory_categories=inventory_categories,
    )
    if not preparation.get("success") or not preparation.get("ready_to_create"):
        return {
            "success": False,
            "phase": "prepare",
            "preparation": preparation,
            "quality_flags": list(preparation.get("quality_flags", [])),
            "error": preparation.get("error") or "Preparation did not produce a creatable artifact.",
        }

    creation = await ox_create_prepared_deal(preparation["prepared_deal_id"])
    if not creation.get("success"):
        return {
            "success": False,
            "phase": "create",
            "preparation": preparation,
            "creation": creation,
            "quality_flags": list(preparation.get("quality_flags", [])) + list(creation.get("quality_flags", [])),
            "error": creation.get("error") or "Creation failed.",
        }

    verification = {"success": False, "error": "Verification did not run."}
    verification_attempts: list[dict[str, str]] = []
    verification_candidates = [
        ("id", creation["deal"].get("id")),
        ("deal_id", creation["deal"].get("deal_id")),
    ]
    seen_verification_ids: set[str] = set()
    for identifier_type, identifier_value in verification_candidates:
        identifier_str = str(identifier_value or "").strip()
        if not identifier_str or identifier_str in seen_verification_ids:
            continue
        seen_verification_ids.add(identifier_str)
        verification_attempts.append({"identifier_type": identifier_type, "identifier_value": identifier_str})
        verification = await ox_get_deal(identifier_str)
        if verification.get("success"):
            break

    deal_url = creation.get("deal_url") or verification.get("deal_url")
    combined_quality_flags = list(preparation.get("quality_flags", [])) + list(creation.get("quality_flags", []))
    if not verification.get("success"):
        combined_quality_flags.append(
            _make_ox_quality_flag(
                "ox_verification_failed",
                verification.get("error") or "OpenX verification re-fetch failed.",
            )
        )
    return {
        "success": creation.get("success", False),
        "phase": "verify",
        "verification_success": verification.get("success", False),
        "deal_url": deal_url,
        "warnings": []
        if verification.get("success")
        else ["Deal creation succeeded, but post-create verification failed."],
        "quality_flags": combined_quality_flags,
        "preparation": preparation,
        "creation": creation,
        "verification": verification,
        "verification_attempts": verification_attempts,
    }


@mcp.tool()
async def ox_list_demand_partners() -> dict[str, Any]:
    """
    Query and return available demand partners from OpenX.

    Use this to discover valid demand partner IDs before creating a deal.

    Returns:
        Dictionary containing:
            - success: Boolean indicating if the query succeeded
            - demand_partners: List of demand partner objects
            - error: Error message (if failed)
    """
    logger.info("list_demand_partners called")

    try:
        client = get_openx_client()
        partners = await client.list_demand_partners()

        logger.info(f"Found {len(partners)} demand partners")
        return {
            "success": True,
            "demand_partners": partners,
        }

    except Exception as e:
        error_msg = f"Failed to list demand partners: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "error": error_msg,
        }


@mcp.tool()
async def ox_list_options_by_path(path: str, filter: dict[str, Any] | None = None) -> dict[str, Any]:
    """
    Query OpenX optionsByPath for allowed values at a schema path.

    This is a debug/discovery helper. Prefer the high-level preparation tools for normal
    agent workflows instead of manually resolving ids from raw schema paths.
    """
    logger.info("ox_list_options_by_path called with path: %s", path)

    try:
        client = get_openx_client()
        options = await client.list_options_by_path(path, filter)
        return {
            "success": True,
            "path": path,
            "filter": filter,
            "options": options,
        }
    except Exception as e:
        error_msg = f"Failed to list options by path: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "path": path,
            "filter": filter,
            "error": error_msg,
        }


@mcp.tool()
async def ox_list_iab_categories(query: str | None = None) -> dict[str, Any]:
    """
    List OpenX IAB category options for generic category discovery.

    Prefer this over hardcoded campaign-specific mappings. It queries the live OpenX
    `deal.package.targeting.domain.categories_iab_v2` path and optionally filters the
    results by normalized id/name substring.

    Args:
        query: Optional search string such as an IAB code fragment or category name

    Returns:
        Dictionary containing:
            - success: Boolean indicating if the query succeeded
            - query: Search query that was applied
            - options: Matching option objects from OpenX
            - total_options: Count of returned options after filtering
            - error: Error message (if failed)
    """
    logger.info("ox_list_iab_categories called with query: %s", query)

    try:
        client = get_openx_client()
        options = await client.list_options_by_path(OPENX_OPTIONS_PATHS["iab_categories"])
        filtered_options = _filter_options_by_query(options, query)
        return {
            "success": True,
            "query": query,
            "options": filtered_options,
            "total_options": len(filtered_options),
        }
    except Exception as e:
        error_msg = f"Failed to list OpenX IAB categories: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "query": query,
            "error": error_msg,
        }


@mcp.tool()
async def ox_translate_iab_categories(values: list[str]) -> dict[str, Any]:
    """
    Resolve IAB category names or IDs against live OpenX options.

    Queries optionsByPath(path: "deal.package.targeting.domain.categories_iab_v2") and
    attempts to match each input value by exact ID, exact name, or unique contains match.

    Legacy IAB codes like "IAB2-5" will not resolve because OpenX uses its own IAB v2
    naming. Use ox_list_iab_categories to discover the correct OpenX names first.

    Args:
        values: OpenX category names or IDs like ["Certified Pre-Owned Cars", "Auto Parts"]

    Returns:
        Dictionary containing:
            - success: Boolean
            - translations: List of {input, resolved_name, resolved_id}
            - unresolved: List of values that could not be matched
            - legacy_codes: List of detected legacy IAB codes that need translation
    """
    logger.info("ox_translate_iab_categories called with values: %s", values)

    try:
        client = get_openx_client()
        options = await client.list_options_by_path(OPENX_OPTIONS_PATHS["iab_categories"])

        translations: list[dict[str, Any]] = []
        unresolved: list[str] = []
        legacy_codes = _detect_legacy_iab_codes(values)

        for raw_value in values:
            if _LEGACY_IAB_CODE_PATTERN.match(raw_value.strip()):
                unresolved.append(raw_value)
                continue
            try:
                matched_option = _choose_unique_option(options, raw_value, "IAB category", allow_contains_match=True)
                translations.append(
                    {
                        "input": raw_value,
                        "resolved_name": matched_option.get("name"),
                        "resolved_id": str(matched_option["id"]),
                    }
                )
            except (LookupError, ValueError):
                unresolved.append(raw_value)

        return {
            "success": True,
            "translations": translations,
            "unresolved": unresolved,
            "legacy_codes": legacy_codes,
        }
    except Exception as e:
        error_msg = f"Failed to translate IAB categories: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "error": error_msg,
        }


@mcp.tool()
async def ox_validate_audience_geo_compatibility(
    audience: str,
    geo: dict[str, Any] | list[str] | str | None = None,
) -> dict[str, Any]:
    """
    Check whether an audience segment is compatible with the requested geographic targeting.

    Audience targeting requires an explicit geographic.includes.country. This tool
    resolves the audience and checks the geo configuration before a live create.

    Args:
        audience: Audience segment name or id
        geo: Geographic targeting as dict, country list, or single country string

    Returns:
        Dictionary containing:
            - success: Boolean
            - compatible: Whether the audience + geo combination should work
            - audience_resolved: The resolved audience id (if found)
            - geo_has_explicit_country: Whether the geo includes an explicit country
            - issues: List of compatibility issues
    """
    logger.info("ox_validate_audience_geo_compatibility called with audience: %s", audience)

    try:
        client = get_openx_client()
        issues: list[str] = []
        audience_resolved: str | None = None
        audience_option: dict[str, Any] | None = None

        try:
            audience_option, audience_warnings = await _resolve_option(
                client,
                path=OPENX_OPTIONS_PATHS["audience"],
                value=audience,
                field_label="audience segment",
                allow_contains_match=True,
            )
            audience_resolved = _normalize_openaudience_id(audience_option["id"])
            issues.extend(audience_warnings)
        except (LookupError, ValueError) as exc:
            issues.append(f"Audience resolution failed: {exc}")

        normalized_geo, _ = _normalize_geographic_targeting(geo or {})
        geo_has_explicit_country = _geographic_includes_explicit_country(normalized_geo)
        geo_is_us_only = _geographic_includes_only_us(normalized_geo)

        if audience_resolved and not geo_has_explicit_country:
            issues.append("Audience targeting requires an explicit geographic.includes.country.")
        if audience_resolved and _audience_export_type_requires_us(audience_option) and not geo_is_us_only:
            issues.append("This OpenX audience export type can target audiences only in US geography.")

        return {
            "success": True,
            "compatible": audience_resolved is not None
            and geo_has_explicit_country
            and (not _audience_export_type_requires_us(audience_option) or geo_is_us_only),
            "audience_resolved": audience_resolved,
            "geo_has_explicit_country": geo_has_explicit_country,
            "issues": issues,
        }
    except Exception as e:
        error_msg = f"Failed to validate audience-geo compatibility: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "error": error_msg,
        }


@mcp.tool()
async def ox_list_fee_partners(query: str | None = None) -> dict[str, Any]:
    """
    List OpenX fee partner options.

    This is an agent-friendly discovery tool for fee partner resolution, so callers do not
    need to use raw optionsByPath directly.
    """
    logger.info("ox_list_fee_partners called with query: %s", query)

    try:
        client = get_openx_client()
        options = await client.list_options_by_path(OPENX_OPTIONS_PATHS["fee_partner"])
        filtered_options = _filter_options_by_query(options, query)
        return {
            "success": True,
            "query": query,
            "options": filtered_options,
            "total_options": len(filtered_options),
        }
    except Exception as e:
        error_msg = f"Failed to list OpenX fee partners: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "query": query,
            "error": error_msg,
        }


@mcp.tool()
async def ox_list_audience_segments(query: str | None = None) -> dict[str, Any]:
    """
    List OpenX audience segment options.

    This is an agent-friendly discovery tool for audience resolution, so callers do not
    need to use raw optionsByPath directly.
    """
    logger.info("ox_list_audience_segments called with query: %s", query)

    try:
        client = get_openx_client()
        options = await client.list_options_by_path(OPENX_OPTIONS_PATHS["audience"])
        filtered_options = _filter_options_by_query(options, query)
        return {
            "success": True,
            "query": query,
            "options": filtered_options,
            "total_options": len(filtered_options),
        }
    except Exception as e:
        error_msg = f"Failed to list OpenX audience segments: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "query": query,
            "error": error_msg,
        }


@mcp.tool()
async def ox_list_states(query: str | None = None, country: str | None = None) -> dict[str, Any]:
    """
    List OpenX state/province targeting options.

    This tool resolves subnational targeting using verified OpenX ids for known US states
    and falls back to live OpenX optionsByPath lookup when needed. Pass country="ca"
    to discover Canadian provinces on the same OpenX geographic.state path.
    """
    logger.info("ox_list_states called with query: %s country: %s", query, country)

    query_text = str(query or "").strip()
    if not query_text:
        return {
            "success": False,
            "query": query,
            "country": country,
            "error": "State discovery requires a non-empty query (for example 'CA' or 'California').",
        }

    try:
        country_hint = _normalize_country_hint(country) or "united states"
        normalized_state, full_name = _normalize_geographic_lookup_input("state", query_text, country_hint)
        verified_state_id = (
            _get_verified_geographic_id("state", normalized_state) if country_hint == "united states" else None
        )
        if verified_state_id:
            options = [
                {
                    "id": verified_state_id,
                    "name": full_name,
                    "path": OPENX_OPTIONS_PATHS["state"],
                    "extra": {
                        "abbreviation": normalized_state,
                        "state": full_name,
                        "country": country_hint,
                        "type": "state",
                        "type_id": f"state-{verified_state_id}",
                    },
                }
            ]
            source = "verified_table"
        else:
            client = get_openx_client()
            options_filter = _build_geographic_options_filter(
                "state",
                full_name,
                country_hint=country_hint,
                include_country_filter=country_hint != "united states",
            )
            options = await client.list_geographic_options("state", options_filter)
            options = [
                option
                for option in options
                if str((option.get("extra") or {}).get("country", "")).lower() == country_hint
            ]
            source = "optionsByPath"
        filtered_options = _filter_options_by_query(options, query)
        return {
            "success": True,
            "query": query,
            "country": country,
            "source": source,
            "options": filtered_options,
            "total_options": len(filtered_options),
        }
    except Exception as e:
        error_msg = f"Failed to list OpenX states: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "query": query,
            "country": country,
            "error": error_msg,
        }


@mcp.tool()
async def ox_list_deals(limit: int = 10, offset: int = 0) -> dict[str, Any]:
    """
    Query and return a list of existing deals from OpenX.

    Use this to review existing deals and their configurations.

    Args:
        limit: Maximum number of deals to return (default: 10)
        offset: Number of deals to skip for pagination (default: 0)

    Returns:
        Dictionary containing:
            - success: Boolean indicating if the query succeeded
            - deals: List of deal objects with basic information
            - error: Error message (if failed)
    """
    logger.info(f"list_deals called (limit={limit}, offset={offset})")

    try:
        client = get_openx_client()
        deals = await client.list_deals(limit=limit, offset=offset)

        logger.info(f"Found {len(deals)} deals")
        return {
            "success": True,
            "deals": deals,
        }

    except Exception as e:
        error_msg = f"Failed to list deals: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "error": error_msg,
        }


@mcp.tool()
async def ox_get_deal(deal_id: str, full_targeting: bool = False) -> dict[str, Any]:
    """
    Fetch full details for a specific deal from OpenX.

    Use this to verify deal configuration or inspect an existing deal.

    Args:
        deal_id: The internal deal id (the numeric dealById id returned at
            creation — NOT the public "OX-..." deal_id string).
        full_targeting: Set True BEFORE a targeting update. Requests every
            documented targeting branch (content, custom, domain, geographic,
            rendering_context, technographic, video, viewability, vtr) so the
            returned package.targeting can be edited and resubmitted to
            ox_update_deal without nulling branches. The deal carries
            _targeting_selection: "full" on success; if the schema rejects the
            full selection it falls back to the legacy subset and marks
            _targeting_selection: "legacy" — in that state a targeting
            resubmit is NOT safe (unfetched branches would be wiped).

    Returns:
        Dictionary containing:
            - success: Boolean indicating if the query succeeded
            - deal: Complete deal object with all fields
            - error: Error message (if failed)
    """
    logger.info(f"get_deal called with deal_id: {deal_id}")

    try:
        client = get_openx_client()
        deal = await client.get_deal(deal_id, full_targeting=full_targeting)

        if not deal:
            return {
                "success": False,
                "error": f"Deal not found: {deal_id}",
            }

        logger.info(f"Retrieved deal: {deal.get('name')}")
        return {
            "success": True,
            "deal": deal,
            "deal_url": _build_openx_deal_url(deal),
        }

    except Exception as e:
        error_msg = f"Failed to get deal: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "error": error_msg,
        }


@mcp.tool()
async def ox_update_deal(
    deal_id: str,
    name: str | None = None,
    deal_price: float | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    status: str | None = None,
    currency: str | None = None,
    pmp_deal_type: str | None = None,
    deal_participants: list[dict[str, Any]] | None = None,
    expected_ad_category: str | None = None,
    third_party_fees_config: dict[str, Any] | list[dict[str, Any]] | None = None,
    url_targeting: dict[str, Any] | None = None,
    package: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """
    Update an existing OpenX deal via the dealUpdate mutation.

    Partial update at the TOP LEVEL of `deal` and `deal.package` only — pass
    just the fields you want changed and OpenX keeps the rest. Use
    status="Paused"/"Active" to pause or resume a deal.

    NESTED OBJECTS ARE FULL-REPLACEMENT: when updating anything inside
    package.targeting, the ENTIRE targeting object must be resubmitted —
    the API sets every unspecified nested field to null. Workflow:
      1. ox_get_deal(deal_id, full_targeting=True)
      2. confirm the returned deal has _targeting_selection == "full"
         (if it says "legacy", STOP — a resubmit would wipe unfetched branches)
      3. edit the returned package.targeting
      4. pass the COMPLETE edited object here as package={"targeting": {...}}

    Propagation: status changes apply in near-realtime; every other change
    can take 1-2 hours to take effect in delivery (per the OpenXSelect API
    guide) — the result's `warnings` reminds you when that applies.

    Args:
        deal_id: The internal deal id (numeric dealById id — NOT the public
            "OX-..." deal_id string). ox_list_deals resolves names to ids.
        name: New deal name.
        deal_price: New CPM price (sent as a 2-decimal string, like create).
        start_date: New start date (ISO 8601, e.g. "2026-07-01 00:00:00").
        end_date: New end date (ISO 8601).
        status: "Active" or "Paused".
        currency: Currency code (e.g. "USD").
        pmp_deal_type: "1"/"2"/"3" or PROGRAMMATIC_GUARANTEED /
            PRIVATE_AUCTION / PREFERRED_DEAL (mapped like create).
        deal_participants: Replacement participants list. Each entry uses
            "demand_partner" (demand_partner_id is auto-renamed). NOTE: this
            replaces the whole list — include every participant the deal
            should keep.
        expected_ad_category: OpenX Expected Sensitive Category (e.g.
            "Politics").
        third_party_fees_config: Fee config dict (or list), same shape as
            create: {partner_id, revenue_method, gross_share, ...}.
        url_targeting: package-level URL targeting. Convenience keys
            allowlist/blocklist convert to type=whitelist/blacklist + urls,
            same as create. This is a top-level package field, so partial
            update is safe — but it replaces the deal's whole url_targeting.
        package: Raw PackageUpdateParams passthrough for package-level edits
            the dedicated args don't cover — most importantly
            {"targeting": <COMPLETE targeting object>} per the workflow
            above. Mutually exclusive with url_targeting when it carries its
            own url_targeting key.

    Returns:
        {"success": True, "deal": <updated deal>, "deal_url": ...,
         "updated_fields": [...], "warnings": [...]}
        or {"success": False, "error": ...}.
    """
    logger.info(f"update_deal called for deal_id: {deal_id}")

    try:
        update_params: dict[str, Any] = {}
        warnings: list[str] = []

        if name is not None:
            update_params["name"] = name
        if deal_price is not None:
            if deal_price <= 0:
                return {"success": False, "error": f"deal_price must be positive, got {deal_price}"}
            update_params["deal_price"] = f"{deal_price:.2f}"
        if start_date is not None:
            update_params["start_date"] = start_date
        if end_date is not None:
            update_params["end_date"] = end_date
        if status is not None:
            if status not in ("Active", "Paused"):
                return {"success": False, "error": f"status must be 'Active' or 'Paused', got {status!r}"}
            update_params["status"] = status
        if currency is not None:
            update_params["currency"] = currency
        if pmp_deal_type is not None:
            update_params["pmp_deal_type"] = PMP_DEAL_TYPE_MAP.get(pmp_deal_type.upper(), pmp_deal_type)
        if deal_participants is not None:
            normalized_participants = []
            for participant in deal_participants:
                normalized = dict(participant)
                if "demand_partner_id" in normalized:
                    normalized["demand_partner"] = normalized.pop("demand_partner_id")
                normalized_participants.append(normalized)
            update_params["deal_participants"] = normalized_participants
            warnings.append("deal_participants replaces the deal's full participant list.")
        if expected_ad_category is not None:
            update_params["expected_ad_category"] = expected_ad_category
        if third_party_fees_config is not None:
            update_params["third_party_fees_config"] = _normalize_third_party_fees_config(third_party_fees_config)

        package_params: dict[str, Any] = dict(package) if package else {}
        if url_targeting is not None:
            if "url_targeting" in package_params:
                return {
                    "success": False,
                    "error": "Pass url_targeting either as the dedicated argument or inside package, not both.",
                }
            normalized_url_targeting = dict(url_targeting)
            if "allowlist" in normalized_url_targeting:
                normalized_url_targeting["type"] = "whitelist"
                normalized_url_targeting["urls"] = normalized_url_targeting.pop("allowlist")
            elif "blocklist" in normalized_url_targeting:
                normalized_url_targeting["type"] = "blacklist"
                normalized_url_targeting["urls"] = normalized_url_targeting.pop("blocklist")
            package_params["url_targeting"] = normalized_url_targeting
            warnings.append("url_targeting replaces the deal's entire URL allow/block list.")
        if package_params:
            if "targeting" in package_params:
                targeting = package_params["targeting"]
                if not isinstance(targeting, dict) or not targeting:
                    return {
                        "success": False,
                        "error": "package.targeting must be the COMPLETE non-empty targeting object — "
                        "nested updates are full-replacement (unspecified fields become null). "
                        "Read it via ox_get_deal(deal_id, full_targeting=True), edit, and resubmit.",
                    }
                warnings.append(
                    "package.targeting was replaced wholesale — verify every branch you meant to keep "
                    "was included (read-modify-write from ox_get_deal full_targeting=True)."
                )
            update_params["package"] = package_params

        if not update_params:
            return {"success": False, "error": "No update fields provided — pass at least one field to change."}

        non_status_fields = sorted(set(update_params.keys()) - {"status"})
        if non_status_fields:
            warnings.append(
                f"Non-status changes ({', '.join(non_status_fields)}) can take 1-2 hours to take effect "
                "in delivery; status changes are near-realtime."
            )

        client = get_openx_client()
        deal = await client.update_deal(deal_id, update_params)

        logger.info(f"Deal updated successfully: {deal.get('deal_id')}")
        return {
            "success": True,
            "deal": deal,
            "deal_url": _build_openx_deal_url(deal),
            "updated_fields": sorted(update_params.keys()),
            "warnings": warnings,
        }

    except Exception as e:
        error_msg = f"Failed to update deal: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "error": error_msg,
        }


@mcp.tool()
async def ox_validate_domains(domains: list[str]) -> dict[str, Any]:
    """
    Validate a list of domain strings for correct formatting.

    This is a local validation tool (no API call) that mimics the UI's
    validation step before submitting domain allowlists or blocklists.

    Valid domain format:
    - Must contain at least one dot
    - Must not contain protocols (http://, https://)
    - Must not contain paths (/)
    - Must use valid characters (alphanumeric, hyphens, dots)
    - Each label (part between dots) must start and end with alphanumeric

    Examples of valid domains:
    - example.com
    - sub.domain.example.com
    - my-site.co.uk

    Examples of invalid domains:
    - https://example.com (contains protocol)
    - example.com/path (contains path)
    - -invalid.com (starts with hyphen)
    - .com (missing domain name)

    Args:
        domains: List of domain strings to validate

    Returns:
        Dictionary containing:
            - valid: List of valid domain strings
            - invalid: List of dictionaries with domain and reason for invalidity
            - summary: Human-readable summary of validation results
    """
    logger.info(f"validate_domains called with {len(domains)} domains")

    valid_domains: list[str] = []
    invalid_domains: list[dict[str, str]] = []

    for domain in domains:
        domain = domain.strip().lower()

        if not domain:
            invalid_domains.append({"domain": domain, "reason": "Empty domain"})
            continue

        # Check for protocol
        if domain.startswith(("http://", "https://", "//")):
            invalid_domains.append({"domain": domain, "reason": "Contains protocol (remove http:// or https://)"})
            continue

        # Check for path
        if "/" in domain:
            invalid_domains.append({"domain": domain, "reason": "Contains path (remove everything after /)"})
            continue

        # Check for query string
        if "?" in domain:
            invalid_domains.append(
                {"domain": domain, "reason": "Contains query string (remove ? and everything after)"}
            )
            continue

        # Check for port
        if ":" in domain:
            invalid_domains.append({"domain": domain, "reason": "Contains port number (remove :port)"})
            continue

        # Check for whitespace
        if " " in domain:
            invalid_domains.append({"domain": domain, "reason": "Contains whitespace"})
            continue

        # Check against accepted url_targeting shapes (web FQDN, Android-style
        # reverse-DNS app bundle, or numeric Apple App Store id).
        if not _is_acceptable_url_targeting_value(domain):
            invalid_domains.append({"domain": domain, "reason": "Invalid domain or app-bundle format"})
            continue

        valid_domains.append(domain)

    # Generate summary
    total = len(domains)
    valid_count = len(valid_domains)
    invalid_count = len(invalid_domains)

    if invalid_count == 0:
        summary = f"All {total} domains are valid."
    elif valid_count == 0:
        summary = f"All {total} domains are invalid. Please review and correct them."
    else:
        summary = (
            f"{valid_count} of {total} domains are valid. {invalid_count} domains have issues that need correction."
        )

    logger.info(summary)

    return {
        "valid": valid_domains,
        "invalid": invalid_domains,
        "summary": summary,
    }


@mcp.tool()
async def ox_introspect_type(type_name: str) -> dict[str, Any]:
    """
    Run a GraphQL introspection query to discover fields of any named type.

    This is a debug/discovery helper. Prefer ox_prepare_deal_from_prompt_inputs,
    ox_prepare_deal_from_brief, or ox_execute_deal_from_prompt_inputs for normal
    deal workflows. Use introspection when the high-level preparation flow returns
    unresolved mapping blockers or when explicit schema evidence is needed.

    Use this tool to discover the exact field names for GraphQL input types
    before constructing payloads. This helps prevent schema mismatches.

    Common types to introspect:
    - TargetingCreateParams: Fields for targeting configuration
    - TargetingGeographicItemCreateParams: Fields for geographic targeting items
    - DealCreateParams: Fields for deal creation
    - PackageCreateParams: Fields for package configuration

    Args:
        type_name: The name of the GraphQL type to introspect

    Returns:
        Dictionary containing:
            - success: Boolean indicating if introspection succeeded
            - type_info: The introspected type with name, kind, and inputFields
            - error: Error message (if failed)
    """
    logger.info(f"ox_introspect_type called with type_name: {type_name}")

    try:
        client = get_openx_client()
        type_info = await client.introspect_type(type_name)

        if not type_info:
            return {
                "success": False,
                "error": f"Type not found: {type_name}",
            }

        logger.info(f"Successfully introspected type: {type_name}")
        return {
            "success": True,
            "type_info": type_info,
            "mapping_hints": _mapping_hints_for_type(type_name, type_info),
        }

    except Exception as e:
        error_msg = f"Failed to introspect type: {str(e)}"
        logger.error(error_msg)
        return {
            "success": False,
            "error": error_msg,
        }


# =============================================================================
# Main Entry Point
# =============================================================================

if __name__ == "__main__":
    logger.info("Starting OpenX MCP Server")

    # Check for OpenX credentials
    has_openx = bool(os.environ.get("OPENX_API_KEY"))

    if not has_openx:
        logger.warning("OpenX API key not configured. Set OPENX_API_KEY to enable OpenX deal creation.")

    try:
        # Use stdio transport (default for FastMCP)
        mcp.run(transport="stdio")
    except Exception as e:
        logger.error(f"Failed to start server: {e}")
        sys.exit(1)
